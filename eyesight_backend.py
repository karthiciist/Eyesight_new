import shutil

from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
import pytesseract
from tesserocr import PyTessBaseAPI
import cv2
import numpy as np
from PIL import Image, ImageChops, ImageDraw, ImageGrab
import io
import tesserocr
import base64
import re
from imageio import imread
import pyodbc
import json
from imutils.object_detection import non_max_suppression
from difflib import SequenceMatcher
import time
import os
import camelot
import matplotlib.pyplot as plt
import pandas as pd
from pdf2image import convert_from_path
from pathlib import Path
from typing import Union
# from camelot import utils
import zipfile
from flask import render_template
import pdfquery
import csv
from os import listdir
from os.path import isfile, join
from threading import Thread
# from pdf2image import convert_from_path
import pdf2image
import PyPDF2
from pdfminer.pdfpage import PDFPage
import img2pdf
from shutil import copyfile
import xml.etree.ElementTree as ET
from pdfquery.cache import FileCache
from lxml import etree
import string
import fitz
import math
from datetime import date, datetime
import xlsxwriter
from openpyxl import load_workbook
import imutils
from skimage.metrics import structural_similarity as ssim
import requests
from pyzbar import pyzbar

app = Flask(__name__)

CORS(app)
local_save_input_file = 'reconstructed.jpg'

global local_save_input_test_file
global local_save_input_test_pdf

global local_save_input_initial_pdf

local_save_cropped_file = 'croppedpic.jpg'
language = 'eng'

headerfound = 0
leftindexfound = 0

# root_path = 'C:/Users/salil/Desktop/eyesight V1/eyesight V1/'
root_path = 'C:/Users/ET437GL/OneDrive - EY/Documents/EYESIGHT/Current_Version_Sourcecode/eyesight/'

# databaseserver = 'Driver={ODBC Driver 17 for SQL Server};''Server=10.0.10.72,8033;''Database=eyesight;''Trusted_Connection=no;''UID=aspaceadmin;''PWD=eyaidb@4321;'
# databaseserverpdf = 'Driver={ODBC Driver 17 for SQL Server};''Server=10.0.10.72,8033;''Database=eyesightpdfs;''Trusted_Connection=no;''UID=aspaceadmin;''PWD=eyaidb@4321;'

# database_server_ocr_solutions = 'Driver={ODBC Driver 17 for SQL Server};''Server=10.0.10.72,8033;''Database=eyesight_ocr_solutions;''Trusted_Connection=no;''UID=aspaceadmin;''PWD=eyaidb@4321;'
# database_server_batches = 'Driver={ODBC Driver 17 for SQL Server};''Server=10.0.10.72,8033;''Database=eyesight_batches;''Trusted_Connection=no;''UID=aspaceadmin;''PWD=eyaidb@4321;'
# database_server_ccass = 'Driver={ODBC Driver 17 for SQL Server};''Server=10.0.10.72,8033;''Database=eyesight_ccass;''Trusted_Connection=no;''UID=aspaceadmin;''PWD=eyaidb@4321;'

#pytesseract.pytesseract.tesseract_cmd = 'C:/Users/ET437GL/AppData/Local/Tesseract-OCR/tesseract.exe'
# pytesseract.pytesseract.tesseract_cmd = 'C:/Users/ET437GL/AppData/Local/Programs/Tesseract-OCR'

@app.route('/status', methods=['GET'])
def healthz():
    return sendresponse("EYESIGHT running...", 200)


# Frontend pages serving

@app.route("/")
def loginpage():
    return render_template('login.html')

@app.route('/dashboardpage', methods=['GET', 'POST'])
def dashboardpage():
    return render_template('dashboard.html')

@app.route('/consumepage', methods=['GET', 'POST'])
def consumepage():
    return render_template('consume.html')

@app.route('/createmodelwithmodalpage', methods=['GET', 'POST'])
def createmodelwithmodalpage():
    return render_template('createmodelwithmodal.html')

@app.route('/createmodelwithmodalkeypage', methods=['GET', 'POST'])
def createmodelwithmodalkeypage():
    return render_template('createmodelwithmodalkey.html')

@app.route('/createmodelwithmodalsimplepage', methods=['GET', 'POST'])
def createmodelwithmodalsimplepage():
    return render_template('createmodelwithmodalsimple.html')

@app.route('/createmodelwithmodalzonalpage', methods=['GET', 'POST'])
def createmodelwithmodalzonalpage():
    return render_template('createmodelwithmodalzonal.html')

@app.route('/helpcenterpage', methods=['GET', 'POST'])
def helpcenterpage():
    return render_template('helpcenter.html')

@app.route('/modelstoretrainpage', methods=['GET', 'POST'])
def modelstoretrainpage():
    return render_template('modelstoretrain.html')

@app.route('/ocrsolutionspage', methods=['GET', 'POST'])
def ocrsolutionspage():
    return render_template('ocrsolutions.html')

@app.route('/createocrsolutionpage', methods=['GET', 'POST'])
def createocrsolutionpage():
    return render_template('createocrsolution.html')

@app.route('/pdfconsumepage', methods=['GET', 'POST'])
def pdfconsumepage():
    return render_template('pdfconsume.html')

@app.route('/pdfoperationspage', methods=['GET', 'POST'])
def pdfoperationspage():
    return render_template('pdfoperations.html')

@app.route('/retrainwithmodalpage', methods=['GET', 'POST'])
def retrainwithmodalpage():
    return render_template('retrainwithmodal.html')

@app.route('/trainedmodelspage', methods=['GET', 'POST'])
def trainedmodelspage():
    return render_template('trainedmodels.html')

@app.route('/trainedpdfmodelspage', methods=['GET', 'POST'])
def trainedpdfmodelspage():
    return render_template('trainedpdfmodels.html')

@app.route('/trainedclassifierspage', methods=['GET', 'POST'])
def trainedclassifierspage():
    return render_template('trainedclassifiers.html')

@app.route('/createclassifierpage', methods=['GET', 'POST'])
def createclassifierpage():
    return render_template('createclassifier.html')

@app.route('/classifiertestpage', methods=['GET', 'POST'])
def classifiertestpage():
    return render_template('classifiertest.html')

@app.route('/modelspage', methods=['GET', 'POST'])
def modelspage():
    return render_template('models.html')

@app.route('/custommodelspage', methods=['GET', 'POST'])
def custommodelspage():
    return render_template('custommodels.html')

@app.route('/batchespage', methods=['GET', 'POST'])
def batchespage():
    return render_template('batches.html')

@app.route('/createmodelpage', methods=['GET', 'POST'])
def createmodelpage():
    return render_template('createmodel.html')





# Backend logic starts here

@app.route('/extracttext', methods=['POST'])
def extracttext():
    try:
        data = request.get_json()
        if data is None:
            print("No valid request body, json missing!")
            return jsonify({'error': 'No valid request body, json missing!'})
        else:
            img_data = data['thumbnail']
            corrstring = img_data + "=="
            corrstring1 = re.sub(r'.*,', ',', corrstring)
            corrstring2 = corrstring1.replace(",", "")

            stringToRGB(corrstring2, local_save_input_file)

        image1 = Image.open(local_save_input_file)
        resultstring = tesserocr.image_to_text(image1, lang=language)  # print ocr text from image
        resultstring = resultstring.strip()
        print(resultstring)
        return sendresponse(resultstring, 200)
    except Exception as e:
        resultstring = 'failed to get text!!!' + str(e)
        print(resultstring)
        return sendresponse(resultstring, 201)


# @app.route('/createmodel', methods = ['POST'])
# def createmodel():
#     modelname = request.form['modelname']
#     processedimg = request.form['processedimg']
#     processedimgwithboxes = request.form['processedimgwithboxes']
#
#     path = root_path + 'static/imagesusedfortraining/' + modelname
#     os.mkdir(path)
#
#     local_save_input_test_image_file_1 = 'static/imagesusedfortraining/' + modelname + '/withoutboxes.jpg'
#     local_save_input_test_image_file_2 = 'static/imagesusedfortraining/' + modelname + '/withboxes.jpg'
#
#     imgdata1 = base64.b64decode(processedimg)
#     with open(local_save_input_test_image_file_1, 'wb') as f:
#         f.write(imgdata1)
#
#     imgdata2 = base64.b64decode(processedimgwithboxes)
#     with open(local_save_input_test_image_file_2, 'wb') as f:
#         f.write(imgdata2)
#
#
#     if modelname == "":
#         return sendresponse("Model name missing", 201)
#     else:
#         conn = pyodbc.connect(databaseserver)
#         cursor = conn.cursor()
#         query = "CREATE TABLE " + modelname + "(sno int identity primary key, xaxis float, yaxis float, width float, " \
#                                               "height float, rotate float, keywidth varchar(255), keyheight varchar(255), " \
#                                               "fullwidth float, fullheight float, label varchar(255), text varchar(" \
#                                               "255), header varchar(255), hposition varchar(255), htol varchar(255), " \
#                                               "footer varchar(255), fposition varchar(255), ftol varchar(255), " \
#                                               "isthre varchar(255), colorcut varchar(255), threshold varchar(255), textnature " \
#                                               "varchar(255), language varchar(255), iskey varchar(255), " \
#                                               "kposition varchar(255), ktol varchar(255), keyval varchar(255), " \
#                                               "leftindex varchar(255), liposition varchar(255), litol varchar(255), kori varchar(255), koriper varchar(255), regex varchar(255)); "
#         print (query)
#         cursor.execute(query)
#         conn.commit()
#         return sendresponse("Model created", 200)
#

# @app.route('/trainmodel', methods = ['POST'])
# def trainmodel():
#     if request.is_json:
#         try:
#             # print (request.headers["modelname"])
#             modelname = request.headers["modelname"]
#             # fullheight = request.headers["fullheight"]
#             # fullwidth = request.headers["fullwidth"]
#             # noofparam = request.headers["noofparam"]
#             # header = request.headers["header"]
#             # footer = request.headers["footer"]
#             # leftindex = request.headers["leftindex"]
#             # defthre = request.headers["defthre"]
#             # defcolcut = request.headers["defcolcut"]
#             # defeyerangex = request.headers["defeyerangex"]
#             # defeyerangey = request.headers["defeyerangey"]
#             # ocrtype = request.headers["ocrtype"]
#
#             # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#             #                       'Server=IN2371790W1\SQLEXPRESS;'
#             #                       'Database=eyesight;'
#             #                       'Trusted_Connection=yes;')
#             conn = pyodbc.connect(databaseserver)
#             cursor = conn.cursor()
#             # modelquery = 'INSERT INTO eyesight.dbo.' + 'models' + '(model, ocr, fullheight, fullwidth, noofparam, header, footer, ' \
#             #                                                       'leftindex, defthre, defcolcut, defeyerangex, defeyerangey) VALUES (' + ' \'' + modelname + '\' , \'' + ocrtype + '\'' + ',\'' + fullheight + '\',\'' + fullwidth + '\',\'' + noofparam + '\',\'' + header + '\',\'' + footer + '\',\'' + leftindex + '\',\'' + defthre + '\',\'' + defcolcut + '\',\'' + defeyerangex + '\',\'' + defeyerangey + '\',\'' + ');'
#
#             # print(modelquery)
#             # cursor.execute(modelquery)
#
#             content = request.get_json()
#             jtopy = json.dumps(content)
#             json_dictionary = json.loads(jtopy)
#             for key in json_dictionary:
#                 indi = json_dictionary[key]
#                 xaxisl = indi["xaxis"]
#                 yaxisl = indi["yaxis"]
#                 widthl = indi["width"]
#                 heightl = indi["height"]
#                 rotatel = indi["rotate"]
#                 keywidth = indi["keywidth"]
#                 keyheight = indi["keyheight"]
#                 fullwidthl = indi["fullwidth"]
#                 fullheightl = indi["fullheight"]
#                 labell = indi["lable"]
#                 textl = indi["text"]
#                 headerl = indi["header"]
#                 hpositionl = indi["hposition"]
#                 htoll = indi["htol"]
#                 footerl = indi["footer"]
#                 fpositionl = indi["fposition"]
#                 ftoll = indi["ftol"]
#                 isthre = indi["isthre"]
#                 colorcut = indi["colorcut"]
#                 threshold = indi["threshold"]
#                 textnature = indi["textnature"]
#                 language = indi["language"]
#                 iskey = indi["iskey"]
#                 kposition = indi["kposition"]
#                 ktol = indi["ktol"]
#                 keyval = indi["keyval"]
#                 leftindex = indi["leftindex"]
#                 liposition = indi["liposition"]
#                 litol = indi["litol"]
#                 kori = indi["kori"]
#                 koriper = indi["koriper"]
#                 regex = indi["regex"]
#
#                 regex = regex.replace(",", "|")
#
#                 print (regex)
#
#                 if "'" in keyval:
#                     keyval = keyval.replace("'", "")
#
#                 cursor = conn.cursor()
#                 # query = 'INSERT INTO eyesight.dbo.' + modelname + '(xaxis, yaxis, width, height, rotate, scalex,
#                 # scaley, ' \ 'fullwidth, fullheight, label, text, header, hposition,' \ 'htol, footer, fposition, ftol,
#                 # isthre, colorcut, ' \ 'threshold, textnature, language, iskey, kposition, ' \ 'ktol) VALUES (' + xaxisl
#                 # + \ ',' + yaxisl + ',' + widthl + ',' + heightl + ',' + rotatel + ',' + scalexl + ',' + scaleyl + ',
#                 # ' + fullwidthl + ',' + fullheightl + ', \'' + labell + '\' , \'' + textl + '\'' + ',\'' + headerl +
#                 # '\',\'' + hpositionl + '\',\'' + htoll + '\',\'' + footerl + '\',\'' + fpositionl + '\',\'' + ftoll +
#                 # '\'' + '); '
#
#                 # query = 'INSERT INTO eyesight.dbo.' + modelname + '(xaxis, yaxis, width, height, rotate, scalex, scaley, ' \
#                 #                                                   'fullwidth, fullheight, label, text, header, hposition,' \
#                 #                                                   'htol, footer, fposition, ftol, isthre, colorcut, ' \
#                 #                                                   'threshold, textnature, language, iskey, kposition, ' \
#                 #                                                   'ktol, keyval) VALUES (' + xaxisl + ',' + yaxisl + ',' + widthl + ',' + heightl + ',' + rotatel + ',' + scalexl + ',' + scaleyl + ',' + fullwidthl + ',' + fullheightl + ', \'' + labell + '\' , \'' + textl + '\'' + ',\'' + headerl + '\',\'' + hpositionl + '\',\'' + htoll + '\',\'' + footerl + '\',\'' + fpositionl + '\',\'' + ftoll + '\',\'' + isthre + '\',\'' + colorcut + '\',\'' + threshold + '\',\'' + textnature + '\',\'' + language + '\',\'' + iskey + '\',\'' + kposition + '\',\'' + ktol + '\',\'' + keyval + '\'' +');'
#
#                 query = 'INSERT INTO ' + modelname + '(xaxis, yaxis, width, height, rotate, keywidth, keyheight, ' \
#                                                                   'fullwidth, fullheight, label, text, header, hposition,' \
#                                                                   'htol, footer, fposition, ftol, isthre, colorcut, ' \
#                                                                   'threshold, textnature, language, iskey, kposition, ' \
#                                                                   'ktol, keyval, leftindex, liposition, litol, kori, koriper, regex) VALUES (' + xaxisl + ',' + yaxisl + ',' + widthl + ',' + heightl + ',' + rotatel + ', \'' + keywidth + ' \', \'' + keyheight + ' \', ' + fullwidthl + ',' + fullheightl + ', \'' + labell + '\' , \'' + textl + '\'' + ',\'' + headerl + '\',\'' + hpositionl + '\',\'' + htoll + '\',\'' + footerl + '\',\'' + fpositionl + '\',\'' + ftoll + '\',\'' + isthre + '\',\'' + colorcut + '\',\'' + threshold + '\',\'' + textnature + '\',\'' + language + '\',\'' + iskey + '\',\'' + kposition + '\',\'' + ktol + '\',\'' + keyval + '\',\'' + leftindex + '\',\'' + liposition + '\',\'' + litol + '\',\'' + kori + '\',\'' + koriper + '\',\'' + regex + '\'' + ');'
#
#                 print(query)
#                 cursor.execute(query)
#                 conn.commit()
#             return sendresponse("Model trained", 200)
#         except Exception as e:
#             return sendresponse("Failed to train model : " + e, 200)
#
#     else:
#         return sendresponse("Failed to train model", 201)

#
# @app.route('/updatemodelstable', methods = ['POST'])
# def updatemodelstable():
#     if request.is_json:
#         modelname = request.headers["modelname"]
#         ocrtype = request.headers["ocrtype"]
#         fullheight = request.headers["fullheight"]
#         fullwidth = request.headers["fullwidth"]
#         noofparam = request.headers["noofparam"]
#         header = request.headers["header"]
#         footer = request.headers["footer"]
#         leftindex = request.headers["leftindex"]
#         defthre = request.headers["defthre"]
#         defcolcut = request.headers["defcolcut"]
#
#         defdilate = request.headers["defdilate"]
#         deferode = request.headers["deferode"]
#
#         defeyerangex = request.headers["defeyerangex"]
#         defeyerangey = request.headers["defeyerangey"]
#
#         apiexposed = request.headers["apiexposed"]
#         apikey = request.headers["apikey"]
#         apiurl = request.headers["apiurl"]
#
#         headval = request.headers["headval"]
#         footval = request.headers["footval"]
#         lival = request.headers["lival"]
#         description = request.headers["descr"]
#
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor = conn.cursor()
#         modelquery = 'INSERT INTO eyesight.aspaceadmin.' + 'models' + '(model, ocr, fullheight, fullwidth, noofparam, header, footer, ' \
#                                                               'leftindex, defthre, defcolcut, defeyerangex, ' \
#                                                               'defeyerangey, apiexposed, apikey, apiurl, defdilate, deferode, headval, footval, lival, descr) VALUES (' + \
#                      ' \'' + modelname + '\' , \'' + ocrtype + '\'' + ',\'' + fullheight + '\',\'' + fullwidth + '\',\'' + noofparam + '\',\'' + header + '\',\'' + footer + '\',\'' + leftindex + '\',\'' + defthre + '\',\'' + defcolcut + '\',\'' + defeyerangex + '\',\'' + defeyerangey + '\',\'' + apiexposed + '\',\'' + apikey + '\',\'' + apiurl + '\',\'' + defdilate + '\',\'' + deferode + '\',\'' + headval + '\',\'' + footval + '\',\'' + lival + '\',\'' + description + '\'' + '); '
#
#         print(modelquery)
#         cursor.execute(modelquery)
#         conn.commit()
#         return sendresponse("Models table populated", 200)
#
#     else:
#         return sendresponse("Failed to train model", 201)
#
# @app.route('/fetchmodels', methods = ['POST'])
# def fetchmodels():
#     querystring1 = "select model, descr from models;"
#     models = []
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     cursor.execute(querystring1)
#     rows = cursor.fetchall()
#     print('Total Row(s):', cursor.rowcount)
#     for row in rows:
#         # models.append([x for x in row])
#         models.append(list(row))
#         # models.append(string)
#         print(type(row))
#         print(row)
#     print(models)
#     modelsjson = json.dumps(models)
#     return modelsjson
#     # querystring = "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE' AND TABLE_CATALOG='eyesight'"
#     # models = []
#     # # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     # #                       'Server=IN2371790W1\SQLEXPRESS;'
#     # #                       'Database=eyesight;'
#     # #                       'Trusted_Connection=yes;')
#     # conn = pyodbc.connect(databaseserver)
#     # cursor = conn.cursor()
#     # cursor.execute(querystring)
#     # rows = cursor.fetchall()
#     # print('Total Row(s):', cursor.rowcount)
#     # for row in rows:
#     #     string = str(row)
#     #     string = string.replace("(", "")
#     #     string = string.replace(")", "")
#     #     string = string.replace(" ", "")
#     #     string = string.replace(",", "")
#     #     string = string.replace("'", "")
#     #     models.append(string)
#     #     print(string)
#     # modelsjson = json.dumps(models)
#     # return modelsjson
#
#
# @app.route('/gettitles', methods = ['POST'])
# def gettitles():
#     modelname = request.form['model']
#     models = []
#     querystring = "select label from " + modelname + " where label != '-';"
#
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     cursor.execute(querystring)
#     rows = cursor.fetchall()
#     for row in rows:
#         string = str(row[0])
#         models.append(string)
#
#     return str(models)


# @app.route('/receiveimage', methods = ['POST'])
# def receiveimage():
#     try:
#         modelname = request.form['model']
#         print("model being consumed - " + modelname)
#         file = request.files['file']
#         if 'file' not in request.files:
#             print('No file part')
#             return sendresponse("file not received", 201)
#         else:
#             timestamp = time.time()
#             timestamp = str(timestamp).split('.')
#             timestamp = str(timestamp[0])
#             print(timestamp)
#
#             local_save_input_test_file = 'testimage' + timestamp + '.jpg'
#
#             finaltext = []
#             file = request.files['file']
#             file.save(local_save_input_test_file)
#             print("file received")
#
#
#             # 1st step - fetch model information from db
#             modelsmetadata = fetchmodelinfo(modelname)
#
#
#             # get dimensions, threshold, colorcut values of training image from modelmetadata
#             modeldetails = fetchmodeldetails(modelname)
#             print (modeldetails)
#             originalimagesize = modelsmetadata[0]
#             originalimagesizewidth = originalimagesize[8]
#             originalimagesizeheight = originalimagesize[9]
#
#             originalthreshold = modeldetails[8]
#             originalcolorcut = modeldetails[9]
#             originaleyerangex = int (modeldetails[10])
#             originaleyerangey = int (modeldetails[11])
#
#             originaldilate = int (modeldetails[15])
#             originalerode = int (modeldetails[16])
#
#
#             # fixing threshold and colorcut
#             try:
#                 inputimage = cv2.imread(local_save_input_test_file)
#                 inputimage = Image.fromarray(inputimage)
#                 processedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, inputimage)
#                 processedimage.save(local_save_input_test_file)
#             except Exception as e:
#                 print(e)
#
#
#             # 2nd step - resize incoming image to match train image dimension
#             resizedimage = resizeimage(local_save_input_test_file, originalimagesizewidth, originalimagesizeheight)
#             resizedimage = Image.fromarray(resizedimage)
#             try:
#                 resizedimage = resizedimage.convert("RGB")
#                 resizedimage.save(local_save_input_test_file)
#             except Exception as e:
#                 if "RGBA" in str(e):
#                     resizedimage = resizedimage.convert("RGB")
#                     resizedimage.save(local_save_input_test_file)
#                 else:
#                     print(e)
#
#             # Removing horizontal and verical lines
#             linesremoved = removelines(local_save_input_test_file)
#             if (linesremoved == "done"):
#                 print("lines removed")
#             elif (linesremoved == "failed"):
#                 print("failed to remove lines")
#
#             # 3rd step - crop area of image before header, after footer and to the left of left index
#             cleanedimage = crophfl(local_save_input_test_file, modelsmetadata, originalimagesizewidth, originalimagesizeheight, originalthreshold, originalcolorcut, originaldilate, originalerode)
#             print (cleanedimage)
#
#
#             # 3rd step - draw bounding boxes and get metadata of all boxes
#             textboxes = gettextboxes(local_save_input_test_file, originalimagesizewidth, originalimagesizeheight, originaldilate, originalerode)
#             print("--------------")
#             print (textboxes)
#             print("--------------")
#
#
#             # 4th step - loop through every row in modelsmetadata to zero down at required bounding box in textboxes
#
#
#             for parameter in modelsmetadata:
#                 print(str(parameter))
#                 label = parameter[10]
#                 if label != "-":
#
#                     # false = "true"
#                     # if false == "false":
#                     if parameter[23] == 'true':
#                         foundtxtfromkey = keybased(parameter, local_save_input_test_file, originaldilate, originalerode)
#                         if (foundtxtfromkey == ""):
#                             foundtxtfromkey == "-"
#                         # foundtxtfromkey= foundtxtfromkey.replace("x0c", "")
#                         foundtxtfromkey = foundtxtfromkey.rstrip("\n\x0c")
#                         print ("foundtxt ------------")
#                         print (foundtxtfromkey)
#                         finaltext.append('"'+foundtxtfromkey+'"')
#
#                     elif parameter[1] == 0:
#                         foundtxtfromfullocr = fullocrbased(parameter, local_save_input_test_file, originalthreshold, originalcolorcut)
#                         if (foundtxtfromfullocr == ""):
#                             foundtxtfromfullocr == "-"
#                         foundtxtfromfullocr = foundtxtfromfullocr.rstrip("\n")
#                         print("foundtxt ------------")
#                         print(foundtxtfromfullocr)
#                         finaltext.append('"'+foundtxtfromfullocr+'"')
#
#                     else :
#                         xaxis = int(parameter[1]) + leftindexfound
#                         yaxis = int(parameter[2]) + headerfound
#                         width = int(parameter[3])
#                         height = int(parameter[4])
#                         label = parameter[10]
#                         threshold = parameter[20]
#                         colorcut = parameter[19]
#                         xalignedboxes = findxalignedboxes(xaxis, textboxes, originaleyerangex)
#                         zeroeddownboxes = findyalignedboxes(yaxis, xalignedboxes, originaleyerangey)
#
#                         # 5th step - if zeroed down boxes are more than one, find the one closest to training data from db
#                         noofzeroeddownboxes = len(zeroeddownboxes)
#                         if (noofzeroeddownboxes > 1):
#                             correctedbox = getcorrectzeroddownbox(zeroeddownboxes, yaxis)
#                         elif (noofzeroeddownboxes == 0):
#                             print("cannot find parameter " + label)
#                             # return sendresponse("cannot find parameter " + label, 201)
#                             continue
#                         else:
#                             correctedbox = zeroeddownboxes[0]
#
#                         # 6th step - crop the correctedbox from the received test image
#                         left = correctedbox[0]
#                         top = correctedbox[1]
#                         right = width + left
#                         bottom = height + top
#                         # right = correctedbox[2]
#                         # bottom = correctedbox[3]
#                         cropcleanedimage = cv2.imread(local_save_input_test_file)
#                         cropcleanedimage = Image.fromarray(cropcleanedimage)
#                         # cropcleanedimage.show()
#
#                         croppedimage = cropimage(cropcleanedimage, left, top, right, bottom)
#                         # croppedimage.show()
#
#                         # 7th step - send the cropped text to adjust thresholding and color cut
#                         enhancedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, croppedimage)
#
#                         # 8th step - send the cropped image to tesseract engine to get text
#                         textfound = invoketesseract(enhancedimage)
#                         textfound = textfound.replace("\'", "")
#                         textfound = textfound.replace("(", "")
#                         textfound = textfound.replace(")", "")
#                         textfound = textfound.replace("\"", "")
#                         print(textfound)
#
#                         if textfound == "":
#                             opencvImage = cv2.cvtColor(np.array(croppedimage), cv2.COLOR_RGB2BGR)
#                             gray = cv2.cvtColor(opencvImage, cv2.COLOR_BGR2GRAY)
#                             # cv2.imshow('Enhanced image', gray)
#                             # cv2.waitKey(0)
#                             # cv2.destroyAllWindows()
#                             im_pil = Image.fromarray(gray)
#                             textfound = invoketesseract(im_pil)
#                             textfound = textfound.replace("\'", "")
#                             textfound = textfound.replace("(", "")
#                             textfound = textfound.replace(")", "")
#                             textfound = textfound.replace("\"", "")
#                             print(textfound)
#                             if (textfound == ""):
#                                 textfound == "-"
#                             textfound = textfound.rstrip("\n")
#                             print("foundtxt ------------")
#                             print(textfound)
#                             finaltext.append('"'+textfound+'"')
#                         else:
#                             if (textfound == ""):
#                                 textfound == "-"
#                             textfound = textfound.rstrip("\n")
#                             print("foundtxt ------------")
#                             print(textfound)
#                             finaltext.append('"'+textfound+'"')
#
#             return sendresponse(str(finaltext), 200)
#
#
#     #         response = gettextfromcroppedpic(modelname, textboxes)
#     #         response = str(response)
#     #         print(response)
#     #         titles = gettitles(modelname)
#     #         return sendresponse(response, 200, titles)
#     except Exception as e:
#         resultstring = 'Something wrong - ' + str(e)
#         return sendresponse(resultstring, 201)


@app.route('/receiveimage', methods = ['POST'])
def receiveimage():

    try:
        modelname = request.form['model']
        filetype = request.form['filetype']
        print("model being consumed - " + modelname)
        file = request.files['file']
        print ("filetype ---- " + filetype)
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            local_save_input_test_file = 'testimage' + timestamp + '.jpg'
            file = request.files['file']

            if (filetype == "pdf"):
                pdftoimgstatus = convertpdftoimgruntime(file, local_save_input_test_file)
                print (pdftoimgstatus)
                if (pdftoimgstatus == "failed"):
                    sendresponse("failed to convert pdf into image", 201)
            else:
                file.save(local_save_input_test_file)
                print("file received")


            # 1st step - fetch model information from db
            modelsmetadata = fetchmodelinfo(modelname)


            # get dimensions, threshold, colorcut values of training image from modelmetadata
            modeldetails = fetchmodeldetails(modelname)
            print (modeldetails)
            originalimagesize = modelsmetadata[0]
            originalimagesizewidth = originalimagesize[8]
            originalimagesizeheight = originalimagesize[9]

            originalthreshold = modeldetails[8]
            originalcolorcut = modeldetails[9]
            originaleyerangex = int (modeldetails[10])
            originaleyerangey = int (modeldetails[11])

            originaldilate = int (modeldetails[15])
            originalerode = int (modeldetails[16])


            # fixing threshold and colorcut
            try:
                inputimage = cv2.imread(local_save_input_test_file)
                inputimage = Image.fromarray(inputimage)
                processedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, inputimage)
                processedimage.save(local_save_input_test_file)
            except Exception as e:
                print(e)


            # 2nd step - resize incoming image to match train image dimension
            resizedimage = resizeimage(local_save_input_test_file, originalimagesizewidth, originalimagesizeheight)
            resizedimage = Image.fromarray(resizedimage)
            try:
                resizedimage = resizedimage.convert("RGB")
                resizedimage.save(local_save_input_test_file)
            except Exception as e:
                if "RGBA" in str(e):
                    resizedimage = resizedimage.convert("RGB")
                    resizedimage.save(local_save_input_test_file)
                else:
                    print(e)

            # deskewing image
            try:
                skewcorrected = skewcorrect(local_save_input_test_file)
                skewcorrected.save(local_save_input_test_file)
            except Exception as e:
                print (e)

            # Removing horizontal and verical lines
            linesremoved = removelines(local_save_input_test_file)
            if (linesremoved == "done"):
                print("lines removed")
            elif (linesremoved == "failed"):
                print("failed to remove lines")

            # 3rd step - crop area of image before header, after footer and to the left of left index
            cleanedimage = crophfl(local_save_input_test_file, modelsmetadata, originalimagesizewidth, originalimagesizeheight, originalthreshold, originalcolorcut, originaldilate, originalerode)
            print (cleanedimage)


            # 3rd step - draw bounding boxes and get metadata of all boxes
            textboxes = gettextboxes(local_save_input_test_file, originalimagesizewidth, originalimagesizeheight, originaldilate, originalerode)
            print("--------------")
            print (textboxes)
            print("--------------")


            # 4th step - loop through every row in modelsmetadata to zero down at required bounding box in textboxes

            # outputs = []

            threads = []
            finaltext = {}

            for parameter in modelsmetadata:
                print(str(parameter))
                process = Thread(target=process_check, args=(parameter, originaldilate, originalerode, originaleyerangex, originaleyerangey, originalthreshold,originalcolorcut, textboxes, local_save_input_test_file, finaltext,))
                time.sleep(2)
                process.start()
                threads.append(process)

            for process in threads:
                process.join()

            print (finaltext)
            finaltext = json.dumps(finaltext)
            print (finaltext)

        return sendresponse(finaltext, 200)


    #         response = gettextfromcroppedpic(modelname, textboxes)
    #         response = str(response)
    #         print(response)
    #         titles = gettitles(modelname)
    #         return sendresponse(response, 200, titles)
    except Exception as e:
        resultstring = 'Something wrong - ' + str(e)
        return sendresponse(resultstring, 201)


def process_check(parameter,originaldilate,originalerode,originaleyerangex,originaleyerangey,originalthreshold, originalcolorcut,textboxes, local_save_input_test_file, finaltext):
    label = parameter[10]

    if label != "-":

        if parameter[23] == 'true':
            foundtxtfromkey = keybased(parameter, local_save_input_test_file, originaldilate, originalerode)
            foundtxtfromkey.replace("\n", " ")
            if (foundtxtfromkey == ""):
                foundtxtfromkey == "-"
            # finaltext.append('"' + foundtxtfromkey + '"')
            # finaltext.append(foundtxtfromkey)
            finaltext[label] = foundtxtfromkey
            return foundtxtfromkey

        elif parameter[1] == 0:
            foundtxtfromfullocr = fullocrbased(parameter, local_save_input_test_file, originalthreshold,
                                               originalcolorcut)
            foundtxtfromfullocr.replace("\n", " ")
            if (foundtxtfromfullocr == ""):
                foundtxtfromfullocr == "-"
            # finaltext.append('"' + foundtxtfromfullocr + '"')
            # finaltext.append(foundtxtfromfullocr)
            finaltext[label] = foundtxtfromfullocr
            return foundtxtfromfullocr

        else:
            xaxis = int(parameter[1]) + leftindexfound
            yaxis = int(parameter[2]) + headerfound
            width = int(parameter[3])
            height = int(parameter[4])
            label = parameter[10]
            threshold = parameter[20]
            colorcut = parameter[19]
            xalignedboxes = findxalignedboxes(xaxis, textboxes, originaleyerangex)
            zeroeddownboxes = findyalignedboxes(yaxis, xalignedboxes, originaleyerangey)

            # 5th step - if zeroed down boxes are more than one, find the one closest to training data from db
            noofzeroeddownboxes = len(zeroeddownboxes)
            if (noofzeroeddownboxes > 1):
                correctedbox = getcorrectzeroddownbox(zeroeddownboxes, yaxis)
            elif (noofzeroeddownboxes == 0):
                print("cannot find parameter " + label)
                # return sendresponse("cannot find parameter " + label, 201)
                # continue
            else:
                correctedbox = zeroeddownboxes[0]

            # 6th step - crop the correctedbox from the received test image
            left = correctedbox[0]
            top = correctedbox[1]
            right = width + left
            bottom = height + top
            # right = correctedbox[2]
            # bottom = correctedbox[3]
            cropcleanedimage = cv2.imread(local_save_input_test_file)
            cropcleanedimage = Image.fromarray(cropcleanedimage)
            # cropcleanedimage.show()

            croppedimage = cropimage(cropcleanedimage, left, top, right, bottom)
            # croppedimage.show()

            # 7th step - send the cropped text to adjust thresholding and color cut
            enhancedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, croppedimage)

            # 8th step - send the cropped image to tesseract engine to get text
            textfound = invoketesseract(enhancedimage)
            textfound = textfound.replace("\'", "")
            textfound = textfound.replace("(", "")
            textfound = textfound.replace(")", "")
            textfound = textfound.replace("\"", "")
            print(textfound)

            if textfound == "":
                opencvImage = cv2.cvtColor(np.array(croppedimage), cv2.COLOR_RGB2BGR)
                gray = cv2.cvtColor(opencvImage, cv2.COLOR_BGR2GRAY)
                # cv2.imshow('Enhanced image', gray)
                # cv2.waitKey(0)
                # cv2.destroyAllWindows()
                im_pil = Image.fromarray(gray)
                textfound = invoketesseract(im_pil)
                textfound = textfound.replace("\'", "")
                textfound = textfound.replace("(", "")
                textfound = textfound.replace(")", "")
                textfound = textfound.replace("\"", "")
                print(textfound)
                if (textfound == ""):
                    textfound == "-"
                # finaltext.append('"' + textfound + '"')
                # finaltext.append(textfound)
                textfound.replace("\n", " ")
                finaltext[label] = textfound
                return textfound
            else:
                if (textfound == ""):
                    textfound == "-"
                # finaltext.append('"' + textfound + '"')
                # finaltext.append(textfound)
                textfound.replace("\n", " ")
                finaltext[label] = textfound
                return textfound



    # outputs.append(finaltext)
    print ("thred output - " + str(finaltext))
    # return finaltext


@app.route('/convertpdftoimage', methods = ['POST'])
def convertpdftoimage():
    try:
        file = request.files['file']
        print (file)
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:

            # returnimage = null

            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            local_save_input_pdf_to_img = 'initialpdftoimg' + timestamp
            pdfpath = 'static/pdftoimg/pdfs/' + local_save_input_pdf_to_img + '.pdf'
            imagepath = 'static/pdftoimg/images/' + local_save_input_pdf_to_img + '.jpg'
            print('here')
            # file = request.files['file']
            file.save(pdfpath)
            print("file received")
            # return local_save_input_initial_pdf

            try:
                images = convert_from_path(pdfpath)
                for img in images:
                    img.save(imagepath, 'JPEG')
                    # returnimage = img
            except Exception as e:
                print(e)

            return local_save_input_pdf_to_img + '.jpg'
            # return returnimage

    except Exception as e:
        print (e)
        return sendresponse("failed", 200)


def convertpdftoimgruntime(file, local_save_input_test_file):
    try:
        print (file)

        # returnimage = null

        timestamp = time.time()
        timestamp = str(timestamp).split('.')
        timestamp = str(timestamp[0])
        print(timestamp)

        local_save_input_pdf_to_img = 'initialpdftoimg' + timestamp
        pdfpath = 'static/pdftoimg/pdfs/' + local_save_input_pdf_to_img + '.pdf'
        # imagepath = 'static/pdftoimg/images/' + local_save_input_pdf_to_img + '.jpg'
        print('here')
        # file = request.files['file']
        file.save(pdfpath)
        print("file received")
        # return local_save_input_initial_pdf

        try:
            images = convert_from_path(pdfpath)
            for img in images:
                img.save(local_save_input_test_file, 'JPEG')
                # returnimage = img
        except Exception as e:
            print(e)

        return "done"
        # return returnimage

    except Exception as e:
        print (e)
        return "failed"



@app.route('/processcroppedimg', methods = ['POST'])
def processcroppedimg():
    try:
        data = request.get_json()
        if data is None:
            print("No valid request body, json missing!")
            return jsonify({'error': 'No valid request body, json missing!'})
        else:
            img_data = data['thumbnail']
            threshold = int(data['threshold'])
            colorcut = int(data['colorcut'])
            corrstring = img_data + "=="
            # print(corrstring)
            img = imread(io.BytesIO(base64.b64decode(corrstring)))
            # corrstring1 = re.sub(r'.*,', ',', corrstring)
            # corrstring2 = corrstring1.replace(",", "")
            cv2_img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            cv2.imwrite(local_save_cropped_file, cv2_img)

            # stringToRGB(img, local_save_cropped_file)
            # final = enhancecroppedpic(threshold, colorcut)
            final = enhancecroppedpicruntime(threshold, colorcut, cv2_img)
            final = np.array(final)
            string = base64.b64encode(cv2.imencode('.jpg', final)[1]).decode()

            message = {"status": 200, "imagestring": string, "updatedresultstring": getupdatedimgstring()}
            response = jsonify(message)
            response.status_code = 200
            return response

        # image1 = Image.open(local_save_cropped_file)
        # image1.show()
        # resultstring = tesserocr.image_to_text(image1, lang="eng")  # print ocr text from image
        # resultstring = resultstring.strip()
        # print(resultstring)
        # return sendresponse(resultstring, 200)
    except Exception as e:
        resultstring = 'failed to get text!!!' + str(e)
        print(resultstring)
        return sendresponse(resultstring, 201)


# @app.route('/updatemodeledtails', methods = ['POST'])
# def updatemodeledtails():
#     if request.is_json:
#         modelname = request.headers["modelname"]
#         apiexposed = request.headers["apiexposed"]
#         defthre = request.headers["defthre"]
#         defcolcut = request.headers["defcolcut"]
#         defeyerangex = request.headers["defeyerangex"]
#         defeyerangey = request.headers["defeyerangey"]
#
#         defdilate = request.headers["defdilate"]
#         deferode = request.headers["deferode"]
#
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor = conn.cursor()
#         # modelquery = 'INSERT INTO eyesight.dbo.' + 'models' + '(model, ocr, fullheight, fullwidth, noofparam, header, footer, ' \
#         #                                                       'leftindex, defthre, defcolcut, defeyerangex, ' \
#         #                                                       'defeyerangey, apiexposed, apikey, apiurl) VALUES (' + \
#         #              ' \'' + modelname + '\' , \'' + ocrtype + '\'' + ',\'' + fullheight + '\',\'' + fullwidth + '\',\'' + noofparam + '\',\'' + header + '\',\'' + footer + '\',\'' + leftindex + '\',\'' + defthre + '\',\'' + defcolcut + '\',\'' + defeyerangex + '\',\'' + defeyerangey + '\',\'' + apiexposed + '\',\'' + apikey + '\',\'' + apiurl + '\'' + '); '
#
#         # modelquery = 'UPDATE models SET apiexposed = ' + '\'' + apiexposed + '\'' + ', ' + 'defthre=' + ' \'' + defthre + '\'' + ', ' + 'defcolcut=' + ' \'' + defcolcut + '\'' + ', ' + 'defeyerangex=' + ' \'' + defeyerangex + '\'' + ', ' + 'defeyerangey=' + ' \'' + defeyerangey + '\'' + ' WHERE model = ' + '\'' + modelname + '\';'
#         modelquery = 'UPDATE models SET apiexposed = ' + '\'' + apiexposed + '\'' + ', ' + 'defthre=' + ' \'' + defthre + '\'' + ', ' + 'defcolcut=' + ' \'' + defcolcut + '\'' + ', ' + 'defeyerangex=' + ' \'' + defeyerangex + '\'' + ', ' + 'defeyerangey=' + ' \'' + defeyerangey + '\''+ ', ' + 'defdilate=' + ' \'' + defdilate + '\''+ ', ' + 'deferode=' + ' \'' + deferode + '\''+ ' WHERE model = ' + '\'' + modelname + '\';'
#         print(modelquery)
#         cursor.execute(modelquery)
#         conn.commit()
#         return sendresponse("Model updated", 200)
#
#     else:
#         return sendresponse("Failed to update model", 201)

# pass a cropped image to this method to enhance it
def enhancecroppedpic(threshold, colorcut):
    # colorcut = 200
    # threshold = 125

    image = cv2.imread(local_save_cropped_file)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # cv2.imwrite(local_save_cropped_file, gray)

    # img = cv2.imread(local_save_cropped_file)
    # img = gray
    gray[(gray > colorcut)] = 255  # change everything to white where pixel is not black

    thresh, enhancedimg = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY)

    cv2.imwrite(local_save_cropped_file, enhancedimg)

    message = {"status": 200, "imagestring": RGBTostring(), "updatedresultstring": getupdatedimgstring()}
    response = jsonify(message)
    response.status_code = 200
    return response


# pass cropped pic to enhance it with threshold and colorcut during runtime
def enhancecroppedpicruntime(threshold, colorcut, croppedpic):
    opencvImage = cv2.cvtColor(np.array(croppedpic), cv2.COLOR_RGB2BGR)
    if threshold == "-":
        threshold = 200
    else:
        threshold = int(threshold)

    if colorcut == "-":
        colorcut = 200
    else:
        colorcut = int(colorcut)

    # image = cv2.imread(local_save_cropped_file)
    gray = cv2.cvtColor(opencvImage, cv2.COLOR_BGR2GRAY)
    gray[(gray > colorcut)] = 255  # change everything to white where pixel is not black
    thresh, enhancedimg = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY)
    # cv2.imwrite(local_save_cropped_file, enhancedimg)

    # message = {"status": 200, "imagestring": RGBTostring(), "updatedresultstring": getupdatedimgstring()}
    # response = jsonify(message)
    # response.status_code = 200
    # cv2.imshow('Enhanced image', enhancedimg)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    im_pil = Image.fromarray(enhancedimg)
    return im_pil


def getupdatedimgstring():
    try:
        # croppedimage = imread(local_save_cropped_file)
        croppedimage = Image.open(local_save_cropped_file)
        resultstring = tesserocr.image_to_text(croppedimage, lang=language)  # print ocr text from image
        resultstring = resultstring.strip()
        if resultstring == "":
            resultstring = "Cannot extract text"
        print(resultstring)
        return resultstring
    except Exception as e:
        print(e)
        return str(e)


# pass a model name to this method to get model meta data in list of lists format
# def fetchmodelinfo(modelname):
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     query = 'select * from ' + modelname + ';'
#     cursor.execute(query)
#     modelmetadata = []
#     for row in cursor:
#         modelmetadata.append(row)
#     return modelmetadata


# call this method and pass width and height to resize an image
def resizeimage(local_save_input_test_file, width, height):
    width = int(width)
    height = int(height)
    img = cv2.imread(local_save_input_test_file, cv2.IMREAD_UNCHANGED)
    print('Original Dimensions : ', img.shape)
    dim = (width, height)

    try:
        resized = cv2.resize(img, dim, interpolation=cv2.INTER_AREA)
    except Exception as e:
        print(e)
        return e

    print('Resized Dimensions : ', resized.shape)
    return resized


# call this method to get boundary boxes aligned to x axis of training image in list of lists format
def findxalignedboxes(xaxis, textboxes, originaleyerangex):
    xalignedboxes = []
    for textbox in textboxes:
        startx = int(textbox[0])
        diff = abs(startx - xaxis)
        if (diff <= originaleyerangex):
            xalignedboxes.append(textbox)
    print ("x-aligned boxes - " + str(xalignedboxes))
    return xalignedboxes


# call this method to get boundary boxes aligned to y axis of training image in list of lists format
def findyalignedboxes(yaxis, xalignedboxes, originaleyerangey):
    yalignedboxes = []
    for xalignedbox in xalignedboxes:
        starty = int(xalignedbox[1])
        diff = abs(starty - yaxis)
        if (diff <= originaleyerangey):
            yalignedboxes.append(xalignedbox)
    print("y-aligned boxes - " + str(yalignedboxes))
    return yalignedboxes


# call this method to get correct box if more than one zerod down boxes are received
def getcorrectzeroddownbox(zeroeddownboxes, yaxis):
    zeroeddownboxyaxislist = []
    for zeroeddownbox in zeroeddownboxes:
        i = zeroeddownbox[1]
        zeroeddownboxyaxislist.append(i)
    correctbox = min(zeroeddownboxyaxislist, key=lambda x: abs(x - yaxis))
    correctindex = zeroeddownboxyaxislist.index(correctbox)
    correctedbox = zeroeddownboxes[correctindex]
    print (correctedbox)
    return correctedbox


# call this method with correct parameters to crop an image
def cropimage(image, left, top, right, bottom):
    cropped = image.crop((left, top, right, bottom))
    # cropped.show()
    cropped = cropped.convert('RGB')
    return cropped


# call this method and pass cropped image to invoke tesseract engine and get text found in that image
def invoketesseract(croppedimage):
    try:
        resultstring = tesserocr.image_to_text(croppedimage, lang=language)  # print ocr text from image
        resultstring = resultstring.strip()
        print(resultstring)
        return resultstring
    except Exception as e:
        print(e)
        return str(e)


# call this method to get the list of all bounding boxes around texts in the image in list of lists format
def gettextboxes(local_save_input_test_file, fullwidth, fullheight, originaldilate, originalerode):
    # Load the image
    img = cv2.imread(local_save_input_test_file)

    # convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # smooth the image to avoid noises
    gray = cv2.medianBlur(gray, 5)

    # Apply adaptive threshold
    thresh = cv2.adaptiveThreshold(gray, 255, 1, 1, 11, 2)
    thresh_color = cv2.cvtColor(thresh, cv2.COLOR_GRAY2BGR)

    # apply some dilation and erosion to join the gaps - change iteration to detect more or less area's
    thresh = cv2.dilate(thresh, None, iterations=originaldilate)
    thresh = cv2.erode(thresh, None, iterations=originalerode)

    # Find the contours
    contours, hierarchy = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # size of the image
    dimensions = img.shape
    height = img.shape[0]
    width = img.shape[1]

    percentage = .5
    xthreshold = percentage * (width / 100)
    ythreshold = percentage * (height / 100)

    # For each contour, find the bounding rectangle and draw it
    boxes = []
    for cnt in contours:
        xypadding = 5
        hwpadding = 10
        x, y, w, h = cv2.boundingRect(cnt)
        if (w > xthreshold):
            if (h > ythreshold):
                x = x - xypadding
                y = y - xypadding
                w = w + hwpadding
                h = h + hwpadding
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.rectangle(thresh_color, (x, y), (x + w, y + h), (0, 255, 0), 2)
                box = [x, y, x + w, y + h]
                boxes.append(box)

    # Finally show the image
    print(boxes)
    # cv2.imshow('img', img)
    # orig = Image.fromarray(img)
    # orig.save("C:/Users/ET437GL/Pictures/Eyesight sample images/passportsamples/10.jpg")
    # cv2.imshow('res', thresh_color)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    return boxes














    # image = cv2.imread(local_save_input_test_file)
    # height, width, channels = image.shape
    #
    # if (fullwidth == 0) & (fullheight == 0):
    #     fullwidth = width
    #     fullheight = height
    #
    # orig = image.copy()
    # (H, W) = image.shape[:2]
    #
    # (newW, newH) = (640, 640)
    # rW = W / float(newW)
    # rH = H / float(newH)
    #
    # image = cv2.resize(image, (newW, newH))
    # (H, W) = image.shape[:2]
    #
    # layerNames = [
    #     "feature_fusion/Conv_7/Sigmoid",
    #     "feature_fusion/concat_3"]
    #
    # print("[INFO] loading EAST text detector...")
    #
    # net = cv2.dnn.readNet("frozen_east_text_detection.pb")
    # blob = cv2.dnn.blobFromImage(image, 1.0, (W, H),
    #                              (123.68, 116.78, 103.94), swapRB=True, crop=False)
    # net.setInput(blob)
    # (scores, geometry) = net.forward(layerNames)
    #
    # (numRows, numCols) = scores.shape[2:4]
    # rects = []
    # confidences = []
    #
    # textboxesxxx = []
    #
    # for y in range(0, numRows):
    #     scoresData = scores[0, 0, y]
    #     xData0 = geometry[0, 0, y]
    #     xData1 = geometry[0, 1, y]
    #     xData2 = geometry[0, 2, y]
    #     xData3 = geometry[0, 3, y]
    #     anglesData = geometry[0, 4, y]
    #
    #     for x in range(0, numCols):
    #         if scoresData[x] < 0.05:
    #             continue
    #
    #         (offsetX, offsetY) = (x * 4.0, y * 4.0)
    #
    #         angle = anglesData[x]
    #         cos = np.cos(angle)
    #         sin = np.sin(angle)
    #
    #         h = xData0[x] + xData2[x]
    #         w = xData1[x] + xData3[x]
    #
    #         # extend the bounding box to 2% of total width and height
    #         # widthextension = (fullwidth / 100) * 2
    #         # heightextension = (fullheight / 100) * 2
    #
    #         widthextension = 10
    #         heightextension = 10
    #
    #         endX = (int(offsetX + (cos * xData1[x]) + (sin * xData2[x]))) + widthextension
    #         endY = (int(offsetY - (sin * xData1[x]) + (cos * xData2[x]))) + heightextension
    #         startX = (int(endX - w)) - widthextension
    #         startY = (int(endY - h)) - heightextension
    #
    #         if (startX < 0):
    #             startX = 0
    #         if (startY < 0):
    #             startY = 0
    #
    #         # print(str(startX) + " " + str(startY) + " " + str(endX) + " " + str(endY))
    #
    #         rects.append((startX, startY, endX, endY))
    #         confidences.append(scoresData[x])
    #
    # boxes = non_max_suppression(np.array(rects), probs=confidences)
    #
    # for (startX, startY, endX, endY) in boxes:
    #     startX = int(startX * rW)
    #     startY = int(startY * rH)
    #     endX = int(endX * rW)
    #     endY = int(endY * rH)
    #
    #     # print(str(startX) + " " + str(startY) + " " + str(endX) + " " + str(endY))
    #     cv2.rectangle(orig, (startX, startY), (endX, endY), (0, 255, 0), 2)
    #
    #     textboxx = [startX, startY, endX, endY]
    #     textboxesxxx.append(textboxx)
    #
    # # cv2.imshow("Text Detection", orig)
    # # cv2.waitKey(0)
    # # print(textboxesxxx)
    # return textboxesxxx


# call this method and pass model name to get the dimensions of train image of a model
# @app.route('/getimagedimensions', methods = ['POST'])
# def getimagedimensions():
#     modelname = request.form['modelname']
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     query = 'select top 1 * from ' + modelname + ';'
#     cursor.execute(query)
#     dimensions = []
#     # row = cursor[0]
#     # width = row[8]
#     # height = row[9]
#     # dimensions.append(width)
#     # dimensions.append(height)
#
#     for row in cursor:
#         width = int(row[8])
#         height = int (row[9])
#         dimensions.append(width)
#         dimensions.append(height)
#     print(dimensions)
#     return str(dimensions)


# call this method and pass model name to get the model metadata table in list of lists format
# @app.route('/fetchexistingmodel', methods = ['POST'])
# def fetchexistingmodel():
#     modelname = request.form['modelname']
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     query = 'select * from ' + modelname + ';'
#     cursor.execute(query)
#     rows = []
#     for row in cursor:
#         rows.append(row)
#     rows = str(rows)
#     # rows.replace("(", "[")
#     # rows.replace(")", "]")
#     # response = json.dumps(rows)
#     print(rows)
#     return rows


# call this method and pass model name to get the model metadata table in list of lists format
# @app.route('/fetchexistingmodelmetadata', methods = ['POST'])
# def fetchexistingmodelmetadata():
#     modelname = request.form['modelname']
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     query = 'select * from models where model = ' + '\'' + modelname + '\'' + ';'
#     cursor.execute(query)
#     rows = []
#     for row in cursor:
#         rows.append(row)
#     rows = str(rows)
#     # rows.replace("(", "[")
#     # rows.replace(")", "]")
#     # response = json.dumps(rows)
#     print(rows)
#     return rows
#
#
# # call this method to retrain the model
# @app.route('/retrain', methods = ['POST'])
# def retrain():
#     try:
#         model = (request.form['model']).strip()
#         label = (request.form['label']).strip()
#         xaxisnew = (request.form['xaxis']).strip()
#         yaxisnew = (request.form['yaxis']).strip()
#         widthnew = (request.form['width']).strip()
#         heightnew = (request.form['height']).strip()
#
#         rotatenew = (request.form['rotate']).strip()
#         keywidthlnew = (request.form['keywidthl']).strip()
#         keyheightlnew = (request.form['keyheightl']).strip()
#         fullwidthnew = (request.form['fullwidth']).strip()
#         fullheightnew = (request.form['fullheight']).strip()
#
#         labelnew = (request.form['label']).strip()
#         textnew = (request.form['text']).strip()
#         headernew = (request.form['header']).strip()
#         hpositionnew = (request.form['hposition']).strip()
#         htolnew = (request.form['htol']).strip()
#         footernew = (request.form['footer']).strip()
#
#         fpositionnew = (request.form['fposition']).strip()
#         ftolnew = (request.form['ftol']).strip()
#         isthrenew = (request.form['isthre']).strip()
#         colorcutnew = (request.form['colorcut']).strip()
#         thresholdnew = (request.form['threshold']).strip()
#
#         textnaturenew = (request.form['textnature']).strip()
#         languagenew = (request.form['language']).strip()
#         iskeynew = (request.form['iskey']).strip()
#         kpositionnew = (request.form['kposition']).strip()
#         ktolnew = (request.form['ktol']).strip()
#
#         keyvalnew = (request.form['keyval']).strip()
#         leftindexnew = (request.form['leftindex']).strip()
#         lipositionnew = (request.form['liposition']).strip()
#         litolnew = (request.form['litol']).strip()
#
#         # if keywidthlnew == "NaN":
#         #     keywidthlnew = 0
#         #
#         # if fullheightnew == "NaN":
#         #     fullheightnew = 0
#
#         print ("here")
#
#         korinew = (request.form['kori']).strip()
#         koripernew = (request.form['koriper']).strip()
#         regexnew = (request.form['regex']).strip()
#
#         print(regexnew)
#
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor1 = conn.cursor()
#         # table = 'dbo.' + model
#         table = model
#         updatequery = 'UPDATE ' + table + ' SET xaxis = ' + xaxisnew + ', yaxis= ' + yaxisnew + ', width= ' + widthnew + ', height= ' + heightnew + ', rotate= ' + rotatenew + ', keywidth= \'' + keywidthlnew + '\', keyheight= \'' + keyheightlnew + '\', fullwidth= ' + fullwidthnew + ', fullheight= ' + fullheightnew + ', label= ' + '\'' + labelnew + '\'' + ', text= ' + '\'' + textnew + '\'' + ', header= ' + '\'' + headernew + '\'' + ', hposition= ' + '\'' + hpositionnew + '\'' + ', htol= ' + '\'' + htolnew + '\'' + ', footer= ' + '\'' + footernew + '\'' + ', fposition= ' + '\'' + fpositionnew + '\'' + ', ftol= ' + '\'' + ftolnew + '\'' + ', isthre= ' + '\'' + isthrenew + '\'' + ', colorcut= ' + '\'' + colorcutnew + '\'' + ', threshold= ' + '\'' +thresholdnew + '\'' + ', textnature= ' + '\'' + textnaturenew + '\'' + ', language= ' + '\'' + languagenew + '\'' + ', iskey= ' + '\'' + iskeynew + '\'' + ', kposition= ' + '\'' + kpositionnew + '\'' + ', ktol= ' + '\'' + ktolnew + '\'' + ', keyval= ' + '\'' + keyvalnew + '\'' + ', leftindex= ' + '\'' + leftindexnew + '\'' + ', liposition= ' + '\'' + lipositionnew + '\'' + ', litol= ' + '\'' + litolnew + '\'' + ', kori= ' + '\'' + korinew + '\'' + ', koriper= ' + '\'' + koripernew + '\'' + ', regex= ' + '\'' + regexnew + '\'' + ' WHERE label = ' + '\'' + label + '\'' + ';'
#
#         print (updatequery)
#
#         cursor1.execute(updatequery)
#         conn.commit()
#         conn.close()
#         return sendresponse("model retrained", 200)
#
#     except Exception as e:
#         print("error : " + str(e))
#         return sendresponse("failed to train model", 201)
#
#
# # call this method to retrain the header
# @app.route('/retrainheader', methods = ['POST'])
# def retrainheader():
#     try:
#         model = (request.form['model']).strip()
#         label = (request.form['label']).strip()
#         xaxisnew = (request.form['xaxis']).strip()
#         yaxisnew = (request.form['yaxis']).strip()
#         widthnew = (request.form['width']).strip()
#         heightnew = (request.form['height']).strip()
#
#         rotatenew = (request.form['rotate']).strip()
#         scalexnew = (request.form['scalex']).strip()
#         scaleynew = (request.form['scaley']).strip()
#         fullwidthnew = (request.form['fullwidth']).strip()
#         fullheightnew = (request.form['fullheight']).strip()
#
#         textnew = (request.form['text']).strip()
#         headernew = (request.form['header']).strip()
#         hpositionnew = (request.form['hposition']).strip()
#         htolnew = (request.form['htol']).strip()
#         footernew = (request.form['footer']).strip()
#
#         fpositionnew = (request.form['fposition']).strip()
#         ftolnew = (request.form['ftol']).strip()
#         isthrenew = (request.form['isthre']).strip()
#         colorcutnew = (request.form['colorcut']).strip()
#         thresholdnew = (request.form['threshold']).strip()
#
#         textnaturenew = (request.form['textnature']).strip()
#         languagenew = (request.form['language']).strip()
#         iskeynew = (request.form['iskey']).strip()
#         kpositionnew = (request.form['kposition']).strip()
#         ktolnew = (request.form['ktol']).strip()
#
#         keyvalnew = (request.form['keyval']).strip()
#         leftindexnew = (request.form['leftindex']).strip()
#         lipositionnew = (request.form['liposition']).strip()
#         litolnew = (request.form['litol']).strip()
#
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor1 = conn.cursor()
#         # table = 'dbo.' + model
#         table = model
#         updatequery = 'UPDATE ' + table + ' SET xaxis = ' + xaxisnew + ', yaxis= ' + yaxisnew + ', width= ' + widthnew + ', height= ' + heightnew + ', rotate= ' + rotatenew + ', scalex= ' + scalexnew + ', scaley= ' + scaleynew + ', fullwidth= ' + fullwidthnew + ', fullheight= ' + fullheightnew + ', text= ' + '\'' + textnew + '\'' + ', header= ' + '\'' + headernew + '\'' + ', hposition= ' + '\'' + hpositionnew + '\'' + ', htol= ' + '\'' + htolnew + '\'' + ', footer= ' + '\'' + footernew + '\'' + ', fposition= ' + '\'' + fpositionnew + '\'' + ', ftol= ' + '\'' + ftolnew + '\'' + ', isthre= ' + '\'' + isthrenew + '\'' + ', colorcut= ' + '\'' + colorcutnew + '\'' + ', threshold= ' + '\'' +thresholdnew + '\'' + ', textnature= ' + '\'' + textnaturenew + '\'' + ', language= ' + '\'' + languagenew + '\'' + ', iskey= ' + '\'' + iskeynew + '\'' + ', kposition= ' + '\'' + kpositionnew + '\'' + ', ktol= ' + '\'' + ktolnew + '\'' + ', keyval= ' + '\'' + keyvalnew + '\'' + ', leftindex= ' + '\'' + leftindexnew + '\'' + ', liposition= ' + '\'' + lipositionnew + '\'' + ', litol= ' + '\'' + litolnew + '\'' + ' WHERE header = ' + '\'' + headernew + '\'' + ';'
#
#         cursor1.execute(updatequery)
#         conn.commit()
#         conn.close()
#         return sendresponse("model retrained", 200)
#
#     except Exception as e:
#         print("error : " + str(e))
#         return sendresponse("failed to train model", 201)
#
#
# # call this method to retrain the footer
# @app.route('/retrainfooter', methods = ['POST'])
# def retrainfooter():
#     try:
#         model = (request.form['model']).strip()
#         label = (request.form['label']).strip()
#         xaxisnew = (request.form['xaxis']).strip()
#         yaxisnew = (request.form['yaxis']).strip()
#         widthnew = (request.form['width']).strip()
#         heightnew = (request.form['height']).strip()
#
#         rotatenew = (request.form['rotate']).strip()
#         scalexnew = (request.form['scalex']).strip()
#         scaleynew = (request.form['scaley']).strip()
#         fullwidthnew = (request.form['fullwidth']).strip()
#         fullheightnew = (request.form['fullheight']).strip()
#
#         textnew = (request.form['text']).strip()
#         headernew = (request.form['header']).strip()
#         hpositionnew = (request.form['hposition']).strip()
#         htolnew = (request.form['htol']).strip()
#         footernew = (request.form['footer']).strip()
#
#         fpositionnew = (request.form['fposition']).strip()
#         ftolnew = (request.form['ftol']).strip()
#         isthrenew = (request.form['isthre']).strip()
#         colorcutnew = (request.form['colorcut']).strip()
#         thresholdnew = (request.form['threshold']).strip()
#
#         textnaturenew = (request.form['textnature']).strip()
#         languagenew = (request.form['language']).strip()
#         iskeynew = (request.form['iskey']).strip()
#         kpositionnew = (request.form['kposition']).strip()
#         ktolnew = (request.form['ktol']).strip()
#
#         keyvalnew = (request.form['keyval']).strip()
#         leftindexnew = (request.form['leftindex']).strip()
#         lipositionnew = (request.form['liposition']).strip()
#         litolnew = (request.form['litol']).strip()
#
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor1 = conn.cursor()
#         # table = 'dbo.' + model
#         table = model
#         updatequery = 'UPDATE ' + table + ' SET xaxis = ' + xaxisnew + ', yaxis= ' + yaxisnew + ', width= ' + widthnew + ', height= ' + heightnew + ', rotate= ' + rotatenew + ', scalex= ' + scalexnew + ', scaley= ' + scaleynew + ', fullwidth= ' + fullwidthnew + ', fullheight= ' + fullheightnew + ', text= ' + '\'' + textnew + '\'' + ', header= ' + '\'' + headernew + '\'' + ', hposition= ' + '\'' + hpositionnew + '\'' + ', htol= ' + '\'' + htolnew + '\'' + ', footer= ' + '\'' + footernew + '\'' + ', fposition= ' + '\'' + fpositionnew + '\'' + ', ftol= ' + '\'' + ftolnew + '\'' + ', isthre= ' + '\'' + isthrenew + '\'' + ', colorcut= ' + '\'' + colorcutnew + '\'' + ', threshold= ' + '\'' +thresholdnew + '\'' + ', textnature= ' + '\'' + textnaturenew + '\'' + ', language= ' + '\'' + languagenew + '\'' + ', iskey= ' + '\'' + iskeynew + '\'' + ', kposition= ' + '\'' + kpositionnew + '\'' + ', ktol= ' + '\'' + ktolnew + '\'' + ', keyval= ' + '\'' + keyvalnew + '\'' + ', leftindex= ' + '\'' + leftindexnew + '\'' + ', liposition= ' + '\'' + lipositionnew + '\'' + ', litol= ' + '\'' + litolnew + '\'' + ' WHERE footer = ' + '\'' + footernew + '\'' + ';'
#
#         cursor1.execute(updatequery)
#         conn.commit()
#         conn.close()
#         return sendresponse("model retrained", 200)
#
#     except Exception as e:
#         print("error : " + str(e))
#         return sendresponse("failed to train model", 201)
#
#
# # call this method to retrain the left index
# @app.route('/retrainleftindex', methods = ['POST'])
# def retrainleftindex():
#     try:
#         model = (request.form['model']).strip()
#         label = (request.form['label']).strip()
#         xaxisnew = (request.form['xaxis']).strip()
#         yaxisnew = (request.form['yaxis']).strip()
#         widthnew = (request.form['width']).strip()
#         heightnew = (request.form['height']).strip()
#
#         rotatenew = (request.form['rotate']).strip()
#         scalexnew = (request.form['scalex']).strip()
#         scaleynew = (request.form['scaley']).strip()
#         fullwidthnew = (request.form['fullwidth']).strip()
#         fullheightnew = (request.form['fullheight']).strip()
#
#         textnew = (request.form['text']).strip()
#         headernew = (request.form['header']).strip()
#         hpositionnew = (request.form['hposition']).strip()
#         htolnew = (request.form['htol']).strip()
#         footernew = (request.form['footer']).strip()
#
#         fpositionnew = (request.form['fposition']).strip()
#         ftolnew = (request.form['ftol']).strip()
#         isthrenew = (request.form['isthre']).strip()
#         colorcutnew = (request.form['colorcut']).strip()
#         thresholdnew = (request.form['threshold']).strip()
#
#         textnaturenew = (request.form['textnature']).strip()
#         languagenew = (request.form['language']).strip()
#         iskeynew = (request.form['iskey']).strip()
#         kpositionnew = (request.form['kposition']).strip()
#         ktolnew = (request.form['ktol']).strip()
#
#         keyvalnew = (request.form['keyval']).strip()
#         leftindexnew = (request.form['leftindex']).strip()
#         lipositionnew = (request.form['liposition']).strip()
#         litolnew = (request.form['litol']).strip()
#
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor1 = conn.cursor()
#         # table = 'dbo.' + model
#         table = model
#         updatequery = 'UPDATE ' + table + ' SET xaxis = ' + xaxisnew + ', yaxis= ' + yaxisnew + ', width= ' + widthnew + ', height= ' + heightnew + ', rotate= ' + rotatenew + ', scalex= ' + scalexnew + ', scaley= ' + scaleynew + ', fullwidth= ' + fullwidthnew + ', fullheight= ' + fullheightnew + ', text= ' + '\'' + textnew + '\'' + ', header= ' + '\'' + headernew + '\'' + ', hposition= ' + '\'' + hpositionnew + '\'' + ', htol= ' + '\'' + htolnew + '\'' + ', footer= ' + '\'' + footernew + '\'' + ', fposition= ' + '\'' + fpositionnew + '\'' + ', ftol= ' + '\'' + ftolnew + '\'' + ', isthre= ' + '\'' + isthrenew + '\'' + ', colorcut= ' + '\'' + colorcutnew + '\'' + ', threshold= ' + '\'' +thresholdnew + '\'' + ', textnature= ' + '\'' + textnaturenew + '\'' + ', language= ' + '\'' + languagenew + '\'' + ', iskey= ' + '\'' + iskeynew + '\'' + ', kposition= ' + '\'' + kpositionnew + '\'' + ', ktol= ' + '\'' + ktolnew + '\'' + ', keyval= ' + '\'' + keyvalnew + '\'' + ', leftindex= ' + '\'' + leftindexnew + '\'' + ', liposition= ' + '\'' + lipositionnew + '\'' + ', litol= ' + '\'' + litolnew + '\'' + ' WHERE leftindex = ' + '\'' + leftindexnew + '\'' + ';'
#
#         cursor1.execute(updatequery)
#         conn.commit()
#         conn.close()
#         return sendresponse("model retrained", 200)
#
#     except Exception as e:
#         print("error : " + str(e))
#         return sendresponse("failed to train model", 201)
#

# call this method to authenticate users
@app.route('/login', methods = ['GET', 'POST'])
def login():
    # id = request.args.get('id')
    # password = request.args.get('password')
    id = request.form['id']
    password = request.form['password']

    print(id + " " + password)

    if (id == "admin"):
        if (password == "ey@123"):
            # return "authenticated"
            return sendresponse("authenticated", 200)
        else:
            # return "password wrong"
            return sendresponse("password wrong", 201)
    else:
        # return "invalid user"
        return sendresponse("invalid user", 201)


# call this method to crop header, footer and left index
@app.route('/crophfl', methods = ['GET', 'POST'])
def crophfl(local_save_input_test_file, modelsmetadata, originalimagesizewidth, originalimagesizeheight, originalthreshold, originalcolorcut, originaldilate, originalerode):
    global headerfound
    global leftindexfound

    try:
        inputimage = cv2.imread(local_save_input_test_file)
        inputimage = Image.fromarray(inputimage)
        # processedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, inputimage)
        # processedimage.save(local_save_input_test_file)
        boxes = gettextboxes(local_save_input_test_file, originalimagesizewidth, originalimagesizeheight, originaldilate, originalerode)
        boxes.sort(key=lambda lis: lis[1])
        print ("Ascending sorted:")
        print (boxes)
    except Exception as e:
        print(e)

    for model in modelsmetadata:
        model = [x for x in model]

        if model[12] == "true":
            originalheader = model[11]
            originalheaderxaxis = model[1]
            originalheaderyaxis = model[2]
            originalheaderheight = model[4]
            originalheaderwidth = model[3]
            headerfound = int(originalheaderyaxis)

            # currently not in use
            if model[18] == "true":
                originalheaderthreshold = model[20]
                originalheadercolorcut = model[19]
            else:
                originalheaderthreshold = "-"
                originalheadercolorcut = "-"

            try:
                for [startX, startY, endX, endY] in boxes:
                    left = startX
                    top = startY
                    right = left + originalheaderwidth
                    bottom = top + originalheaderheight
                    cropped_image = cropimage(processedimage, left, top, right, bottom)
                    result_text = invoketesseract(cropped_image)
                    similarityratio = similar(originalheader, result_text)

                    if (similarityratio > 0.8):
                        final_image = cropimage(inputimage, 0, top, originalimagesizewidth, originalimagesizeheight)
                        # cv2.imshow('Enhanced image', final_image)
                        # cv2.waitKey(0)
                        final_image.save(local_save_input_test_file)
                        headerfound = 0
                        break
                    else:
                        print("Cannot find header in the provided image")
            except Exception as e:
                print(e)

        elif model[15] == "true":
            originalfooter = model[11]
            originalfooterheight = model[4]
            originalfooterwidth = model[3]
            if model[18] == "true":
                originalfooterthreshold = model[20]
                originalfootercolorcut = model[19]
            else:
                originalfooterthreshold = "-"
                originalfootercolorcut = "-"

            try:
                inputimage = cv2.imread(local_save_input_test_file)
                headcroppedimgheight, headcroppedimgwidth, channels = inputimage.shape
                inputimage = Image.fromarray(inputimage)
                processedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, inputimage)
                processedimage.save(local_save_input_test_file)
                boxes = gettextboxes(local_save_input_test_file, 0, 0, originaldilate, originalerode)
                boxes.sort(key=lambda lis: lis[1], reverse=True)
                print("Descending sorted:")
                print(boxes)
                for [startX, startY, endX, endY] in boxes:
                    left = startX
                    top = startY
                    right = left + originalfooterwidth
                    bottom = top + originalfooterheight
                    cropped_image = cropimage(processedimage, left, top, right, bottom)
                    result_text = invoketesseract(cropped_image)

                    similarityratio = similar(originalfooter, result_text)

                    if (similarityratio > 0.8):
                        final_image = cropimage(inputimage, 0, 0, headcroppedimgwidth, bottom)
                        # cv2.imshow('Enhanced image', final_image)
                        # cv2.waitKey(0)
                        final_image.save(local_save_input_test_file)
                        break
                    else:
                        print("Cannot find footer in the provided image")
            except Exception as e:
                print(e)

        elif model[27] == "true":
            originalleftindex = model[11]
            originalleftindexxaxis = model[1]
            originalleftindexyaxis = model[2]
            originalleftindexheight = model[4]
            originalleftindexwidth = model[3]
            leftindexfound = int(originalleftindexxaxis)
            if model[18] == "true":
                originalleftindexthreshold = model[20]
                originalleftindexcolorcut = model[19]
            else:
                originalleftindexthreshold = "-"
                originalleftindexcolorcut = "-"

            try:
                inputimage = cv2.imread(local_save_input_test_file)
                footcroppedimgheight, footcroppedimgwidth, channels = inputimage.shape
                inputimage = Image.fromarray(inputimage)
                processedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, inputimage)
                processedimage.save(local_save_input_test_file)
                boxes = gettextboxes(local_save_input_test_file, 0, 0, originaldilate, originalerode)
                boxes.sort(key=lambda lis: lis[0])
                print("X Ascending sorted:")
                print(boxes)
                for [startX, startY, endX, endY] in boxes:
                    left = startX
                    top = startY
                    right = left + originalleftindexwidth
                    bottom = top + originalleftindexheight
                    cropped_image = cropimage(processedimage, left, top, right, bottom)
                    result_text = invoketesseract(cropped_image)

                    similarityratio = similar(originalleftindex, result_text)

                    if (similarityratio > 0.8):
                        final_image = cropimage(inputimage, left, 0, footcroppedimgwidth, footcroppedimgheight)
                        # cv2.imshow('Enhanced image', final_image)
                        # cv2.waitKey(0)
                        final_image.save(local_save_input_test_file)
                        leftindexfound = 0
                        break
                    else:
                        print("Cannot find left index in the provided image")
            except Exception as e:
                print(e)


# # call this method with a model name to delete it in DB
# @app.route('/deletemodel', methods = ['POST'])
# def deletemodel():
#     try:
#         modelname = request.form['model']
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor = conn.cursor()
#         # query = 'select * from eyesight.dbo.' + modelname + ';'
#         query = 'DROP TABLE ' + modelname + ';'
#         query1 = 'DELETE from models WHERE model = ' + '\'' + modelname + '\'' +';'
#         print (query1)
#         cursor.execute(query)
#         cursor.execute(query1)
#         conn.commit()
#         conn.close()
#         return sendresponse("model deleted", 200)
#     except Exception as e:
#         print("error : " + str(e))
#         return sendresponse("cannot delete model", 201)
#
#
# # call this method to get previos header and left index coordinates
# @app.route('/getoldheadleft', methods = ['POST'])
# def getoldheadleft():
#     try:
#         modelname = request.form['model']
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor = conn.cursor()
#
#         try:
#             query = '''select yaxis from ''' + modelname + ''' where header = 'true';'''
#             # querystring = "select label from eyesight.dbo." + modelname + " where label != '-';"
#             cursor.execute(query)
#             for row in cursor:
#                 headeryaxis = row
#             headeryaxis = str(headeryaxis)
#             headeryaxis = headeryaxis.replace('(', '')
#             headeryaxis = headeryaxis.replace(')', '')
#             headeryaxis = headeryaxis.replace(',', '')
#             headeryaxis = headeryaxis.replace(' ', '')
#             headeryaxis = headeryaxis.replace('\'', '')
#         except Exception as e:
#             print("error : " + str(e))
#
#         try:
#             query = '''select xaxis from ''' + modelname + ''' where leftindex = 'true';'''
#             cursor.execute(query)
#             for row in cursor:
#                 leftindexxaxis = row
#             leftindexxaxis = str(leftindexxaxis)
#             leftindexxaxis = leftindexxaxis.replace('(', '')
#             leftindexxaxis = leftindexxaxis.replace(')', '')
#             leftindexxaxis = leftindexxaxis.replace(',', '')
#             leftindexxaxis = leftindexxaxis.replace(' ', '')
#             headeryaxis = headeryaxis.replace('\'', '')
#         except Exception as e:
#             print("error : " + str(e))
#
#         conn.commit()
#         conn.close()
#
#         result = str ([headeryaxis, leftindexxaxis])
#
#         return sendresponse(result, 200)
#     except Exception as e:
#         print("error : " + str(e))
#         return sendresponse(str(e), 201)
#
#
# call this method to get details of a model
@app.route('/getmodeledtails', methods = ['POST'])
def getmodeledtails():
    try:
        modelname = request.form['model']
        # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
        #                       'Server=IN2371790W1\SQLEXPRESS;'
        #                       'Database=eyesight;'
        #                       'Trusted_Connection=yes;')
        conn = pyodbc.connect(databaseserver)
        cursor = conn.cursor()
        query = 'select * from models where model = \'' + modelname + '\';'
        # query = 'DROP TABLE dbo.' + modelname + ';'
        cursor.execute(query)
        rows = []
        for row in cursor:
            rows.append(row)
        rows = str(rows)
        conn.commit()
        conn.close()
        return sendresponse(rows, 200)
    except Exception as e:
        print("error : " + str(e))
        return sendresponse("cannot fetch model details", 201)


# call this method to get initial threshold and colorcut corrected image
@app.route('/initialimageprocess', methods = ['POST'])
def initialimageprocess():
    try:
        data = request.get_json()
        if data is None:
            print("No valid request body, json missing!")
            return jsonify({'error': 'No valid request body, json missing!'})
        else:
            img_data = data['thumbnail']
            threshold = int(data['threshold'])
            colorcut = int(data['colorcut'])
            erode = int(data['erode'])
            dilate = int(data['dilate'])
            corrstring = img_data + "=="
            indent = ""

            try:
                indent = data['indent']
                originalwidth = data['originalwidth']
                originalheight = data['originalheight']
            except Exception as e:
                print (e)


            # print (corrstring)

            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            # print(timestamp)

            # corrstring = re.sub(r'.*,', ',', corrstring)
            # corrstring = corrstring.replace(",", "")

            # stringToRGB(corrstring, local_save_cropped_file)
            img = imread(io.BytesIO(base64.b64decode(corrstring)))
            cv2_img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            final = enhancecroppedpicruntime(threshold, colorcut, cv2_img)
            # print (type(final))
            # final.show()

            local_save_input_initial_file = 'initialimage' + timestamp + '.jpg'
            final.save(local_save_input_initial_file)

            # correct skewness
            try:
                skewcorrected = skewcorrect(local_save_input_initial_file)
                skewcorrected.save(local_save_input_initial_file)
            except Exception as e:
                print (e)

            # try:
            #     orientationcorrected = orientationcorrect(local_save_input_initial_file)
            #     if (orientationcorrected == "done"):
            #         print ("orientation corrected")
            #     else:
            #         print("failed to correct orientation")
            # except Exception as e:
            #     print(e)

            # print("not-retrain")

            if (indent == "retrain"):
                dim = (originalwidth, originalheight)
                print (dim)
                try:
                    im = Image.open(local_save_input_initial_file)
                    im = im.resize(dim)
                    # local_save_input_initial_file = 'initialimage' + timestamp + '.jpg'
                    im.save(local_save_input_initial_file)
                except Exception as e:
                    print(e)

            withoutboxes = ""

            # removing lines
            try:
                # timestamp = time.time()
                # timestamp = str(timestamp).split('.')
                # timestamp = str(timestamp[0])
                # print(timestamp)
                # local_save_input_initial_file = 'initialimage' + timestamp + '.jpg'
                # final.save(local_save_input_initial_file)
                final = removelines2(local_save_input_initial_file)
                final.save(local_save_input_initial_file)
                withoutboxes = final
            except Exception as e:
                print (e)
                pass

            try:
                # # Load the image
                # img = cv2.imread('C:/Users/ET437GL/Documents/EYESIGHT/richa/images/newtel1.jpg')

                # convert to grayscale
                final = cv2.imread(local_save_input_initial_file)
                final = np.array(final)
                gray = cv2.cvtColor(final, cv2.COLOR_BGR2GRAY)

                # smooth the image to avoid noises
                gray = cv2.medianBlur(gray, 5)

                # Apply adaptive threshold
                thresh = cv2.adaptiveThreshold(gray, 255, 1, 1, 11, 2)
                thresh_color = cv2.cvtColor(thresh, cv2.COLOR_GRAY2BGR)

                # apply some dilation and erosion to join the gaps - change iteration to detect more or less area's
                thresh = cv2.dilate(thresh, None, iterations=dilate)
                thresh = cv2.erode(thresh, None, iterations=erode)

                # Find the contours
                contours, hierarchy = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                # size of the image
                dimensions = img.shape
                height = img.shape[0]
                width = img.shape[1]

                percentage = 0.5
                xthreshold = percentage * (width / 100)
                ythreshold = percentage * (height / 100)

                # For each contour, find the bounding rectangle and draw it
                boxes = []
                for cnt in contours:

                    xypadding = 5
                    hwpadding = 10
                    x, y, w, h = cv2.boundingRect(cnt)

                    if (w > xthreshold):
                        if (h > ythreshold):
                            x = x - xypadding
                            y = y - xypadding
                            w = w + hwpadding
                            h = h + hwpadding
                            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 1)
                            cv2.rectangle(final, (x, y), (x + w, y + h), (0, 255, 0), 1)
                            box = [x, y, x + w, y + h]
                            boxes.append(box)

                # Finally show the image
                print(boxes)
                # cv2.imshow('img', img)
                orig = Image.fromarray(final)
                final = orig
                # orig.save("C:/Users/ET437GL/Documents/EYESIGHT/richa/images/newtel 2.png")
                # cv2.imshow('res',thresh_color)
                # cv2.waitKey(0)
                # cv2.destroyAllWindows()
            except Exception as e:
                print (e)
                pass



            final = np.array(final)
            withoutboxes = np.array(withoutboxes)
            string = base64.b64encode(cv2.imencode('.jpg', final)[1]).decode()
            withoutboxesstring = base64.b64encode(cv2.imencode('.jpg', withoutboxes)[1]).decode()
            message = {"status": 200, "imagestring": withoutboxesstring, "withboxes": string}
            response = jsonify(message)
            response.status_code = 200
            return response
            # cv2.imshow('Received image', final)
            # cv2.waitKey(0)
            # return final
            # return sendresponse("ok", 200)

    except Exception as e:
        resultstring = 'failed to get text!!!' + str(e)
        print(resultstring)
        return sendresponse(resultstring, 201)


# call this method for full image OCR
@app.route('/fullocr', methods = ['POST'])
def fullocr():
    try:
        data = request.get_json()
        if data is None:
            print("No valid request body, json missing!")
            return jsonify({'error': 'No valid request body, json missing!'})
        else:
            img_data = data['thumbnail']
            threshold = int(data['threshold'])
            colorcut = int(data['colorcut'])
            inputformat = data['inputformat']
            pdftoimagefile = data['pdftoimagefile']
            corrstring = img_data + "=="
            print (img_data)
            print (pdftoimagefile)

            # print (corrstring)

            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            # print(timestamp)

            # corrstring = re.sub(r'.*,', ',', corrstring)
            # corrstring = corrstring.replace(",", "")

            # stringToRGB(corrstring, local_save_cropped_file)
            print("here")

            if (inputformat == "pdf"):
                img = cv2.imread(pdftoimagefile)
            else:
                img = imread(io.BytesIO(base64.b64decode(corrstring)))

            # try:
            #     img = imread(io.BytesIO(base64.b64decode(corrstring)))
            # except Exception as e:
            #     print (e)
            #
            # try:
            #     # img = base64.b64decode(img_data)
            #     img = stringToRGBnosave(img_data)
            # except Exception as e:
            #     print (e)

            print("here1")
            cv2_img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            print("here2")
            final = enhancecroppedpicruntime(threshold, colorcut, cv2_img)

            skewcorrected = skewcorrectforfullocr(final)
            print (skewcorrected)
            # extracted_text = pytesseract.image_to_string(final, lang='eng', config='digits')
            # extracted_text = pytesseract.image_to_string(final, lang='eng', config='--psm 11 --oem 3 -c tessedit_char_whitelist=0123456789QWERTYUIOPLKJHGFDSAZXCVBNMqwertyuioplkjhgfdsazxcvbnm~!@#$%^&*()_+=-{}|\][:;<>?/., ')
            extracted_text = pytesseract.image_to_string(skewcorrected, lang=language, config='--psm 1 --oem 3')
            print(extracted_text)

            extracted_text = extracted_text.replace('\n\n\n', '\n')
            extracted_text = extracted_text.replace('\n\n', '\n')

            return sendresponse(extracted_text, 200)

        # image1 = Image.open(local_save_cropped_file)
        # image1.show()
        # resultstring = tesserocr.image_to_text(image1, lang="eng")  # print ocr text from image
        # resultstring = resultstring.strip()
        # print(resultstring)
        # return sendresponse(resultstring, 200)
    except Exception as e:
        resultstring = 'failed to get text!!!' + str(e)
        print(resultstring)
        return sendresponse(resultstring, 201)


def fullocrbased(parameter, local_save_input_test_file, originalthreshold, originalcolorcut):
    regex = parameter[32]

    if (regex == "-"):
        inputimage = cv2.imread(local_save_input_test_file)
        inputimage = Image.fromarray(inputimage)
        processedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, inputimage)
        # processedimage.save(local_save_input_test_file)
        extracted_text = pytesseract.image_to_string(processedimage, lang=language, config='--psm 11')
        extracted_text = extracted_text.replace('"', '')
        extracted_text = extracted_text.replace('“', '')
        extracted_text = extracted_text.replace('\\ ', 'n')
        string = extracted_text.replace('   ', '')
    else:
        regex = regex.split("|")
        print(regex)
        inputimage = cv2.imread(local_save_input_test_file)
        inputimage = Image.fromarray(inputimage)
        processedimage = enhancecroppedpicruntime(originalthreshold, originalcolorcut, inputimage)
        # processedimage.save(local_save_input_test_file)
        extracted_text = pytesseract.image_to_string(processedimage, lang=language, config='--psm 11')
        extracted_text = extracted_text.replace('"', '')
        extracted_text = extracted_text.replace('“', '')
        extracted_text = extracted_text.replace('\\ ', 'n')
        extracted_text = extracted_text.replace('   ', '')
        extracted_text = extracted_text.replace("\n", " ")

        regexnature = regex[0]
        regexnaturealphabets = regex[1]
        textentered = regex[2]
        textentered = textentered.replace("'", "")

        if (regexnature == "4"):
            string = []
            words = extracted_text.split(" ")
            for word in words:
                print(word)
                exactword = word
                r = re.search('[a-zA-Z]', exactword)
                print(r)
                if (r != None):
                    r2 = re.search(r'\d', exactword)
                    if (r2 != None):
                        if (regexnaturealphabets == "1"):
                            if (len(exactword) >= int(textentered)):
                                string.append(exactword)
                        if (regexnaturealphabets == "3"):
                            if (len(exactword) == int(textentered)):
                                string.append(exactword)

            print(string)

        elif (regexnature == "2"):
            string = []
            words = extracted_text.split(" ")
            for word in words:
                print(word)
                exactword = word
                exactword = re.sub('[^0-9]','', exactword)
                r = re.search(r'\d', exactword)
                print(r)
                if (r != None):
                    if (regexnaturealphabets == "1"):
                        if (len(exactword) >= int(textentered)):
                            string.append(exactword)

            print(string)


    string = str(string)
    string = string.replace("[", "")
    string = string.replace("]", "")
    string = string.replace("'", "")
    return string



def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


# sending API response back to client
def sendresponse(message, statuscode):
    message1 = {"status": statuscode, "response": message}
    response = jsonify(message1)
    response.status_code = statuscode
    return response


# get image base64 and save it locally
def stringToRGB(base64_string, filename):
    img = imread(io.BytesIO(base64.b64decode(base64_string)))
    cv2_img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    cv2.imwrite(filename, cv2_img)


# get image base64 and save it locally
def stringToRGBnosave(base64_string):
    print("here4")
    # imgdata = base64.b64decode(str(base64_string))
    imgdata = Image.open(io.BytesIO(base64.b64decode(str(base64_string))))
    print("here5")
    # image = imread(io.BytesIO(imgdata))
    return imgdata



# get image and convert it into base64
def RGBTostring():
    img = cv2.imread(local_save_cropped_file)
    string = base64.b64encode(cv2.imencode('.jpg', img)[1]).decode()
    return string


# image processing steps
# 1st method to call
def trim(im):
    bg = Image.new(im.mode, im.size, im.getpixel((0,0)))
    diff = ImageChops.difference(im, bg)
    diff = ImageChops.add(diff, diff, 2.0, -100)
    bbox = diff.getbbox()
    if bbox:
        return im.crop(bbox)

#
# def gettextfromcroppedpic(local_save_input_test_file, modelname, textboxesl):
#     try:
#         extractedtext = []
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesight;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserver)
#         cursor = conn.cursor()
#         query = 'select * from ' + modelname + ';'
#         cursor.execute(query)
#         for row in cursor:
#             xaxis = row[1]
#             yaxis = row[2]
#             width = row[3]
#             height = row[4]
#             fullwidthl = row[8]
#             fullheightl = row[9]
#             label = row[10]
#             print(row)
#
#             im = Image.open(local_save_input_test_file)
#             fullwidth, fullheight = im.size
#
#             fullwidthl = int(fullwidthl)
#             fullheightl = int(fullheightl)
#
#             resizedimage = resizeimage(fullwidthl, fullheightl)
#             resizedimage = Image.fromarray(resizedimage)
#             try:
#                 resizedimage.save(local_save_input_test_file)
#             except Exception as e:
#                 if "RGBA" in str(e):
#                     resizedimage = resizedimage.convert("RGB")
#                     resizedimage.save(local_save_input_test_file)
#                     print(e)
#                 else:
#                     print(e)
#
#             left = int(xaxis)
#             top = int(yaxis)
#             right = int(width + xaxis)
#             bottom = int(height + yaxis)
#
#             textboxes = textboxesl
#
#             xalignedboxes = []
#
#             for textbox in textboxes:
#                 startx = int(textbox[0])
#                 diff = abs(startx - left)
#                 if (diff <= 5):
#                     xalignedboxes.append(textbox)
#
#             yalignedboxes = []
#
#             for xalignedbox in xalignedboxes:
#                 starty = int(xalignedbox[1])
#                 diff = abs(starty - top)
#                 if (diff <= 5):
#                     yalignedboxes.append(xalignedbox)
#
#             print("final box - " + str(yalignedboxes))
#
#             noofpoints = len(yalignedboxes)
#
#             if (noofpoints > 1):
#                 yaxislist = []
#                 for yaxis in yalignedboxes:
#                     i = yaxis[1]
#                     yaxislist.append(i)
#                 correct = min(yaxislist, key=lambda x: abs(x - top))
#                 correctindex = yaxislist.index(correct)
#                 correctlist = yalignedboxes[correctindex]
#             else:
#                 correctlist = yalignedboxes[0]
#
#
#             # leftbox = yalignedboxes[0]
#             left1 = correctlist[0]
#             top1 = correctlist[1]
#             right1 = width + left1
#             bottom1 = height + top1
#
#
#             im1 = resizedimage.crop((left1, top1, right1, bottom1))
#             # im1.show()
#             rgb = im1.convert('RGB')
#
#             try:
#                 resultstring = tesserocr.image_to_text(rgb, lang=language)  # print ocr text from image
#                 resultstring = resultstring.strip()
#                 print(label + "-" + resultstring)
#                 result = label + "-" + resultstring
#                 extractedtext.append(result)
#             except Exception as e:
#                 print(e)
#         print(extractedtext)
#         nullstring = " "
#         filledstring = nullstring.join(str(x) for x in extractedtext)
#         # return sendresponse(filledstring, 200)
#         return filledstring
#     except Exception as e:
#         return str(e) + "failed to extract text"
#
#
# def fetchmodeldetails(model):
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     query = 'select * from models where model = ' + '\'' + model + '\'' + ';'
#     cursor.execute(query)
#     modelmetadata = []
#     for row in cursor:
#         # modelmetadata.append(row)
#         for element in row:
#             modelmetadata.append(element)
#     print ("-------------")
#     print (modelmetadata)
#     print ("-------------")
#     return modelmetadata


def keybased(model, local_save_input_test_file, originaldilate, originalerode):
    inputimage = cv2.imread(local_save_input_test_file)
    inputimage = Image.fromarray(inputimage)
    processedimage = enhancecroppedpicruntime("-", "-", inputimage)
    processedimage.save(local_save_input_test_file)
    txtboxx = []
    # for model in modelsmetadata:
    model = [x for x in model]

    originaltextwidth = model[3]
    originaltextheight = model[4]
    originalimagesizewidth = model[8]
    originalimagesizeheight = model[9]
    keyval = model[26]
    kposition = model[24]

    kori = model[30]
    koriper = model[31]

    # uncomment these lines after adding these columns in db
    key_width = model[6]
    key_height = model[7]

    boxes = gettextboxes(local_save_input_test_file, originalimagesizewidth, originalimagesizeheight, originaldilate, originalerode)
    boxes.sort(key=lambda x: x[1])

    for [startX, startY, endX, endY] in boxes:
        cropped_image = cropimage(processedimage, int(startX), int(startY), int(startX) + int(key_width), int(startY) + int(key_height))
        # cropped_image.show()
        result_text = invoketesseract(cropped_image)
        similarityratio = similar(keyval, result_text)

        if similarityratio > 0.8:
            keybox = [startX, startY, endX, endY]
            keyxaxis = keybox[0]
            keyyaxis = keybox[1]

            if kposition == 'below':
                boxes.sort(key=lambda a: a[1])
                print("sorted boxes:")
                print(boxes)
                # print(sortedbox)
                index = boxes.index(keybox)
                print(index)

                # getting boxes below key value box
                boxesbelowkey = boxes[(index + 1):(len(boxes))]
                print("boxes below key - " + str(boxesbelowkey))

                yfiltered = []

                ypercentage = 2
                yvision = originalimagesizeheight / 100
                yvision = yvision * ypercentage
                print("vision - " + str(yvision))

                for textbox in boxesbelowkey:
                    txtyaxis = int(textbox[1])
                    diff = abs(txtyaxis - keyyaxis)
                    if (diff > yvision):
                        yfiltered.append(textbox)

                print("y-aligned boxes - " + str(yfiltered))

                # get all the boxes that are aligned to x axis of the key box
                xalignedboxes = []

                percentage = float(koriper)
                vision = originalimagesizewidth / 100
                vision = vision * percentage
                print("vision - " + str(vision))


                for textbox in yfiltered:
                    startx = int(textbox[0])
                    diff = abs(startx - keyxaxis)

                    if (diff <= vision):
                        xalignedboxes.append(textbox)
                print ("x-aligned boxes - " + str(xalignedboxes))

                xalignedboxes2 = []

                for textbox in xalignedboxes:
                    startx1 = int(textbox[0])
                    if (kori == "left"):
                        if (startx1 < keyxaxis):
                            xalignedboxes2.append(textbox)
                    if (kori == "right"):
                        if (startx1 > keyxaxis):
                            xalignedboxes2.append(textbox)
                    if (kori == "immediate"):
                        xalignedboxes2.append(textbox)


                # sort the x aligned boxed in ascending oreder of y axis
                xalignedboxes2.sort(key=lambda y: y[1])
                print (xalignedboxes2)
                
                # take first 3 boxes from x aligned boxes sorted in ascending order of y axis
                xalignedboxes2 = xalignedboxes2[:3]

                # sort again based on x axis
                xalignedboxes2.sort(key=lambda y: y[0])
                print(xalignedboxes2)


                zeroeddownbox1 = xalignedboxes2[0]
                yofzeroeddownbox1 = zeroeddownbox1[1]
                keydiff1 = abs(yofzeroeddownbox1 - keyyaxis)

                zeroeddownbox2 = xalignedboxes2[1]
                yofzeroeddownbox2 = zeroeddownbox2[1]
                keydiff2 = abs(yofzeroeddownbox2 - keyyaxis)

                zeroeddownbox3 = xalignedboxes2[2]
                yofzeroeddownbox3 = zeroeddownbox3[1]
                keydiff3 = abs(yofzeroeddownbox3 - keyyaxis)


                if (keydiff1 < keydiff2):
                    if (keydiff1 < keydiff3):
                        # get the first box in x aligned boxes, which will be the immediate below box to the key box
                        zeroeddownbox = xalignedboxes2[0]
                        print(zeroeddownbox)
                    else:
                        # get the first box in x aligned boxes, which will be the immediate below box to the key box
                        zeroeddownbox = xalignedboxes2[2]
                        print(zeroeddownbox)
                elif (keydiff1 > keydiff2):
                    if (keydiff2 < keydiff3):
                        # get the first box in x aligned boxes, which will be the immediate below box to the key box
                        zeroeddownbox = xalignedboxes2[1]
                        print(zeroeddownbox)
                    else:
                        # get the first box in x aligned boxes, which will be the immediate below box to the key box
                        zeroeddownbox = xalignedboxes2[2]
                        print(zeroeddownbox)
                    # get the first box in x aligned boxes, which will be the immediate below box to the key box
                    # zeroeddownbox = xalignedboxes2[1]
                    # print(zeroeddownbox)


                txtStartX = zeroeddownbox[0]
                txtStartY = zeroeddownbox[1]

                cropped_image = cropimage(processedimage, txtStartX, txtStartY, txtStartX + originaltextwidth, txtStartY + originaltextheight)
                # cropped_image.show()

                # result_text = invoketesseract(cropped_image)
                result_text = pytesseract.image_to_string(cropped_image, lang=language)
                # sendresponse(result_text, 200)
                print(result_text)
                # return (result_text)

            elif kposition == 'top':
                boxes.sort(key=lambda a: a[1])
                print("sorted boxes:")
                print(boxes)
                # print(sortedbox)
                index = boxes.index(keybox)
                print(index)

                # getting boxes above key value box
                boxesabovekey = boxes[0:(index)]
                print("boxes above key - " + str(boxesabovekey))

                samelinefiltered = []

                ypercentage = 2
                yvision = originalimagesizeheight / 100
                yvision = yvision * ypercentage
                print("vision - " + str(yvision))

                for textbox in boxesabovekey:
                    txtyaxis = int(textbox[1])
                    diff = abs(txtyaxis - keyyaxis)
                    if (diff > yvision):
                        samelinefiltered.append(textbox)

                print("y-aligned boxes - " + str(samelinefiltered))

                if (len(samelinefiltered) == 0):
                    result_text = "-"
                    break



                # writing new logic
                koriperyboxes = []
                percentage = float(koriper)
                vision = originalimagesizeheight / 100
                vision = vision * percentage
                print("vision for koriperyboxes - " + str(vision))
                for textbox in samelinefiltered:
                    starty = int(textbox[1])
                    diff = abs(starty - keyyaxis)
                    if (diff <= vision):
                        koriperyboxes.append(textbox)
                print("key orientation Y boxes - " + str(koriperyboxes))

                if (len(koriperyboxes) == 0):
                    result_text = "-"
                    break



                koriperxboxes = []
                percentage = float(koriper)
                vision = originalimagesizewidth / 100
                vision = vision * percentage
                print("vision for koriperxboxes - " + str(vision))
                for textbox in koriperyboxes:
                    startx = int(textbox[0])
                    diff = abs(startx - keyxaxis)
                    if (diff <= vision):
                        koriperxboxes.append(textbox)
                print("key orientation X boxes - " + str(koriperxboxes))

                if (len(koriperxboxes) == 0):
                    result_text = "-"
                    break



                # sort the x aligned boxed in ascending oreder of y axis
                koriperxboxes.sort(key=lambda y: y[1], reverse=True)
                print("Descending order sorted boxes - " + str(koriperxboxes))

                # take first 3 boxes from x aligned boxes sorted in ascending order of y axis
                if (len(koriperxboxes) > 2):
                    koriperxboxes = koriperxboxes[:3]

                xdifferences = []
                for boxes in koriperxboxes:
                    xvalue = boxes[0]
                    diff = abs(xvalue - keyxaxis)
                    xdifferences.append(diff)

                finalindex = xdifferences.index(min(xdifferences))
                finalbox = koriperxboxes[finalindex]

                txtStartX = finalbox[0]
                txtStartY = finalbox[1]

                cropped_image = cropimage(processedimage, txtStartX, txtStartY, txtStartX + originaltextwidth, txtStartY + originaltextheight)
                # cropped_image.show()

                result_text = pytesseract.image_to_string(cropped_image, lang=language)
                # sendresponse(result_text, 200)
                print("output - " + result_text)
                if (result_text == ""):
                    result_text = "-"
                break


            elif kposition == 'right':
                boxes.sort(key=lambda a: a[0])
                print("sorted boxes:")
                print(boxes)
                # print(sortedbox)
                index = boxes.index(keybox)
                print(index)

                # getting boxes to the right of key value box
                boxestotherightofkey = boxes[(index + 1):(len(boxes))]
                print("boxes to the right of key - " + str(boxestotherightofkey))

                keyrighttip = float(keyxaxis) + float(key_width)

                absoluteboxestotherightofkey = []

                for box in boxestotherightofkey:
                    if (box[0] > keyrighttip):
                        absoluteboxestotherightofkey.append(box)

                print (absoluteboxestotherightofkey)

                # xfiltered = []
                #
                # xpercentage = 2
                # xvision = originalimagesizewidth / 100
                # xvision = xvision * xpercentage
                # print("vision - " + str(xvision))
                #
                # for textbox in absoluteboxestotherightofkey:
                #     txtxaxis = int(textbox[0])
                #     diff = abs(txtxaxis - keyxaxis)
                #     if (diff > xvision):
                #         xfiltered.append(textbox)
                #
                # print("x-aligned boxes - " + str(xfiltered))

                # get all the boxes that are aligned to x axis of the key box
                xalignedboxes = []

                percentage = float(koriper)
                vision = originalimagesizewidth / 100
                vision = vision * percentage
                print("vision - " + str(vision))

                for textbox in absoluteboxestotherightofkey:
                    starty = int(textbox[1])
                    diff = abs(starty - int(keyyaxis))

                    if (diff <= vision):
                        xalignedboxes.append(textbox)
                print("x-aligned boxes - " + str(xalignedboxes))

                # # get all the boxes that are aligned to x axis of the key box
                # xalignedboxes3 = []
                #
                # percentage = float(koriper)
                # vision = originalimagesizewidth / 100
                # vision = vision * percentage
                # print("vision - " + str(vision))
                #
                # for textbox in xalignedboxes:
                #     startx = int(textbox[0])
                #     diff = abs(startx - (keyxaxis + int(key_width)))
                #
                #     if (diff <= vision):
                #         xalignedboxes3.append(textbox)
                # print("x-aligned boxes - " + str(xalignedboxes3))

                xalignedboxes2 = []

                for textbox in xalignedboxes:
                    startx1 = int(textbox[0])
                    if (kori == "left"):
                        if (startx1 < keyxaxis):
                            xalignedboxes2.append(textbox)
                    if (kori == "right"):
                        if (startx1 > keyxaxis):
                            xalignedboxes2.append(textbox)
                    if (kori == "immediate"):
                        xalignedboxes2.append(textbox)
                    if (kori == "auto"):
                        xalignedboxes2.append(textbox)



                # sort the x aligned boxed in ascending oreder of y axis
                xalignedboxes2.sort(key=lambda y: y[0])
                print(xalignedboxes2)



                # take first 3 boxes from x aligned boxes sorted in ascending order of y axis
                if (len(xalignedboxes2) > 2):
                    xalignedboxes2 = xalignedboxes2[:3]

                # sort again based on y axis
                # xalignedboxes2.sort(key=lambda y: y[1])
                # print(xalignedboxes2)


                ydifferences = []
                for boxes in xalignedboxes2:
                    xvalue = boxes[0]
                    diff = abs(xvalue - keyxaxis)
                    ydifferences.append(diff)

                finalindex = ydifferences.index(min(ydifferences))

                finalboxes = xalignedboxes2[finalindex]

                # try:
                #     zeroeddownbox1 = xalignedboxes2[0]
                #     yofzeroeddownbox1 = zeroeddownbox1[1]
                #     keydiff1 = abs(yofzeroeddownbox1 - keyyaxis)
                #
                #     zeroeddownbox2 = xalignedboxes2[1]
                #     yofzeroeddownbox2 = zeroeddownbox2[1]
                #     keydiff2 = abs(yofzeroeddownbox2 - keyyaxis)
                #
                #     zeroeddownbox3 = xalignedboxes2[2]
                #     yofzeroeddownbox3 = zeroeddownbox3[1]
                #     keydiff3 = abs(yofzeroeddownbox3 - keyyaxis)
                # except Exception as e:
                #     print("Error: " + e)
                #
                # if (keydiff1 < keydiff2):
                #     if (keydiff1 < keydiff3):
                #         # get the first box in x aligned boxes, which will be the immediate below box to the key box
                #         zeroeddownbox = xalignedboxes2[0]
                #         print(zeroeddownbox)
                #     else:
                #         # get the first box in x aligned boxes, which will be the immediate below box to the key box
                #         zeroeddownbox = xalignedboxes2[2]
                #         print(zeroeddownbox)
                # elif (keydiff1 > keydiff2):
                #     if (keydiff2 < keydiff3):
                #         # get the first box in x aligned boxes, which will be the immediate below box to the key box
                #         zeroeddownbox = xalignedboxes2[1]
                #         print(zeroeddownbox)
                #     else:
                #         # get the first box in x aligned boxes, which will be the immediate below box to the key box
                #         zeroeddownbox = xalignedboxes2[2]
                #         print(zeroeddownbox)
                #
                #
                #
                #
                #
                # finalboxes = []
                #
                # fpercentage = 5
                # fvision = originalimagesizewidth / 100
                # fvision = fvision * fpercentage
                # print("fvision - " + str(fvision))
                #
                # for textbox in xalignedboxes2:
                #     starty = int(textbox[1])
                #     diff = abs(starty - keyyaxis)
                #
                #     if (diff <= fvision):
                #         finalboxes.append(textbox)
                print("final box - " + str(finalboxes))

                    # get the first box in x aligned boxes, which will be the immediate below box to the key box
                    # zeroeddownbox = xalignedboxes2[1]
                    # print(zeroeddownbox)

                zeroeddownbox = finalboxes

                txtStartX = zeroeddownbox[0]
                txtStartY = zeroeddownbox[1]

                cropped_image = cropimage(processedimage, txtStartX, txtStartY, txtStartX + originaltextwidth, txtStartY + originaltextheight)
                # cropped_image.show()

                # result_text = invoketesseract(cropped_image)
                result_text = pytesseract.image_to_string(cropped_image, lang=language, config='--psm 11')
                # sendresponse(result_text, 200)
                print(result_text)
                if (result_text == ""):
                    result_text = "-"
                break

            elif kposition == 'left':
                boxes.sort(key=lambda lis: lis[0])
                index = boxes.index(keybox)
                value = sorted([boxes[i] if boxes[i][0] in range(boxes[index][0] - 0.05 * originalimagesizewidth,
                                                                 boxes[index][3] - 0.05 * originalimagesizewidth) or
                                            boxes[i][1] in range(boxes[index][1] - 0.05 * originalimagesizeheight,
                                                                 boxes[index][
                                                                     1] + 0.05 * originalimagesizeheight) else
                                boxes[index - 1] for i in range(0, index)], key=lambda x: x[0])
                [startX, startY, endX, endY] = value[0]
                cropped_image = cropimage(processedimage, startX, startY, startX + originaltextwidth,
                                          startY + originaltextheight)
                result_text = invoketesseract(cropped_image)
                # sendresponse(result_text, 200)
                print(result_text)
                # return (result_text)

            else:
                return "Value position missing"
                # sendresponse("Value position is missing", 201)
        else:
            print("Searching for the key - " + keyval)
            # sendresponse("cannot find value for this key", 201)

    return result_text


def removelines(local_save_input_test_file):
    try:
        image = cv2.imread(local_save_input_test_file)
        result = image.copy()
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

        # Remove horizontal lines
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (80, 2))
        remove_horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
        cnts = cv2.findContours(remove_horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = cnts[0] if len(cnts) == 2 else cnts[1]
        for c in cnts:
            cv2.drawContours(result, [c], -1, (255, 255, 255), 5)

        # Remove vertical lines
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 80))
        remove_vertical = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
        cnts = cv2.findContours(remove_vertical, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = cnts[0] if len(cnts) == 2 else cnts[1]
        for c in cnts:
            cv2.drawContours(result, [c], -1, (255, 255, 255), 5)

        # cv2.imshow('thresh', thresh)
        # cv2.imshow('result', result)
        # cv2.imwrite('result.png', result)
        result = Image.fromarray(result)
        result.save(local_save_input_test_file)
        # cv2.waitKey()
        return "done"
    except Exception as e:
        return "failed"


def removelines2(local_save_input_initial_file):
    try:
        image = cv2.imread(local_save_input_initial_file)
        result = image.copy()
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

        # Remove horizontal lines
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (80, 2))
        remove_horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
        cnts = cv2.findContours(remove_horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = cnts[0] if len(cnts) == 2 else cnts[1]
        for c in cnts:
            cv2.drawContours(result, [c], -1, (255, 255, 255), 5)

        # Remove vertical lines
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 80))
        remove_vertical = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
        cnts = cv2.findContours(remove_vertical, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = cnts[0] if len(cnts) == 2 else cnts[1]
        for c in cnts:
            cv2.drawContours(result, [c], -1, (255, 255, 255), 5)

        # cv2.imshow('thresh', thresh)
        # cv2.imshow('result', result)
        # cv2.imwrite('result.png', result)
        result = Image.fromarray(result)
        # result.save(local_save_input_initial_file)
        # cv2.waitKey()
        return result
    except Exception as e:
        return "failed"


def skewcorrect(local_save_input_initial_file):
    try:
        image = cv2.imread(local_save_input_initial_file)
        # Correcting colorcut and thresholding
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.bitwise_not(gray)
        thresh = cv2.threshold(gray, 0, 255,
                               cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        # Detecting angle of rotatio and adjusting the text lines
        coords = np.column_stack(np.where(thresh > 0))
        angle = cv2.minAreaRect(coords)[-1]

        if angle < -45:
            angle = -(90 + angle)

        else:
            angle = -angle

        (h, w) = image.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)

        # Outputing the corrected image
        # cv2.putText(rotated, "Angle: {:.2f} degrees".format(angle), (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
        #             (0, 0, 255), 2)
        print("[INFO] angle: {:.3f}".format(angle))
        # cv2.imshow("Input", image)
        # cv2.imshow("Rotated", rotated)
        # cv2.waitKey(0)
        return Image.fromarray(rotated)
    except Exception as e:
        return "failed"


def skewcorrectforfullocr(pilimg):
    try:
        print (type(pilimg))
        image = cv2.cvtColor(np.array(pilimg), cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.bitwise_not(gray)
        thresh = cv2.threshold(gray, 0, 255,
                               cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        # Detecting angle of rotatio and adjusting the text lines
        coords = np.column_stack(np.where(thresh > 0))
        angle = cv2.minAreaRect(coords)[-1]

        if angle < -45:
            angle = -(90 + angle)

        else:
            angle = -angle

        (h, w) = image.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)

        # Outputing the corrected image
        # cv2.putText(rotated, "Angle: {:.2f} degrees".format(angle), (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
        #             (0, 0, 255), 2)
        print("[INFO] angle: {:.3f}".format(angle))
        # cv2.imshow("Input", image)
        # cv2.imshow("Rotated", rotated)
        # cv2.waitKey(0)
        return Image.fromarray(rotated)
    except Exception as e:
        return e


def orientationcorrect(local_save_input_initial_file):
    imPath = local_save_input_initial_file
    im = cv2.imread(str(imPath), cv2.IMREAD_COLOR)
    newdata = pytesseract.image_to_osd(im)
    rotation = re.search('(?<=Orientation in degrees: )\d+', newdata).group(0)
    print (newdata)
    if (int(rotation) > 0):
        angle = 360 - int(re.search('(?<=Rotate: )\d+', pytesseract.image_to_osd(im)).group(0))
        print(angle)
        (h, w) = im.shape[:2]
        center = (w / 2, h / 2)
        scale = 1.0
        # Perform the rotation
        M = cv2.getRotationMatrix2D(center, angle, scale)
        rotated = cv2.warpAffine(im, M, (w, h))
        orig = Image.fromarray(rotated)
        # cv2.imshow('res',thresh_color)
        # cv2.waitKey(0)
        orig.save(local_save_input_initial_file)
        return "done"
    else:
        return "done"



# pdf logics
# call this method for initial saving pdf
@app.route('/receiveinitialpdf', methods = ['POST'])
def receiveinitialpdf():
    try:
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            local_save_input_initial_pdf = 'initialpdf' + timestamp + '.pdf'

            file = request.files['file']
            file.save('static/initialreceivepdf/' + local_save_input_initial_pdf)
            print("file received")
            return local_save_input_initial_pdf
    except Exception as e:
        print (e)
        return sendresponse("failed", 200)


# call this method for table extraction from pdf
# @app.route('/processpdf', methods = ['POST'])
# def processpdf():
#
#     try:
#         filename = request.form['filename']
#         # filename = 'initialreceivepdf/' + filename
#         inputpages = request.form['pages']
#         flavor = request.form['flavor']
#         processbglines = request.form['processbglines']
#         cuttext = request.form['cuttext']
#         superscripts = request.form['superscripts']
#         smalllines = int(request.form['smalllines'])
#         grouprows = int(request.form['grouprows'])
#         groupcolumns = int(request.form['groupcolumns'])
#         edgetol = int(request.form['edgetol'])
#         tableposition = request.form['tableposition']
#
#         tablepositionlist = []
#         tablepositionlist.append(tableposition)
#
#         layout, dim = utils.get_page_layout('static/initialreceivepdf/' + filename)
#         print (layout.width)
#         print (layout.height)
#         dim = (int(layout.width), int(layout.height))
#
#
#
#         #delete file if already present
#         try:
#             deletepath = "static/eyesightpdfvision/" + filename[:filename.index(".")]
#             deletepath2 = "static/outputtables/" + filename + ".html"
#             # os.rmdir(deletepath)
#             os.remove(deletepath2)
#             del_dir(deletepath)
#             # del_dir()
#
#         except Exception as e:
#             print("no folder found - " + str(e))
#
#         pd.set_option('display.max_colwidth', -1)
#
#         foldername = filename[:filename.index(".")]
#         print("foldername")
#         print(foldername)
#         print(filename)
#         path = root_path + '/static/eyesightpdfvision/' + foldername
#         os.mkdir(path)
#
#         pages = convert_from_path('static/initialreceivepdf/' + filename, 200)
#
#         i=1
#
#
#         print(pages)
#
#         for page in pages:
#             imgpath = path + '/' + str(i) + '.jpg'
#
#             page.save(imgpath, 'JPEG')
#             img = cv2.imread(imgpath, cv2.IMREAD_UNCHANGED)
#             resized = cv2.resize(img, dim, interpolation=cv2.INTER_AREA)
#             print('Resized Dimensions : ', resized.shape)
#             cv2.imwrite(imgpath,resized)
#
#             i = i+1
#
#         print("here")
#
#         if (flavor == "lattice"):
#             if (tableposition == "['-,-,-,-']"):
#                 table_list = camelot.read_pdf('static/initialreceivepdf/' + filename, flavor=flavor, strip=' \n', pages=inputpages, process_background=processbglines, split_text=cuttext, flag_size=superscripts, line_scale=smalllines)
#             else:
#                 table_list = camelot.read_pdf('static/initialreceivepdf/' + filename, flavor=flavor, strip=' \n', pages=inputpages, process_background=processbglines, split_text=cuttext, flag_size=superscripts, table_areas=tablepositionlist, line_scale=smalllines)
#         elif (flavor == "stream"):
#             if (tableposition == "['-,-,-,-']"):
#                 table_list = camelot.read_pdf('static/initialreceivepdf/' + filename, flavor=flavor, strip=' \n', pages=inputpages, split_text=cuttext, flag_size=superscripts, edge_tol=edgetol, row_tol=grouprows)
#             else:
#                 table_list = camelot.read_pdf('static/initialreceivepdf/' + filename, flavor=flavor, strip=' \n', pages=inputpages, split_text=cuttext, flag_size=superscripts, table_areas=tablepositionlist, edge_tol=edgetol, row_tol=grouprows)
#
#
#         nooftables = len(table_list)
#
#         print ("nooftables")
#         print (nooftables)
#
#         fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"
#
#         for x in range(nooftables):
#
#             # camelot.plot(table_list[x], kind='text').show()
#             # f = plt.gcf()
#             # size = f.get_size_inches()
#             # f.set_size_inches(size[0] * 4, size[1] * 4, forward=True)
#             # plot = camelot.plot(table_list[x], kind='text').show()
#
#             # foldername = filename[:filename.index(".")]
#             #
#             # pathtest = 'C:/Users/ET437GL/Documents/ASpace/updated/model_v1/model_v1/Image-Border/ImageBorder/eyesightpdfvision/test' + foldername
#             # os.mkdir(pathtest)
#             # plot.savefig(pathtest + "/" + str(x) + ".png")
#
#             df = table_list[x].df
#             df.rename(columns=df.iloc[0]).drop(df.index[0])
#             if (x == 0):
#                 html = df.to_html()
#                 fullhtml += html
#             else:
#                 tabletitle = "<h1>Table " + str(x-1) + "</h1>"
#                 # tabletitle = "<h1>Extracted Table</h1>"
#                 fullhtml += tabletitle
#                 html = df.to_html()
#                 fullhtml += html
#
#         print(fullhtml)
#
#         # return render_template('simple.html',  tables=[df.to_html(classes='data')], titles=df.columns.values)
#         with open('static/outputtables/' + filename + '.html', 'w', encoding='utf-8') as file:
#             file.write(fullhtml)
#
#         return fullhtml
#     except Exception as e:
#         print ("no file found" + str(e))
#         return "failed"
#

# call this method to create a model without table areas
# @app.route('/createpdfmodel', methods = ['POST'])
# def createpdfmodel():
#     pages = request.form['pages']
#     flavortxt = request.form['flavortxt']
#     bglinestxt = request.form['bglinestxt']
#     cuttexttxt = request.form['cuttexttxt']
#     superscripttxt = request.form['superscripttxt']
#     smalllines = request.form['smalllines']
#     grouprows = request.form['grouprows']
#     groupcolumns = request.form['groupcolumns']
#     edgetol = request.form['edgetol']
#     tables = request.form['tables']
#     modelname = request.form['modelname']
#     description = request.form['description']
#
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesightpdfs;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserverpdf)
#     cursor = conn.cursor()
#     query = "CREATE TABLE " + modelname + " (sno int identity primary key, pages varchar(255), flavortxt varchar(255), bglinestxt varchar(255), " \
#                                           "cuttexttxt varchar(255), superscripttxt varchar(255), smalllines varchar(255), grouprows varchar(255), " \
#                                           "groupcolumns varchar(255), edgetol varchar(255), tables varchar(255)); "
#     print(query)
#     cursor.execute(query)
#
#     query2 = 'INSERT INTO ' + modelname + '(pages, flavortxt, bglinestxt, cuttexttxt, superscripttxt, smalllines, ' \
#                                                       'grouprows, groupcolumns, edgetol, tables) VALUES (' + '\'' + pages + '\', \'' + flavortxt + '\', \'' + bglinestxt + '\', \'' + cuttexttxt + '\', \'' + superscripttxt + '\', \'' + smalllines + '\', \'' + grouprows + '\', \'' + groupcolumns + '\', \'' + edgetol + '\', \'' + tables + '\'' + ');'
#
#     print(query2)
#     cursor.execute(query2)
#
#     query3 = 'INSERT INTO pdfmodels (model, descr) VALUES (' + '\'' + modelname + '\', \'' + description + '\'' + ');'
#
#     print(query3)
#     cursor.execute(query3)
#
#     conn.commit()
#     return sendresponse("Model created", 200)
#

# @app.route('/fetchpdfmodels', methods = ['POST'])
# def fetchpdfmodels():
#     querystring1 = "select model, descr from pdfmodels;"
#     models = []
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesight;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserverpdf)
#     cursor = conn.cursor()
#     cursor.execute(querystring1)
#     rows = cursor.fetchall()
#     print('Total Row(s):', cursor.rowcount)
#     for row in rows:
#         # models.append([x for x in row])
#         models.append(list(row))
#         # models.append(string)
#         print(type(row))
#         print(row)
#     print(models)
#     modelsjson = json.dumps(models)
#     return modelsjson
#
#     # querystring = "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE' AND TABLE_CATALOG='eyesightpdfs'"
#     # models = []
#     # # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     # #                       'Server=IN2371790W1\SQLEXPRESS;'
#     # #                       'Database=eyesightpdfs;'
#     # #                       'Trusted_Connection=yes;')
#     # conn = pyodbc.connect(databaseserverpdf)
#     # cursor = conn.cursor()
#     # cursor.execute(querystring)
#     # rows = cursor.fetchall()
#     # print('Total Row(s):', cursor.rowcount)
#     # for row in rows:
#     #     string = str(row)
#     #     string = string.replace("(", "")
#     #     string = string.replace(")", "")
#     #     string = string.replace(" ", "")
#     #     string = string.replace(",", "")
#     #     string = string.replace("'", "")
#     #     models.append(string)
#     #     print(string)
#     # modelsjson = json.dumps(models)
#     # return modelsjson
#
#
# @app.route('/trainpdfmodel', methods = ['POST'])
# def trainpdfmodel():
#     if request.is_json:
#         try:
#             modelname = request.headers["modelname"]
#
#             conn = pyodbc.connect(databaseserverpdf)
#             cursor = conn.cursor()
#
#             query = "CREATE TABLE " + modelname + " (sno int identity primary key, pages varchar(255), flavortxt varchar(255), bglinestxt varchar(255), " \
#                                                   "cuttexttxt varchar(255), superscripttxt varchar(255), smalllines varchar(255), grouprows varchar(255), " \
#                                                   "groupcolumns varchar(255), edgetol varchar(255), tables varchar(255), tablename varchar(255)); "
#
#             print(query)
#             cursor.execute(query)
#             conn.commit()
#
#             content = request.get_json()
#             jtopy = json.dumps(content)
#             json_dictionary = json.loads(jtopy)
#             for key in json_dictionary:
#                 indi = json_dictionary[key]
#                 page = indi["page"]
#                 flavor = indi["flavor"]
#                 bglines = indi["bglines"]
#                 cuttext = indi["cuttext"]
#                 superscript = indi["superscript"]
#                 smalllines = indi["smalllines"]
#                 grouprow = indi["grouprow"]
#                 groupcol = indi["groupcol"]
#                 edgetol = indi["edgetol"]
#                 tablearea = indi["tablearea"]
#                 tablename = indi["tablename"]
#
#                 cursor1 = conn.cursor()
#
#                 query1 = 'INSERT INTO ' + modelname + ' (pages, flavortxt, bglinestxt, cuttexttxt, superscripttxt, smalllines, grouprows, ' \
#                                                                   'groupcolumns, edgetol, tables, tablename) VALUES (' + '\'' + page + ' \', \'' + flavor + ' \', \'' + bglines + '\', \'' + cuttext + '\', \'' + superscript + '\' , \'' + smalllines + '\'' + ',\'' + grouprow + '\',\'' + groupcol + '\',\'' + edgetol + '\',\'' + tablearea + '\',\'' + tablename + '\');'
#
#                 print(query1)
#                 cursor1.execute(query1)
#                 conn.commit()
#             return sendresponse("Model trained", 200)
#         except Exception as e:
#             return sendresponse("Failed to train model : " + e, 200)
#
#     else:
#         return sendresponse("Failed to train model", 201)
#
#
# # call this method with a model name to delete it in DB
# @app.route('/deletepdfmodel', methods = ['POST'])
# def deletepdfmodel():
#     try:
#         modelname = request.form['model']
#         # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#         #                       'Server=IN2371790W1\SQLEXPRESS;'
#         #                       'Database=eyesightpdfs;'
#         #                       'Trusted_Connection=yes;')
#         conn = pyodbc.connect(databaseserverpdf)
#         cursor = conn.cursor()
#         # query = 'select * from eyesight.dbo.' + modelname + ';'
#         query = 'DROP TABLE ' + modelname + ';'
#         # query1 = 'DELETE from dbo.models WHERE model = ' + '\'' + modelname + '\'' +';'
#         # print (query1)
#         query1 = 'DELETE from pdfmodels WHERE model = ' + '\'' + modelname + '\'' +';'
#         cursor.execute(query)
#         cursor.execute(query1)
#         # cursor.execute(query1)
#         conn.commit()
#         conn.close()
#         return sendresponse("model deleted", 200)
#     except Exception as e:
#         print("error : " + str(e))
#         return sendresponse("cannot delete model", 201)


@app.route('/receivepdf', methods = ['POST'])
def receivepdf():
    try:
        modelname = request.form['model']
        modelname = modelname.strip()
        outputformat = request.form['outputformat']
        print("model being consumed - " + modelname)
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            local_save_input_test_pdf = 'testpdf' + timestamp + '.pdf'

            finaltext = []
            file = request.files['file']
            file.save(local_save_input_test_pdf)
            print("file received")

            # print("here0")
            # 1st step - fetch model information from db
            modelsmetadata = fetchpdfmodelinfo(modelname)
            # print("here3")
            # alltablehtml = '<div style="margin: 10px;"></div><center><input type="button" id="downloadbtn" class="btn-send" value="Download" onclick="downloaddata()" align="center" /></center>'

            # output_zip_path = 'file:///C:/Users/ET437GL/Documents/ASpace/updated/model_v1/model_v1/Image-Border/ImageBorder/static/finaloutputtables/' + local_save_input_test_pdf[:local_save_input_test_pdf.index(".")] + '/' + 'eyesight_output.zip'
            output_zip_path = '../finaloutputtables/' + local_save_input_test_pdf[:local_save_input_test_pdf.index(".")] + '/' + 'eyesight_output.zip'

            # output_zip_path_without_static = 'finaloutputtables/' + local_save_input_test_pdf[:local_save_input_test_pdf.index(".")] + '/' + 'eyesight_output.zip'

            # href = "{{url_for('static', filename = output_zip_path_without_static)}}"
            # href = "static/" + output_zip_path_without_static

            alltablehtml = '<head><link rel=''stylesheet'' href=''https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css''></head><body><div style="margin: 10px;"></div><center><a href="' + output_zip_path + '" download>Download</a></center></body>'
            print ("here")
            # alltablehtml = '<div style="margin: 10px;"></div><center><a href="{{url_for(\'static\', filename = ' + output_zip_path_without_static + ')}}">Download</a></center>'

            # print("here1")

            for parameter in modelsmetadata:
                print(str(parameter))
                label = parameter[10]

                if parameter[10] == '-':
                    inputpages = parameter[1]
                    flavor = parameter[2]
                    processbglines = parameter[3]
                    cuttext = parameter[4]
                    superscripts = parameter[5]
                    smalllines = int(parameter[6])
                    grouprows = int(parameter[7])
                    groupcolumns = int(parameter[8])
                    edgetol = int(parameter[9])
                    tableposition = parameter[10]

                    if (flavor == "lattice"):
                        table_list = camelot.read_pdf(local_save_input_test_pdf, flavor=flavor, strip=' \n',
                                                      pages=inputpages,
                                                      process_background=processbglines, split_text=cuttext,
                                                      flag_size=superscripts, line_scale=smalllines)
                    elif (flavor == "stream"):
                        table_list = camelot.read_pdf(local_save_input_test_pdf, flavor=flavor, strip=' \n',
                                                      pages=inputpages,
                                                      split_text=cuttext, flag_size=superscripts, edge_tol=edgetol,
                                                      row_tol=grouprows)

                    nooftables = len(table_list)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                    for x in range(nooftables):
                        df = table_list[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        if (x == 0):
                            html = df.to_html()
                            fullhtml += html
                        else:
                            tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                            # tabletitle = "<h1>Extracted Table</h1>"
                            fullhtml += tabletitle
                            html = df.to_html()
                            fullhtml += html

                    print(fullhtml)

                    # return render_template('simple.html',  tables=[df.to_html(classes='data')], titles=df.columns.values)
                    with open('static/runtimeoutputtables/' + local_save_input_test_pdf + '.html', 'w', encoding='utf-8') as file:
                        file.write(fullhtml)

                    return local_save_input_test_pdf

                else:
                    # print ("table space given")
                    inputpages = parameter[1]
                    flavor = parameter[2].replace(" ", "")
                    processbglines = parameter[3]
                    cuttext = parameter[4]
                    superscripts = parameter[5]
                    smalllines = int(parameter[6])
                    grouprows = int(parameter[7])
                    groupcolumns = int(parameter[8])
                    edgetol = int(parameter[9])
                    tableposition = parameter[10]
                    tablename = parameter[11]

                    tablepositionlist = []
                    tablepositionlist.append(tableposition)

                    if (flavor == "lattice"):
                        table_list = camelot.read_pdf(local_save_input_test_pdf, flavor=flavor, strip=' \n', pages=inputpages,
                                                      process_background=processbglines, split_text=cuttext,
                                                      flag_size=superscripts, table_areas=tablepositionlist,
                                                      line_scale=smalllines)

                    elif (flavor == "stream"):
                        table_list = camelot.read_pdf(local_save_input_test_pdf, flavor=flavor, strip=' \n', pages=inputpages,
                                                      split_text=cuttext, flag_size=superscripts,
                                                      table_areas=tablepositionlist, edge_tol=edgetol,
                                                      row_tol=grouprows)

                    foldername = local_save_input_test_pdf[:local_save_input_test_pdf.index(".")]
                    path = root_path + 'static/finaloutputtables/' + foldername

                    try:
                        os.mkdir(path)
                    except Exception as e:
                        print (e)

                    csvname = root_path + "static/finaloutputtables/" + foldername + "/" + tablename + "." + outputformat

                    if (outputformat == "xlsx"):
                        table_list[0].to_excel(csvname)
                    else:
                        table_list.export(csvname, f=outputformat, compress=False)  # json, excel, html

                    # nooftables = len(table_list)

                    # fullhtml = "<body background='http://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"
                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"
                    df = table_list[0].df
                    df.rename(columns=df.iloc[0]).drop(df.index[0])
                    html = df.to_html()
                    fullhtml += html

                    alltablehtml += fullhtml

                    print(fullhtml)

            # print("here2")

            # json processing
            pdf = pdfquery.PDFQuery(local_save_input_test_pdf)
            pdf.load()

            model_json_file = 'pdfmodels/' + modelname + '.json'
            with open(model_json_file) as f:
                modeljson = json.load(f)

            values = modeljson['values']

            output_dict = {}

            for i in range(len(values)):
                keyword1 = values[i]
                keyword1_json = json.dumps(keyword1)

                values1 = json.loads(keyword1_json)
                parameter_group = values1['keyword']

                parameters = json.dumps(parameter_group)
                parameters = json.loads(parameters)

                captured_value = parameters['capturedvalue']
                keyword = re.escape(str(parameters['keyword']))
                position = parameters['position']
                width = int(parameters['width'])
                height = int(parameters['height'])

                if (position == "Bottom"):
                    parameter = 'LTTextLineHorizontal:contains(' + keyword + ')'
                    print(parameter)
                    label = pdf.pq(parameter)
                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))
                    captured_value = pdf.pq(
                        'LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0, y0 - height, x1, y1 - 21)).text()
                    print(captured_value)
                    captured_value = captured_value.replace(":", "")
                    captured_value = captured_value.replace(",", "")
                    output_dict[keyword.replace("\\", "")] = captured_value
                elif (position == "Right"):
                    parameter = 'LTTextLineHorizontal:contains(' + keyword + ')'
                    label = pdf.pq(parameter)
                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))
                    captured_value = pdf.pq(
                        'LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0 + width, y0, x1 + width, y1)).text()
                    print(captured_value)
                    captured_value = captured_value.replace(":", "")
                    captured_value = captured_value.replace(",", "")
                    output_dict[keyword.replace("\\", "")] = captured_value
                elif (position == "Left"):
                    parameter = 'LTTextLineHorizontal:contains(' + keyword + ')'
                    label = pdf.pq(parameter)
                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))
                    captured_value = pdf.pq(
                        'LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0 - width, y0, x1 - width, y1)).text()
                    print(captured_value)
                    captured_value = captured_value.replace(":", "")
                    captured_value = captured_value.replace(",", "")
                    output_dict[keyword.replace("\\", "")] = captured_value
                elif (position == "Top"):
                    parameter = 'LTTextLineHorizontal:contains(' + keyword + ')'
                    label = pdf.pq(parameter)
                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))
                    captured_value = pdf.pq(
                        'LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0, y0 + height, x1, y1 + height)).text()
                    print(captured_value)
                    captured_value = captured_value.replace(":", "")
                    captured_value = captured_value.replace(",", "")
                    output_dict[keyword.replace("\\", "")] = captured_value

            print("Updated Dict is: ", output_dict)

            with open('static/finaloutputtables/' + local_save_input_test_pdf[
                                                    :local_save_input_test_pdf.index(".")] + '/capturedvalues.csv',
                      'w') as f:
                for key in output_dict.keys():
                    f.write("%s,%s\n" % (key, output_dict[key]))

            csv_to_html_table = pd.read_csv('static/finaloutputtables/' + local_save_input_test_pdf[
                                                                          :local_save_input_test_pdf.index(
                                                                              ".")] + '/capturedvalues.csv')
            html_table = csv_to_html_table.to_html()
            # fullhtml = "<h1>Captured values</h1>"
            # fullhtml = "<body background='http://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured values</h1></body>"
            fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Values</h1></body>"
            fullhtml += html_table
            alltablehtml += fullhtml

            with open('static/runtimeoutputtables/' + local_save_input_test_pdf[:local_save_input_test_pdf.index(".")] + '.html', 'w',
                      encoding='utf-8') as file:
                file.write(alltablehtml)

            output_zip = zipfile.ZipFile('static/finaloutputtables/' + local_save_input_test_pdf[:local_save_input_test_pdf.index(".")] + '/' + 'eyesight_output.zip', 'w')
            for folder, subfolders, files in os.walk('static/finaloutputtables/' + local_save_input_test_pdf[:local_save_input_test_pdf.index(".")]):
                for file in files:
                    if not (file.endswith('.zip')):
                        output_zip.write(os.path.join(folder, file),
                                          os.path.relpath(os.path.join(folder, file), 'static/finaloutputtables/' + local_save_input_test_pdf[:local_save_input_test_pdf.index(".")]),
                                          compress_type=zipfile.ZIP_DEFLATED)

            output_zip.close()

            return local_save_input_test_pdf[:local_save_input_test_pdf.index(".")]

    except Exception as e:
        resultstring = 'Something wrong - ' + str(e)
        return sendresponse(resultstring, 201)


# pass a model name to this method to get model meta data in list of lists format
# def fetchpdfmodelinfo(modelname):
#     # conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
#     #                       'Server=IN2371790W1\SQLEXPRESS;'
#     #                       'Database=eyesightpdfs;'
#     #                       'Trusted_Connection=yes;')
#     conn = pyodbc.connect(databaseserverpdf)
#     cursor = conn.cursor()
#     query = 'select * from ' + modelname + ';'
#     cursor.execute(query)
#     modelmetadata = []
#     for row in cursor:
#         modelmetadata.append(row)
#     return modelmetadata
#

# call this method to get initial pdf key value try
@app.route('/trycurrentvalue', methods = ['POST'])
def trycurrentvalue():
    filenameinserver = request.form['filenameinserver']
    keyword = request.form['keyword']
    position = request.form['position']
    width = int(request.form['width'])
    height = int(request.form['height'])

    pdf = pdfquery.PDFQuery('static/initialreceivepdf/' + filenameinserver)
    pdf.load()

    if (position == "Right"):
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        # foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0+width, y0+height, x1+width, y1+height)).text()
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
            x1, y0 - height, x1 + width, y1 + height)).text()
        foundtext = foundtext.replace(keyword, "")
        print("text - " + foundtext)
    elif (position == "Left"):
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0+width, y0+height, x1+width, y1+height)).text()
        print("text - " + foundtext)
    elif (position == "Top"):
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0 + width, y0 + height, x1 + width, y1 + height)).text()
        print("text - " + foundtext)
    elif (position == "Bottom"):
        print ("here")
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        print (label.attr('x0'))
        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        # foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0 - width, y0 - height, x1 + width, y1 - 20)).text()
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
        x0 - width, y0 - height, x1 + width, y1 - (y1 - y0))).text()
        foundtext = foundtext.replace(keyword, "")
        print("text - " + foundtext)

    print (foundtext)
    return foundtext


# call this method to create a pdf model using json
@app.route('/createpdfmodelfromjson', methods = ['POST'])
def createpdfmodelfromjson():
    modeljson = request.form['modeljson']
    title = request.form['title']
    title = title.strip()
    print (json.dumps(modeljson))
    with open('pdfmodels/' + title +'.json', 'w') as json_file:
        # json.dump(modeljson, json_file)
        json_file.write(modeljson)
    return "done"


# @app.route('/updatepdfmodelstable', methods = ['POST'])
# def updatepdfmodelstable():
#     if request.is_json:
#         modelname = request.headers["modelname"]
#         description = request.headers["descr"]
#
#         conn = pyodbc.connect(databaseserverpdf)
#         cursor = conn.cursor()
#         # modelquery = 'INSERT INTO pdfmodels(model, descr) VALUES (' + modelname + ', ' + description + '');
#         modelquery = 'INSERT INTO pdfmodels (model, descr) VALUES (' + '\'' + modelname + ' \', \'' + description + '\');'
#
#         print(modelquery)
#         cursor.execute(modelquery)
#         conn.commit()
#         return sendresponse("Models table populated", 200)
#
#     else:
#         return sendresponse("Failed to train model", 201)


@app.route('/fetchmodelsforclassifiers', methods = ['POST'])
def fetchmodelsforclassifiers():
    querystring1 = "select model from pdfmodels;"
    querystring2 = "select model from models;"

    models = []

    conn = pyodbc.connect(databaseserverpdf)
    cursor = conn.cursor()
    cursor.execute(querystring1)
    rows = cursor.fetchall()
    print('Total Row(s):', cursor.rowcount)
    for row in rows:
        # models.append([x for x in row])
        models.append(list(row))
        # models.append(string)
        print(type(row))
        print(row)
    print(models)
    # modelsjson = json.dumps(models)
    # return modelsjson


    conn1 = pyodbc.connect(databaseserver)
    cursor1 = conn1.cursor()
    cursor1.execute(querystring2)
    rows = cursor1.fetchall()
    print('Total Row(s):', cursor.rowcount)
    for row in rows:
        # models.append([x for x in row])
        models.append(list(row))
        # models.append(string)
        print(type(row))
        print(row)
    print(models)
    modelsjson = json.dumps(models)
    return modelsjson


@app.route('/createclassifier', methods = ['POST'])
def createclassifier():
    try:
        classifiername = request.headers["classifiername"]
        classifierjson = request.headers["classifierjson"]
        classifiername = classifiername.strip()

        with open('classifiers/' + classifiername + '.json', 'w') as json_file:
            json_file.write(classifierjson)

        return sendresponse("Classifier created", 200)

    except Exception as e:
        print (e)
        return sendresponse("Failed to create classifier : " + str(e), 200)


@app.route('/fetchclassifiers', methods = ['POST'])
def fetchclassifiers():

    path = root_path + "classifiers/"
    classifiers = [f for f in listdir(path) if isfile(join(path, f))]
    print (classifiers)
    return str(classifiers)


@app.route('/receiveimageforclassifier', methods = ['POST'])
def receiveimageforclassifier():
    try:
        classifiername = request.form['classifier']
        selectedfilename = request.form['selectedfilename']
        print("classifier being tested - " + classifiername)
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            local_save_input_classifier_file = 'classifiertestimage' + timestamp + '.jpg'

            file = request.files['file']
            file.save(local_save_input_classifier_file)
            print("file received")

            # run fullpage ocr and get text
            img = imread(local_save_input_classifier_file)
            cv2_img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
            final = enhancecroppedpicruntime(120, 120, cv2_img)

            extracted_text = pytesseract.image_to_string(final, config='--psm 1 --oem 3')
            extracted_text = extracted_text.replace('\n', ' ')
            # print(extracted_text)

            # fetching classifier json file
            f = open ('classifiers/' + classifiername + '.json',)
            classifier = json.load(f)
            print(classifier)

            modelnamelist = []
            modelscorelist = []

            for modelname in classifier:
                modelnamelist.append(modelname)
                keywords = classifier[modelname]
                print (modelname)
                keywords = keywords.split(",")
                print(keywords)

                noofkeys = len(keywords)
                perkeyweightage = 100/noofkeys

                modelscore = 0

                for keyword in keywords:
                    if keyword in extracted_text:
                        print (keyword + " Found!")
                        modelscore = modelscore + perkeyweightage
                    else:
                        print(keyword + " Not found!")

                modelscorelist.append(modelscore)

            print (modelnamelist)
            print (modelscorelist)

            largestscore = max(modelscorelist)
            largestscoreindex = modelscorelist.index(largestscore)

            print(largestscore)
            print(largestscoreindex)

            finalresponsejson = {}

            if (int(largestscore) == 0):
                finalmodelname = "unidentified"
            else:
                finalmodelname = modelnamelist[int(largestscoreindex)]

            finalresponsejson[selectedfilename] = finalmodelname

            f.close()

            return sendresponse(str(finalresponsejson), 200)

    except Exception as e:
        resultstring = 'Something wrong - ' + str(e)
        return sendresponse(resultstring, 201)


@app.route('/fetchmodelsfromclassifiers', methods = ['POST'])
def fetchmodelsfromclassifiers():
    selectedclassifiers = request.form['selectedclassifiers']

    selectedclassifiers = selectedclassifiers.split(",")
    # print(type(selectedclassifiers))

    modelslist = []

    for selectedclassifier in selectedclassifiers:

        f = open( root_path + 'classifiers/' + selectedclassifier + '.json', )
        data = json.load(f)
        modellist = []
        for i in data:
            print(i)
            modellist.append(i)
        modelslist.extend(modellist)
        f.close()

    print (modelslist)

    return str(modelslist)


# @app.route('/fetchparametersfrommodels', methods = ['POST'])
# def fetchparametersfrommodels():
#     selectedmodel = request.form['selectedmodel']
#
#     querystring = "select label from " + selectedmodel + ";"
#     parameters = []
#     conn = pyodbc.connect(databaseserver)
#     cursor = conn.cursor()
#     cursor.execute(querystring)
#     rows = cursor.fetchall()
#     print('Total Row(s):', cursor.rowcount)
#     for row in rows:
#         parameters.append(list(row))
#         print(type(row))
#         print(row)
#     print(parameters)
#     parametersjson = json.dumps(parameters)
#     return parametersjson
#

@app.route('/createocrsolution', methods = ['POST'])
def createocrsolution():
    try:
        ocrsolutionname = request.headers["ocrsolutionname"]
        description = request.headers["description"]
        ocrsolutionjson = request.headers["ocrsolutionjson"]
        ocrsolutionname = ocrsolutionname.strip()

        filename = ocrsolutionname + "$" + description

        with open(root_path + 'ocrsolutions/' + filename + '.json', 'w') as json_file:
            json_file.write(ocrsolutionjson)

        ocrsolutionjson_object = json.loads(ocrsolutionjson)
        keys = ocrsolutionjson_object['parameters'].keys()
        keys_list = []
        for key in keys:
            keys_list.append(key)
        print(keys_list)

        column_names_full_string = ""

        for key in keys_list:
            column_names_string = str(key) + " varchar(255), "
            column_names_full_string = column_names_full_string + column_names_string

        print(column_names_full_string)

        # try:
        #     conn = pyodbc.connect(database_server_ocr_solutions)
        #     cursor = conn.cursor()
        #     ocrsolutionname_for_db = ocrsolutionname.replace(" ", "_")
        #
        #     query = "CREATE TABLE " + ocrsolutionname_for_db + " (" + column_names_full_string + ")"
        #
        #     print(query)
        #     cursor.execute(query)
        #     conn.commit()
        # except Exception as e:
        #     print ("cannot create database " + str(e))
        #     return sendresponse("Cannot create OCR solution - DB error", 201)



        return sendresponse("OCR solution created", 200)

    except Exception as e:
        return sendresponse("Failed to create OCR solution : " + str(e), 201)





def del_dir(target: Union[Path, str], only_if_empty: bool = False):
    target = Path(target).expanduser()
    assert target.is_dir()
    for p in sorted(target.glob('**/*'), reverse=True):
        if not p.exists():
            continue
        p.chmod(0o666)
        if p.is_dir():
            p.rmdir()
        else:
            if only_if_empty:
                raise RuntimeError(f'{p.parent} is not empty!')
            p.unlink()
    target.rmdir()


@app.route('/gettableareafromheaderfooter', methods = ['POST'])
def gettableareafromheaderfooter():
    try:
        pdfpath = request.form['pdfpath']
        header = request.form['header']
        header_included_in_table = request.form['header_included_in_table']
        footer = request.form['footer']
        footer_included_in_table = request.form['footer_included_in_table']

        print (pdfpath)
        print (header)
        print (header_included_in_table)
        print (footer)
        print (footer_included_in_table)

        pdfpath = 'static/initialreceivepdf/' + pdfpath

        pdf = pdfquery.PDFQuery(pdfpath)
        pdf.load()

        header_label = pdf.pq('LTTextLineHorizontal:contains("' + header + '")')
        header_starting_point = float(header_label.attr('y0'))
        header_ending_point = float(header_label.attr('y1'))

        footer_label = pdf.pq('LTTextLineHorizontal:contains("' + footer + '")')
        footer_starting_point = float(footer_label.attr('y0'))
        footer_ending_point = float(footer_label.attr('y1'))

        final_area = []

        if (header_included_in_table == "True"):
            final_area.append(0)
            final_area.append(float(header_starting_point) - 10)
        else:
            final_area.append(0)
            final_area.append(float(header_ending_point) + 10)

        if (footer_included_in_table == "True"):
            final_area.append(5000)
            final_area.append(float(footer_ending_point))
        else:
            final_area.append(5000)
            final_area.append(float(footer_starting_point))

        return str(final_area)
    except Exception as e:
        print (e)
        return e





# New logic for new studio

@app.route('/receivefile', methods = ['POST'])
def receivefile():
    try:
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            folderpath = root_path + '/static/initialreceivepdf/'
            foldername = timestamp
            filename = timestamp + '.pdf'

            folderpath = os.path.join(folderpath, foldername)
            print(folderpath)
            try:
                os.mkdir(folderpath)
            except Exception as e:
                print("create folder")
                print(e)

            file.save(folderpath + "/" + filename)
            print("file received")
            return timestamp
    except Exception as e:
        print (e)
        return sendresponse("failed", 200)


@app.route('/processsentfile', methods = ['POST'])
def processsentfile():
    try:
        timestamp = request.form['timestamp']
        language = request.form['lang']
        if (timestamp == ""):
            print('No filename in request')
            return sendresponse("filename not received", 201)
        else:
            # noofpagesinpdf = 0

            folderpath = root_path + '/static/initialreceivepdf/'
            foldername = timestamp
            filename = timestamp + '.pdf'

            inputpdfpath = folderpath + timestamp + '/' + timestamp + '.pdf'
            savepath = folderpath + timestamp + '/'
            linesremovedpath = folderpath + timestamp + '/linesremoved/'
            pdfpath = folderpath + timestamp + '/pdfs/'
            consolidatedpdffilename = timestamp + '.pdf'

            pdfstoresizepath = pdfpath + '/pdfstoresize/'

            folderpath = os.path.join(savepath, "linesremoved")
            os.mkdir(folderpath)
            folderpath = os.path.join(savepath, "pdfs")
            os.mkdir(folderpath)

            # copyfile(inputpdfpath, pdfpath + consolidatedpdffilename)

            start = time.time()

            final_response_json = {}

            pdf_format = is_pdf_txt_or_img(inputpdfpath)
            print (pdf_format)
            if (pdf_format == "img"):
                convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs(savepath)
                    print("binarize_imgs_status - " + "done")
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(savepath)
                        print("skew_correct_status - " + "done")
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(savepath)
                            print("remove_lines_status - " + "done")
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                                print("get_txt_boxes_status - " + "done")
                                if (get_txt_boxes_status == "done"):
                                    convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath, language)
                                    if (convert_imgs_into_pdf_status == "done"):
                                        print("convert_imgs_into_pdf_status - " + "done")
                                        merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                        if (merge_pdfs_status == "done"):

                                            page_dims = []

                                            for filename_in_path in os.listdir(savepath):
                                                if filename_in_path.endswith(".jpg") or filename_in_path.endswith(".png"):
                                                    if filename_in_path.startswith("o"):
                                                        current_page_dim = []
                                                        im = Image.open(savepath + filename_in_path)
                                                        width, height = im.size
                                                        current_page_dim.append(width)
                                                        current_page_dim.append(height)
                                                        page_dims.append(current_page_dim)

                                            print("page_dims - " + str(page_dims))

                                            resize_pages_status = resize_pages(pdfpath, filename, pdfstoresizepath, page_dims, filename)

                                            if (resize_pages_status != "done"):
                                                return "failed in rescaling pdf pages"

                                            # object = PyPDF2.PdfFileReader(pdfpath + consolidatedpdffilename)
                                            # NumPages = object.getNumPages()
                                            #
                                            # print ("NumPages - " + str(NumPages))
                                            #
                                            # for i in range(0, NumPages):
                                            #     page_dim = page_dims[i]
                                            #     resize_width = page_dim[0]
                                            #     resize_height = page_dim[1]
                                            #     pdffile = PyPDF2.PdfFileReader(pdfpath + consolidatedpdffilename)
                                            #     page = pdffile.getPage(i)
                                            #     page.scaleTo(resize_width, resize_height)
                                            #     writer = PyPDF2.PdfFileWriter()
                                            #     writer.addPage(page)
                                            #     with open(pdfpath + consolidatedpdffilename, "wb+") as f:
                                            #         writer.write(f)

                                            final_response_json["dimensions"] = str(page_dims)

                                            pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                            pdf.load()

                                            pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                            with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                                text_found_in_pdf = ""
                                                for page in doc:
                                                    text_found_in_pdf += page.getText()

                                            print(text_found_in_pdf,  file=open(pdfpath + 'fulltext.txt', 'w'))

                                            print("fully done")
                                            end = time.time()
                                            timetaken = end - start
                                            print(f"Runtime of the program is {end - start}")
                                            print(timestamp)
                                            with open(root_path + '/static/initialreceivepdf/' + timestamp + '/final_json.json') as f:
                                                data = f.read()
                                            jsondata = json.dumps(data)

                                            final_response_json["jsondata"] = str(jsondata)
                                            json_object = json.dumps(final_response_json, indent=4)

                                            # print (json_object)
                                            return str(json_object)
                                        else:
                                            print("failed in merge_pdfs")
                                            return "failed in merge_pdfs"
                                    else:
                                        print("failed in convert_imgs_into_pdf")
                                        return "failed in convert_imgs_into_pdf"
                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"
            elif (pdf_format == "txt"):
                copyfile(inputpdfpath, pdfpath + consolidatedpdffilename)
                convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs(savepath)
                    print("binarize_imgs_status - " + binarize_imgs_status)
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(savepath)
                        print("skew_correct_status - " + skew_correct_status)
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(savepath)
                            print("remove_lines_status - " + remove_lines_status)
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                                print("get_txt_boxes_status - " + get_txt_boxes_status)
                                if (get_txt_boxes_status == "done"):



                                    page_dims = []

                                    for filename_in_path in os.listdir(savepath):
                                        if filename_in_path.endswith(".jpg") or filename_in_path.endswith(".png"):
                                            if filename_in_path.startswith("o"):
                                                current_page_dim = []
                                                im = Image.open(savepath + filename_in_path)
                                                width, height = im.size
                                                current_page_dim.append(width)
                                                current_page_dim.append(height)
                                                page_dims.append(current_page_dim)

                                    print("page_dims - " + str(page_dims))

                                    resize_pages_status = resize_pages(pdfpath, filename, pdfstoresizepath, page_dims,
                                                                       filename)

                                    if (resize_pages_status != "done"):
                                        return "failed in rescaling pdf pages"

                                    # object = PyPDF2.PdfFileReader(pdfpath + consolidatedpdffilename)
                                    # NumPages = object.getNumPages()
                                    #
                                    # print("NumPages - " + str(NumPages))
                                    #
                                    # for i in range(0, NumPages):
                                    #     page_dim = page_dims[i]
                                    #
                                    #     resize_width = page_dim[0]
                                    #     resize_height = page_dim[1]
                                    #
                                    #     pdffile = PyPDF2.PdfFileReader(pdfpath + consolidatedpdffilename)
                                    #     print (pdfpath + consolidatedpdffilename)
                                    #     page = pdffile.getPage(i)
                                    #     page.scaleTo(resize_width, resize_height)
                                    #     writer = PyPDF2.PdfFileWriter()
                                    #     writer.addPage(page)
                                    #     with open(pdfpath + "consolidatedpdffilename.pdf", "wb+") as f:
                                    #         writer.write(f)

                                    # with open(pdfpath + consolidatedpdffilename, 'rb') as pdf_file:
                                    #     pdf_reader = PyPDF2.PdfFileReader(pdf_file)
                                    #     pdf_writer = PyPDF2.PdfFileWriter()
                                    #
                                    #     page_dim = page_dims[0]
                                    #     resize_width = page_dim[0]
                                    #     resize_height = page_dim[1]
                                    #
                                    #     for page_num in range(pdf_reader.numPages):
                                    #         pdf_page = pdf_reader.getPage(page_num)
                                    #         # pdf_page.rotateClockwise(90)  # rotateCounterClockwise()
                                    #         pdf_page.scaleTo(resize_width, resize_height)
                                    #         pdf_writer.addPage(pdf_page)
                                    #
                                    #     with open(pdfpath + "consolidatedpdffilename.pdf", 'wb') as pdf_file_rotated:
                                    #         pdf_writer.write(pdf_file_rotated)





                                    final_response_json["dimensions"] = str(page_dims)


                                    print ("tam")
                                    pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                    pdf.load()

                                    pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                    with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                        text_found_in_pdf = ""
                                        for page in doc:
                                            text_found_in_pdf += page.getText()
                                    print("tam1")
                                    print(text_found_in_pdf, file=open(pdfpath + 'fulltext.txt', 'w', encoding='utf-8'))

                                    print("fully done")
                                    end = time.time()
                                    timetaken = end - start
                                    print(f"Runtime of the program is {end - start}")
                                    print(timestamp)
                                    with open(
                                            root_path + '/static/initialreceivepdf/' + timestamp + '/final_json.json') as f:
                                        data = f.read()
                                    jsondata = json.dumps(data)
                                    # print(jsondata)

                                    final_response_json["jsondata"] = str(jsondata)
                                    json_object = json.dumps(final_response_json, indent=4)

                                    # print(json_object)
                                    return str(json_object)

                                    # return str(jsondata)


                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"

    except Exception as e:
        print(e)
        return sendresponse("failed", 200)


@app.route('/receiveimagefile', methods = ['POST'])
def receiveimagefile():
    try:
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            folderpath = root_path + '/static/initialreceivepdf/'
            foldername = timestamp
            filename = timestamp + '.jpg'

            folderpath = os.path.join(folderpath, foldername)
            print(folderpath)
            try:
                os.mkdir(folderpath)
            except Exception as e:
                print(e)

            file = remove_transparency(file, bg_colour=(255, 255, 255))

            file.save(folderpath + "/" + filename)
            print("file received")
            return timestamp
    except Exception as e:
        print (e)
        return sendresponse("failed", 200)


@app.route('/processsentimage', methods = ['POST'])
def processsentimage():
    try:
        timestamp = request.form['timestamp']
        language = request.form['lang']
        if (timestamp == ""):
            print('No filename in request')
            return sendresponse("filename not received", 201)
        else:

            folderpath = root_path + '/static/initialreceivepdf/'
            foldername = timestamp
            filename = timestamp + '.jpg'

            inputimgpath = folderpath + timestamp + '/' + timestamp + '.jpg'
            inputpdfpath = folderpath + timestamp + '/' + timestamp + '.pdf'
            savepath = folderpath + timestamp + '/'
            linesremovedpath = folderpath + timestamp + '/linesremoved/'
            pdfpath = folderpath + timestamp + '/pdfs/'
            consolidatedpdffilename = timestamp + '.pdf'

            pdfstoresizepath = pdfpath + '/pdfstoresize/'

            folderpath = os.path.join(savepath, "linesremoved")
            os.mkdir(folderpath)
            folderpath = os.path.join(savepath, "pdfs")
            os.mkdir(folderpath)

            start = time.time()

            final_response_json = {}

            convert_img_to_pdf_status = convert_img_to_pdf(inputimgpath, savepath, timestamp)
            if (convert_img_to_pdf_status == "done"):
                convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs(savepath)
                    print(binarize_imgs_status)
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(savepath)
                        print(skew_correct_status)
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(savepath)
                            print(remove_lines_status)
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                                print(get_txt_boxes_status)
                                if (get_txt_boxes_status == "done"):
                                    convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath, language)
                                    if (convert_imgs_into_pdf_status == "done"):
                                        print("done")
                                        merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                        if (merge_pdfs_status == "done"):

                                            page_dims = []
                                            # page_dims.append(original_width)
                                            # page_dims.append(original_height)

                                            for filename_in_path in os.listdir(savepath):
                                                if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
                                                        ".png"):
                                                    if filename_in_path.startswith("o"):
                                                        current_page_dim = []
                                                        im = Image.open(savepath + filename_in_path)
                                                        width, height = im.size
                                                        current_page_dim.append(width)
                                                        current_page_dim.append(height)
                                                        page_dims.append(current_page_dim)

                                            print("page_dims - " + str(page_dims))

                                            resize_pages_status = resize_pages(pdfpath, consolidatedpdffilename, pdfstoresizepath,
                                                                               page_dims,
                                                                               consolidatedpdffilename)

                                            print ("resize_pages_status - " + str(resize_pages_status))

                                            if (resize_pages_status != "done"):
                                                return "failed in rescaling pdf pages"

                                            # for filename in os.listdir(savepath):
                                            #     if filename.endswith(".jpg") or filename.endswith(".png"):
                                            #         if filename.startswith("o"):
                                            #             current_page_dim = []
                                            #             im = Image.open(savepath + filename)
                                            #             width, height = im.size
                                            #             current_page_dim.append(width)
                                            #             current_page_dim.append(height)
                                            #             page_dims.append(current_page_dim)


                                            # print("page_dims - " + str(page_dims))
                                            final_response_json["dimensions"] = str(page_dims)

                                            pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                            pdf.load()

                                            pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                            with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                                text_found_in_pdf = ""
                                                for page in doc:
                                                    text_found_in_pdf += page.getText()

                                            print(text_found_in_pdf,  file=open(pdfpath + 'fulltext.txt', 'w'))

                                            print("fully done")
                                            end = time.time()
                                            timetaken = end - start
                                            print(f"Runtime of the program is {end - start}")
                                            print(timestamp)
                                            with open(
                                                    root_path + '/static/initialreceivepdf/' + timestamp + '/final_json.json') as f:
                                                data = f.read()
                                            jsondata = json.dumps(data)
                                            final_response_json["jsondata"] = str(jsondata)
                                            json_object = json.dumps(final_response_json, indent=4)

                                            # print(json_object)
                                            return str(json_object)
                                        else:
                                            print("failed in merge_pdfs")
                                            return "failed in merge_pdfs"
                                    else:
                                        print("failed in convert_imgs_into_pdf")
                                        return "failed in convert_imgs_into_pdf"
                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"
            else:
                print("failed in binarize_imgs")
                return "failed in binarize_imgs"

    except Exception as e:
        print (e)
        return sendresponse("failed", 200)


@app.route('/trycurrentcondition', methods = ['POST'])
def trycurrentcondition():
    try:
        conditions = request.get_json()
        print (conditions)
        if conditions is None:
            print("No valid request body, json missing!")
            return jsonify({'error': 'No valid request body, json missing!'})
        else:
            noofconditions = len(conditions)
            print ("noofconditions " + str(noofconditions))
            if (noofconditions == 3):
                timestamp = conditions['timestamp']
                page_from_user = conditions['page']

                folderpath = root_path + '/static/initialreceivepdf/' + timestamp + '/pdfs/'
                filename = timestamp + '.pdf'
                pdf = pdfquery.PDFQuery(folderpath + filename)

                if (page_from_user != ""):
                    pdf.load((int(page_from_user) - 1))
                else:
                    pdf.load()



                pdf.tree.write(folderpath + "/pdfxml.xml", pretty_print=True)

                xml_doc = etree.parse(folderpath + "/pdfxml.xml")
                root = xml_doc.getroot()

                condition = conditions['condition1']

                # page = condition['page']
                keyword = condition['keyword']
                position = condition['position']
                width = int(condition['width'])
                height = int(condition['height'])


                searchbox = []
                if (position == "right"):
                    parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    print ("parameter " + parameter)
                    label = pdf.pq(parameter)

                    print("label - " + str(label))
                    print (label.attr('bbox'))

                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))

                    print ("x0 - " + str(x0))
                    print ("y0 - " + str(y0))
                    print ("x1 - " + str(x1))
                    print ("y1 - " + str(y1))
                    print ("height - " + str(height))
                    print ("width - " + str(width))

                    foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
                    x1, y0 - height, x1 + width, y1 + height)).text()
                    searchbox.append(x1)
                    searchbox.append(y0 - height)
                    searchbox.append(x1 + width)
                    searchbox.append(y1 + height)
                    foundtext = foundtext.replace(keyword, "")
                    print("text - " + foundtext)
                elif (position == "left"):
                    parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    label = pdf.pq(parameter)
                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))
                    foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
                    x0 + width, y0 + height, x1 + width, y1 + height)).text()
                    print("text - " + foundtext)
                elif (position == "top"):
                    parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    label = pdf.pq(parameter)
                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))
                    foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
                    x0 + width, y0 + height, x1 + width, y1 + height)).text()
                    print("text - " + foundtext)
                elif (position == "bottom"):
                    print("here")
                    parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    label = pdf.pq(parameter)
                    print(label.attr('x0'))
                    x0 = int(float(label.attr('x0')))
                    y0 = int(float(label.attr('y0')))
                    x1 = int(float(label.attr('x1')))
                    y1 = int(float(label.attr('y1')))
                    print (str(x0 - width) + "," + str(y1 - 1) + "," + str(x1 + width) + "," + str(y1 - height))
                    foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0 - width, y0 - height, x1 + width, y1 - (y1-y0))).text()
                    # foundtext = pdf.pq('LTTextLineHorizontal:in_bbox("%s, %s, %s, %s")' % (x0, y0 - 30, x0 + 150, y0)).text()
                    foundtext = foundtext.replace(keyword, "")
                    print("text - " + foundtext)

                print("foundtext " + foundtext)
                returnjson = '{"searchbox":"' + str(searchbox) + '","foundtext":"' + foundtext + '"}'
                return returnjson
            else:
                timestamp = conditions['timestamp']
                folderpath = root_path + '/static/initialreceivepdf/' + timestamp + '/pdfs/'
                filename = timestamp + '.pdf'
                pdf = pdfquery.PDFQuery(folderpath + filename)
                pdf.load()

                pdf.tree.write(folderpath + "/pdfxml.xml", pretty_print=True)
                print ("here")
                # for condition in conditions:
                formatted_conditions = [["-", "-"], ["-", "-"], ["-", "-"], ["-", "-"]]
                for condition in range(1, (noofconditions - 2)):
                    print (condition)
                    condition = conditions['condition' + str(condition)]
                    print(condition)

                    relationship_keyword = condition["keyword"]
                    relationship_position = condition["position"]

                    relationship_condition = []
                    relationship_condition.insert(0, relationship_keyword)
                    relationship_condition.insert(1, relationship_position)
                    if (relationship_position == "top"):
                        formatted_conditions[0] = relationship_condition
                    elif (relationship_position == "bottom"):
                        formatted_conditions[1] = relationship_condition
                    elif (relationship_position == "left"):
                        formatted_conditions[2] = relationship_condition
                    elif (relationship_position == "right"):
                        formatted_conditions[3] = relationship_condition

                print ("formatted_conditions " + str(formatted_conditions))

                master_condition = conditions['condition' + str(noofconditions - 2)]
                print ("master_condition " + str(master_condition))

                print("-------------------")

                master_keyword = master_condition["keyword"]
                master_position = master_condition["position"]
                master_width = master_condition["width"]
                master_height = master_condition["height"]

                xml_doc = etree.parse(folderpath + "/pdfxml.xml")
                root = xml_doc.getroot()

                final_text_found = get_final_text_found(pdf, root, formatted_conditions, master_keyword, master_position, master_width, master_height)
                print (final_text_found)

                returnjson = '{"searchbox":"' + "-" + '","foundtext":"' + str(final_text_found) + '"}'
                return returnjson




                    # conditions = [["-", "-"], ["Description of Goods", "bottom"], ["-", "-"], ["-", "-"]]
                    # for key in condition






                    # keyword = condition['keyword']
                    # position = condition['position']
                    # width = condition['width']
                    # height = condition['height']
                    # print("here")

                    # if (width == ""):
                    #     parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    #     label = pdf.pq(parameter)
                    #     print(label)
                    #     x0 = int(float(label.attr('x0')))
                    #     y0 = int(float(label.attr('y0')))
                    #     x1 = int(float(label.attr('x1')))
                    #     y1 = int(float(label.attr('y1')))
                    #     print(x0)
                    #     print(y0)
                    #     print(x1)
                    #     print(y1)
                    #     updatedxml = getupdatedxml(folderpath, position, x0, y0, x1, y1)
                    #     print ("updatedxml")
                    #     print(updatedxml)
                    # else :
                    #     searchbox = []
                    #     if (position == "right"):
                    #         parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    #         label = pdf.pq(parameter)
                    #         print(label)
                    #         x0 = int(float(label.attr('x0')))
                    #         y0 = int(float(label.attr('y0')))
                    #         x1 = int(float(label.attr('x1')))
                    #         y1 = int(float(label.attr('y1')))
                    #         foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
                    #             x1, y0 - height, x1 + width, y1 + height)).text()
                    #         searchbox.append(x1)
                    #         searchbox.append(y0 - height)
                    #         searchbox.append(x1 + width)
                    #         searchbox.append(y1 + height)
                    #         foundtext = foundtext.replace(keyword, "")
                    #         print("text - " + foundtext)
                    #     elif (position == "left"):
                    #         parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    #         label = pdf.pq(parameter)
                    #         x0 = int(float(label.attr('x0')))
                    #         y0 = int(float(label.attr('y0')))
                    #         x1 = int(float(label.attr('x1')))
                    #         y1 = int(float(label.attr('y1')))
                    #         foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
                    #             x0 + width, y0 + height, x1 + width, y1 + height)).text()
                    #         print("text - " + foundtext)
                    #     elif (position == "top"):
                    #         parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    #         label = pdf.pq(parameter)
                    #         x0 = int(float(label.attr('x0')))
                    #         y0 = int(float(label.attr('y0')))
                    #         x1 = int(float(label.attr('x1')))
                    #         y1 = int(float(label.attr('y1')))
                    #         foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
                    #             x0 + width, y0 + height, x1 + width, y1 + height)).text()
                    #         print("text - " + foundtext)
                    #     elif (position == "bottom"):
                    #         parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
                    #         label = pdf.pq(parameter)
                    #         x0 = int(float(label.attr('x0')))
                    #         y0 = int(float(label.attr('y0')))
                    #         x1 = int(float(label.attr('x1')))
                    #         y1 = int(float(label.attr('y1')))
                    #         print(str(x0 - width) + "," + str(y1 - 1) + "," + str(x1 + width) + "," + str(y1 - height))
                    #         foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
                    #         x0 - width, y0 - height, x1 + width, y1 - (y1 - y0))).text()
                    #         # foundtext = pdf.pq('LTTextLineHorizontal:in_bbox("%s, %s, %s, %s")' % (x0, y0 - 30, x0 + 150, y0)).text()
                    #         foundtext = foundtext.replace(keyword, "")
                    #         print("text - " + foundtext)
                    #
                    #     print(foundtext)
                    #     returnjson = '{"searchbox":"' + str(searchbox) + '","foundtext":"' + foundtext + '"}'
                    #     return returnjson


            return "done"


    except Exception as e:
        # resultstring = 'failed to get text!!!' + str(e)
        print(e)
        # return str(e)
        return sendresponse(str(e), 201)


@app.route('/trycurrentpostprocessingcondition', methods = ['POST'])
def trycurrentpostprocessingcondition():
    try:
        conditions = request.get_json()
        print(conditions)
        # noofconditions = len(conditions)
        # print(noofconditions)

        captured_text = conditions['captured_text']
        all_conditions_list = conditions['all_conditions_list']

        print (captured_text)
        print (len(all_conditions_list))

        for i in range(len(all_conditions_list)):
            condition_list = all_conditions_list[i];
            condi1 = condition_list[0]
            condi2 = condition_list[1]
            condi3 = condition_list[2]

            print (condi1)
            print (condi2)
            print (condi3)

            if (condi1 == "add"):
                captured_text = add (captured_text, condi2, condi3)
            elif (condi1 == "remove"):
                captured_text = remove (captured_text, condi2, condi3)
            elif (condi1 == "removeall"):
                captured_text = removeall (captured_text, condi2, condi3)
            elif (condi1 == "replace"):
                captured_text = replace (captured_text, condi2, condi3)
            elif (condi1 == "replaceall"):
                captured_text = replaceall (captured_text, condi2, condi3)
            elif (condi1 == "translate"):
                captured_text = translate (captured_text, condi2, condi3)
            elif (condi1 == "runregex"):
                captured_text = runregex (captured_text, condi2, condi3)
            elif (condi1 == "extractonly"):
                captured_text = extractonly (captured_text, condi2, condi3)
                print (captured_text)

        print (captured_text)

        captured_text = str(captured_text).replace("[", "")
        captured_text = captured_text.replace("]", "")
        captured_text = captured_text.replace("'", "")

        return str(captured_text)

    except Exception as e:
        print(e)
        return e


@app.route('/trycurrentfullocrcondition', methods = ['POST'])
def trycurrentfullocrcondition():
    try:
        conditions = request.get_json()
        print(conditions)
        # noofconditions = len(conditions)
        # print(noofconditions)

        timestamp = conditions['timestamp']
        all_conditions_list = conditions['all_conditions_list']

        print (timestamp)
        print (all_conditions_list)

        f = open( root_path + "/static/initialreceivepdf/" + timestamp + "/pdfs/fulltext.txt", "r")
        captured_text = f.read()
        # print(captured_text)

        for i in range(len(all_conditions_list)):
            condition_list = all_conditions_list[i];
            condi1 = condition_list[0]
            condi2 = condition_list[1]
            condi3 = condition_list[2]

            print (condi1)
            print (condi2)
            print (condi3)

            if (condi1 == "add"):
                captured_text = add (captured_text, condi2, condi3)
            elif (condi1 == "remove"):
                captured_text = remove (captured_text, condi2, condi3)
            elif (condi1 == "removeall"):
                captured_text = removeall (captured_text, condi2, condi3)
            elif (condi1 == "replace"):
                captured_text = replace (captured_text, condi2, condi3)
            elif (condi1 == "replaceall"):
                captured_text = replaceall (captured_text, condi2, condi3)
            elif (condi1 == "translate"):
                captured_text = translate (captured_text, condi2, condi3)
            elif (condi1 == "runregex"):
                captured_text = runregex (captured_text, condi2, condi3)
            elif (condi1 == "extractonly"):
                captured_text = extractonly (captured_text, condi2, condi3)
                print (captured_text)

        print ("captured_text " + str(captured_text))

        captured_text = str(captured_text).replace("[", "")
        captured_text = captured_text.replace("]", "")
        captured_text = captured_text.replace("'", "")

        return str(captured_text)

    except Exception as e:
        print(e)
        return e


@app.route('/trycurrentregionbasedcondition', methods = ['POST'])
def trycurrentregionbasedcondition():
    try:
        conditions = request.get_json()
        selected_page = conditions['selected_page']
        timestamp = conditions['timestamp']
        x0 = int(conditions['x0'])
        y0 = int(conditions['y0'])
        x1 = int(conditions['x1'])
        y1 = int(conditions['y1'])

        print(selected_page)
        print(x0)
        print(y0)
        print(x1)
        print(y1)

        corresponding_image = root_path + "/static/initialreceivepdf/" + timestamp + "/output" + str(int(selected_page) - 1) + ".jpg"
        im = Image.open(corresponding_image)
        corresponding_image_width, corresponding_image_height = im.size

        print ("corresponding_image_width - " + str(corresponding_image_width))
        print ("corresponding_image_height - " + str(corresponding_image_height))




        # doc = fitz.open("C:/Users/salil/Desktop/eyesight V1/eyesight V1/static/initialreceivepdf/" + timestamp + "/pdfs/" + timestamp + ".pdf")
        # page = doc.loadPage(int(selected_page)-1)
        # width_of_pdf = int(page.MediaBoxSize[0])
        # height_of_pdf = int(page.MediaBoxSize[1])

        pdf = PyPDF2.PdfFileReader(root_path + "/static/initialreceivepdf/" + timestamp + "/pdfs/" + timestamp + ".pdf")
        page = pdf.getPage(int(selected_page)-1).mediaBox
        print ("page - " + str(page))
        orientation = pdf.getPage(int(selected_page)-1).get('/Rotate')
        print("orientation - " + str(orientation))
        if ((orientation == None) or (orientation == 0)):
            width_of_pdf = int(page[2])
            height_of_pdf = int(page[3])
        else:
            width_of_pdf = int(page[3])
            height_of_pdf = int(page[2])


        print("width_of_pdf - " + str(width_of_pdf))
        print("height_of_pdf - " + str(height_of_pdf))

        # x0_percentage = (int(x0)/corresponding_image_width) * 100
        # x0 = (width_of_pdf/100) * x0_percentage
        #
        # x1_percentage = (int(x1)/corresponding_image_width) * 100
        # x1 = (width_of_pdf/100) * x1_percentage
        #
        # y0_percentage = (int(y0)/corresponding_image_height) * 100
        # y0 = (height_of_pdf/100) * y0_percentage
        #
        # y1_percentage = (int(y1)/corresponding_image_height) * 100
        # y1 = (height_of_pdf/100) * y1_percentage

        # x0_percentage = (int(x0)/900) * 100
        # x0 = (width_of_pdf/100) * x0_percentage
        #
        # x1_percentage = (int(x1)/900) * 100
        # x1 = (width_of_pdf/100) * x1_percentage
        #
        # y0_percentage = (int(y0)/1300) * 100
        # y0 = (height_of_pdf/100) * y0_percentage
        #
        # y1_percentage = (int(y1)/1300) * 100
        # y1 = (height_of_pdf/100) * y1_percentage

        if (y0 > y1):
            y0, y1 = y1, y0

        # print ("x0 " + str(x0))
        # print ("x1 " + str(x1))
        # print ("y0 " + str(y0))
        # print ("y1 " + str(y1))

        # scaled_x0 = int(float(x0) * 2.8355)
        # scaled_x1 = int(float(x1) * 2.8355)
        # scaled_y0 = int(float(y0) * 2.7730)
        # scaled_y1 = int(float(y1) * 2.7730)

        # print(selected_page)
        # print(scaled_x0)
        # print(scaled_x1)
        # print(scaled_y0)
        # print(scaled_y1)

        folderpath = root_path + '/static/initialreceivepdf/' + timestamp + '/pdfs/'
        filename = timestamp + '.pdf'
        pdf = pdfquery.PDFQuery(folderpath + filename)

        if (selected_page != ""):
            pdf.load((int(selected_page) - 1))
        else:
            pdf.load()

        pdf.tree.write(folderpath + "/pdfxml.xml", pretty_print=True)

        xml_doc = etree.parse(folderpath + "/pdfxml.xml")
        root = xml_doc.getroot()

        # xmlfile = open("C:/Users/salil/Desktop/eyesight V1/eyesight V1/static/initialreceivepdf/" + timestamp + "/pdfs/pdfxml.xml", "r", encoding='utf-8')
        # xml = xmlfile.read()
        # tree = ET.fromstring(xml)
        #
        # xml_doc = etree.parse("C:/Users/salil/Desktop/eyesight V1/eyesight V1/static/initialreceivepdf/" + timestamp + "/pdfs/pdfxml.xml")
        # root = xml_doc.getroot()
        # # tree = tree.getroot()


        final_text = ""

        print ("-------------------------------")

        for elem in root.iter():
            if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                try:
                    bbox = elem.attrib.get("bbox")
                    bbox_list = bbox.strip('][').split(', ')
                    # print ("bbox_list - " + str(bbox_list))


                    # pageno = getpagno(bbox, tree)

                    # print(type(pageno))

                    # if (type(pageno) != None):
                    #     pageno = int(getpagno(bbox, tree))+1
                    #     print ("pageno - " + str(pageno))
                    # else:
                    #     break

                    x0_current_box = float(elem.attrib.get("x0"))
                    y0_current_box = float(elem.attrib.get("y0"))
                    x1_current_box = float(elem.attrib.get("x1"))
                    y1_current_box = float(elem.attrib.get("y1"))

                    # print(x0_current_box)
                    # print(y0_current_box)
                    # print(x1_current_box)
                    # print(y1_current_box)

                    # if (str(selected_page) == str(pageno)):
                    if (x0 < x0_current_box):
                        # print ("here1")
                        # print("x1_current_box " + str(x1_current_box))
                        # print("x1 " + str(x1))
                        # foundtext = elem.text
                        # if (foundtext != None):
                        #     print("x1_current_box " + str(x1_current_box))
                        #     print(foundtext)
                        if (x1 > x1_current_box):
                            # print("here2")
                            # print("y0_current_box " + str(y0_current_box))
                            # print("y0 " + str(y0))
                            # foundtext = elem.text
                            # if (foundtext != None):
                            #     print(foundtext)
                            if (y0 < y0_current_box):
                                # print("here3")
                                # print("y1_current_box " + str(y0_current_box))
                                # print("y1 " + str(y0))
                                foundtext = elem.text
                                # if (foundtext != None):
                                #     print(foundtext)
                                if (y1 > y1_current_box):
                                    # print("here4")
                                    foundtext = elem.text
                                    if (foundtext != None):
                                        print (foundtext)
                                        print ("bbox - " + str(bbox))
                                        print (getpagno(bbox, root))
                                        final_text += foundtext

                except Exception as e:
                    print(e)
                    return "failed"

        print("-------------------------------")

        print ("final_text - " + str(final_text))

        return final_text

        # return ("done")

    except Exception as e:
        print(e)
        return e


@app.route('/trycurrentstaticcondition', methods = ['POST'])
def trycurrentstaticcondition():
    try:
        conditions = request.get_json()
        staticlabel = conditions['staticlabel']
        capturedvaluestatic = conditions['capturedvaluestatic']
        staticfirstdropdownid = conditions['staticfirstdropdownid']
        statickeywordid = conditions['statickeywordid']

        final_text = ""

        if (staticfirstdropdownid == "statictext"):
            final_text = statickeywordid
        elif (staticfirstdropdownid == "date"):
            today = date.today()
            d1 = today.strftime("%d/%m/%Y")
            final_text = d1
        elif (staticfirstdropdownid == "day"):
            now = datetime.now()
            today = now.strftime("%A")
            final_text = today
        elif (staticfirstdropdownid == "month"):
            mydate = datetime.now()
            month = mydate.strftime("%B")
            final_text = month
        elif (staticfirstdropdownid == "year"):
            mydate = datetime.now()
            year = mydate.year
            final_text = str(year)
        elif (staticfirstdropdownid == "time"):
            mydate = datetime.now()
            current_time = mydate.strftime("%H:%M:%S")
            final_text = str(current_time)
        elif (staticfirstdropdownid == "timestamp"):
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            final_text = str(timestamp)



        return final_text


    except Exception as e:
        print(e)
        return e


@app.route('/checkdbstatus', methods = ['POST'])
def checkdbstatus():
    try:
        server = request.form['server']
        port = request.form['port']
        database = request.form['database']
        uid = request.form['uid']
        password = request.form['password']
        trustedconnection = request.form['trustedconnection']

        try:
            dbconnect = "Driver={ODBC Driver 17 for SQL Server};Server=" + server + "," + port + ";Database=" + database + ";Trusted_Connection=" + trustedconnection + ";UID=" + uid + ";PWD=" + password + ";"
            print (dbconnect)
            connection = pyodbc.connect(dbconnect)
            print('Connected to SQL Server Successfully')
            return sendresponse("Connected to SQL Server Successfully", 200)
        except Exception as e:
            print(e)
            print('Connection failed to SQL Server')
            return sendresponse("Failed to connect to SQL server", 201)

    except Exception as e:
        print (e)
        return sendresponse("Failed to connect to SQL server", 201)


@app.route('/fetchocrsolutions', methods = ['POST'])
def fetchocrsolutions():

    path = root_path + "/ocrsolutions/"
    ocrsolutions = [f for f in listdir(path) if isfile(join(path, f))]
    # print (ocrsolutions)
    return str(ocrsolutions)


@app.route('/fetchpublishedocrsolutions', methods = ['POST'])
def fetchpublishedocrsolutions():

    path = root_path + "/published_ocr_solutions"
    publishedocrsolutions = [f for f in listdir(path) if isfile(join(path, f))]
    print (publishedocrsolutions)
    return str(publishedocrsolutions)




@app.route('/gettablewithoutborder', methods = ['POST'])
def gettablewithoutborder():
    try:
        timestamp = request.form['timestamp']

        originalwidth = request.form['originalwidth']
        originalheight = request.form['originalheight']

        # originalheight = 1300

        header = request.form['header']
        header_included_in_table = request.form['header_included_in_table']
        footer = request.form['footer']
        footer_included_in_table = request.form['footer_included_in_table']

        multipagetable = request.form['multipagetable']

        top_margin = request.form['top_margin']
        bottom_margin = request.form['bottom_margin']

        group_into_row = request.form['group_into_row']
        group_into_column = request.form['group_into_column']
        detect_superscripts = request.form['detect_superscripts']
        cut_text = request.form['cut_text']
        text_edge_tol = request.form['text_edge_tol']

        header_availability = request.form['header_availability']
        footer_availability = request.form['footer_availability']

        # print (top_margin)
        # print (bottom_margin)
        # print (timestamp)
        # print (header)
        # print (header_included_in_table)
        # print (footer)
        # print (footer_included_in_table)
        # print (type(group_into_row))
        # print (group_into_column)
        # print (detect_superscripts)
        # print (cut_text)
        # print (type(text_edge_tol))
        # print (header_availability)
        # print (footer_availability)

        print ("multipagetable - " + str(multipagetable))

        xml_doc = etree.parse(root_path + "/static/initialreceivepdf/" + timestamp + "/pdfs/pdfxml.xml")
        root = xml_doc.getroot()

        bbox_of_header_key = {}
        for elem in root.iter():
            text = elem.text
            try:
                if (header in text):
                    box = elem.attrib.get("bbox")
                    print ("header box - " + str(box))
                    # bbox_of_key.append(box)
                    pageno = getpagno(box, root)
                    print ("header pageno " + str(pageno))
                    bbox_of_header_key[pageno] = box
            except Exception as e:
                # print (e)
                continue

        print ("bbox_of_header_key - " + str(bbox_of_header_key))

        bbox_of_footer_key = {}
        for elem in root.iter():
            text = elem.text
            try:
                if (footer in text):
                    box = elem.attrib.get("bbox")
                    print ("footer box - " + str(box))
                    # bbox_of_key.append(box)
                    pageno = getpagno(box, root)
                    print("footer pageno " + str(pageno))
                    bbox_of_footer_key[pageno] = box
            except Exception as e:
                # print (e)
                continue
                
        print ("bbox_of_footer_key - " + str(bbox_of_footer_key))

        bbox_of_header_key_list = [(k, v) for k, v in bbox_of_header_key.items()]
        bbox_of_footer_key_list = [(k, v) for k, v in bbox_of_footer_key.items()]

        print ("bbox_of_header_key_list " + str(bbox_of_header_key_list))
        print ("bbox_of_footer_key_list " + str(bbox_of_footer_key_list))

        # multi page table
        if (multipagetable == "True"):
            if ((header_availability == "True") & (footer_availability == "False")):
                print ("header available")

                fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                final_df_list = []

                for i in range(len(bbox_of_header_key_list)):
                    header_block = bbox_of_header_key_list[i]
                    header_page = int(header_block[0])
                    header_rect = header_block[1]
                    header_rect = header_rect.strip('][').split(', ')
                    # print("header_page " + str(header_page))

                    # footer_box = bbox_of_footer_key_list[header_page]
                    # print (footer_box)

                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]
                    for footer_elements in bbox_of_footer_key_list:
                        footer_page = int(footer_elements[0])
                        # print (type(footer_page))
                        if (footer_page == header_page):
                            print ("footer_page == header_page")
                            footer_rect = footer_elements[1]
                            footer_rect_list = footer_rect.strip('][').split(', ')
                            footer_starting_point = footer_rect_list[1]
                            footer_ending_point = footer_rect_list[3]
                            print ("footer_rect " + str(type(footer_rect)))
                            print ("footer_starting_point " + str(footer_starting_point))
                            print ("footer_ending_point " + str(footer_ending_point))
                        else:
                            footer_starting_point = int(bottom_margin) + 10
                            footer_ending_point = int(bottom_margin)

                    # print ("footer_rect " + str(footer_rect))
                    # print ("footer_starting_point " + str(footer_starting_point))

                    final_area = []
                    final_area.append(0)
                    if (header_included_in_table == "True"):
                        final_area.append(float(header_starting_point) - 200)
                    else:
                        final_area.append(float(header_ending_point) + 10)
                    # if (footer_included_in_table == "True"):
                    #     final_area.append(float(footer_starting_point) - 10)
                    # else:
                    #     final_area.append(float(footer_ending_point) + 10)
                    final_area.append(5000)
                    if (footer_included_in_table == "True"):
                        final_area.append(float(footer_ending_point) - 200)
                    else:
                        final_area.append(float(footer_starting_point) + 10)
                    # if (header_included_in_table == "True"):
                    #     final_area.append(float(header_starting_point) - 10)
                    # else:
                    #     final_area.append(float(header_ending_point) + 10)

                    print ("final_area " + str(final_area))

                    final_area = str(final_area)
                    final_area = final_area.replace("[", "")
                    final_area = final_area.replace("]", "")
                    final_area_list = []
                    final_area_list.append(final_area)
                    print("final_area_list " + str(final_area_list))

                    pdf_file_path = root_path + "/static/initialreceivepdf/" + str(
                        timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                    try:
                        print ("------------")
                        print (header_page)
                        print (final_area_list)
                        print (pdf_file_path)
                        print ("------------")

                        table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                                 pages=str(header_page), split_text=cut_text,
                                                                 table_areas=final_area_list,
                                                                 flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                                 row_tol=int(group_into_row))
                        nooftables = len(table_list)
                        print("nooftables_firstpage")
                        print(nooftables)


                        for x in range(nooftables):
                            df = table_list[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            df = df.loc[:, df.isin([' ', 'NULL', '', math.nan]).mean() < .5]
                            # df.dropna(axis=1, how='any')
                            final_df_list.append(df)

                    except Exception as e:
                        print(e)
                        print ("e")
                        continue

                    # footer_block = bbox_of_footer_key_list[0]
                    # footer_page = int(footer_block[0])
                    # footer_rect = footer_block[1]
                    # footer_rect = footer_rect.strip('][').split(', ')
                    # print("footer_page " + str(footer_page))
                    #
                    # header_starting_point = header_rect[1]
                    # header_ending_point = header_rect[3]
                    #
                    # footer_starting_point = footer_rect[1]
                    # footer_ending_point = footer_rect[3]
                    #
                    # final_area = []
                    #
                    # if (header_included_in_table == "True"):
                    #     final_area.append(0)
                    #     final_area.append(float(header_starting_point) - 10)
                    #     if (header_page == footer_page):
                    #         final_area.append(2000)
                    #         final_area.append(int(footer_starting_point))
                    #     else:
                    #         final_area.append(2000)
                    #         final_area.append(int(bottom_margin))
                    # else:
                    #     final_area.append(0)
                    #     final_area.append(float(header_ending_point) + 10)
                    #     if (header_page == footer_page):
                    #         final_area.append(2000)
                    #         final_area.append(int(float(footer_starting_point)))
                    #     else:
                    #         final_area.append(2000)
                    #         final_area.append(int(bottom_margin))
                    #
                    # print("final_area " + str(final_area))
                    #
                    # final_area = str(final_area)
                    # final_area = final_area.replace("[", "")
                    # final_area = final_area.replace("]", "")
                    # final_area_list = []
                    # final_area_list.append(final_area)
                    # print("final_area_first_page_list " + str(final_area_list))

                print("final_df_list")
                print (len(final_df_list))

                result_df = pd.concat(final_df_list)

                for column in result_df:
                    # print(result_df[column])
                    result_df[column] = result_df[column].replace('<s>', '', regex=True)

                for column in result_df:
                    # print(result_df[column])
                    result_df[column] = result_df[column].replace('</s>', '', regex=True)

                result_df = result_df.apply(lambda x: sorted(x, key=pd.isnull), 1)

                csv_file = root_path + '/static/finaloutputtables/' + timestamp + '.csv'
                csv_file_cleaned = root_path + '/static/finaloutputtables/' + timestamp + '_cleaned.csv'

                result_df.to_csv(csv_file)
                with open(csv_file, 'r', encoding="utf8") as infile, \
                        open(csv_file_cleaned, 'w', encoding="utf8") as outfile:
                    data = infile.read()
                    data = data.replace("[", "")
                    data = data.replace("]", "")
                    data = data.replace("'", "")
                    data = data.replace(", nan", " ")
                    data = data.replace("\\n", "")

                    outfile.write(data)
                # print (result_df)
                html = pd.DataFrame(result_df).to_html()
                # fullhtml += html
                fullhtml += html

                with open(root_path + '/static/outputtables/' + timestamp + '.html', 'w') as file:
                    file.write(fullhtml)

                return fullhtml

            # elif ((header_availability == "False") & (footer_availability == "False")):
            #
            #     if (header != ""):
            #         print("header available")
            #         fullhtml = "<head><link rel='stylesheet' href='http://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='http://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"
            #         final_df_list = []
            #
            #         for i in range(len(bbox_of_header_key_list)):
            #             header_block = bbox_of_header_key_list[i]
            #             header_page = int(header_block[0])
            #             header_rect = header_block[1]
            #             header_rect = header_rect.strip('][').split(', ')

        # single page table
        if (multipagetable == "False"):
            if (len(bbox_of_header_key_list) == len(bbox_of_footer_key_list)):

                full_html_for_multiple_page_table = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                for i in range(len(bbox_of_header_key_list)):
                    header_block = bbox_of_header_key_list[i]
                    header_page = int(header_block[0])
                    header_rect = header_block[1]
                    header_rect = header_rect.strip('][').split(', ')
                    print ("header_page " + str(header_page))

                    footer_block = bbox_of_footer_key_list[i]
                    footer_page = int(footer_block[0])
                    footer_rect = footer_block[1]
                    footer_rect = footer_rect.strip('][').split(', ')
                    print ("footer_page " + str(footer_page))

                    if (header_page == footer_page):
                        header_starting_point = header_rect[1]
                        header_ending_point = header_rect[3]

                        print (header_starting_point)
                        print (header_ending_point)

                        footer_starting_point = footer_rect[1]
                        footer_ending_point = footer_rect[3]

                        print(footer_starting_point)
                        print(footer_ending_point)

                        final_area = []



                        if (header_included_in_table == "True"):
                            final_area.append(0)
                            final_area.append(float(header_starting_point) - 10)
                        else:
                            final_area.append(0)
                            final_area.append(float(header_ending_point) + 10)

                        if (footer_included_in_table == "True"):
                            final_area.append(5000)
                            final_area.append(float(footer_ending_point))
                        else:
                            final_area.append(5000)
                            final_area.append(float(footer_starting_point))

                        final_area = str(final_area)
                        final_area = final_area.replace("[", "")
                        final_area = final_area.replace("]", "")
                        final_area_list = []
                        final_area_list.append(final_area)
                        print(type(final_area))
                        print("final_area " + str(final_area_list))
                        # tablepositionlist = ['0, 1064.808, 2000, 44.658']
                        # print (type(tablepositionlist[0]))

                        pdf_file_path = root_path + "/static/initialreceivepdf/" + str(timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                        print("here")

                        table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n', pages=str(header_page), split_text=cut_text, table_areas=final_area_list, flag_size=detect_superscripts, edge_tol=int(text_edge_tol), row_tol=int(group_into_row))

                        nooftables = len(table_list)

                        print("nooftables")
                        print(nooftables)

                        fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                        for x in range(nooftables):

                            df = table_list[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])

                            for column in df:
                                # print(result_df[column])
                                df[column] = df[column].replace('<s>', '', regex=True)

                            for column in df:
                                # print(result_df[column])
                                df[column] = df[column].replace('</s>', '', regex=True)

                            # df = df.apply(lambda x: sorted(x, key=pd.isnull), 1)

                            csv_file = root_path + '/static/finaloutputtables/' + timestamp + '.csv'
                            csv_file_cleaned = root_path + '/static/finaloutputtables/' + timestamp + '_cleaned.csv'

                            df.to_csv(csv_file)
                            with open(csv_file, 'r') as infile, \
                                    open(csv_file_cleaned, 'w') as outfile:
                                data = infile.read()
                                data = data.replace("[", "")
                                data = data.replace("]", "")
                                data = data.replace("'", "")
                                outfile.write(data)

                            if (x == 0):
                                html = pd.DataFrame(df).to_html()
                                fullhtml += html
                            else:
                                tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                                fullhtml += tabletitle
                                html = pd.DataFrame(df).to_html()
                                fullhtml += html

                        with open(root_path + '/static/outputtables/' + timestamp + '.html', 'w', encoding="utf8") as file:
                            file.write(fullhtml)

                        print("here--")

                        return fullhtml

                    elif (footer_page == (header_page + 1)):
                        print ("table spawns in 2 pages")

                        header_starting_point = header_rect[1]
                        header_ending_point = header_rect[3]
                        footer_starting_point = footer_rect[1]
                        footer_ending_point = footer_rect[3]

                        final_area_first_page = []
                        final_area_second_page = []

                        if (header_included_in_table == "True"):
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_starting_point) - 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))
                        else:
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_ending_point) + 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))

                        if (footer_included_in_table == "True"):
                            final_area_second_page.append(0)
                            final_area_second_page.append(int(top_margin))
                            final_area_second_page.append(5000)
                            final_area_second_page.append(float(footer_ending_point))
                        else:
                            final_area_second_page.append(0)
                            final_area_second_page.append(originalheight - (int(top_margin)))
                            final_area_second_page.append(5000)
                            final_area_second_page.append(float(footer_starting_point))

                        print (final_area_first_page)
                        print (final_area_second_page)

                        # print ("tick")

                        final_area_first_page = str(final_area_first_page)
                        final_area_first_page = final_area_first_page.replace("[", "")
                        final_area_first_page = final_area_first_page.replace("]", "")
                        final_area_first_page_list = []
                        final_area_first_page_list.append(final_area_first_page)
                        print(final_area_first_page)

                        final_area_second_page = str(final_area_second_page)
                        final_area_second_page = final_area_second_page.replace("[", "")
                        final_area_second_page = final_area_second_page.replace("]", "")
                        final_area_second_page_list = []
                        final_area_second_page_list.append(final_area_second_page)
                        print(final_area_second_page)



                        pdf_file_path = root_path + "/static/initialreceivepdf/" + str(timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                        table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n', pages=str(header_page), split_text=cut_text, table_areas=final_area_first_page_list, flag_size=detect_superscripts, edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                        nooftables_firstpage = len(table_list_first_page)

                        table_list_second_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n', pages=str(footer_page), split_text=cut_text, table_areas=final_area_second_page_list, flag_size=detect_superscripts, edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                        nooftables_secondpage = len(table_list_second_page)

                        print("tick")

                        print("nooftables_firstpage")
                        print(nooftables_firstpage)

                        print("nooftables_secondpage")
                        print(nooftables_secondpage)

                        fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                        df_list = []

                        for x in range(nooftables_firstpage):

                            df = table_list_first_page[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            df_list.append(df)
                            # if (x == 0):
                            #     html = df.to_html()
                            #     fullhtml += html
                            # else:
                            #     tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                            #     fullhtml += tabletitle
                            #     html = df.to_html()
                            #     fullhtml += html

                        for y in range(nooftables_secondpage):

                            df = table_list_second_page[y].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            df_list.append(df)
                            # if (y == 0):
                            #     html = df.to_html()
                            #     fullhtml += html
                            # else:
                            #     tabletitle = "<h1>Table " + str(y - 1) + "</h1>"
                            #     fullhtml += tabletitle
                            #     html = df.to_html()
                            #     fullhtml += html

                        # for x in range(nooftables_firstpage):
                        #
                        #     df = table_list_first_page[x].df
                        #     df.rename(columns=df.iloc[0]).drop(df.index[0])
                        #     df_list.append(df)
                        #     if (x == 0):
                        #         html = df.to_html()
                        #         fullhtml += html
                        #     else:
                        #         tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                        #         fullhtml += tabletitle
                        #         html = df.to_html()
                        #         fullhtml += html

                        result_df = pd.concat(df_list)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('<s>', '', regex=True)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('</s>', '', regex=True)

                        # print (result_df)
                        html = result_df.to_html()
                        fullhtml += html

                        with open(root_path + '/static/outputtables/' + timestamp + '.html', 'w') as file:
                            file.write(fullhtml)

                        return fullhtml

                    else:
                        print ("table spawns in more than 2 pages")

                        header_starting_point = header_rect[1]
                        header_ending_point = header_rect[3]
                        footer_starting_point = footer_rect[1]
                        footer_ending_point = footer_rect[3]

                        final_area_first_page = []
                        final_area_last_page = []

                        if (header_included_in_table == "True"):
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_starting_point) - 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))
                        else:
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_ending_point) + 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))

                        if (footer_included_in_table == "True"):
                            final_area_last_page.append(0)
                            final_area_last_page.append(int(top_margin))
                            final_area_last_page.append(5000)
                            final_area_last_page.append(float(footer_ending_point))
                        else:
                            final_area_last_page.append(0)
                            final_area_last_page.append(originalheight - (int(top_margin)))
                            final_area_last_page.append(5000)
                            final_area_last_page.append(float(footer_starting_point))

                        print("final_area_first_page " + str(final_area_first_page))
                        print("final_area_last_page " + str(final_area_last_page))

                        print ("tick")

                        final_df_list = []

                        final_area_first_page = str(final_area_first_page)
                        final_area_first_page = final_area_first_page.replace("[", "")
                        final_area_first_page = final_area_first_page.replace("]", "")
                        final_area_first_page_list = []
                        final_area_first_page_list.append(final_area_first_page)
                        print("final_area_first_page_list " + str(final_area_first_page_list))

                        final_area_last_page = str(final_area_last_page)
                        final_area_last_page = final_area_last_page.replace("[", "")
                        final_area_last_page = final_area_last_page.replace("]", "")
                        final_area_last_page_list = []
                        final_area_last_page_list.append(final_area_last_page)
                        print("final_area_last_page_list " + str(final_area_last_page_list))

                        pdf_file_path = root_path + "/static/initialreceivepdf/" + str(
                            timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                        table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                                 pages=str(header_page), split_text=cut_text,
                                                                 table_areas=final_area_first_page_list,
                                                                 flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                                 row_tol=int(group_into_row))
                        nooftables_firstpage = len(table_list_first_page)
                        print("nooftables_firstpage")
                        print(nooftables_firstpage)

                        table_list_last_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                                  pages=str(footer_page), split_text=cut_text,
                                                                  table_areas=final_area_last_page_list,
                                                                  flag_size=detect_superscripts,
                                                                  edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                        nooftables_lastpage = len(table_list_last_page)
                        print("nooftables_secondpage")
                        print(nooftables_lastpage)

                        fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                        for x in range(nooftables_firstpage):
                            df = table_list_first_page[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            final_df_list.append(df)


                        pages_in_between_list = []
                        for pages_in_between in range(header_page, footer_page+1):
                            if (pages_in_between != header_page):
                                if (pages_in_between != (footer_page)):
                                    # print ("pages_in_between " + str(pages_in_between))
                                    pages_in_between_list.append(pages_in_between)

                        print (pages_in_between_list)

                        table_area_in_between_pages = []
                        table_area_in_between_pages.append(0)
                        table_area_in_between_pages.append(originalheight - (int(top_margin)))
                        table_area_in_between_pages.append(5000)
                        table_area_in_between_pages.append(int(bottom_margin))
                        table_area_in_between_pages = str(table_area_in_between_pages)
                        table_area_in_between_pages = table_area_in_between_pages.replace("[", "")
                        table_area_in_between_pages = table_area_in_between_pages.replace("]", "")
                        table_area_in_between_pages_list = []
                        table_area_in_between_pages_list.append(table_area_in_between_pages)
                        print(table_area_in_between_pages_list)

                        for pages_in_between_for_df in pages_in_between_list:
                            table_list_in_between_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                                     pages=str(pages_in_between_for_df), split_text=cut_text,
                                                                     table_areas=table_area_in_between_pages_list,
                                                                     flag_size=detect_superscripts,
                                                                     edge_tol=int(text_edge_tol),
                                                                     row_tol=int(group_into_row))
                            nooftables_inbetweenpage = len(table_list_in_between_page)
                            print(nooftables_inbetweenpage)

                            for x in range(nooftables_inbetweenpage):
                                df = table_list_in_between_page[x].df
                                df.rename(columns=df.iloc[0]).drop(df.index[0])
                                final_df_list.append(df)


                        for y in range(nooftables_lastpage):

                            df = table_list_last_page[y].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            final_df_list.append(df)

                        result_df = pd.concat(final_df_list)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('<s>', '', regex=True)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('</s>', '', regex=True)

                        # print (result_df)
                        html = result_df.to_html()
                        # fullhtml += html
                        full_html_for_multiple_page_table += html

                with open(root_path + 'static/outputtables/' + timestamp + '.html', 'w') as file:
                    file.write(full_html_for_multiple_page_table)

                return full_html_for_multiple_page_table






        # return "done"
    except Exception as e:
        print (e)
        return e


@app.route('/gettablewithborder', methods=['POST'])
def gettablewithborder():
    try:
        timestamp = request.form['timestamp']

        originalwidth = request.form['originalwidth']
        originalheight = request.form['originalheight']

        header = request.form['header']
        header_included_in_table = request.form['header_included_in_table']
        footer = request.form['footer']
        footer_included_in_table = request.form['footer_included_in_table']
        multipagetable = request.form['multipagetable']
        top_margin = request.form['top_margin']
        bottom_margin = request.form['bottom_margin']
        process_bg_lines = request.form['process_bg_lines']
        cuttext = request.form['cuttext']
        detect_superscripts = request.form['detect_superscripts']
        smalllines = request.form['smalllines']
        text_edge_tol = request.form['text_edge_tol']
        header_availability = request.form['header_availability']
        footer_availability = request.form['footer_availability']
        columns_list = request.form['columns_list']

        # print(top_margin)
        # print(bottom_margin)
        #
        # print(timestamp)
        # print(header)
        # print(header_included_in_table)
        # print(footer)
        # print(footer_included_in_table)
        # print(type(process_bg_lines))
        # print(cuttext)
        # print(detect_superscripts)
        # print(smalllines)
        # print(type(text_edge_tol))
        # print(header_availability)
        # print(footer_availability)

        xml_doc = etree.parse(root_path + "/static/initialreceivepdf/" + timestamp + "/pdfs/pdfxml.xml")
        root = xml_doc.getroot()

        bbox_of_header_key = {}
        for elem in root.iter():
            text = elem.text
            try:
                if (header in text):
                    box = elem.attrib.get("bbox")
                    print("header box - " + str(box))
                    # bbox_of_key.append(box)
                    pageno = getpagno(box, root)
                    print("header pageno - " + str(pageno))
                    bbox_of_header_key[pageno] = box
            except Exception as e:
                # print (e)
                continue
        print("bbox_of_header_key - " + str(bbox_of_header_key))

        bbox_of_footer_key = {}
        for elem in root.iter():
            text = elem.text
            try:
                if (footer in text):
                    box = elem.attrib.get("bbox")
                    print("footer box - " + str(box))
                    # bbox_of_key.append(box)
                    pageno = getpagno(box, root)
                    print("footer pageno " + str(pageno))
                    bbox_of_footer_key[pageno] = box
            except Exception as e:
                # print (e)
                continue
        print("bbox_of_footer_key - " + str(bbox_of_footer_key))

        bbox_of_header_key_list = [(k, v) for k, v in bbox_of_header_key.items()]
        bbox_of_footer_key_list = [(k, v) for k, v in bbox_of_footer_key.items()]

        print("bbox_of_header_key_list " + str(bbox_of_header_key_list))
        print("bbox_of_footer_key_list " + str(bbox_of_footer_key_list))

        # multi page table
        if (multipagetable == "False"):
            if ((header_availability == "True") & (footer_availability == "False")):
                print("header available")

                fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                final_df_list = []

                for i in range(len(bbox_of_header_key_list)):
                    header_block = bbox_of_header_key_list[i]
                    header_page = int(header_block[0])
                    header_rect = header_block[1]
                    header_rect = header_rect.strip('][').split(', ')
                    # print("header_page " + str(header_page))

                    # footer_box = bbox_of_footer_key_list[header_page]
                    # print (footer_box)

                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]
                    for footer_elements in bbox_of_footer_key_list:
                        footer_page = int(footer_elements[0])
                        # print (type(footer_page))
                        if (footer_page == header_page):
                            print("footer_page == header_page")
                            footer_rect = footer_elements[1]
                            footer_rect_list = footer_rect.strip('][').split(', ')
                            footer_starting_point = footer_rect_list[1]
                            footer_ending_point = footer_rect_list[3]
                            print("footer_rect " + str(type(footer_rect)))
                            print("footer_starting_point " + str(footer_starting_point))
                            print("footer_ending_point " + str(footer_ending_point))
                        else:
                            footer_starting_point = int(bottom_margin) + 10
                            footer_ending_point = int(bottom_margin)

                    # print ("footer_rect " + str(footer_rect))
                    # print ("footer_starting_point " + str(footer_starting_point))

                    final_area = []
                    final_area.append(0)
                    if (header_included_in_table == "True"):
                        final_area.append(float(header_starting_point) - 10)
                    else:
                        final_area.append(float(header_ending_point) + 10)
                    # if (footer_included_in_table == "True"):
                    #     final_area.append(float(footer_starting_point) - 10)
                    # else:
                    #     final_area.append(float(footer_ending_point) + 10)
                    final_area.append(5000)
                    if (footer_included_in_table == "True"):
                        final_area.append(float(footer_ending_point) - 10)
                    else:
                        final_area.append(float(footer_starting_point) + 10)
                    # if (header_included_in_table == "True"):
                    #     final_area.append(float(header_starting_point) - 10)
                    # else:
                    #     final_area.append(float(header_ending_point) + 10)

                    print("final_area " + str(final_area))

                    final_area = str(final_area)
                    final_area = final_area.replace("[", "")
                    final_area = final_area.replace("]", "")
                    final_area_list = []
                    final_area_list.append(final_area)
                    print("final_area_list " + str(final_area_list))

                    pdf_file_path = root_path + "/static/initialreceivepdf/" + str(
                        timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                    try:
                        print("------------")
                        print(header_page)
                        print(final_area_list)
                        print(pdf_file_path)
                        print("------------")

                        table_list = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n', pages=str(header_page), split_text=cuttext, table_areas=final_area_list, flag_size=detect_superscripts, process_background=process_bg_lines, line_scale=int(smalllines))

                        nooftables = len(table_list)
                        print("nooftables_firstpage")
                        print(nooftables)
                        print (table_list[0])

                        for x in range(nooftables):
                            df = table_list[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            df = df.loc[:, df.isin([' ', 'NULL', '', math.nan]).mean() < .5]
                            # df.dropna(axis=1, how='any')
                            final_df_list.append(df)

                    except Exception as e:
                        print(e)
                        print("e")
                        continue

                    # footer_block = bbox_of_footer_key_list[0]
                    # footer_page = int(footer_block[0])
                    # footer_rect = footer_block[1]
                    # footer_rect = footer_rect.strip('][').split(', ')
                    # print("footer_page " + str(footer_page))
                    #
                    # header_starting_point = header_rect[1]
                    # header_ending_point = header_rect[3]
                    #
                    # footer_starting_point = footer_rect[1]
                    # footer_ending_point = footer_rect[3]
                    #
                    # final_area = []
                    #
                    # if (header_included_in_table == "True"):
                    #     final_area.append(0)
                    #     final_area.append(float(header_starting_point) - 10)
                    #     if (header_page == footer_page):
                    #         final_area.append(2000)
                    #         final_area.append(int(footer_starting_point))
                    #     else:
                    #         final_area.append(2000)
                    #         final_area.append(int(bottom_margin))
                    # else:
                    #     final_area.append(0)
                    #     final_area.append(float(header_ending_point) + 10)
                    #     if (header_page == footer_page):
                    #         final_area.append(2000)
                    #         final_area.append(int(float(footer_starting_point)))
                    #     else:
                    #         final_area.append(2000)
                    #         final_area.append(int(bottom_margin))
                    #
                    # print("final_area " + str(final_area))
                    #
                    # final_area = str(final_area)
                    # final_area = final_area.replace("[", "")
                    # final_area = final_area.replace("]", "")
                    # final_area_list = []
                    # final_area_list.append(final_area)
                    # print("final_area_first_page_list " + str(final_area_list))

                print ("final_df_list")
                print(len(final_df_list))


                result_df = pd.concat(final_df_list)


                for column in result_df:
                    # print(result_df[column])
                    result_df[column] = result_df[column].replace('<s>', '', regex=True)

                for column in result_df:
                    # print(result_df[column])
                    result_df[column] = result_df[column].replace('</s>', '', regex=True)

                try:
                    result_df = result_df.apply(lambda x: sorted(x, key=pd.isnull), 1)
                except Exception as e:
                    print ("error occured")
                    print (e)

                csv_file = root_path + '/static/finaloutputtables/' + timestamp + '.csv'
                csv_file_cleaned = root_path + '/static/finaloutputtables/' + timestamp + '_cleaned.csv'

                print("reached here")

                result_df.to_csv(csv_file)
                with open(csv_file, 'r', encoding="utf8") as infile, \
                        open(csv_file_cleaned, 'w', encoding="utf8") as outfile:
                    data = infile.read()
                    data = data.replace("[", "")
                    data = data.replace("]", "")
                    data = data.replace("'", "")
                    outfile.write(data)
                print (result_df)
                html = pd.DataFrame(result_df).to_html()
                # fullhtml += html
                fullhtml += html



                with open(root_path + '/static/outputtables/' + timestamp + '.html',
                          'w') as file:
                    file.write(fullhtml)

                return fullhtml

            elif ((header_availability == "False") & (footer_availability == "False")):

                if (header != ""):
                    print("header available")
                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"
                    final_df_list = []

                    for i in range(len(bbox_of_header_key_list)):
                        header_block = bbox_of_header_key_list[i]
                        header_page = int(header_block[0])
                        header_rect = header_block[1]
                        header_rect = header_rect.strip('][').split(', ')

        # single page table
        if (multipagetable == "True"):
            if (len(bbox_of_header_key_list) == len(bbox_of_footer_key_list)):

                full_html_for_multiple_page_table = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                for i in range(len(bbox_of_header_key_list)):
                    header_block = bbox_of_header_key_list[i]
                    header_page = int(header_block[0])
                    header_rect = header_block[1]
                    header_rect = header_rect.strip('][').split(', ')
                    print("header_page " + str(header_page))

                    footer_block = bbox_of_footer_key_list[i]
                    footer_page = int(footer_block[0])
                    footer_rect = footer_block[1]
                    footer_rect = footer_rect.strip('][').split(', ')
                    print("footer_page " + str(footer_page))

                    if (header_page == footer_page):
                        header_starting_point = header_rect[1]
                        header_ending_point = header_rect[3]

                        print(header_starting_point)
                        print(header_ending_point)

                        footer_starting_point = footer_rect[1]
                        footer_ending_point = footer_rect[3]

                        print(footer_starting_point)
                        print(footer_ending_point)

                        final_area = []

                        if (header_included_in_table == "True"):
                            final_area.append(0)
                            final_area.append(float(header_starting_point) - 10)
                        else:
                            final_area.append(0)
                            final_area.append(float(header_ending_point) + 10)

                        if (footer_included_in_table == "True"):
                            final_area.append(5000)
                            final_area.append(float(footer_ending_point))
                        else:
                            final_area.append(5000)
                            final_area.append(float(footer_starting_point))

                        final_area = str(final_area)
                        final_area = final_area.replace("[", "")
                        final_area = final_area.replace("]", "")
                        final_area_list = []
                        final_area_list.append(final_area)
                        print(type(final_area))
                        print("final_area " + str(final_area_list))
                        # tablepositionlist = ['0, 1064.808, 2000, 44.658']
                        # print (type(tablepositionlist[0]))

                        pdf_file_path = root_path + "/static/initialreceivepdf/" + str(
                            timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                        print("here")

                        # table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                        #                               pages=str(header_page), split_text=cut_text,
                        #                               table_areas=final_area_list, flag_size=detect_superscripts,
                        #                               edge_tol=int(text_edge_tol), row_tol=int(group_into_row))

                        table_list = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                      pages=str(header_page), split_text=cuttext,
                                                      table_areas=final_area_list,
                                                      flag_size=detect_superscripts,
                                                      process_background=process_bg_lines, line_scale=int(smalllines))

                        nooftables = len(table_list)

                        print("nooftables")
                        print(nooftables)

                        fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                        for x in range(nooftables):

                            df = table_list[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])

                            for column in df:
                                # print(result_df[column])
                                df[column] = df[column].replace('<s>', '', regex=True)

                            for column in df:
                                # print(result_df[column])
                                df[column] = df[column].replace('</s>', '', regex=True)

                            # df = df.apply(lambda x: sorted(x, key=pd.isnull), 1)

                            csv_file = root_path + '/static/finaloutputtables/' + timestamp + '.csv'
                            csv_file_cleaned = root_path + '/static/finaloutputtables/' + timestamp + '_cleaned.csv'

                            df.to_csv(csv_file)
                            with open(csv_file, 'r') as infile, \
                                    open(csv_file_cleaned, 'w') as outfile:
                                data = infile.read()
                                data = data.replace("[", "")
                                data = data.replace("]", "")
                                data = data.replace("'", "")
                                outfile.write(data)

                            if (x == 0):
                                html = pd.DataFrame(df).to_html()
                                fullhtml += html
                            else:
                                tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                                fullhtml += tabletitle
                                html = pd.DataFrame(df).to_html()
                                fullhtml += html

                        with open(
                                root_path + '/static/outputtables/' + timestamp + '.html',
                                'w', encoding="utf8") as file:
                            file.write(fullhtml)

                        return fullhtml

                    elif (footer_page == (header_page + 1)):
                        print("table spawns in 2 pages")

                        header_starting_point = header_rect[1]
                        header_ending_point = header_rect[3]
                        footer_starting_point = footer_rect[1]
                        footer_ending_point = footer_rect[3]

                        final_area_first_page = []
                        final_area_second_page = []

                        if (header_included_in_table == "True"):
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_starting_point) - 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))
                        else:
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_ending_point) + 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))

                        if (footer_included_in_table == "True"):
                            final_area_second_page.append(0)
                            final_area_second_page.append(int(top_margin))
                            final_area_second_page.append(5000)
                            final_area_second_page.append(float(footer_ending_point))
                        else:
                            final_area_second_page.append(0)
                            final_area_second_page.append(originalheight - (int(top_margin)))
                            final_area_second_page.append(5000)
                            final_area_second_page.append(float(footer_starting_point))

                        print(final_area_first_page)
                        print(final_area_second_page)

                        # print ("tick")

                        final_area_first_page = str(final_area_first_page)
                        final_area_first_page = final_area_first_page.replace("[", "")
                        final_area_first_page = final_area_first_page.replace("]", "")
                        final_area_first_page_list = []
                        final_area_first_page_list.append(final_area_first_page)
                        print(final_area_first_page)

                        final_area_second_page = str(final_area_second_page)
                        final_area_second_page = final_area_second_page.replace("[", "")
                        final_area_second_page = final_area_second_page.replace("]", "")
                        final_area_second_page_list = []
                        final_area_second_page_list.append(final_area_second_page)
                        print(final_area_second_page)

                        pdf_file_path = root_path + "/static/initialreceivepdf/" + str(
                            timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                        table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                                 pages=str(header_page), split_text=cuttext,
                                                                 table_region=final_area_first_page_list,
                                                                 flag_size=detect_superscripts,
                                                                 process_background=eval(process_bg_lines), line_scale=int(smalllines))
                        nooftables_firstpage = len(table_list_first_page)

                        table_list_second_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                                  pages=str(footer_page), split_text=cuttext,
                                                                  table_region=final_area_second_page_list,
                                                                  flag_size=detect_superscripts,
                                                                  process_background=eval(process_bg_lines), line_scale=int(smalllines))
                        nooftables_secondpage = len(table_list_second_page)

                        print("tick")

                        print("nooftables_firstpage " + str(nooftables_firstpage))
                        # print()

                        print("nooftables_secondpage " + str(nooftables_secondpage))
                        # print()

                        fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                        df_list = []

                        for x in range(nooftables_firstpage):
                            df = table_list_first_page[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            df_list.append(df)
                            # if (x == 0):
                            #     html = df.to_html()
                            #     fullhtml += html
                            # else:
                            #     tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                            #     fullhtml += tabletitle
                            #     html = df.to_html()
                            #     fullhtml += html

                        for y in range(nooftables_secondpage):
                            df = table_list_second_page[y].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            df_list.append(df)
                            # if (y == 0):
                            #     html = df.to_html()
                            #     fullhtml += html
                            # else:
                            #     tabletitle = "<h1>Table " + str(y - 1) + "</h1>"
                            #     fullhtml += tabletitle
                            #     html = df.to_html()
                            #     fullhtml += html

                        # for x in range(nooftables_firstpage):
                        #
                        #     df = table_list_first_page[x].df
                        #     df.rename(columns=df.iloc[0]).drop(df.index[0])
                        #     df_list.append(df)
                        #     if (x == 0):
                        #         html = df.to_html()
                        #         fullhtml += html
                        #     else:
                        #         tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                        #         fullhtml += tabletitle
                        #         html = df.to_html()
                        #         fullhtml += html

                        result_df = pd.concat(df_list)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('<s>', '', regex=True)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('</s>', '', regex=True)

                        csv_file = root_path + '/static/finaloutputtables/' + timestamp + '.csv'
                        csv_file_cleaned = root_path + '/static/finaloutputtables/' + timestamp + '_cleaned.csv'

                        result_df.to_csv(csv_file)
                        with open(csv_file, 'r') as infile, \
                                open(csv_file_cleaned, 'w') as outfile:
                            data = infile.read()
                            data = data.replace("[", "")
                            data = data.replace("]", "")
                            data = data.replace("'", "")
                            outfile.write(data)

                        # print (result_df)
                        html = result_df.to_html()
                        fullhtml += html

                        with open(
                                root_path + '/static/outputtables/' + timestamp + '.html',
                                'w') as file:
                            file.write(fullhtml)

                        return fullhtml

                    else:
                        print("table spawns in more than 2 pages")

                        header_starting_point = header_rect[1]
                        header_ending_point = header_rect[3]
                        footer_starting_point = footer_rect[1]
                        footer_ending_point = footer_rect[3]

                        final_area_first_page = []
                        final_area_last_page = []

                        if (header_included_in_table == "True"):
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_starting_point) - 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))
                        else:
                            final_area_first_page.append(0)
                            final_area_first_page.append(float(header_ending_point) + 10)
                            final_area_first_page.append(5000)
                            final_area_first_page.append(int(bottom_margin))

                        if (footer_included_in_table == "True"):
                            final_area_last_page.append(0)
                            final_area_last_page.append(int(top_margin))
                            final_area_last_page.append(5000)
                            final_area_last_page.append(float(footer_ending_point))
                        else:
                            final_area_last_page.append(0)
                            final_area_last_page.append(originalheight - (int(top_margin)))
                            final_area_last_page.append(5000)
                            final_area_last_page.append(float(footer_starting_point))

                        print("final_area_first_page " + str(final_area_first_page))
                        print("final_area_last_page " + str(final_area_last_page))

                        print("tick")

                        final_df_list = []

                        final_area_first_page = str(final_area_first_page)
                        final_area_first_page = final_area_first_page.replace("[", "")
                        final_area_first_page = final_area_first_page.replace("]", "")
                        final_area_first_page_list = []
                        final_area_first_page_list.append(final_area_first_page)
                        print("final_area_first_page_list " + str(final_area_first_page_list))

                        final_area_last_page = str(final_area_last_page)
                        final_area_last_page = final_area_last_page.replace("[", "")
                        final_area_last_page = final_area_last_page.replace("]", "")
                        final_area_last_page_list = []
                        final_area_last_page_list.append(final_area_last_page)
                        print("final_area_last_page_list " + str(final_area_last_page_list))

                        pdf_file_path = root_path + "/static/initialreceivepdf/" + str(
                            timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                        table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                                 pages=str(header_page), split_text=cuttext,
                                                                 table_region=final_area_first_page_list,
                                                                 flag_size=detect_superscripts,
                                                                 process_background=eval(process_bg_lines),
                                                                 line_scale=int(smalllines))

                        # table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                        #                                          pages=str(header_page), split_text=cut_text,
                        #                                          table_areas=final_area_first_page_list,
                        #                                          flag_size=detect_superscripts,
                        #                                          edge_tol=int(text_edge_tol),
                        #                                          row_tol=int(group_into_row))
                        nooftables_firstpage = len(table_list_first_page)
                        print("nooftables_firstpage")
                        print(nooftables_firstpage)

                        table_list_last_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                                 pages=str(footer_page), split_text=cuttext,
                                                                 table_region=final_area_last_page_list,
                                                                 flag_size=detect_superscripts,
                                                                 process_background=eval(process_bg_lines),
                                                                 line_scale=int(smalllines))

                        # table_list_last_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                        #                                         pages=str(footer_page), split_text=cut_text,
                        #                                         table_areas=final_area_last_page_list,
                        #                                         flag_size=detect_superscripts,
                        #                                         edge_tol=int(text_edge_tol),
                        #                                         row_tol=int(group_into_row))
                        nooftables_lastpage = len(table_list_last_page)
                        print("nooftables_secondpage")
                        print(nooftables_lastpage)

                        fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                        for x in range(nooftables_firstpage):
                            df = table_list_first_page[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            final_df_list.append(df)

                        pages_in_between_list = []
                        for pages_in_between in range(header_page, footer_page + 1):
                            if (pages_in_between != header_page):
                                if (pages_in_between != (footer_page)):
                                    # print ("pages_in_between " + str(pages_in_between))
                                    pages_in_between_list.append(pages_in_between)

                        print(pages_in_between_list)

                        table_area_in_between_pages = []
                        table_area_in_between_pages.append(0)
                        table_area_in_between_pages.append(originalheight - (int(top_margin)))
                        table_area_in_between_pages.append(5000)
                        table_area_in_between_pages.append(int(bottom_margin))
                        table_area_in_between_pages = str(table_area_in_between_pages)
                        table_area_in_between_pages = table_area_in_between_pages.replace("[", "")
                        table_area_in_between_pages = table_area_in_between_pages.replace("]", "")
                        table_area_in_between_pages_list = []
                        table_area_in_between_pages_list.append(table_area_in_between_pages)
                        print(table_area_in_between_pages_list)

                        for pages_in_between_for_df in pages_in_between_list:
                            table_list_in_between_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                                    pages=str(pages_in_between_for_df), split_text=cuttext,
                                                                    table_region=table_area_in_between_pages_list,
                                                                    flag_size=detect_superscripts,
                                                                    process_background=eval(process_bg_lines),
                                                                    line_scale=int(smalllines))

                            # table_list_in_between_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                            #                                               pages=str(pages_in_between_for_df),
                            #                                               split_text=cut_text,
                            #                                               table_areas=table_area_in_between_pages_list,
                            #                                               flag_size=detect_superscripts,
                            #                                               edge_tol=int(text_edge_tol),
                            #                                               row_tol=int(group_into_row))
                            nooftables_inbetweenpage = len(table_list_in_between_page)
                            print(nooftables_inbetweenpage)

                            for x in range(nooftables_inbetweenpage):
                                df = table_list_in_between_page[x].df
                                df.rename(columns=df.iloc[0]).drop(df.index[0])
                                final_df_list.append(df)

                        for y in range(nooftables_lastpage):
                            df = table_list_last_page[y].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            final_df_list.append(df)

                        result_df = pd.concat(final_df_list)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('<s>', '', regex=True)

                        for column in result_df:
                            # print(result_df[column])
                            result_df[column] = result_df[column].replace('</s>', '', regex=True)

                        # print (result_df)
                        html = result_df.to_html()
                        # fullhtml += html
                        full_html_for_multiple_page_table += html

                with open(root_path + '/static/outputtables/' + timestamp + '.html',
                          'w') as file:
                    file.write(full_html_for_multiple_page_table)

                return full_html_for_multiple_page_table

        # return "done"
    except Exception as e:
        print("exception - " + str(e))
        return e


@app.route('/createnewmodel', methods = ['POST'])
def createnewmodel():
    try:
        modelname = request.headers["modelname"]
        description = request.headers["description"]
        modeljson = request.headers["modeljson"]
        timestamp = request.headers["timestamp"]
        # classifiername = classifiername.strip()

        print ("modelname - " + str(modelname))
        print ("description - " + str(description))
        print ("modeljson - " + str(modeljson))

        with open(root_path + 'modelsnew/' + modelname + '$' + description + '.json', 'w') as json_file:
            json_file.write(modeljson)

        pdf_path = root_path + '/static/initialreceivepdf/' + timestamp + '/' + timestamp + '.pdf'
        image_name = modelname + '$' + description
        images = convert_from_path(pdf_path)
        images[0].save(root_path + 'imagesforclassifier/' + image_name + '.png', 'PNG')

        return sendresponse("Model created", 200)

    except Exception as e:
        print (e)
        return sendresponse("Failed to create model : " + str(e), 200)


@app.route('/fetchnewmodels', methods = ['POST'])
def fetchnewmodels():

    path = root_path + "/modelsnew/"
    classifiers = [f for f in listdir(path) if isfile(join(path, f))]
    print (classifiers)
    return str(classifiers)


@app.route('/deletemodelnew', methods = ['POST'])
def deletemodelnew():
    try:
        modelname = request.form['model']
        print (modelname)

        deletepath = root_path + "/modelsnew/" + modelname + ".json"
        os.remove(deletepath)


        return sendresponse("model deleted", 200)
    except Exception as e:
        print("error : " + str(e))
        return sendresponse("cannot delete model", 201)


# Testing a model logics
@app.route('/receivefilefortest', methods = ['POST'])
def receivefilefortest():
    try:
        extension = request.form['extension']
        modelname = request.form['modelname']
        modelname = modelname.strip()
        print("model being consumed - " + modelname)
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            folderpath = root_path + '/static/testfiles/'
            foldername = timestamp
            filename = timestamp + "." + extension

            folderpath = os.path.join(folderpath, foldername)
            outputfolderpath = os.path.join(folderpath, "outputs")
            print(folderpath)
            try:
                os.mkdir(folderpath)
                os.mkdir(outputfolderpath)
            except Exception as e:
                print("create folder")
                print(e)

            file.save(folderpath + "/" + filename)
            print("file received")
            return timestamp
    except Exception as e:
        print(e)
        return sendresponse("failed", 200)

    # try:
    #     modelname = request.form['model']
    #     modelname = modelname.strip()
    #     print("model being consumed - " + modelname)
    #     file = request.files['file']
    #     if 'file' not in request.files:
    #         print('No file part')
    #         return sendresponse("file not received", 201)
    #     else:
    #         timestamp = time.time()
    #         timestamp = str(timestamp).split('.')
    #         timestamp = str(timestamp[0])
    #         print(timestamp)
    #
    #         test_files_path = root_path + 'static/testfiles/'
    #
    #         received_file_name = 'file_received_' + timestamp + '.pdf'
    #
    #         finaltext = []
    #         file = request.files['file']
    #         file.save(test_files_path + received_file_name)
    #         print("file received")
    #         return sendresponse("file received", 200)
    # except Exception as e:
    #     return sendresponse(str(e), 201)


@app.route('/processsentfiletest', methods = ['POST'])
def processsentfiletest():
    timestamp = request.form['timestamp']
    extension = request.form['extension']
    modelname = request.form['model']
    # language = request.form['lang']
    language = "eng"

    print ("extension " + extension)
    print ("modelname " + modelname)

    modelpath = root_path + 'modelsnew/' + modelname + ".json"
    packagepath = root_path + '/static/testfiles/' + timestamp + "/"
    outputpath = root_path + '/static/testfiles/' + timestamp + "/outputs/"
    outputpath_for_frontend = '/static/testfiles/' + timestamp + "/outputs/"
    pdfpath = packagepath + "pdfs/" + timestamp + ".pdf"

    output_zip_path = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/testfiles/" + timestamp + '/outputs/eyesight_output.zip'
    alltablehtml = '<head><link rel=''stylesheet'' href=''https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css''></head><body><div style="margin: 10px;"></div><center><a href="' + output_zip_path + '" download>Download</a></center></body>'

    print ("modelpath - " + modelpath)

    with open(modelpath) as f:
        model_meta_data = json.load(f)
        print (model_meta_data)

    original_width = model_meta_data['originalwidth']
    original_height = model_meta_data['originalheight']

    if (extension == "pdf"):
        jsondata = processsentfileifpdf(timestamp, original_width, original_height)
    else:
        jsondata = processsentfileifimage(timestamp, original_width, original_height)


    # Extracting key value pairs

    final_keyvalue_pairs = {"Key" : "Captured value"}

    for keyvalue in model_meta_data['keyvalues']:
        print (keyvalue)
        type = keyvalue["type"]
        label = keyvalue["label"]

        if (type == "keybased"):
            print ("value extraction type - keybased")
            captured_value = runtime_keybased_extract(keyvalue, packagepath, timestamp)
            print ("extracted keybased value - " + str (captured_value))

            print ("len(keyvalue['post_processing'])")
            print (len(keyvalue["post_processing"]))

            if (len(keyvalue["post_processing"]) != 0):
                post_processed_text = ""
                for i in range (len(keyvalue["post_processing"])):
                    print ("i - " + str(i))
                    conditionname = "post_processing_condition_" + str(i+1)
                    condition = keyvalue["post_processing"][conditionname]
                    if (post_processed_text == ""):
                        post_processed_text = captured_value
                    post_processed_text = runtime_post_process(condition, post_processed_text)

                print ("post_processed_text - " + str(post_processed_text))
                final_keyvalue_pairs[label] = post_processed_text
            else:
                final_keyvalue_pairs[label] = captured_value

        elif (type == "regionbased"):
            print ("value extraction type - regionbased")
            captured_value = runtime_regionbased_extract(keyvalue, packagepath, timestamp)
            print("extracted regionbased value - " + str(captured_value))

            print("len(keyvalue['post_processing'])")
            print(len(keyvalue["post_processing"]))

            if (len(keyvalue["post_processing"]) != 0):
                post_processed_text = ""
                for i in range(len(keyvalue["post_processing"])):
                    print("i - " + str(i))
                    conditionname = "post_processing_condition_" + str(i + 1)
                    condition = keyvalue["post_processing"][conditionname]
                    if (post_processed_text == ""):
                        post_processed_text = captured_value
                    post_processed_text = runtime_post_process(condition, post_processed_text)

                print("post_processed_text - " + str(post_processed_text))
                final_keyvalue_pairs[label] = post_processed_text
            else:
                final_keyvalue_pairs[label] = captured_value

        elif (type == "fullocr"):
            print ("value extraction type - fullocr")
            captured_value = runtime_fullocr_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

        elif (type == "static"):
            print ("value extraction type - static")
            captured_value = runtime_static_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

        elif (type == "fromdb"):
            print("value extraction type - fromdb")
            captured_value = runtime_fromdb_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

    print (final_keyvalue_pairs)

    with open(outputpath + 'capturedvalues.csv', 'w') as f:
        for key in final_keyvalue_pairs.keys():
            f.write("%s,%s\n" % (key, final_keyvalue_pairs[key]))

    csv_to_html_table = pd.read_csv(outputpath + 'capturedvalues.csv', error_bad_lines=False, encoding='latin1')
    # csv_to_html_table = pd.read_csv(outputpath + 'capturedvalues.csv', sep='delimiter')

    html_table = csv_to_html_table.to_html()

    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Values</h1></body>"
    fullhtml += html_table
    alltablehtml += fullhtml



    # Extracting tables

    fullhtml_for_table = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'></body>"

    for table in model_meta_data['tables']:
        print(table)
        table_type = table["table_type"]

        if (table_type == "withoutborder"):
            final_table_html = gettablewithoutborder_runtime(table, packagepath, timestamp, original_height)
            if (final_table_html != None):
                fullhtml_for_table += final_table_html

        elif (table_type == "withborder"):
            final_table_html = gettablewithborder_runtime(table, packagepath, timestamp, original_height)
            if (final_table_html != None):
                fullhtml_for_table += final_table_html

        alltablehtml += fullhtml_for_table








    with open(outputpath + 'fullhtml.html', 'w', encoding='utf-8') as file:
        file.write(alltablehtml)

    output_zip = zipfile.ZipFile(outputpath + 'eyesight_output.zip', 'w')
    for folder, subfolders, files in os.walk(outputpath):
        for file in files:
            if not (file.endswith('.zip')):
                if not (file.endswith('.html')):
                    output_zip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), outputpath), compress_type=zipfile.ZIP_DEFLATED)

    output_zip.close()

    return outputpath_for_frontend + 'fullhtml.html'








    # return final_keyvalue_pairs


def processsentfileifpdf(timestamp, original_width, original_height):
    try:
        if (timestamp == ""):
            print('No filename in request')
            return sendresponse("filename not received", 201)
        else:
            # noofpagesinpdf = 0

            # pdf = pdfquery.PDFQuery(pdfpath)
            # pdf.load()

            folderpath = root_path + '/static/testfiles/'
            foldername = timestamp
            filename = timestamp + '.pdf'

            inputpdfpath = folderpath + timestamp + '/' + timestamp + '.pdf'
            savepath = folderpath + timestamp + '/'
            linesremovedpath = folderpath + timestamp + '/linesremoved/'
            pdfpath = folderpath + timestamp + '/pdfs/'
            consolidatedpdffilename = timestamp + '.pdf'

            pdfstoresizepath = pdfpath + '/pdfstoresize/'

            folderpath = os.path.join(savepath, "linesremoved")
            os.mkdir(folderpath)
            folderpath = os.path.join(savepath, "pdfs")
            os.mkdir(folderpath)

            # copyfile(inputpdfpath, pdfpath + consolidatedpdffilename)

            start = time.time()

            pdf_format = is_pdf_txt_or_img(inputpdfpath)
            print (pdf_format)
            if (pdf_format == "img"):
                convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs(savepath)
                    print("binarize_imgs_status - " + "done")
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(savepath)
                        print("skew_correct_status - " + "done")
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(savepath)
                            print("remove_lines_status - " + "done")
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                                print("get_txt_boxes_status - " + "done")
                                if (get_txt_boxes_status == "done"):
                                    convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath, language)
                                    if (convert_imgs_into_pdf_status == "done"):
                                        print("convert_imgs_into_pdf_status - " + "done")
                                        merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                        if (merge_pdfs_status == "done"):

                                            page_dims = []
                                            # page_dims.append(original_width)
                                            # page_dims.append(original_height)

                                            for filename_in_path in os.listdir(savepath):
                                                if filename_in_path.endswith(".jpg") or filename_in_path.endswith(".png"):
                                                    if filename_in_path.startswith("o"):
                                                        current_page_dim = []
                                                        current_page_dim.append(original_width)
                                                        current_page_dim.append(original_height)
                                                        page_dims.append(current_page_dim)

                                            print("page_dims - " + str(page_dims))

                                            resize_pages_status = resize_pages(pdfpath, filename, pdfstoresizepath, page_dims, filename)

                                            if (resize_pages_status != "done"):
                                                return "failed in rescaling pdf pages"

                                            pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                            pdf.load()

                                            pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                            with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                                text_found_in_pdf = ""
                                                for page in doc:
                                                    text_found_in_pdf += page.getText()

                                            print(text_found_in_pdf,  file=open(pdfpath + 'fulltext.txt', 'w'))

                                            print("fully done")
                                            end = time.time()
                                            timetaken = end - start
                                            print(f"Runtime of the program is {end - start}")
                                            print(timestamp)
                                            with open(root_path + '/static/testfiles/' + timestamp + '/final_json.json') as f:
                                                data = f.read()
                                            jsondata = json.dumps(data)
                                            # print (jsondata)
                                            return str(jsondata)
                                        else:
                                            print("failed in merge_pdfs")
                                            return "failed in merge_pdfs"
                                    else:
                                        print("failed in convert_imgs_into_pdf")
                                        return "failed in convert_imgs_into_pdf"
                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"
            elif (pdf_format == "txt"):
                copyfile(inputpdfpath, pdfpath + consolidatedpdffilename)
                convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs(savepath)
                    print("binarize_imgs_status - " + binarize_imgs_status)
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(savepath)
                        print("skew_correct_status - " + skew_correct_status)
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(savepath)
                            print("remove_lines_status - " + remove_lines_status)
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                                print("get_txt_boxes_status - " + get_txt_boxes_status)
                                if (get_txt_boxes_status == "done"):

                                    page_dims = []
                                    # page_dims.append(original_width)
                                    # page_dims.append(original_height)

                                    for filename_in_path in os.listdir(savepath):
                                        if filename_in_path.endswith(".jpg") or filename_in_path.endswith(".png"):
                                            if filename_in_path.startswith("o"):
                                                current_page_dim = []
                                                current_page_dim.append(original_width)
                                                current_page_dim.append(original_height)
                                                page_dims.append(current_page_dim)

                                    print("page_dims - " + str(page_dims))

                                    resize_pages_status = resize_pages(pdfpath, filename, pdfstoresizepath, page_dims,
                                                                       filename)

                                    if (resize_pages_status != "done"):
                                        return "failed in rescaling pdf pages"

                                    pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                    pdf.load()

                                    pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                    with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                        text_found_in_pdf = ""
                                        for page in doc:
                                            text_found_in_pdf += page.getText()
                                    print(text_found_in_pdf, file=open(pdfpath + 'fulltext.txt', 'w', encoding='utf-8'))

                                    print("fully done")
                                    end = time.time()
                                    timetaken = end - start
                                    print(f"Runtime of the program is {end - start}")
                                    print(timestamp)
                                    with open(
                                            root_path + 'static/testfiles/' + timestamp + '/final_json.json') as f:
                                        data = f.read()
                                    jsondata = json.dumps(data)
                                    # print(jsondata)
                                    return str(jsondata)

                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"

    except Exception as e:
        print ("no file found")
        print(e)
        return sendresponse("failed", 200)

def processsentfileifimage(timestamp, original_width, original_height):
    try:
        if (timestamp == ""):
            print('No filename in request')
            return sendresponse("filename not received", 201)
        else:

            folderpath = root_path + 'static/testfiles/'
            foldername = timestamp
            filename = timestamp + '.jpg'

            inputimgpath = folderpath + timestamp + '/' + timestamp + '.jpg'
            inputpdfpath = folderpath + timestamp + '/' + timestamp + '.pdf'
            savepath = folderpath + timestamp + '/'
            linesremovedpath = folderpath + timestamp + '/linesremoved/'
            pdfpath = folderpath + timestamp + '/pdfs/'
            consolidatedpdffilename = timestamp + '.pdf'

            pdfstoresizepath = pdfpath + '/pdfstoresize/'

            folderpath = os.path.join(savepath, "linesremoved")
            os.mkdir(folderpath)
            folderpath = os.path.join(savepath, "pdfs")
            os.mkdir(folderpath)

            start = time.time()

            convert_img_to_pdf_status = convert_img_to_pdf(inputimgpath, savepath, timestamp)
            if (convert_img_to_pdf_status == "done"):
                convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs(savepath)
                    print(binarize_imgs_status)
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(savepath)
                        print(skew_correct_status)
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(savepath)
                            print(remove_lines_status)
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                                print(get_txt_boxes_status)
                                if (get_txt_boxes_status == "done"):
                                    convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath, language)
                                    if (convert_imgs_into_pdf_status == "done"):
                                        print("done")
                                        merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                        if (merge_pdfs_status == "done"):

                                            page_dims = []
                                            # page_dims.append(original_width)
                                            # page_dims.append(original_height)

                                            for filename_in_path in os.listdir(savepath):
                                                if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
                                                        ".png"):
                                                    if filename_in_path.startswith("o"):
                                                        current_page_dim = []
                                                        current_page_dim.append(original_width)
                                                        current_page_dim.append(original_height)
                                                        page_dims.append(current_page_dim)

                                            print("page_dims - " + str(page_dims))

                                            resize_pages_status = resize_pages(pdfpath, filename, pdfstoresizepath,
                                                                               page_dims,
                                                                               filename)

                                            if (resize_pages_status != "done"):
                                                return "failed in rescaling pdf pages"



                                            pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                            pdf.load()

                                            pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                            with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                                text_found_in_pdf = ""
                                                for page in doc:
                                                    text_found_in_pdf += page.getText()

                                            print(text_found_in_pdf,  file=open(pdfpath + 'fulltext.txt', 'w'))

                                            print("fully done")
                                            end = time.time()
                                            timetaken = end - start
                                            print(f"Runtime of the program is {end - start}")
                                            print(timestamp)
                                            with open(
                                                    root_path + 'static/testfiles/' + timestamp + '/final_json.json') as f:
                                                data = f.read()
                                            jsondata = json.dumps(data)
                                            # print (jsondata)
                                            return str(jsondata)
                                        else:
                                            print("failed in merge_pdfs")
                                            return "failed in merge_pdfs"
                                    else:
                                        print("failed in convert_imgs_into_pdf")
                                        return "failed in convert_imgs_into_pdf"
                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"
            else:
                print("failed in binarize_imgs")
                return "failed in binarize_imgs"

    except Exception as e:
        print (e)
        return sendresponse("failed", 200)









#OCR Solutions logics
@app.route('/fetchmodelnames', methods = ['POST'])
def fetchmodelnames():

    path = root_path + "/modelsnew/"
    # models = [f for f in listdir(path) if isfile(join(path, f))]
    # print (models)
    models = os.listdir(path)
    print(models)
    models_list = []
    for i in range(len(models)):
        model = models[i]
        # model = model.split('$')[0]
        models_list.append(model)

    return str(models_list)


@app.route('/deleteocrsolution', methods = ['POST'])
def deleteocrsolution():
    try:
        modelname = request.form['model']
        print (modelname)

        deletepath = root_path + "/ocrsolutions/" + modelname + ".json"
        os.remove(deletepath)


        return sendresponse("model deleted", 200)
    except Exception as e:
        print("error : " + str(e))
        return sendresponse("cannot delete model", 201)


@app.route('/getparametersfrommodels', methods = ['POST'])
def getparametersfrommodels():
    # print ("hi")
    selectedmodel = request.form['selectedmodel']
    print ("selectedmodel - " + str(selectedmodel))
    path = root_path + "/modelsnew/"
    models = os.listdir(path)
    print(models)
    models_list = []
    for i in range(len(models)):
        model = models[i]
        models_list.append(model)

    model_full_name = ""
    for j in range(len(models_list)):
        model = models_list[j]
        result = model.startswith(selectedmodel)
        if (result):
            model_full_name = model

    with open(path + model_full_name) as f:
        json_object = json.load(f)

    parameters_list = []

    for element in json_object['keyvalues']:
        parameters_list.append(element['label'])

    print (parameters_list)
    parametersjson = json.dumps(parameters_list)

    return parametersjson


@app.route('/getparametersfrommanymodels', methods = ['POST'])
def getparametersfrommanymodels():
    print ("hi")
    selectedmodel = request.form['selectedmodels']
    model_list = selectedmodel.split(",")
    print ("model_list - " + str(model_list))

    final_result = {}

    for curr_model in model_list:
        print ("selectedmodel - " + str(curr_model))
        path = root_path + "/modelsnew/"
        models = os.listdir(path)
        models_list = []
        for i in range(len(models)):
            model = models[i]
            models_list.append(model)
        print ("models_list - " + str(models_list))

        model_full_name = ""
        for j in range(len(models_list)):
            model = models_list[j]
            result = model.startswith(curr_model)
            if (result):
                model_full_name = model

        print ("model_full_name - " + str(model_full_name))

        with open(path + model_full_name) as f:
            json_object = json.load(f)

        # print (str(json_object))

        parameters_list = []

        for element in json_object['keyvalues']:
            parameters_list.append(element['label'])

        print ("parameters_list - " + str(parameters_list))
        final_result[curr_model] = parameters_list
        # parametersjson = json.dumps(parameters_list)
    print ("final_result - " + str(final_result))
    parametersjson = json.dumps(final_result)
    return parametersjson




@app.route('/gettablesfrommodels', methods = ['POST'])
def gettablesfrommodels():
    # print ("hi")
    selectedmodel = request.form['selectedmodel']
    print ("selectedmodel - " + str(selectedmodel))
    path = root_path + "/modelsnew/"
    models = os.listdir(path)
    print(models)
    models_list = []
    for i in range(len(models)):
        model = models[i]
        models_list.append(model)

    model_full_name = ""
    for j in range(len(models_list)):
        model = models_list[j]
        result = model.startswith(selectedmodel)
        if (result):
            model_full_name = model

    with open(path + model_full_name) as f:
        json_object = json.load(f)

    tables_list = []

    for element in json_object['tables']:
        print (element)
        tables_list.append(element['table_name'])

    print ("tables_list - " + str(tables_list))
    tablesjson = json.dumps(tables_list)

    return tablesjson


@app.route('/receivefiles', methods = ['POST'])
def receivefiles():

    ocrsolution = request.form["ocrsolution"]
    ocrsolution_for_db = ocrsolution.split('$')[0]

    print("ocrsolution - " + ocrsolution)

    if (ocrsolution == ""):
        return sendresponse("Empty OCR solution", 201)

    # create folders for every file received
    ocr_list = []
    for filename in os.listdir(root_path + '/ocrsolutions/'):
        print("filename - ", filename)
        ocrsolution_name = filename.split('$')[0]
        print ("ocrsolution_name - ", ocrsolution_name)
        if (ocrsolution_name == ocrsolution):
            ocr_list.append(ocrsolution_name)

    print (ocr_list)

    if (len(ocr_list) == 0):
        return sendresponse("Incorrect OCR solution", 201)

    # sourcedocument = request.form["sourcedocument"]
    # companyname = request.form["companyname"]
    # country = request.form["country"]
    # sitename = request.form["sitename"]
    # topic = request.form["topic"]
    # subtopic = request.form["subtopic"]
    # indicator = request.form["indicator"]
    # subindicator = request.form["subindicator"]





    # .form['modelname']

    timestamp = time.time()
    timestamp = str(timestamp).split('.')
    timestamp = str(timestamp[0])

    folderpath = root_path + 'static/initialreceivefiles_ocr_solution'
    foldername = timestamp
    folderpath = os.path.join(folderpath, foldername)
    folderpath = folderpath + "/"
    input_files_path = folderpath + "/inputfiles/"

    try:
        os.mkdir(folderpath)
        os.mkdir(input_files_path)
    except Exception as e:
        print(e)


    # get files and save all in input folder
    files = request.files.getlist("file")
    for file in files:
        file.save(input_files_path + file.filename)


    # create folders for every file received
    for filename in os.listdir(input_files_path):
        foldername_to_be_created = os.path.splitext(filename)[0]
        os.mkdir(folderpath + foldername_to_be_created)


    # create excel sheet for output
    ocr_path = root_path + '/ocrsolutions/'
    for filename in os.listdir(ocr_path):
        try:
            if filename.startswith(ocrsolution):
                ocrsolution = filename
            else:
                continue
        except Exception as e:
            print (e)


    # ocrsolution = ocrsolution + '.json'
    ocr_sol_path = root_path + '/ocrsolutions/' + ocrsolution

    with open(ocr_sol_path) as f:
        ocrsolution_metadata = json.load(f)

    key_list = []
    parameters = ocrsolution_metadata["parameters"]
    models = ocrsolution_metadata["models"]
    print ("models - ", models)

    for key in parameters:
        key_list.append(key)
    print ("key_list - " + str(key_list))
    workbook = xlsxwriter.Workbook(folderpath + 'eyesight_output.xlsx')
    worksheet = workbook.add_worksheet()
    row = 0
    col = 0
    for item in key_list:
        worksheet.write(row, col, item)
        col += 1

    print ("key_list - " + str(key_list))

    workbook.close()

    final_return_json = {}
    # final_keyvalue_json = {}

    # process every file received
    for filename in os.listdir(input_files_path):
        try:
            if filename.endswith(".jpg") or filename.endswith(".PNG") or filename.endswith(".JPG") or filename.endswith(".png"):
                print(filename + " to image processing")
                individual_foldername = os.path.splitext(filename)[0]
                model_meta_data_list = ocr_solution_initial_process_img(input_files_path, filename, folderpath, timestamp, models)
                print("model_meta_data - ", model_meta_data_list)

                for i in range(len(model_meta_data_list)):
                    model_meta_data = model_meta_data_list[i]

                    if (model_meta_data != "failed"):
                        output_dict = ocr_solution_get_data_from_img(model_meta_data, folderpath+individual_foldername, timestamp)
                        print ("output_dict - " + str(output_dict))
                        model_name = model_meta_data["modelname"]
                        remapped_dict = {}
                        for current_param in output_dict:
                            current_value = output_dict[current_param]

                            for parameter in parameters:
                                values = parameters[parameter]
                                for key in values:
                                    corres_value = values[key]
                                    if (key == model_name):
                                        if (corres_value == current_param):
                                            remapped_dict[parameter] = current_value

                        print ("remapped_dict - " + str(remapped_dict))

                        # key_list - ['legal', 'trade', 'reg']
                        dict_to_excel = {}

                        for col_name in key_list:
                            try:
                                fetched_value_list = []
                                fetched_value = remapped_dict[col_name]
                                # print("fetched_value_before_none - " + str(fetched_value))
                                if (fetched_value == None):
                                    fetched_value = "-"
                                # print("fetched_value_after_none - " + str(fetched_value))
                                fetched_value_list.append(fetched_value)
                                dict_to_excel[col_name] = fetched_value_list
                            except Exception as e:
                                print ("exception - ")
                                print (e)
                                fetched_value_list.append("-")
                                dict_to_excel[col_name] = fetched_value_list
                                continue


                        print ("dict_to_excel - " + str(dict_to_excel))

                        final_return_json[filename + "(" + str(i) + ")"] = dict_to_excel


                        # # sending values to DB table
                        # keys = dict_to_excel.keys()
                        # keys_list = []
                        # for key in keys:
                        #     key = key.replace(" ", "_")
                        #     key = key.replace("/", "")
                        #     keys_list.append(key)
                        # keys_list = str(keys_list)
                        # keys_list = keys_list.replace("[", "")
                        # keys_list = keys_list.replace("]", "")
                        # keys_list = keys_list.replace("'", "")
                        # keys_list = "(" + keys_list + ")"
                        # print("keys_list - " + str(keys_list))
                        #
                        # values = dict_to_excel.values()
                        # values_list = []
                        # for value in values:
                        #     value = str(value)
                        #     value = value.replace(" ", "_")
                        #     value = value.replace("/", "")
                        #     values_list.append(value)
                        # values_list = str(values_list)
                        # values_list = values_list.replace("[", "")
                        # values_list = values_list.replace("]", "")
                        # values_list = values_list.replace('"', '')
                        # values_list = "(" + values_list + ")"
                        # print("values_list - " + str(values_list))
                        # # print(values_list)
                        #
                        # # try:
                        # #     conn = pyodbc.connect(database_server_ocr_solutions)
                        # #     cursor = conn.cursor()
                        # #     query = 'INSERT INTO ' + ocrsolution_for_db + ' ' + keys_list + 'VALUES' + values_list + ';'
                        # #     print(query)
                        # #     cursor.execute(query)
                        # #     conn.commit()
                        # # except Exception as e:
                        # #     print ("cannot append to DB - " + str(e))
                        #
                        # # df = pd.DataFrame({'reg': 'E', 'legal': 100})
                        df = pd.DataFrame(dict_to_excel)

                        writer = pd.ExcelWriter(folderpath + 'eyesight_output.xlsx', engine='openpyxl')
                        writer.book = load_workbook(folderpath + 'eyesight_output.xlsx')
                        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                        reader = pd.read_excel(folderpath + 'eyesight_output.xlsx')
                        df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)
                        writer.close()

                    else:
                        sendresponse("failed to process file", 201)

            elif filename.endswith(".pdf") or filename.endswith(".PDF"):
                print(filename + " to pdf processing")
                individual_foldername = os.path.splitext(filename)[0]

                model_meta_data_list = ocr_solution_initial_process_pdf(input_files_path, filename, folderpath, timestamp, models)
                print ("model_meta_data - ", model_meta_data_list)
                # model_meta_data = model_meta_data[1]

                for i in range(len(model_meta_data_list)):
                    model_meta_data = model_meta_data_list[i]

                    if (model_meta_data != "failed"):
                        final_keyvalue_json = {}
                        output_dict = ocr_solution_get_data_from_pdf(model_meta_data, folderpath + individual_foldername,
                                                                     timestamp, individual_foldername)
                        print ("output_dict - " + str(output_dict))
                        print ("model_meta_data - ", model_meta_data)
                        model_name = model_meta_data["modelname"]
                        print ("model_name - ", model_name)

                        remapped_dict = {}

                        key_values = output_dict["keyvalues"]

                        for current_param in key_values:
                            current_value = key_values[current_param]
                            # print ("current_value - ", current_value)
                            for parameter in parameters:
                                values = parameters[parameter]
                                # print ("values - ", values)
                                for key in values:
                                    corres_value = values[key]
                                    # print("corres_value - ", corres_value)
                                    if (key == model_name):
                                        # print("key - ", key)
                                        if (corres_value == current_param):
                                            # print("parameter - ", parameter)
                                            remapped_dict[parameter] = current_value

                        print("remapped_dict - " + str(remapped_dict))
                        print ("key_list - " + str(key_list))

                        dict_to_excel = {}

                        for col_name in key_list:
                            try:
                                fetched_value_list = []
                                fetched_value = remapped_dict[col_name]
                                # print("fetched_value_before_none - " + str(fetched_value))
                                if (fetched_value == None):
                                    fetched_value = "-"
                                # print("fetched_value_after_none - " + str(fetched_value))
                                fetched_value_list.append(fetched_value)
                                dict_to_excel[col_name] = fetched_value_list
                            except Exception as e:
                                print ("exception - ")
                                print (e)
                                fetched_value_list.append("-")
                                dict_to_excel[col_name] = fetched_value_list
                                continue

                        print("dict_to_excel - " + str(dict_to_excel))
                        final_keyvalue_json["keyvalues"] = dict_to_excel
                        final_keyvalue_json["tables"] = output_dict["tables"]
                        final_keyvalue_json["ocrsolution"] = ocrsolution.split('$')[0]
                        final_keyvalue_json["consumedmodel"] = model_name

                        final_return_json[filename + "(" + str(i) + ")"] = final_keyvalue_json


                        # # sending values to DB table
                        # keys = dict_to_excel.keys()
                        # keys_list = []
                        # for key in keys:
                        #     key = key.strip();
                        #     key = key.replace(" ", "_")
                        #     key = key.replace("/", "")
                        #     key = key.replace("-", "")
                        #     keys_list.append(key)
                        #
                        # # keys_list.append("sourcedocument")
                        # # keys_list.append("companyname")
                        # # keys_list.append("country")
                        # # keys_list.append("sitename")
                        # # keys_list.append("topic")
                        # # keys_list.append("subtopic")
                        # # keys_list.append("indicator")
                        # # keys_list.append("subindicator")
                        #
                        # keys_list_1 = str(keys_list)
                        # keys_list_1 = keys_list_1.replace("[", "")
                        # keys_list_1 = keys_list_1.replace("]", "")
                        # keys_list_1 = keys_list_1.replace("'", "")
                        # keys_list_1 = "(" + keys_list_1 + ")"
                        # print(keys_list_1)
                        #
                        # values = dict_to_excel.values()
                        # values_list = []
                        # for value in values:
                        #     value = str(value)
                        #     # value = value.replace(" ", "_")
                        #     # value = value.replace("/", "")
                        #     values_list.append(value)
                        #
                        # # values_list.append(sourcedocument)
                        # # values_list.append(companyname)
                        # # values_list.append(country)
                        # # values_list.append(sitename)
                        # # values_list.append(topic)
                        # # values_list.append(subtopic)
                        # # values_list.append(indicator)
                        # # values_list.append(subindicator)
                        #
                        # values_list_1 = str(values_list)
                        # values_list_1 = values_list_1.replace("[", "")
                        # values_list_1 = values_list_1.replace("]", "")
                        # values_list_1 = values_list_1.replace('"', '')
                        # values_list_1 = "(" + values_list_1 + ")"
                        # print(values_list_1)

                        # try:
                        #     conn = pyodbc.connect(database_server_ocr_solutions)
                        #     cursor = conn.cursor()
                        #
                        #     query = 'INSERT INTO ' + ocrsolution_for_db + ' ' + keys_list_1 + 'VALUES' + values_list_1 + ';'
                        #     print(query)
                        #     cursor.execute(query)
                        #     conn.commit()
                        #
                        #     keys_list.append("sourcedocument")
                        #     keys_list.append("companyname")
                        #     keys_list.append("country")
                        #     keys_list.append("sitename")
                        #     keys_list.append("topic")
                        #     keys_list.append("subtopic")
                        #     keys_list.append("indicator")
                        #     keys_list.append("subindicator")
                        #
                        #     keys_list_2 = str(keys_list)
                        #     keys_list_2 = keys_list_2.replace("[", "")
                        #     keys_list_2 = keys_list_2.replace("]", "")
                        #     keys_list_2 = keys_list_2.replace("'", "")
                        #     keys_list_2 = "(" + keys_list_2 + ")"
                        #     print(keys_list_2)
                        #
                        #     values_list.append(sourcedocument)
                        #     values_list.append(companyname)
                        #     values_list.append(country)
                        #     values_list.append(sitename)
                        #     values_list.append(topic)
                        #     values_list.append(subtopic)
                        #     values_list.append(indicator)
                        #     values_list.append(subindicator)
                        #
                        #     values_list_2 = str(values_list)
                        #     values_list_2 = values_list_2.replace("[", "")
                        #     values_list_2 = values_list_2.replace("]", "")
                        #     values_list_2 = values_list_2.replace('"', '')
                        #     values_list_2 = "(" + values_list_2 + ")"
                        #     print(values_list_2)
                        #
                        #     conn1 = pyodbc.connect(database_server_ccass)
                        #     cursor1 = conn1.cursor()
                        #
                        #     query1 = 'INSERT INTO ' + 'raw' + ' ' + keys_list_2 + 'VALUES' + values_list_2 + ';'
                        #     print(query1)
                        #     cursor1.execute(query1)
                        #     conn1.commit()
                        #
                        # except Exception as e:
                        #     print("cannot append to DB - " + str(e))

                        # df = pd.DataFrame({'reg': 'E', 'legal': 100})
                        df = pd.DataFrame(dict_to_excel)

                        writer = pd.ExcelWriter(folderpath + 'eyesight_output.xlsx', engine='openpyxl')
                        writer.book = load_workbook(folderpath + 'eyesight_output.xlsx')
                        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                        reader = pd.read_excel(folderpath + 'eyesight_output.xlsx')
                        df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)
                        writer.close()
                    else:
                        sendresponse("failed to process file", 201)

            else:
                print("unsupported file format")
        except Exception as e:
            print (e)
            continue

    download_path = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/initialreceivefiles_ocr_solution/" + timestamp + "/eyesight_output.xlsx"
    final_return_json["download_path"] = download_path
    print ("final_return_json - " + str(final_return_json))

    # return "done"

    # return sendresponse(str(final_return_json), 200)
    return sendresponse(final_return_json, 200)



@app.route('/getocrsolutionmetadata', methods = ['POST'])
def getocrsolutionmetadata():

    ocrsolname = request.headers["ocrsolname"]
    ocr_sol_path = root_path + 'ocrsolutions/' + ocrsolname + '.json'
    with open(ocr_sol_path) as f:
        ocrsolution_metadata = json.load(f)

    return ocrsolution_metadata



@app.route('/moveocrsoltopublished', methods = ['POST'])
def moveocrsoltopublished():
    ocrsolution = request.form["ocrsolution"]
    ocrsolution_cleaned = ocrsolution.split('$')[0]

    try:
        original = root_path + "/ocrsolutions/" + ocrsolution + ".json"
        target = root_path + "/published_ocr_solutions/" + ocrsolution + ".json"
        shutil.copyfile(original, target)

        url = "https://demo.aspace.ai/aspace/checkSolutionStatus?tool=eyesight&toolId=" + ocrsolution_cleaned

        payload = {}
        headers = {}

        response = requests.request("GET", url, headers=headers, data=payload, verify=False)

        print(response.text)

        json_string = json.loads(response.text)
        json_string = json.loads(json_string)
        print(json_string)
        print(json_string["SolutionPublished"])
        published_status = json_string["SolutionPublished"]

        if (published_status == "False"):
            src = root_path + "eyesight_docker"
            dst = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned
            shutil.copytree(src, dst)

            json_file_path = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned + "/configurations.json"
            with open(json_file_path, 'r+') as f:
                data = {}
                data['ocrsolution'] = ocrsolution_cleaned
                json.dump(data, f, indent=4)

            # shutil.make_archive(ocrsolution_cleaned, 'zip', "static/aspace_published_ocr_solutions/")
            zip_src = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned
            zip_dest = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned + ".zip"

            try:
                os.remove(zip_dest)
                print("Zip file replaced")
            except Exception as e:
                print ("Zip file placed")

            make_archive(zip_src, zip_dest)

            shutil.rmtree(zip_src)

            return_dict = {}
            return_dict["tool"] = "eyesight"
            return_dict["toolid"] = ocrsolution_cleaned
            return_dict["modelFileUrl"] = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/aspace_published_ocr_solutions/" + ocrsolution_cleaned + ".zip"
            return_dict["returnUrl"] = "http://aspace-web.eastus.cloudapp.azure.com/eyesight/ocrsolutionspage"

            return_json = json.dumps(return_dict, indent=4)

            print (return_json)

        else:
            print (json_string)
            AspaceSolutionId = json_string["AspaceSolutionId"]
            src = root_path + "eyesight_docker"
            dst = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned
            shutil.copytree(src, dst)

            json_file_path = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned + "/configurations.json"
            with open(json_file_path, 'r+') as f:
                data = {}
                data['ocrsolution'] = ocrsolution_cleaned
                json.dump(data, f, indent=4)

            # shutil.make_archive(ocrsolution_cleaned, 'zip', "static/aspace_published_ocr_solutions/")
            zip_src = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned
            zip_dest = root_path + "static/aspace_published_ocr_solutions/" + ocrsolution_cleaned + ".zip"

            try:
                os.remove(zip_dest)
                print("Zip file replaced")
            except Exception as e:
                print("Zip file placed")

            make_archive(zip_src, zip_dest)

            shutil.rmtree(zip_src)

            return_dict = {}
            return_dict["tool"] = "eyesight"
            return_dict["toolid"] = ocrsolution_cleaned
            return_dict["AspaceSolutionId"] = AspaceSolutionId
            return_dict["modelFileUrl"] = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/" + zip_dest
            return_dict["returnUrl"] = "http://aspace-web.eastus.cloudapp.azure.com/eyesight/ocrsolutionspage"

            return_json = json.dumps(return_dict, indent=4)

            print(return_json)

        print ("OCR solution published successfully")
        # return "OCR solution published successfully"
        return return_json
    except Exception as e:
        print (e)
        return "Cannot publish OCR solution"


@app.route('/removeocrsolfrompublished', methods = ['POST'])
def removeocrsolfrompublished():
    ocrsolution = request.form["ocrsolution"]
    ocrsolution_cleaned = ocrsolution.split('$')[0]

    try:
        original = root_path + "/published_ocr_solutions/" + ocrsolution + ".json"
        aspace_original = root_path + "/static/aspace_published_ocr_solutions/" + ocrsolution_cleaned + ".zip"
        os.remove(original)
        os.remove(aspace_original)
        print ("OCR solution unpublished successfully")
        return "OCR solution unpublished successfully"
    except Exception as e:
        print (e)
        return "Cannot unpublish OCR solution"



@app.route('/checkpublishstatus', methods = ['POST'])
def checkpublishstatus():
    ocrsolution = request.form["ocrsolution"]
    original = root_path + "/published_ocr_solutions/"
    files_list = os.listdir(original)
    for file in files_list:
        result = file.startswith(ocrsolution)
        if (result):
            return ("published")
        else:
            continue
    return ("unpublished")











def ocr_solution_initial_process_img(input_files_path, filename, folderpath, timestamp, models):
    try:
        # folderpath = root_path + 'static/testfiles/'
        # foldername = timestamp
        # filename = timestamp + '.jpg'

        filename_without_extension = os.path.splitext(filename)[0]
        folder_for_individual_file = folderpath + filename_without_extension + "/"
        # os.mkdir(folder_for_individual_file)

        copyfile(input_files_path + filename, folder_for_individual_file + filename)

        print ("input_files_path - " + input_files_path)
        print ("filename - " + filename)
        print ("folderpath - " + folderpath)
        print ("timestamp - " + timestamp)

        folderpath = root_path + "/static/initialreceivefiles_ocr_solution/"
        inputimgpath = folderpath + timestamp + '/' + filename_without_extension + '/' + filename
        inputpdfpath = folderpath + timestamp + '/' + filename_without_extension + '/' + timestamp + '.pdf'
        savepath = folderpath + timestamp + '/' + filename_without_extension + '/'
        linesremovedpath = folderpath + timestamp + '/' + filename_without_extension + '/linesremoved/'
        pdfpath = folderpath + timestamp + '/' + filename_without_extension + '/pdfs/'
        consolidatedpdffilename = timestamp + '.pdf'
        pdfstoresizepath = pdfpath + '/pdfstoresize/'

        folderpath_linesremoved = os.path.join(savepath, "linesremoved")
        os.mkdir(folderpath_linesremoved)
        folderpath_pdfs = os.path.join(savepath, "pdfs")
        os.mkdir(folderpath_pdfs)



        start = time.time()

        convert_img_to_pdf_status = convert_img_to_pdf(inputimgpath, savepath, timestamp)
        if (convert_img_to_pdf_status == "done"):
            convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
            if (convert_pdf_to_img_status == "done"):
                binarize_imgs_status = binarize_imgs(savepath)
                print(binarize_imgs_status)
                if (binarize_imgs_status == "done"):
                    skew_correct_status = skew_correct(savepath)
                    print(skew_correct_status)
                    if (skew_correct_status == "done"):
                        remove_lines_status = remove_lines(savepath)
                        print(remove_lines_status)
                        if (remove_lines_status == "done"):
                            get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                            print(get_txt_boxes_status)
                            if (get_txt_boxes_status == "done"):
                                convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath, language)
                                if (convert_imgs_into_pdf_status == "done"):
                                    print("done")
                                    merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                    if (merge_pdfs_status == "done"):

                                        # get model name from classifier
                                        # model_name = invoke_classifier(pdfpath + consolidatedpdffilename)
                                        model_names = invoke_classifier(savepath + "/output0.jpg",
                                                                       root_path + "/imagesforclassifier/", models)

                                        model_name = model_names[0]

                                        print("model name from classifier - " + str(model_name))
                                        modelpath = root_path + 'modelsnew/' + model_name

                                        with open(modelpath) as f:
                                            model_meta_data = json.load(f)
                                            print(model_meta_data)

                                        original_width = model_meta_data['originalwidth']
                                        original_height = model_meta_data['originalheight']

                                        page_dims = []

                                        for filename_in_path in os.listdir(savepath):
                                            if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
                                                    ".png"):
                                                if filename_in_path.startswith("o"):
                                                    current_page_dim = []
                                                    current_page_dim.append(original_width)
                                                    current_page_dim.append(original_height)
                                                    page_dims.append(current_page_dim)

                                        print("page_dims - " + str(page_dims))

                                        resize_pages_status = resize_pages(pdfpath, consolidatedpdffilename, pdfstoresizepath,
                                                                           page_dims,
                                                                           filename)

                                        if (resize_pages_status != "done"):
                                            return "failed in rescaling pdf pages"

                                        pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                        pdf.load()

                                        pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                        with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                            text_found_in_pdf = ""
                                            for page in doc:
                                                text_found_in_pdf += page.getText()

                                        print(text_found_in_pdf,  file=open(pdfpath + 'fulltext.txt', 'w'))

                                        print("fully done")
                                        # return "done"
                                        end = time.time()
                                        timetaken = end - start
                                        print(f"Runtime of the program is {end - start}")
                                        print(timestamp)
                                        with open(
                                                root_path + 'static/initialreceivefiles_ocr_solution/' + timestamp + '/' + filename_without_extension + '/final_json.json') as f:
                                            data = f.read()
                                        jsondata = json.dumps(data)

                                        return_model_metadata = []
                                        for modelname in model_names:
                                            model_path = root_path + 'modelsnew/' + modelname

                                            with open(model_path) as f:
                                                return_model_meta_data = json.load(f)
                                                return_model_metadata.append(return_model_meta_data)

                                        print("return_model_metadata - ", return_model_metadata)
                                        # return model_meta_data
                                        return return_model_metadata

                                        # print (jsondata)
                                        # return model_meta_data
                                        # return sendresponse("done", 200)
                                    else:
                                        print("failed in merge_pdfs")
                                        return "failed in merge_pdfs"
                                else:
                                    print("failed in convert_imgs_into_pdf")
                                    return "failed in convert_imgs_into_pdf"
                            else:
                                print("failed in get_txt_boxes")
                                return "failed in get_txt_boxes"
                        else:
                            print("failed in remove_lines")
                            return "failed in remove_lines"
                    else:
                        print("failed in skew_correct")
                        return "failed in skew_correct"
                else:
                    print("failed in binarize_imgs")
                    return "failed in binarize_imgs"
            else:
                print("failed in convert_pdf_to_img")
                return "failed in convert_pdf_to_img"
        else:
            print("failed in binarize_imgs")
            return "failed in binarize_imgs"

    except Exception as e:
        print (e)
        return "failed"


# def ocr_solution_initial_process_pdf(input_files_path, filename, folderpath, timestamp):
#     try:
#         # folderpath = root_path + 'static/testfiles/'
#         # foldername = timestamp
#         # filename = timestamp + '.jpg'
#
#         filename_without_extension = os.path.splitext(filename)[0]
#         folder_for_individual_file = folderpath + filename_without_extension + "/"
#         # os.mkdir(folder_for_individual_file)
#
#         copyfile(input_files_path + filename, folder_for_individual_file + timestamp + ".pdf")
#
#         print ("input_files_path - " + input_files_path)
#         print ("filename - " + filename)
#         print ("folderpath - " + folderpath)
#         print ("timestamp - " + timestamp)
#
#         folderpath = "C:/Users/salil/Desktop/eyesight V1/eyesight V1/static/initialreceivefiles_ocr_solution/"
#         inputimgpath = folderpath + timestamp + '/' + filename_without_extension + '/' + filename
#         inputpdfpath = folderpath + timestamp + '/' + filename_without_extension + '/' + timestamp + '.pdf'
#         inputpdfpath_for_pdf_process = folderpath + timestamp + '/' + filename_without_extension + '/' + filename_without_extension + '.pdf'
#         savepath = folderpath + timestamp + '/' + filename_without_extension + '/'
#         linesremovedpath = folderpath + timestamp + '/' + filename_without_extension + '/linesremoved/'
#         pdfpath = folderpath + timestamp + '/' + filename_without_extension + '/pdfs/'
#         consolidatedpdffilename = timestamp + '.pdf'
#         pdfstoresizepath = pdfpath + '/pdfstoresize/'
#
#         folderpath_linesremoved = os.path.join(savepath, "linesremoved")
#         os.mkdir(folderpath_linesremoved)
#         folderpath_pdfs = os.path.join(savepath, "pdfs")
#         os.mkdir(folderpath_pdfs)
#
#
#
#         start = time.time()
#
#         pdf_format = is_pdf_txt_or_img(inputpdfpath)
#         print(pdf_format)
#         if (pdf_format == "img"):
#             convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
#             if (convert_pdf_to_img_status == "done"):
#                 binarize_imgs_status = binarize_imgs(savepath)
#                 print("binarize_imgs_status - " + "done")
#                 if (binarize_imgs_status == "done"):
#                     skew_correct_status = skew_correct(savepath)
#                     print("skew_correct_status - " + "done")
#                     if (skew_correct_status == "done"):
#                         remove_lines_status = remove_lines(savepath)
#                         print("remove_lines_status - " + "done")
#                         if (remove_lines_status == "done"):
#                             get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
#                             print("get_txt_boxes_status - " + "done")
#                             if (get_txt_boxes_status == "done"):
#                                 convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath,
#                                                                                      language)
#                                 if (convert_imgs_into_pdf_status == "done"):
#                                     print("convert_imgs_into_pdf_status - " + "done")
#                                     merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
#                                     if (merge_pdfs_status == "done"):
#
#                                         # get model name from classifier
#                                         # model_name = invoke_classifier(pdfpath + consolidatedpdffilename)
#                                         model_name = invoke_classifier(savepath + "/output0.jpg",
#                                                                        root_path + "/imagesforclassifier/")
#                                         print("model name from classifier - " + str(model_name))
#                                         modelpath = root_path + 'modelsnew/' + model_name
#
#                                         with open(modelpath) as f:
#                                             model_meta_data = json.load(f)
#                                             print(model_meta_data)
#
#                                         original_width = model_meta_data['originalwidth']
#                                         original_height = model_meta_data['originalheight']
#
#                                         page_dims = []
#
#                                         for filename_in_path in os.listdir(savepath):
#                                             if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
#                                                     ".png"):
#                                                 if filename_in_path.startswith("o"):
#                                                     current_page_dim = []
#                                                     current_page_dim.append(original_width)
#                                                     current_page_dim.append(original_height)
#                                                     page_dims.append(current_page_dim)
#
#                                         print("page_dims - " + str(page_dims))
#
#                                         resize_pages_status = resize_pages(pdfpath, consolidatedpdffilename,
#                                                                            pdfstoresizepath,
#                                                                            page_dims,
#                                                                            filename)
#
#                                         if (resize_pages_status != "done"):
#                                             return "failed in rescaling pdf pages"
#
#                                         pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
#                                         pdf.load()
#
#                                         pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")
#
#                                         with fitz.open(pdfpath + consolidatedpdffilename) as doc:
#                                             text_found_in_pdf = ""
#                                             for page in doc:
#                                                 text_found_in_pdf += page.getText()
#
#                                         print(text_found_in_pdf, file=open(pdfpath + 'fulltext.txt', 'w'))
#
#                                         print("fully done")
#                                         # return "done"
#                                         end = time.time()
#                                         timetaken = end - start
#                                         print(f"Runtime of the program is {end - start}")
#                                         print(timestamp)
#                                         with open(
#                                                 root_path + 'static/initialreceivefiles_ocr_solution/' + timestamp + '/' + filename_without_extension + '/final_json.json') as f:
#                                             data = f.read()
#                                         jsondata = json.dumps(data)
#                                         # print(jsondata)
#                                         return model_meta_data
#
#                                     else:
#                                         print("failed in merge_pdfs")
#                                         return "failed in merge_pdfs"
#                                 else:
#                                     print("failed in convert_imgs_into_pdf")
#                                     return "failed in convert_imgs_into_pdf"
#                             else:
#                                 print("failed in get_txt_boxes")
#                                 return "failed in get_txt_boxes"
#                         else:
#                             print("failed in remove_lines")
#                             return "failed in remove_lines"
#                     else:
#                         print("failed in skew_correct")
#                         return "failed in skew_correct"
#                 else:
#                     print("failed in binarize_imgs")
#                     return "failed in binarize_imgs"
#             else:
#                 print("failed in convert_pdf_to_img")
#                 return "failed in convert_pdf_to_img"
#         elif (pdf_format == "txt"):
#             copyfile(inputpdfpath, pdfpath + consolidatedpdffilename)
#             convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
#             if (convert_pdf_to_img_status == "done"):
#                 binarize_imgs_status = binarize_imgs(savepath)
#                 print("binarize_imgs_status - " + binarize_imgs_status)
#                 if (binarize_imgs_status == "done"):
#                     skew_correct_status = skew_correct(savepath)
#                     print("skew_correct_status - " + skew_correct_status)
#                     if (skew_correct_status == "done"):
#                         remove_lines_status = remove_lines(savepath)
#                         print("remove_lines_status - " + remove_lines_status)
#                         if (remove_lines_status == "done"):
#                             get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
#                             print("get_txt_boxes_status - " + get_txt_boxes_status)
#                             if (get_txt_boxes_status == "done"):
#
#                                 # get model name from classifier
#                                 # model_name = invoke_classifier(pdfpath + consolidatedpdffilename)
#                                 model_name = invoke_classifier(savepath + "/output0.jpg", root_path + "/imagesforclassifier/")
#                                 print ("model name from classifier - " + str(model_name))
#                                 modelpath = root_path + 'modelsnew/' + model_name
#
#                                 with open(modelpath) as f:
#                                     model_meta_data = json.load(f)
#                                     print(model_meta_data)
#
#                                 original_width = model_meta_data['originalwidth']
#                                 original_height = model_meta_data['originalheight']
#
#                                 page_dims = []
#
#                                 for filename_in_path in os.listdir(savepath):
#                                     if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
#                                             ".png"):
#                                         if filename_in_path.startswith("o"):
#                                             current_page_dim = []
#                                             current_page_dim.append(original_width)
#                                             current_page_dim.append(original_height)
#                                             page_dims.append(current_page_dim)
#
#                                 print("page_dims - " + str(page_dims))
#
#                                 resize_pages_status = resize_pages(pdfpath, consolidatedpdffilename, pdfstoresizepath,
#                                                                    page_dims,
#                                                                    filename)
#
#                                 if (resize_pages_status != "done"):
#                                     return "failed in rescaling pdf pages"
#
#                                 pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
#                                 pdf.load()
#
#                                 pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")
#
#                                 with fitz.open(pdfpath + consolidatedpdffilename) as doc:
#                                     text_found_in_pdf = ""
#                                     for page in doc:
#                                         text_found_in_pdf += page.getText()
#
#                                 print(text_found_in_pdf, file=open(pdfpath + 'fulltext.txt', 'w'))
#
#                                 print("fully done")
#                                 # return "done"
#                                 end = time.time()
#                                 timetaken = end - start
#                                 print(f"Runtime of the program is {end - start}")
#                                 print(timestamp)
#                                 with open(
#                                         root_path + 'static/initialreceivefiles_ocr_solution/' + timestamp + '/' + filename_without_extension + '/final_json.json') as f:
#                                     data = f.read()
#                                 jsondata = json.dumps(data)
#                                 # print(jsondata)
#                                 return model_meta_data
#
#                             else:
#                                 print("failed in get_txt_boxes")
#                                 return "failed in get_txt_boxes"
#                         else:
#                             print("failed in remove_lines")
#                             return "failed in remove_lines"
#                     else:
#                         print("failed in skew_correct")
#                         return "failed in skew_correct"
#                 else:
#                     print("failed in binarize_imgs")
#                     return "failed in binarize_imgs"
#             else:
#                 print("failed in convert_pdf_to_img")
#                 return "failed in convert_pdf_to_img"
#
#
#     except Exception as e:
#         print (e)
#         return sendresponse("failed", 201)


def ocr_solution_initial_process_pdf(input_files_path, filename, folderpath, timestamp, models):
    try:
        # folderpath = root_path + 'static/testfiles/'
        # foldername = timestamp
        # filename = timestamp + '.jpg'

        filename_without_extension = os.path.splitext(filename)[0]
        folder_for_individual_file = folderpath + filename_without_extension + "/"
        # os.mkdir(folder_for_individual_file)

        copyfile(input_files_path + filename, folder_for_individual_file + timestamp + ".pdf")

        print ("input_files_path - " + input_files_path)
        print ("filename - " + filename)
        print ("folderpath - " + folderpath)
        print ("timestamp - " + timestamp)

        folderpath = root_path + "/static/initialreceivefiles_ocr_solution/"
        inputimgpath = folderpath + timestamp + '/' + filename_without_extension + '/' + filename
        inputpdfpath = folderpath + timestamp + '/' + filename_without_extension + '/' + timestamp + '.pdf'
        inputpdfpath_for_pdf_process = folderpath + timestamp + '/' + filename_without_extension + '/' + filename_without_extension + '.pdf'
        savepath = folderpath + timestamp + '/' + filename_without_extension + '/'
        linesremovedpath = folderpath + timestamp + '/' + filename_without_extension + '/linesremoved/'
        pdfpath = folderpath + timestamp + '/' + filename_without_extension + '/pdfs/'
        consolidatedpdffilename = timestamp + '.pdf'
        pdfstoresizepath = pdfpath + '/pdfstoresize/'

        folderpath_linesremoved = os.path.join(savepath, "linesremoved")
        os.mkdir(folderpath_linesremoved)
        folderpath_pdfs = os.path.join(savepath, "pdfs")
        os.mkdir(folderpath_pdfs)



        start = time.time()

        pdf_format = is_pdf_txt_or_img(inputpdfpath)
        print(pdf_format)
        if (pdf_format == "img"):
            convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
            if (convert_pdf_to_img_status == "done"):
                binarize_imgs_status = binarize_imgs(savepath)
                print("binarize_imgs_status - " + "done")
                if (binarize_imgs_status == "done"):
                    skew_correct_status = skew_correct(savepath)
                    print("skew_correct_status - " + "done")
                    if (skew_correct_status == "done"):
                        remove_lines_status = remove_lines(savepath)
                        print("remove_lines_status - " + "done")
                        if (remove_lines_status == "done"):
                            get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                            print("get_txt_boxes_status - " + "done")
                            if (get_txt_boxes_status == "done"):
                                convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath,
                                                                                     language)
                                if (convert_imgs_into_pdf_status == "done"):
                                    print("convert_imgs_into_pdf_status - " + "done")
                                    merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                    if (merge_pdfs_status == "done"):

                                        # get model name from classifier
                                        # model_name = invoke_classifier(pdfpath + consolidatedpdffilename)
                                        model_names = invoke_classifier(savepath + "/output0.jpg",
                                                                       root_path + "/imagesforclassifier/", models)

                                        model_name = model_names[0]

                                        print("model name from classifier - " + str(model_name))
                                        modelpath = root_path + 'modelsnew/' + model_name

                                        with open(modelpath) as f:
                                            model_meta_data = json.load(f)
                                            print(model_meta_data)

                                        original_width = model_meta_data['originalwidth']
                                        original_height = model_meta_data['originalheight']

                                        page_dims = []

                                        for filename_in_path in os.listdir(savepath):
                                            if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
                                                    ".png"):
                                                if filename_in_path.startswith("o"):
                                                    current_page_dim = []
                                                    current_page_dim.append(original_width)
                                                    current_page_dim.append(original_height)
                                                    page_dims.append(current_page_dim)

                                        print("page_dims - " + str(page_dims))

                                        resize_pages_status = resize_pages(pdfpath, consolidatedpdffilename,
                                                                           pdfstoresizepath,
                                                                           page_dims,
                                                                           filename)

                                        if (resize_pages_status != "done"):
                                            return "failed in rescaling pdf pages"

                                        pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                        pdf.load()

                                        pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                        with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                            text_found_in_pdf = ""
                                            for page in doc:
                                                text_found_in_pdf += page.getText()

                                        print(text_found_in_pdf, file=open(pdfpath + 'fulltext.txt', 'w'))

                                        print("fully done")
                                        # return "done"
                                        end = time.time()
                                        timetaken = end - start
                                        print(f"Runtime of the program is {end - start}")
                                        print(timestamp)
                                        with open(
                                                root_path + 'static/initialreceivefiles_ocr_solution/' + timestamp + '/' + filename_without_extension + '/final_json.json') as f:
                                            data = f.read()
                                        jsondata = json.dumps(data)
                                        # print(jsondata)

                                        return_model_metadata = []
                                        for modelname in model_names:
                                            model_path = root_path + 'modelsnew/' + modelname

                                            with open(model_path) as f:
                                                return_model_meta_data = json.load(f)
                                                return_model_metadata.append(return_model_meta_data)

                                        print ("return_model_metadata - ", return_model_metadata)
                                        # return model_meta_data
                                        return return_model_metadata

                                    else:
                                        print("failed in merge_pdfs")
                                        return "failed in merge_pdfs"
                                else:
                                    print("failed in convert_imgs_into_pdf")
                                    return "failed in convert_imgs_into_pdf"
                            else:
                                print("failed in get_txt_boxes")
                                return "failed in get_txt_boxes"
                        else:
                            print("failed in remove_lines")
                            return "failed in remove_lines"
                    else:
                        print("failed in skew_correct")
                        return "failed in skew_correct"
                else:
                    print("failed in binarize_imgs")
                    return "failed in binarize_imgs"
            else:
                print("failed in convert_pdf_to_img")
                return "failed in convert_pdf_to_img"
        elif (pdf_format == "txt"):
            copyfile(inputpdfpath, pdfpath + consolidatedpdffilename)
            convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, savepath)
            if (convert_pdf_to_img_status == "done"):
                binarize_imgs_status = binarize_imgs(savepath)
                print("binarize_imgs_status - " + binarize_imgs_status)
                if (binarize_imgs_status == "done"):
                    skew_correct_status = skew_correct(savepath)
                    print("skew_correct_status - " + skew_correct_status)
                    if (skew_correct_status == "done"):
                        remove_lines_status = remove_lines(savepath)
                        print("remove_lines_status - " + remove_lines_status)
                        if (remove_lines_status == "done"):
                            get_txt_boxes_status = get_txt_boxes(linesremovedpath, savepath, language)
                            print("get_txt_boxes_status - " + get_txt_boxes_status)
                            if (get_txt_boxes_status == "done"):

                                # get model name from classifier
                                # model_name = invoke_classifier(pdfpath + consolidatedpdffilename)
                                model_names = invoke_classifier(savepath + "/output0.jpg", root_path + "/imagesforclassifier/", models)

                                model_name = model_names[0]
                                print ("model_names - " + str(model_names))
                                print ("model name from classifier - " + str(model_name))
                                modelpath = root_path + 'modelsnew/' + model_name

                                with open(modelpath) as f:
                                    model_meta_data = json.load(f)
                                    print(model_meta_data)

                                original_width = model_meta_data['originalwidth']
                                original_height = model_meta_data['originalheight']

                                page_dims = []

                                for filename_in_path in os.listdir(savepath):
                                    if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
                                            ".png"):
                                        if filename_in_path.startswith("o"):
                                            current_page_dim = []
                                            current_page_dim.append(original_width)
                                            current_page_dim.append(original_height)
                                            page_dims.append(current_page_dim)

                                print("page_dims - " + str(page_dims))

                                resize_pages_status = resize_pages(pdfpath, consolidatedpdffilename, pdfstoresizepath,
                                                                   page_dims,
                                                                   filename)

                                if (resize_pages_status != "done"):
                                    return "failed in rescaling pdf pages"

                                pdf = pdfquery.PDFQuery(pdfpath + consolidatedpdffilename)
                                pdf.load()

                                pdf.tree.write(pdfpath + "pdfxml.xml", pretty_print=True, encoding="utf-8")

                                with fitz.open(pdfpath + consolidatedpdffilename) as doc:
                                    text_found_in_pdf = ""
                                    for page in doc:
                                        text_found_in_pdf += page.getText()

                                print(text_found_in_pdf, file=open(pdfpath + 'fulltext.txt', 'w'))

                                print("fully done")
                                # return "done"
                                end = time.time()
                                timetaken = end - start
                                print(f"Runtime of the program is {end - start}")
                                print(timestamp)
                                with open(
                                        root_path + 'static/initialreceivefiles_ocr_solution/' + timestamp + '/' + filename_without_extension + '/final_json.json') as f:
                                    data = f.read()
                                jsondata = json.dumps(data)
                                # print(jsondata)

                                return_model_metadata = []
                                for modelname in model_names:
                                    model_path = root_path + 'modelsnew/' + modelname

                                    with open(model_path) as f:
                                        return_model_meta_data = json.load(f)
                                        return_model_metadata.append(return_model_meta_data)

                                print("return_model_metadata - ", return_model_metadata)
                                # return model_meta_data
                                return return_model_metadata



                            else:
                                print("failed in get_txt_boxes")
                                return "failed in get_txt_boxes"
                        else:
                            print("failed in remove_lines")
                            return "failed in remove_lines"
                    else:
                        print("failed in skew_correct")
                        return "failed in skew_correct"
                else:
                    print("failed in binarize_imgs")
                    return "failed in binarize_imgs"
            else:
                print("failed in convert_pdf_to_img")
                return "failed in convert_pdf_to_img"


    except Exception as e:
        print (e)
        return sendresponse("failed", 201)


def ocr_solution_get_data_from_img(model_meta_data, packagepath, timestamp):
    final_keyvalue_pairs = {}

    for keyvalue in model_meta_data['keyvalues']:
        print(keyvalue)
        type = keyvalue["type"]
        label = keyvalue["label"]

        if (type == "keybased"):
            print("value extraction type - keybased")
            captured_value = runtime_keybased_extract(keyvalue, packagepath, timestamp)
            print("extracted keybased value - " + str(captured_value))

            print("len(keyvalue['post_processing'])")
            print(len(keyvalue["post_processing"]))

            if (len(keyvalue["post_processing"]) != 0):
                post_processed_text = ""
                for i in range(len(keyvalue["post_processing"])):
                    print("i - " + str(i))
                    conditionname = "post_processing_condition_" + str(i + 1)
                    condition = keyvalue["post_processing"][conditionname]
                    if (post_processed_text == ""):
                        post_processed_text = captured_value
                    post_processed_text = runtime_post_process(condition, post_processed_text)

                print("post_processed_text - " + str(post_processed_text))
                final_keyvalue_pairs[label] = post_processed_text
            else:
                final_keyvalue_pairs[label] = captured_value

        elif (type == "regionbased"):
            print("value extraction type - regionbased")
            captured_value = runtime_regionbased_extract(keyvalue, packagepath, timestamp)
            print("extracted regionbased value - " + str(captured_value))

            print("len(keyvalue['post_processing'])")
            print(len(keyvalue["post_processing"]))

            if (len(keyvalue["post_processing"]) != 0):
                post_processed_text = ""
                for i in range(len(keyvalue["post_processing"])):
                    print("i - " + str(i))
                    conditionname = "post_processing_condition_" + str(i + 1)
                    condition = keyvalue["post_processing"][conditionname]
                    if (post_processed_text == ""):
                        post_processed_text = captured_value
                    post_processed_text = runtime_post_process(condition, post_processed_text)

                print("post_processed_text - " + str(post_processed_text))
                final_keyvalue_pairs[label] = post_processed_text
            else:
                final_keyvalue_pairs[label] = captured_value

        elif (type == "fullocr"):
            print("value extraction type - fullocr")
            captured_value = runtime_fullocr_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

        elif (type == "static"):
            print("value extraction type - static")
            captured_value = runtime_static_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

        elif (type == "fromdb"):
            print("value extraction type - fromdb")
            captured_value = runtime_fromdb_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

    # print(final_keyvalue_pairs)
    return final_keyvalue_pairs


def ocr_solution_get_data_from_pdf(model_meta_data, packagepath, timestamp, individual_foldername, table_output=None):
    original_height = model_meta_data['originalheight']

    final_return_json = {}


    final_keyvalue_pairs = {}
    for keyvalue in model_meta_data['keyvalues']:
        print(keyvalue)
        type = keyvalue["type"]
        label = keyvalue["label"]

        if (type == "keybased"):
            print("value extraction type - keybased")
            # captured_value = runtime_keybased_extract(keyvalue, packagepath, timestamp)
            captured_value = runtime_keybased_extract(keyvalue, packagepath, individual_foldername)
            print("extracted keybased value - " + str(captured_value))

            print("len(keyvalue['post_processing'])")
            print(len(keyvalue["post_processing"]))

            if (len(keyvalue["post_processing"]) != 0):
                post_processed_text = ""
                for i in range(len(keyvalue["post_processing"])):
                    print("i - " + str(i))
                    conditionname = "post_processing_condition_" + str(i + 1)
                    condition = keyvalue["post_processing"][conditionname]
                    if (post_processed_text == ""):
                        post_processed_text = captured_value
                    post_processed_text = runtime_post_process(condition, post_processed_text)

                print("post_processed_text - " + str(post_processed_text))
                final_keyvalue_pairs[label] = post_processed_text
            else:
                final_keyvalue_pairs[label] = captured_value

        elif (type == "regionbased"):
            print("value extraction type - regionbased")
            captured_value = runtime_regionbased_extract(keyvalue, packagepath, timestamp)
            # captured_value = runtime_regionbased_extract(keyvalue, packagepath, individual_foldername)
            print("extracted regionbased value - " + str(captured_value))

            print("len(keyvalue['post_processing'])")
            print(len(keyvalue["post_processing"]))

            if (len(keyvalue["post_processing"]) != 0):
                post_processed_text = ""
                for i in range(len(keyvalue["post_processing"])):
                    print("i - " + str(i))
                    conditionname = "post_processing_condition_" + str(i + 1)
                    condition = keyvalue["post_processing"][conditionname]
                    if (post_processed_text == ""):
                        post_processed_text = captured_value
                    post_processed_text = runtime_post_process(condition, post_processed_text)

                print("post_processed_text - " + str(post_processed_text))
                final_keyvalue_pairs[label] = post_processed_text
            else:
                final_keyvalue_pairs[label] = captured_value

        elif (type == "fullocr"):
            print("value extraction type - fullocr")
            captured_value = runtime_fullocr_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

        elif (type == "static"):
            print("value extraction type - static")
            captured_value = runtime_static_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

        elif (type == "fromdb"):
            print("value extraction type - fromdb")
            captured_value = runtime_fromdb_extract(keyvalue, packagepath, timestamp)
            final_keyvalue_pairs[label] = captured_value

    # print ("final_keyvalue_pairs - ", final_keyvalue_pairs)
    final_return_json["keyvalues"] = final_keyvalue_pairs

    final_table_json = {}
    for table in model_meta_data['tables']:
        table_type = table["table_type"]
        table_name = table["table_name"]
        column_details = table["columns_list"]

        column_list = []
        for column in column_details:
            column_name = column[3]
            column_list.append(column_name)

        print ("column_list - ", column_list)

        if (table_type == "withoutborder"):
            print ("without border")
            table_output = gettablewithoutborder_ocrsolution_runtime(table, packagepath, timestamp, original_height, column_list)
            # print ("gettablewithoutborder_ocrsolution_runtime - ", table_output)
            final_table_json[table_name] = table_output
        elif (table_type == "withborder"):
            print ("with border")
    final_return_json["tables"] = final_table_json







    print("final_return_json - ", final_return_json)
    return final_return_json





def invoke_classifier(input_image, directory, models):
    print ("models in classifier - ", models)
    scoreFileDict = {}
    if input_image.endswith(".pdf"):
        pdftrainFileImages = pdf2image.convert_from_path(input_image)
        for page in pdftrainFileImages:
            for count in range(0, 1):
                page.save('input_image.png', 'PNG')
                input_image = 'input_image.png'

    for testImage in os.listdir(directory):

        image_title = str(testImage).replace(".png", "")
        image_title = image_title.split('$')[0]
        print("image_title - ", image_title)

        if (image_title in models):

            imageB = cv2.imread(os.path.join(directory, testImage))
            if (testImage.endswith(".pdf") or testImage.endswith(".PDF")):
                pdftestFileImages = convert_from_path(os.path.join(directory, testImage))
                for page in pdftestFileImages:
                    for count in range(0, 1):
                        page.save(os.path.join(directory, testImage + '.png'), 'PNG')
                        imageB = cv2.imread(os.path.join(directory, (testImage + '.png')))
            # load the two input images
            imageA = cv2.imread(input_image)
            inputImageResizeA = cv2.resize(imageA, (800, 800))
            existingImageResizeB = cv2.resize(imageB, (800, 800))
            # convert the images to grayscale
            grayA = cv2.cvtColor(inputImageResizeA, cv2.COLOR_BGR2GRAY)
            grayB = cv2.cvtColor(existingImageResizeB, cv2.COLOR_BGR2GRAY)
            # compute the Structural Similarity Index (SSIM) between the two
            # images, ensuring that the difference image is returned
            (score, diff) = ssim(grayA, grayB, full=True)
            diff = (diff * 255).astype("uint8")
            scoreFileDict[testImage] = score

            # threshold the difference image, followed by finding contours to
            # obtain the regions of the two input images that differ
            thresh = cv2.threshold(diff, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
            cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = imutils.grab_contours(cnts)
            # loop over the contours
            for c in cnts:
                # compute the bounding box of the contour and then draw the
                # bounding box on both input images to represent where the two
                # images differ
                (x, y, w, h) = cv2.boundingRect(c)
                cv2.rectangle(imageA, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.rectangle(imageB, (x, y), (x + w, y + h), (0, 0, 255), 2)
    mostMatchedScore = max(scoreFileDict, key=scoreFileDict.get)
    print ("mostMatchedScore - ", mostMatchedScore)


    final_list_of_models = []
    values = list(scoreFileDict.values())
    max_score_value = max(values)
    for key in scoreFileDict:
        value = scoreFileDict[key]
        if (str(value).strip() == str(max_score_value).strip()):
            key = key.replace(".png", ".json")
            final_list_of_models.append(key)

    print ("final_list_of_models - ", final_list_of_models)
    print ("scoreFileDict")
    print(scoreFileDict)

    mostMatchedScore = mostMatchedScore.replace(".png", ".json")
    # return (mostMatchedScore)
    return (final_list_of_models)


def gettablewithoutborder_ocrsolution_runtime(table, packagepath, timestamp, originalheight, column_list):
    table_name = table["table_name"]
    header = table["header"].strip()
    header_included_in_table = table["header_included_in_table"]
    footer = table["footer"].strip()
    footer_included_in_table = table["footer_included_in_table"]
    header_availability = table["header_availability"]
    footer_availability = table["footer_availability"]
    is_multipage = table["is_multipage"]
    top_margin = table["top_margin"]
    bottom_margin = table["bottom_margin"]
    group_into_row = table["group_into_row"]
    group_into_column = table["group_into_column"]
    detect_superscripts = table["detect_superscripts"]
    cut_text = table["cut_text"]
    text_edge_tol = table["text_edge_tol"]

    print ("header - " + str(header))
    print ("footer - " + str(footer))

    print ("packagepath - ", packagepath)

    pdf = pdfquery.PDFQuery(packagepath + "/" + timestamp + ".pdf")
    pdf.load()
    pdf.tree.write(packagepath + "/pdfs/pdfxml.xml", pretty_print=True, encoding="utf-8")

    xml_doc = etree.parse(packagepath + "/pdfs/pdfxml.xml")
    print (xml_doc)
    root = xml_doc.getroot()

    bbox_of_header_key = {}
    for elem in root.iter():
        text1 = elem.text
        # print ("header text - " + str(text1))
        try:
            if (header in text1):
                # print("header text - " + str(text1))
                box = elem.attrib.get("bbox")
                # print("header box - " + str(box))
                # bbox_of_key.append(box)
                pageno = getpagno(box, root)
                # print("header pageno " + str(pageno))
                bbox_of_header_key[pageno] = box
        except Exception as e:
            print (e)
            continue
    print("bbox_of_header_key - " + str(bbox_of_header_key))

    bbox_of_footer_key = {}
    for elem in root.iter():
        text = elem.text
        # print("footer text - " + str(text))
        try:
            if (footer in text):
                box = elem.attrib.get("bbox")
                # print("footer box - " + str(box))
                # bbox_of_key.append(box)
                pageno = getpagno(box, root)
                # print("footer pageno " + str(pageno))
                bbox_of_footer_key[pageno] = box
        except Exception as e:
            # print (e)
            continue
    print("bbox_of_footer_key - " + str(bbox_of_footer_key))

    bbox_of_header_key_list = [(k, v) for k, v in bbox_of_header_key.items()]
    bbox_of_footer_key_list = [(k, v) for k, v in bbox_of_footer_key.items()]

    print("bbox_of_header_key_list " + str(bbox_of_header_key_list))
    print("bbox_of_footer_key_list " + str(bbox_of_footer_key_list))

    os.mkdir(packagepath + "tableoutputs")

    # multi page table
    if (is_multipage == "True"):
        if ((header_availability == "True") & (footer_availability == "False")):
            print("header available")

            fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

            final_df_list = []

            for i in range(len(bbox_of_header_key_list)):
                header_block = bbox_of_header_key_list[i]
                header_page = int(header_block[0])
                header_rect = header_block[1]
                header_rect = header_rect.strip('][').split(', ')
                # print("header_page " + str(header_page))

                # footer_box = bbox_of_footer_key_list[header_page]
                # print (footer_box)

                header_starting_point = header_rect[1]
                header_ending_point = header_rect[3]
                for footer_elements in bbox_of_footer_key_list:
                    footer_page = int(footer_elements[0])
                    # print (type(footer_page))
                    if (footer_page == header_page):
                        print("footer_page == header_page")
                        footer_rect = footer_elements[1]
                        footer_rect_list = footer_rect.strip('][').split(', ')
                        footer_starting_point = footer_rect_list[1]
                        footer_ending_point = footer_rect_list[3]
                        print("footer_rect " + str(type(footer_rect)))
                        print("footer_starting_point " + str(footer_starting_point))
                        print("footer_ending_point " + str(footer_ending_point))
                    else:
                        footer_starting_point = int(bottom_margin) + 10
                        footer_ending_point = int(bottom_margin)

                # print ("footer_rect " + str(footer_rect))
                # print ("footer_starting_point " + str(footer_starting_point))

                final_area = []
                final_area.append(0)
                if (header_included_in_table == "True"):
                    final_area.append(float(header_starting_point) - 200)
                else:
                    final_area.append(float(header_ending_point) + 10)
                # if (footer_included_in_table == "True"):
                #     final_area.append(float(footer_starting_point) - 10)
                # else:
                #     final_area.append(float(footer_ending_point) + 10)
                final_area.append(5000)
                if (footer_included_in_table == "True"):
                    final_area.append(float(footer_ending_point) - 200)
                else:
                    final_area.append(float(footer_starting_point) + 10)
                # if (header_included_in_table == "True"):
                #     final_area.append(float(header_starting_point) - 10)
                # else:
                #     final_area.append(float(header_ending_point) + 10)

                print("final_area " + str(final_area))

                final_area = str(final_area)
                final_area = final_area.replace("[", "")
                final_area = final_area.replace("]", "")
                final_area_list = []
                final_area_list.append(final_area)
                print("final_area_list " + str(final_area_list))

                pdf_file_path = packagepath + "/pdfs/" + str(timestamp) + ".pdf"

                try:
                    print("------------")
                    print(header_page)
                    print(final_area_list)
                    print(pdf_file_path)
                    print("------------")

                    table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                  pages=str(header_page), split_text=cut_text,
                                                  table_areas=final_area_list,
                                                  flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                  row_tol=int(group_into_row))
                    nooftables = len(table_list)
                    print("nooftables_firstpage")
                    print(nooftables)

                    for x in range(nooftables):
                        df = table_list[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        df = df.loc[:, df.isin([' ', 'NULL', '', math.nan]).mean() < .5]
                        # df.dropna(axis=1, how='any')
                        final_df_list.append(df)

                except Exception as e:
                    print(e)
                    print("e")
                    continue

                # footer_block = bbox_of_footer_key_list[0]
                # footer_page = int(footer_block[0])
                # footer_rect = footer_block[1]
                # footer_rect = footer_rect.strip('][').split(', ')
                # print("footer_page " + str(footer_page))
                #
                # header_starting_point = header_rect[1]
                # header_ending_point = header_rect[3]
                #
                # footer_starting_point = footer_rect[1]
                # footer_ending_point = footer_rect[3]
                #
                # final_area = []
                #
                # if (header_included_in_table == "True"):
                #     final_area.append(0)
                #     final_area.append(float(header_starting_point) - 10)
                #     if (header_page == footer_page):
                #         final_area.append(2000)
                #         final_area.append(int(footer_starting_point))
                #     else:
                #         final_area.append(2000)
                #         final_area.append(int(bottom_margin))
                # else:
                #     final_area.append(0)
                #     final_area.append(float(header_ending_point) + 10)
                #     if (header_page == footer_page):
                #         final_area.append(2000)
                #         final_area.append(int(float(footer_starting_point)))
                #     else:
                #         final_area.append(2000)
                #         final_area.append(int(bottom_margin))
                #
                # print("final_area " + str(final_area))
                #
                # final_area = str(final_area)
                # final_area = final_area.replace("[", "")
                # final_area = final_area.replace("]", "")
                # final_area_list = []
                # final_area_list.append(final_area)
                # print("final_area_first_page_list " + str(final_area_list))

            print(len(final_df_list))

            result_df = pd.concat(final_df_list)

            for column in result_df:
                # print(result_df[column])
                result_df[column] = result_df[column].replace('<s>', '', regex=True)

            for column in result_df:
                # print(result_df[column])
                result_df[column] = result_df[column].replace('</s>', '', regex=True)

            result_df = result_df.apply(lambda x: sorted(x, key=pd.isnull), 1)

            # os.mkdir(packagepath + "tableoutputs")

            csv_file = packagepath + "tableoutputs/" + timestamp + '.csv'
            csv_file_cleaned = packagepath + "tableoutputs/" + timestamp + '_cleaned.csv'

            result_df.to_csv(csv_file)
            with open(csv_file, 'r', encoding="utf8") as infile, \
                    open(csv_file_cleaned, 'w', encoding="utf8") as outfile:
                data = infile.read()
                data = data.replace("[", "")
                data = data.replace("]", "")
                data = data.replace("'", "")
                data = data.replace(", nan", " ")
                data = data.replace("\\n", "")

                outfile.write(data)
            copyfile(csv_file_cleaned, packagepath + "/outputs/capturedtables.csv")
            # print (result_df)
            html = pd.DataFrame(result_df).to_html()
            # fullhtml += html
            fullhtml += html

            with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                file.write(fullhtml)

            return fullhtml

        elif ((header_availability == "False") & (footer_availability == "False")):

            if (header != ""):
                print("header available")
                fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"
                final_df_list = []

                for i in range(len(bbox_of_header_key_list)):
                    header_block = bbox_of_header_key_list[i]
                    header_page = int(header_block[0])
                    header_rect = header_block[1]
                    header_rect = header_rect.strip('][').split(', ')

    # single page table
    if (is_multipage == "False"):
        if (len(bbox_of_header_key_list) == len(bbox_of_footer_key_list)):

            # full_html_for_multiple_page_table = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

            for i in range(len(bbox_of_header_key_list)):
                header_block = bbox_of_header_key_list[i]
                header_page = int(header_block[0])
                header_rect = header_block[1]
                header_rect = header_rect.strip('][').split(', ')
                print("header_page " + str(header_page))

                footer_block = bbox_of_footer_key_list[i]
                footer_page = int(footer_block[0])
                footer_rect = footer_block[1]
                footer_rect = footer_rect.strip('][').split(', ')
                print("footer_page " + str(footer_page))

                if (header_page == footer_page):
                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]

                    print(header_starting_point)
                    print(header_ending_point)

                    footer_starting_point = footer_rect[1]
                    footer_ending_point = footer_rect[3]

                    print(footer_starting_point)
                    print(footer_ending_point)

                    final_area = []

                    if (header_included_in_table == "True"):
                        final_area.append(0)
                        final_area.append(float(header_starting_point) - 10)
                    else:
                        final_area.append(0)
                        final_area.append(float(header_ending_point) + 10)

                    if (footer_included_in_table == "True"):
                        final_area.append(5000)
                        final_area.append(float(footer_ending_point))
                    else:
                        final_area.append(5000)
                        final_area.append(float(footer_starting_point))

                    final_area = str(final_area)
                    final_area = final_area.replace("[", "")
                    final_area = final_area.replace("]", "")
                    final_area_list = []
                    final_area_list.append(final_area)
                    print(type(final_area))
                    print("final_area " + str(final_area_list))
                    # tablepositionlist = ['0, 1064.808, 2000, 44.658']
                    # print (type(tablepositionlist[0]))

                    pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"

                    print("here")

                    table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n', pages=str(header_page),
                                                  split_text=cut_text, table_areas=final_area_list,
                                                  flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                  row_tol=int(group_into_row))

                    nooftables = len(table_list)

                    print("nooftables")
                    print(nooftables)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

                    for x in range(nooftables):

                        df = table_list[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])

                        for column in df:
                            # print(result_df[column])
                            df[column] = df[column].replace('<s>', '', regex=True)

                        for column in df:
                            # print(result_df[column])
                            df[column] = df[column].replace('</s>', '', regex=True)

                        # df = df.apply(lambda x: sorted(x, key=pd.isnull), 1)

                        # csv_file = root_path + 'static/finaloutputtables/' + timestamp + '.csv'
                        # csv_file_cleaned = root_path + 'static/finaloutputtables/' + timestamp + '_cleaned.csv'

                        # os.mkdir(packagepath + "tableoutputs")

                        csv_file = packagepath + "tableoutputs/" + timestamp + '.csv'
                        csv_file_cleaned = packagepath + "tableoutputs/" + timestamp + '_cleaned.csv'

                        df.to_csv(csv_file)
                        with open(csv_file, 'r') as infile, \
                                open(csv_file_cleaned, 'w') as outfile:
                            data = infile.read()
                            data = data.replace("[", "")
                            data = data.replace("]", "")
                            data = data.replace("'", "")
                            outfile.write(data)

                        # copyfile(csv_file_cleaned, packagepath + "/outputs/capturedtables.csv")
                        #
                        # if (x == 0):
                        #     html = pd.DataFrame(df).to_html()
                        #     fullhtml += html
                        # else:
                        #     tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                        #     fullhtml += tabletitle
                        #     html = pd.DataFrame(df).to_html()
                        #     fullhtml += html

                    # with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                    #     file.write(fullhtml)
                    #

                    df.columns = column_list

                    # print(df.to_dict('records'))
                    return df.to_dict('records')

                # elif (footer_page == (header_page + 1)):
                #     print("table spawns in 2 pages")
                #
                #     header_starting_point = header_rect[1]
                #     header_ending_point = header_rect[3]
                #     footer_starting_point = footer_rect[1]
                #     footer_ending_point = footer_rect[3]
                #
                #     final_area_first_page = []
                #     final_area_second_page = []
                #
                #     if (header_included_in_table == "True"):
                #         final_area_first_page.append(0)
                #         final_area_first_page.append(float(header_starting_point) - 10)
                #         final_area_first_page.append(5000)
                #         final_area_first_page.append(int(bottom_margin))
                #     else:
                #         final_area_first_page.append(0)
                #         final_area_first_page.append(float(header_ending_point) + 10)
                #         final_area_first_page.append(5000)
                #         final_area_first_page.append(int(bottom_margin))
                #
                #     if (footer_included_in_table == "True"):
                #         final_area_second_page.append(0)
                #         final_area_second_page.append(int(top_margin))
                #         final_area_second_page.append(5000)
                #         final_area_second_page.append(float(footer_ending_point))
                #     else:
                #         final_area_second_page.append(0)
                #         final_area_second_page.append(originalheight - (int(top_margin)))
                #         final_area_second_page.append(5000)
                #         final_area_second_page.append(float(footer_starting_point))
                #
                #     print(final_area_first_page)
                #     print(final_area_second_page)
                #
                #     # print ("tick")
                #
                #     final_area_first_page = str(final_area_first_page)
                #     final_area_first_page = final_area_first_page.replace("[", "")
                #     final_area_first_page = final_area_first_page.replace("]", "")
                #     final_area_first_page_list = []
                #     final_area_first_page_list.append(final_area_first_page)
                #     print(final_area_first_page)
                #
                #     final_area_second_page = str(final_area_second_page)
                #     final_area_second_page = final_area_second_page.replace("[", "")
                #     final_area_second_page = final_area_second_page.replace("]", "")
                #     final_area_second_page_list = []
                #     final_area_second_page_list.append(final_area_second_page)
                #     print(final_area_second_page)
                #
                #     pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"
                #
                #     table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                #                                              pages=str(header_page), split_text=cut_text,
                #                                              table_areas=final_area_first_page_list,
                #                                              flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                #                                              row_tol=int(group_into_row))
                #     nooftables_firstpage = len(table_list_first_page)
                #
                #     table_list_second_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                #                                               pages=str(footer_page), split_text=cut_text,
                #                                               table_areas=final_area_second_page_list,
                #                                               flag_size=detect_superscripts,
                #                                               edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                #     nooftables_secondpage = len(table_list_second_page)
                #
                #     print("tick")
                #
                #     print("nooftables_firstpage")
                #     print(nooftables_firstpage)
                #
                #     print("nooftables_secondpage")
                #     print(nooftables_secondpage)
                #
                #     fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"
                #
                #     df_list = []
                #
                #     for x in range(nooftables_firstpage):
                #         df = table_list_first_page[x].df
                #         df.rename(columns=df.iloc[0]).drop(df.index[0])
                #         df_list.append(df)
                #         # if (x == 0):
                #         #     html = df.to_html()
                #         #     fullhtml += html
                #         # else:
                #         #     tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                #         #     fullhtml += tabletitle
                #         #     html = df.to_html()
                #         #     fullhtml += html
                #
                #     for y in range(nooftables_secondpage):
                #         df = table_list_second_page[y].df
                #         df.rename(columns=df.iloc[0]).drop(df.index[0])
                #         df_list.append(df)
                #         # if (y == 0):
                #         #     html = df.to_html()
                #         #     fullhtml += html
                #         # else:
                #         #     tabletitle = "<h1>Table " + str(y - 1) + "</h1>"
                #         #     fullhtml += tabletitle
                #         #     html = df.to_html()
                #         #     fullhtml += html
                #
                #     # for x in range(nooftables_firstpage):
                #     #
                #     #     df = table_list_first_page[x].df
                #     #     df.rename(columns=df.iloc[0]).drop(df.index[0])
                #     #     df_list.append(df)
                #     #     if (x == 0):
                #     #         html = df.to_html()
                #     #         fullhtml += html
                #     #     else:
                #     #         tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                #     #         fullhtml += tabletitle
                #     #         html = df.to_html()
                #     #         fullhtml += html
                #
                #     result_df = pd.concat(df_list)
                #
                #     for column in result_df:
                #         # print(result_df[column])
                #         result_df[column] = result_df[column].replace('<s>', '', regex=True)
                #
                #     for column in result_df:
                #         # print(result_df[column])
                #         result_df[column] = result_df[column].replace('</s>', '', regex=True)
                #
                #     # print (result_df)
                #     html = result_df.to_html()
                #     fullhtml += html
                #
                #     with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                #         file.write(fullhtml)
                #
                #     return fullhtml
                #
                # else:
                #     print("table spawns in more than 2 pages")
                #
                #     header_starting_point = header_rect[1]
                #     header_ending_point = header_rect[3]
                #     footer_starting_point = footer_rect[1]
                #     footer_ending_point = footer_rect[3]
                #
                #     final_area_first_page = []
                #     final_area_last_page = []
                #
                #     if (header_included_in_table == "True"):
                #         final_area_first_page.append(0)
                #         final_area_first_page.append(float(header_starting_point) - 10)
                #         final_area_first_page.append(5000)
                #         final_area_first_page.append(int(bottom_margin))
                #     else:
                #         final_area_first_page.append(0)
                #         final_area_first_page.append(float(header_ending_point) + 10)
                #         final_area_first_page.append(5000)
                #         final_area_first_page.append(int(bottom_margin))
                #
                #     if (footer_included_in_table == "True"):
                #         final_area_last_page.append(0)
                #         final_area_last_page.append(int(top_margin))
                #         final_area_last_page.append(5000)
                #         final_area_last_page.append(float(footer_ending_point))
                #     else:
                #         final_area_last_page.append(0)
                #         final_area_last_page.append(originalheight - (int(top_margin)))
                #         final_area_last_page.append(5000)
                #         final_area_last_page.append(float(footer_starting_point))
                #
                #     print("final_area_first_page " + str(final_area_first_page))
                #     print("final_area_last_page " + str(final_area_last_page))
                #
                #     print("tick")
                #
                #     final_df_list = []
                #
                #     final_area_first_page = str(final_area_first_page)
                #     final_area_first_page = final_area_first_page.replace("[", "")
                #     final_area_first_page = final_area_first_page.replace("]", "")
                #     final_area_first_page_list = []
                #     final_area_first_page_list.append(final_area_first_page)
                #     print("final_area_first_page_list " + str(final_area_first_page_list))
                #
                #     final_area_last_page = str(final_area_last_page)
                #     final_area_last_page = final_area_last_page.replace("[", "")
                #     final_area_last_page = final_area_last_page.replace("]", "")
                #     final_area_last_page_list = []
                #     final_area_last_page_list.append(final_area_last_page)
                #     print("final_area_last_page_list " + str(final_area_last_page_list))
                #
                #     pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"
                #
                #     table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                #                                              pages=str(header_page), split_text=cut_text,
                #                                              table_areas=final_area_first_page_list,
                #                                              flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                #                                              row_tol=int(group_into_row))
                #     nooftables_firstpage = len(table_list_first_page)
                #     print("nooftables_firstpage")
                #     print(nooftables_firstpage)
                #
                #     table_list_last_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                #                                             pages=str(footer_page), split_text=cut_text,
                #                                             table_areas=final_area_last_page_list,
                #                                             flag_size=detect_superscripts,
                #                                             edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                #     nooftables_lastpage = len(table_list_last_page)
                #     print("nooftables_secondpage")
                #     print(nooftables_lastpage)
                #
                #     fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"
                #
                #     for x in range(nooftables_firstpage):
                #         df = table_list_first_page[x].df
                #         df.rename(columns=df.iloc[0]).drop(df.index[0])
                #         final_df_list.append(df)
                #
                #     pages_in_between_list = []
                #     for pages_in_between in range(header_page, footer_page + 1):
                #         if (pages_in_between != header_page):
                #             if (pages_in_between != (footer_page)):
                #                 # print ("pages_in_between " + str(pages_in_between))
                #                 pages_in_between_list.append(pages_in_between)
                #
                #     print(pages_in_between_list)
                #
                #     table_area_in_between_pages = []
                #     table_area_in_between_pages.append(0)
                #     table_area_in_between_pages.append(originalheight - (int(top_margin)))
                #     table_area_in_between_pages.append(5000)
                #     table_area_in_between_pages.append(int(bottom_margin))
                #     table_area_in_between_pages = str(table_area_in_between_pages)
                #     table_area_in_between_pages = table_area_in_between_pages.replace("[", "")
                #     table_area_in_between_pages = table_area_in_between_pages.replace("]", "")
                #     table_area_in_between_pages_list = []
                #     table_area_in_between_pages_list.append(table_area_in_between_pages)
                #     print(table_area_in_between_pages_list)
                #
                #     for pages_in_between_for_df in pages_in_between_list:
                #         table_list_in_between_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                #                                                       pages=str(pages_in_between_for_df),
                #                                                       split_text=cut_text,
                #                                                       table_areas=table_area_in_between_pages_list,
                #                                                       flag_size=detect_superscripts,
                #                                                       edge_tol=int(text_edge_tol),
                #                                                       row_tol=int(group_into_row))
                #         nooftables_inbetweenpage = len(table_list_in_between_page)
                #         print(nooftables_inbetweenpage)
                #
                #         for x in range(nooftables_inbetweenpage):
                #             df = table_list_in_between_page[x].df
                #             df.rename(columns=df.iloc[0]).drop(df.index[0])
                #             final_df_list.append(df)
                #
                #     for y in range(nooftables_lastpage):
                #         df = table_list_last_page[y].df
                #         df.rename(columns=df.iloc[0]).drop(df.index[0])
                #         final_df_list.append(df)
                #
                #     result_df = pd.concat(final_df_list)
                #
                #     for column in result_df:
                #         # print(result_df[column])
                #         result_df[column] = result_df[column].replace('<s>', '', regex=True)
                #
                #     for column in result_df:
                #         # print(result_df[column])
                #         result_df[column] = result_df[column].replace('</s>', '', regex=True)
                #
                #     # print (result_df)
                #     html = result_df.to_html()
                #     # fullhtml += html
                #     full_html_for_multiple_page_table += html










def make_archive(source, destination):
    base_name = '.'.join(destination.split('.')[:-1])
    format = destination.split('.')[-1]
    root_dir = os.path.dirname(source)
    base_dir = os.path.basename(source.strip(os.sep))
    shutil.make_archive(base_name, format, root_dir, base_dir)

















# Runtime keyvalue extraction methods

def runtime_keybased_extract(keyvalues, packagepath, timestamp):
    label = keyvalues["label"]
    page = keyvalues["page"]
    no_of_main_conditions = len(keyvalues["main_conditions"])

    print("label " + label)
    print("no_of_main_conditions " + str(no_of_main_conditions))

    if (no_of_main_conditions == 1):
        main_condition = keyvalues["main_conditions"]["condition_1"]

        keyword = main_condition[0].strip()
        position = main_condition[1]
        width = int(main_condition[2])
        height = int(main_condition[3])

        print("keyword - " + keyword)
        print("position - " + position)
        print("width - " + str(width))
        print("height - " + str(height))

        filename = '/pdfs/' + timestamp + '.pdf'

        pdf = pdfquery.PDFQuery(packagepath + '/' + filename)

        if (page != "-"):
            pdf.load((int(page) - 1))
        else:
            pdf.load()

        pdf.tree.write(packagepath + "/pdfs/pdfxml.xml", pretty_print=True)

        xml_doc = etree.parse(packagepath + "/pdfs/pdfxml.xml")
        root = xml_doc.getroot()

    if (position == "right"):
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        print("parameter " + parameter)
        label = pdf.pq(parameter)

        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
            x1, y0 - height, x1 + width, y1 + height)).text()
        foundtext = foundtext.replace(keyword, "")
    elif (position == "left"):
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
            x0 + width, y0 + height, x1 + width, y1 + height)).text()
        foundtext = foundtext.replace(keyword, "")
    elif (position == "top"):
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
            x0 + width, y0 + height, x1 + width, y1 + height)).text()
        foundtext = foundtext.replace(keyword, "")
    elif (position == "bottom"):
        print("here")
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        print(label.attr('x0'))
        x0 = int(float(label.attr('x0')))
        y0 = int(float(label.attr('y0')))
        x1 = int(float(label.attr('x1')))
        y1 = int(float(label.attr('y1')))
        print(str(x0 - width) + "," + str(y1 - 1) + "," + str(x1 + width) + "," + str(y1 - height))
        foundtext = pdf.pq('LTTextLineHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (
        x0 - width, y0 - height, x1 + width, y1 - (y1 - y0))).text()
        # foundtext = pdf.pq('LTTextLineHorizontal:in_bbox("%s, %s, %s, %s")' % (x0, y0 - 30, x0 + 150, y0)).text()
        foundtext = foundtext.replace(keyword, "")

    foundtext = foundtext.replace(",", "")

    print("foundtext - " + foundtext)

    return foundtext


def runtime_regionbased_extract(keyvalue, packagepath, timestamp):
    label = keyvalue["label"]
    selected_page = keyvalue["page"]
    x0 = int(keyvalue["x0"])
    y0 = int(keyvalue["y0"])
    x1 = int(keyvalue["x1"])
    y1 = int(keyvalue["y1"])

    corresponding_image = root_path + "/static/testfiles/" + timestamp + "/output" + str(
        int(selected_page) - 1) + ".jpg"
    im = Image.open(corresponding_image)
    corresponding_image_width, corresponding_image_height = im.size

    print("corresponding_image_width - " + str(corresponding_image_width))
    print("corresponding_image_height - " + str(corresponding_image_height))

    pdf = PyPDF2.PdfFileReader(
        root_path + "/static/testfiles/" + timestamp + "/pdfs/" + timestamp + ".pdf")
    page = pdf.getPage(int(selected_page) - 1).mediaBox
    orientation = pdf.getPage(int(selected_page) - 1).get('/Rotate')
    if (orientation == None):
        width_of_pdf = int(page[2])
        height_of_pdf = int(page[3])
    else:
        width_of_pdf = int(page[3])
        height_of_pdf = int(page[2])

    print("width_of_pdf - " + str(width_of_pdf))
    print("height_of_pdf - " + str(height_of_pdf))

    if (y0 > y1):
        y0, y1 = y1, y0

    folderpath = root_path + 'static/testfiles/' + timestamp + '/pdfs/'
    filename = timestamp + '.pdf'
    pdf = pdfquery.PDFQuery(folderpath + filename)

    if (selected_page != ""):
        pdf.load((int(selected_page) - 1))
    else:
        pdf.load()

    pdf.tree.write(folderpath + "/pdfxml.xml", pretty_print=True)

    xml_doc = etree.parse(folderpath + "/pdfxml.xml")
    root = xml_doc.getroot()



    # doc = fitz.open(packagepath + "/pdfs/" + timestamp + ".pdf")
    # page = doc.loadPage(int(selected_page) - 1)
    # width_of_pdf = int(page.MediaBoxSize[0])
    # height_of_pdf = int(page.MediaBoxSize[1])
    #
    # x0_percentage = (int(x0) / 900) * 100
    # x0 = (width_of_pdf / 100) * x0_percentage
    #
    # x1_percentage = (int(x1) / 900) * 100
    # x1 = (width_of_pdf / 100) * x1_percentage
    #
    # y0_percentage = (int(y0) / 1300) * 100
    # y0 = (height_of_pdf / 100) * y0_percentage
    #
    # y1_percentage = (int(y1) / 1300) * 100
    # y1 = (height_of_pdf / 100) * y1_percentage
    #
    # if (y0 > y1):
    #     y0, y1 = y1, y0

    # print("x0 " + str(x0))
    # print("x1 " + str(x1))
    # print("y0 " + str(y0))
    # print("y1 " + str(y1))

    xmlfile = open(packagepath + "/pdfs/pdfxml.xml", "r", encoding='utf-8')
    xml = xmlfile.read()
    tree = ET.fromstring(xml)
    # tree = tree.getroot()

    final_text = ""

    print("-------------------------------")

    for elem in root.iter():
        if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
            try:
                bbox = elem.attrib.get("bbox")
                bbox_list = bbox.strip('][').split(', ')
                # print (bbox_list)

                x0_current_box = float(elem.attrib.get("x0"))
                y0_current_box = float(elem.attrib.get("y0"))
                x1_current_box = float(elem.attrib.get("x1"))
                y1_current_box = float(elem.attrib.get("y1"))

                # print(x0_current_box)
                # print(y0_current_box)
                # print(x1_current_box)
                # print(y1_current_box)

                if (x0 < x0_current_box):
                    # print ("here1")
                    # print("x1_current_box " + str(x1_current_box))
                    # print("x1 " + str(x1))
                    if (x1 > x1_current_box):
                        # print("here2")
                        # print("y0_current_box " + str(y0_current_box))
                        # print("y0 " + str(y0))
                        # foundtext = elem.text
                        # print(foundtext)
                        if (y0 < y0_current_box):
                            # print("here3")
                            # print("y1_current_box " + str(y0_current_box))
                            # print("y1 " + str(y0))
                            # foundtext = elem.text
                            # print(foundtext)
                            if (y1 > y1_current_box):
                                # print("here4")
                                foundtext = elem.text
                                print(foundtext)
                                if (foundtext != None):
                                    final_text += foundtext

            except Exception as e:
                print(e)
                return "failed"

    print ("-------------------------------")

    print("final_text - " + str(final_text))

    final_text = final_text.replace(",", "")

    return final_text


def runtime_fullocr_extract(keyvalue, packagepath, timestamp):
    label = keyvalue["label"]

    f = open(packagepath + "pdfs/fulltext.txt", "r")
    captured_text = f.read()

    no_of_main_conditions = len(keyvalue["main_conditions"])

    print("label " + label)
    print("no_of_main_conditions " + str(no_of_main_conditions))

    post_processed_text = ""

    for i in range(no_of_main_conditions):

        if (post_processed_text == ""):
            post_processed_text = captured_text

        condition_name = "condition_" + str(i+1)
        main_condition = keyvalue["main_conditions"][condition_name]

        condi1 = main_condition[0]
        condi2 = main_condition[1]
        condi3 = main_condition[2]

        if (condi1 == "add"):
            captured_text = add (post_processed_text, condi2, condi3)
        elif (condi1 == "remove"):
            captured_text = remove (post_processed_text, condi2, condi3)
        elif (condi1 == "removeall"):
            captured_text = removeall (post_processed_text, condi2, condi3)
        elif (condi1 == "replace"):
            captured_text = replace (post_processed_text, condi2, condi3)
        elif (condi1 == "replaceall"):
            captured_text = replaceall (post_processed_text, condi2, condi3)
        elif (condi1 == "translate"):
            captured_text = translate (post_processed_text, condi2, condi3)
        elif (condi1 == "runregex"):
            captured_text = runregex (post_processed_text, condi2, condi3)
        elif (condi1 == "extractonly"):
            captured_text = extractonly (post_processed_text, condi2, condi3)
            print (captured_text)

        print ("captured_text " + str(captured_text))

        post_processed_text = str(captured_text)

        post_processed_text = str(post_processed_text).replace("[", "")
        post_processed_text = post_processed_text.replace("]", "")
        post_processed_text = post_processed_text.replace("'", "")

    print("post_processed_text - " + str(post_processed_text))

    post_processed_text = post_processed_text.replace(",", "")

    return post_processed_text


def runtime_static_extract(keyvalue, packagepath, timestamp):
    label = keyvalue["label"]
    staticfirstdropdownvalue = keyvalue["staticfirstdropdownvalue"]
    statickeywordvalue = keyvalue["statickeywordvalue"]

    final_text = ""

    if (staticfirstdropdownvalue == "statictext"):
        final_text = statickeywordvalue
    elif (staticfirstdropdownvalue == "date"):
        today = date.today()
        d1 = today.strftime("%d/%m/%Y")
        final_text = d1
    elif (staticfirstdropdownvalue == "day"):
        now = datetime.now()
        today = now.strftime("%A")
        final_text = today
    elif (staticfirstdropdownvalue == "month"):
        mydate = datetime.now()
        month = mydate.strftime("%B")
        final_text = month
    elif (staticfirstdropdownvalue == "year"):
        mydate = datetime.now()
        year = mydate.year
        final_text = str(year)
    elif (staticfirstdropdownvalue == "time"):
        mydate = datetime.now()
        current_time = mydate.strftime("%H:%M:%S")
        final_text = str(current_time)
    elif (staticfirstdropdownvalue == "timestamp"):
        timestamp = time.time()
        timestamp = str(timestamp).split('.')
        timestamp = str(timestamp[0])
        final_text = str(timestamp)

    print("post_processed_text - " + str(final_text))

    final_text = final_text.replace(",", "")

    return final_text


def runtime_fromdb_extract(keyvalues, packagepath, timestamp):
    label = keyvalues["label"]
    print("label " + label)


def runtime_post_process(condition, post_processed_text):
    print (type(condition))
    condi1 = condition[0]
    condi2 = condition[1]
    condi3 = condition[2]

    if (condi1 == "add"):
        captured_text = add(post_processed_text, condi2, condi3)
    elif (condi1 == "remove"):
        captured_text = remove(post_processed_text, condi2, condi3)
    elif (condi1 == "removeall"):
        captured_text = removeall(post_processed_text, condi2, condi3)
    elif (condi1 == "replace"):
        captured_text = replace(post_processed_text, condi2, condi3)
    elif (condi1 == "replaceall"):
        captured_text = replaceall(post_processed_text, condi2, condi3)
    elif (condi1 == "translate"):
        captured_text = translate(post_processed_text, condi2, condi3)
    elif (condi1 == "runregex"):
        captured_text = runregex(post_processed_text, condi2, condi3)
    elif (condi1 == "extractonly"):
        captured_text = extractonly(post_processed_text, condi2, condi3)

    print(captured_text)

    captured_text = str(captured_text).replace("[", "")
    captured_text = captured_text.replace("]", "")
    captured_text = captured_text.replace("'", "")

    return str(captured_text)



# Runtime table extraction methods

def gettablewithoutborder_runtime(table, packagepath, timestamp, originalheight):
    table_name = table["table_name"]
    header = table["header"].strip()
    header_included_in_table = table["header_included_in_table"]
    footer = table["footer"].strip()
    footer_included_in_table = table["footer_included_in_table"]
    header_availability = table["header_availability"]
    footer_availability = table["footer_availability"]
    is_multipage = table["is_multipage"]
    top_margin = table["top_margin"]
    bottom_margin = table["bottom_margin"]
    group_into_row = table["group_into_row"]
    group_into_column = table["group_into_column"]
    detect_superscripts = table["detect_superscripts"]
    cut_text = table["cut_text"]
    text_edge_tol = table["text_edge_tol"]

    print ("header - " + str(header))
    print ("footer - " + str(footer))

    pdf = pdfquery.PDFQuery(packagepath + "pdfs/" + timestamp + ".pdf")
    pdf.load()
    pdf.tree.write(packagepath + "pdfs/pdfxml.xml", pretty_print=True, encoding="utf-8")

    xml_doc = etree.parse(packagepath + "pdfs/pdfxml.xml")
    print (xml_doc)
    root = xml_doc.getroot()

    bbox_of_header_key = {}
    for elem in root.iter():
        text1 = elem.text
        print ("header text - " + str(text1))
        try:
            if (header in text1):
                print("header text - " + str(text1))
                box = elem.attrib.get("bbox")
                print("header box - " + str(box))
                # bbox_of_key.append(box)
                pageno = getpagno(box, root)
                print("header pageno " + str(pageno))
                bbox_of_header_key[pageno] = box
        except Exception as e:
            # print (e)
            continue
    print("bbox_of_header_key - " + str(bbox_of_header_key))

    bbox_of_footer_key = {}
    for elem in root.iter():
        text = elem.text
        # print("footer text - " + str(text))
        try:
            if (footer in text):
                box = elem.attrib.get("bbox")
                print("footer box - " + str(box))
                # bbox_of_key.append(box)
                pageno = getpagno(box, root)
                print("footer pageno " + str(pageno))
                bbox_of_footer_key[pageno] = box
        except Exception as e:
            # print (e)
            continue
    print("bbox_of_footer_key - " + str(bbox_of_footer_key))

    bbox_of_header_key_list = [(k, v) for k, v in bbox_of_header_key.items()]
    bbox_of_footer_key_list = [(k, v) for k, v in bbox_of_footer_key.items()]

    print("bbox_of_header_key_list " + str(bbox_of_header_key_list))
    print("bbox_of_footer_key_list " + str(bbox_of_footer_key_list))

    os.mkdir(packagepath + "tableoutputs")

    # multi page table
    if (is_multipage == "True"):
        if ((header_availability == "True") & (footer_availability == "False")):
            print("header available")

            fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

            final_df_list = []

            for i in range(len(bbox_of_header_key_list)):
                header_block = bbox_of_header_key_list[i]
                header_page = int(header_block[0])
                header_rect = header_block[1]
                header_rect = header_rect.strip('][').split(', ')
                # print("header_page " + str(header_page))

                # footer_box = bbox_of_footer_key_list[header_page]
                # print (footer_box)

                header_starting_point = header_rect[1]
                header_ending_point = header_rect[3]
                for footer_elements in bbox_of_footer_key_list:
                    footer_page = int(footer_elements[0])
                    # print (type(footer_page))
                    if (footer_page == header_page):
                        print("footer_page == header_page")
                        footer_rect = footer_elements[1]
                        footer_rect_list = footer_rect.strip('][').split(', ')
                        footer_starting_point = footer_rect_list[1]
                        footer_ending_point = footer_rect_list[3]
                        print("footer_rect " + str(type(footer_rect)))
                        print("footer_starting_point " + str(footer_starting_point))
                        print("footer_ending_point " + str(footer_ending_point))
                    else:
                        footer_starting_point = int(bottom_margin) + 10
                        footer_ending_point = int(bottom_margin)

                # print ("footer_rect " + str(footer_rect))
                # print ("footer_starting_point " + str(footer_starting_point))

                final_area = []
                final_area.append(0)
                if (header_included_in_table == "True"):
                    final_area.append(float(header_starting_point) - 200)
                else:
                    final_area.append(float(header_ending_point) + 10)
                # if (footer_included_in_table == "True"):
                #     final_area.append(float(footer_starting_point) - 10)
                # else:
                #     final_area.append(float(footer_ending_point) + 10)
                final_area.append(5000)
                if (footer_included_in_table == "True"):
                    final_area.append(float(footer_ending_point) - 200)
                else:
                    final_area.append(float(footer_starting_point) + 10)
                # if (header_included_in_table == "True"):
                #     final_area.append(float(header_starting_point) - 10)
                # else:
                #     final_area.append(float(header_ending_point) + 10)

                print("final_area " + str(final_area))

                final_area = str(final_area)
                final_area = final_area.replace("[", "")
                final_area = final_area.replace("]", "")
                final_area_list = []
                final_area_list.append(final_area)
                print("final_area_list " + str(final_area_list))

                pdf_file_path = packagepath + "/pdfs/" + str(timestamp) + ".pdf"

                try:
                    print("------------")
                    print(header_page)
                    print(final_area_list)
                    print(pdf_file_path)
                    print("------------")

                    table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                  pages=str(header_page), split_text=cut_text,
                                                  table_areas=final_area_list,
                                                  flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                  row_tol=int(group_into_row))
                    nooftables = len(table_list)
                    print("nooftables_firstpage")
                    print(nooftables)

                    for x in range(nooftables):
                        df = table_list[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        df = df.loc[:, df.isin([' ', 'NULL', '', math.nan]).mean() < .5]
                        # df.dropna(axis=1, how='any')
                        final_df_list.append(df)

                except Exception as e:
                    print(e)
                    print("e")
                    continue

                # footer_block = bbox_of_footer_key_list[0]
                # footer_page = int(footer_block[0])
                # footer_rect = footer_block[1]
                # footer_rect = footer_rect.strip('][').split(', ')
                # print("footer_page " + str(footer_page))
                #
                # header_starting_point = header_rect[1]
                # header_ending_point = header_rect[3]
                #
                # footer_starting_point = footer_rect[1]
                # footer_ending_point = footer_rect[3]
                #
                # final_area = []
                #
                # if (header_included_in_table == "True"):
                #     final_area.append(0)
                #     final_area.append(float(header_starting_point) - 10)
                #     if (header_page == footer_page):
                #         final_area.append(2000)
                #         final_area.append(int(footer_starting_point))
                #     else:
                #         final_area.append(2000)
                #         final_area.append(int(bottom_margin))
                # else:
                #     final_area.append(0)
                #     final_area.append(float(header_ending_point) + 10)
                #     if (header_page == footer_page):
                #         final_area.append(2000)
                #         final_area.append(int(float(footer_starting_point)))
                #     else:
                #         final_area.append(2000)
                #         final_area.append(int(bottom_margin))
                #
                # print("final_area " + str(final_area))
                #
                # final_area = str(final_area)
                # final_area = final_area.replace("[", "")
                # final_area = final_area.replace("]", "")
                # final_area_list = []
                # final_area_list.append(final_area)
                # print("final_area_first_page_list " + str(final_area_list))

            print(len(final_df_list))

            result_df = pd.concat(final_df_list)

            for column in result_df:
                # print(result_df[column])
                result_df[column] = result_df[column].replace('<s>', '', regex=True)

            for column in result_df:
                # print(result_df[column])
                result_df[column] = result_df[column].replace('</s>', '', regex=True)

            result_df = result_df.apply(lambda x: sorted(x, key=pd.isnull), 1)

            # os.mkdir(packagepath + "tableoutputs")

            csv_file = packagepath + "tableoutputs/" + timestamp + '.csv'
            csv_file_cleaned = packagepath + "tableoutputs/" + timestamp + '_cleaned.csv'

            result_df.to_csv(csv_file)
            with open(csv_file, 'r', encoding="utf8") as infile, \
                    open(csv_file_cleaned, 'w', encoding="utf8") as outfile:
                data = infile.read()
                data = data.replace("[", "")
                data = data.replace("]", "")
                data = data.replace("'", "")
                data = data.replace(", nan", " ")
                data = data.replace("\\n", "")

                outfile.write(data)
            copyfile(csv_file_cleaned, packagepath + "/outputs/capturedtables.csv")
            # print (result_df)
            html = pd.DataFrame(result_df).to_html()
            # fullhtml += html
            fullhtml += html

            with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                file.write(fullhtml)

            return fullhtml

        elif ((header_availability == "False") & (footer_availability == "False")):

            if (header != ""):
                print("header available")
                fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"
                final_df_list = []

                for i in range(len(bbox_of_header_key_list)):
                    header_block = bbox_of_header_key_list[i]
                    header_page = int(header_block[0])
                    header_rect = header_block[1]
                    header_rect = header_rect.strip('][').split(', ')

    # single page table
    if (is_multipage == "False"):
        if (len(bbox_of_header_key_list) == len(bbox_of_footer_key_list)):

            full_html_for_multiple_page_table = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

            for i in range(len(bbox_of_header_key_list)):
                header_block = bbox_of_header_key_list[i]
                header_page = int(header_block[0])
                header_rect = header_block[1]
                header_rect = header_rect.strip('][').split(', ')
                print("header_page " + str(header_page))

                footer_block = bbox_of_footer_key_list[i]
                footer_page = int(footer_block[0])
                footer_rect = footer_block[1]
                footer_rect = footer_rect.strip('][').split(', ')
                print("footer_page " + str(footer_page))

                if (header_page == footer_page):
                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]

                    print(header_starting_point)
                    print(header_ending_point)

                    footer_starting_point = footer_rect[1]
                    footer_ending_point = footer_rect[3]

                    print(footer_starting_point)
                    print(footer_ending_point)

                    final_area = []

                    if (header_included_in_table == "True"):
                        final_area.append(0)
                        final_area.append(float(header_starting_point) - 10)
                    else:
                        final_area.append(0)
                        final_area.append(float(header_ending_point) + 10)

                    if (footer_included_in_table == "True"):
                        final_area.append(5000)
                        final_area.append(float(footer_ending_point))
                    else:
                        final_area.append(5000)
                        final_area.append(float(footer_starting_point))

                    final_area = str(final_area)
                    final_area = final_area.replace("[", "")
                    final_area = final_area.replace("]", "")
                    final_area_list = []
                    final_area_list.append(final_area)
                    print(type(final_area))
                    print("final_area " + str(final_area_list))
                    # tablepositionlist = ['0, 1064.808, 2000, 44.658']
                    # print (type(tablepositionlist[0]))

                    pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"

                    print("here")

                    table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n', pages=str(header_page),
                                                  split_text=cut_text, table_areas=final_area_list,
                                                  flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                  row_tol=int(group_into_row))

                    nooftables = len(table_list)

                    print("nooftables")
                    print(nooftables)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

                    for x in range(nooftables):

                        df = table_list[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])

                        for column in df:
                            # print(result_df[column])
                            df[column] = df[column].replace('<s>', '', regex=True)

                        for column in df:
                            # print(result_df[column])
                            df[column] = df[column].replace('</s>', '', regex=True)

                        # df = df.apply(lambda x: sorted(x, key=pd.isnull), 1)

                        # csv_file = root_path + 'static/finaloutputtables/' + timestamp + '.csv'
                        # csv_file_cleaned = root_path + 'static/finaloutputtables/' + timestamp + '_cleaned.csv'

                        # os.mkdir(packagepath + "tableoutputs")

                        csv_file = packagepath + "tableoutputs/" + timestamp + '.csv'
                        csv_file_cleaned = packagepath + "tableoutputs/" + timestamp + '_cleaned.csv'

                        df.to_csv(csv_file)
                        with open(csv_file, 'r') as infile, \
                                open(csv_file_cleaned, 'w') as outfile:
                            data = infile.read()
                            data = data.replace("[", "")
                            data = data.replace("]", "")
                            data = data.replace("'", "")
                            outfile.write(data)

                        copyfile(csv_file_cleaned, packagepath + "/outputs/capturedtables.csv")

                        if (x == 0):
                            html = pd.DataFrame(df).to_html()
                            fullhtml += html
                        else:
                            tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                            fullhtml += tabletitle
                            html = pd.DataFrame(df).to_html()
                            fullhtml += html

                    with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                        file.write(fullhtml)

                    return fullhtml

                elif (footer_page == (header_page + 1)):
                    print("table spawns in 2 pages")

                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]
                    footer_starting_point = footer_rect[1]
                    footer_ending_point = footer_rect[3]

                    final_area_first_page = []
                    final_area_second_page = []

                    if (header_included_in_table == "True"):
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_starting_point) - 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))
                    else:
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_ending_point) + 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))

                    if (footer_included_in_table == "True"):
                        final_area_second_page.append(0)
                        final_area_second_page.append(int(top_margin))
                        final_area_second_page.append(5000)
                        final_area_second_page.append(float(footer_ending_point))
                    else:
                        final_area_second_page.append(0)
                        final_area_second_page.append(originalheight - (int(top_margin)))
                        final_area_second_page.append(5000)
                        final_area_second_page.append(float(footer_starting_point))

                    print(final_area_first_page)
                    print(final_area_second_page)

                    # print ("tick")

                    final_area_first_page = str(final_area_first_page)
                    final_area_first_page = final_area_first_page.replace("[", "")
                    final_area_first_page = final_area_first_page.replace("]", "")
                    final_area_first_page_list = []
                    final_area_first_page_list.append(final_area_first_page)
                    print(final_area_first_page)

                    final_area_second_page = str(final_area_second_page)
                    final_area_second_page = final_area_second_page.replace("[", "")
                    final_area_second_page = final_area_second_page.replace("]", "")
                    final_area_second_page_list = []
                    final_area_second_page_list.append(final_area_second_page)
                    print(final_area_second_page)

                    pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"

                    table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                             pages=str(header_page), split_text=cut_text,
                                                             table_areas=final_area_first_page_list,
                                                             flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                             row_tol=int(group_into_row))
                    nooftables_firstpage = len(table_list_first_page)

                    table_list_second_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                              pages=str(footer_page), split_text=cut_text,
                                                              table_areas=final_area_second_page_list,
                                                              flag_size=detect_superscripts,
                                                              edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                    nooftables_secondpage = len(table_list_second_page)

                    print("tick")

                    print("nooftables_firstpage")
                    print(nooftables_firstpage)

                    print("nooftables_secondpage")
                    print(nooftables_secondpage)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

                    df_list = []

                    for x in range(nooftables_firstpage):
                        df = table_list_first_page[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        df_list.append(df)
                        # if (x == 0):
                        #     html = df.to_html()
                        #     fullhtml += html
                        # else:
                        #     tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                        #     fullhtml += tabletitle
                        #     html = df.to_html()
                        #     fullhtml += html

                    for y in range(nooftables_secondpage):
                        df = table_list_second_page[y].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        df_list.append(df)
                        # if (y == 0):
                        #     html = df.to_html()
                        #     fullhtml += html
                        # else:
                        #     tabletitle = "<h1>Table " + str(y - 1) + "</h1>"
                        #     fullhtml += tabletitle
                        #     html = df.to_html()
                        #     fullhtml += html

                    # for x in range(nooftables_firstpage):
                    #
                    #     df = table_list_first_page[x].df
                    #     df.rename(columns=df.iloc[0]).drop(df.index[0])
                    #     df_list.append(df)
                    #     if (x == 0):
                    #         html = df.to_html()
                    #         fullhtml += html
                    #     else:
                    #         tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                    #         fullhtml += tabletitle
                    #         html = df.to_html()
                    #         fullhtml += html

                    result_df = pd.concat(df_list)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('<s>', '', regex=True)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('</s>', '', regex=True)

                    # print (result_df)
                    html = result_df.to_html()
                    fullhtml += html

                    with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                        file.write(fullhtml)

                    return fullhtml

                else:
                    print("table spawns in more than 2 pages")

                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]
                    footer_starting_point = footer_rect[1]
                    footer_ending_point = footer_rect[3]

                    final_area_first_page = []
                    final_area_last_page = []

                    if (header_included_in_table == "True"):
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_starting_point) - 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))
                    else:
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_ending_point) + 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))

                    if (footer_included_in_table == "True"):
                        final_area_last_page.append(0)
                        final_area_last_page.append(int(top_margin))
                        final_area_last_page.append(5000)
                        final_area_last_page.append(float(footer_ending_point))
                    else:
                        final_area_last_page.append(0)
                        final_area_last_page.append(originalheight - (int(top_margin)))
                        final_area_last_page.append(5000)
                        final_area_last_page.append(float(footer_starting_point))

                    print("final_area_first_page " + str(final_area_first_page))
                    print("final_area_last_page " + str(final_area_last_page))

                    print("tick")

                    final_df_list = []

                    final_area_first_page = str(final_area_first_page)
                    final_area_first_page = final_area_first_page.replace("[", "")
                    final_area_first_page = final_area_first_page.replace("]", "")
                    final_area_first_page_list = []
                    final_area_first_page_list.append(final_area_first_page)
                    print("final_area_first_page_list " + str(final_area_first_page_list))

                    final_area_last_page = str(final_area_last_page)
                    final_area_last_page = final_area_last_page.replace("[", "")
                    final_area_last_page = final_area_last_page.replace("]", "")
                    final_area_last_page_list = []
                    final_area_last_page_list.append(final_area_last_page)
                    print("final_area_last_page_list " + str(final_area_last_page_list))

                    pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"

                    table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                             pages=str(header_page), split_text=cut_text,
                                                             table_areas=final_area_first_page_list,
                                                             flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                                                             row_tol=int(group_into_row))
                    nooftables_firstpage = len(table_list_first_page)
                    print("nooftables_firstpage")
                    print(nooftables_firstpage)

                    table_list_last_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                            pages=str(footer_page), split_text=cut_text,
                                                            table_areas=final_area_last_page_list,
                                                            flag_size=detect_superscripts,
                                                            edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                    nooftables_lastpage = len(table_list_last_page)
                    print("nooftables_secondpage")
                    print(nooftables_lastpage)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                    for x in range(nooftables_firstpage):
                        df = table_list_first_page[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        final_df_list.append(df)

                    pages_in_between_list = []
                    for pages_in_between in range(header_page, footer_page + 1):
                        if (pages_in_between != header_page):
                            if (pages_in_between != (footer_page)):
                                # print ("pages_in_between " + str(pages_in_between))
                                pages_in_between_list.append(pages_in_between)

                    print(pages_in_between_list)

                    table_area_in_between_pages = []
                    table_area_in_between_pages.append(0)
                    table_area_in_between_pages.append(originalheight - (int(top_margin)))
                    table_area_in_between_pages.append(5000)
                    table_area_in_between_pages.append(int(bottom_margin))
                    table_area_in_between_pages = str(table_area_in_between_pages)
                    table_area_in_between_pages = table_area_in_between_pages.replace("[", "")
                    table_area_in_between_pages = table_area_in_between_pages.replace("]", "")
                    table_area_in_between_pages_list = []
                    table_area_in_between_pages_list.append(table_area_in_between_pages)
                    print(table_area_in_between_pages_list)

                    for pages_in_between_for_df in pages_in_between_list:
                        table_list_in_between_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                                                                      pages=str(pages_in_between_for_df),
                                                                      split_text=cut_text,
                                                                      table_areas=table_area_in_between_pages_list,
                                                                      flag_size=detect_superscripts,
                                                                      edge_tol=int(text_edge_tol),
                                                                      row_tol=int(group_into_row))
                        nooftables_inbetweenpage = len(table_list_in_between_page)
                        print(nooftables_inbetweenpage)

                        for x in range(nooftables_inbetweenpage):
                            df = table_list_in_between_page[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            final_df_list.append(df)

                    for y in range(nooftables_lastpage):
                        df = table_list_last_page[y].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        final_df_list.append(df)

                    result_df = pd.concat(final_df_list)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('<s>', '', regex=True)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('</s>', '', regex=True)

                    # print (result_df)
                    html = result_df.to_html()
                    # fullhtml += html
                    full_html_for_multiple_page_table += html

            with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                file.write(full_html_for_multiple_page_table)

            return full_html_for_multiple_page_table


def gettablewithborder_runtime(table, packagepath, timestamp, originalheight):
    table_name = table["table_name"]
    header = table["header"]
    header_included_in_table = table["header_included_in_table"]
    footer = table["footer"]
    footer_included_in_table = table["footer_included_in_table"]
    header_availability = table["header_availability"]
    footer_availability = table["footer_availability"]
    is_multipage = table["is_multipage"]
    top_margin = table["top_margin"]
    bottom_margin = table["bottom_margin"]
    process_bg_lines = table["process_bg_lines"]
    cuttext = table["cuttext"]
    detect_superscripts = table["detect_superscripts"]
    smalllines = table["smalllines"]
    text_edge_tol = table["text_edge_tol"]

    print("header - " + str(header))
    print("footer - " + str(footer))

    pdf = pdfquery.PDFQuery(packagepath + "pdfs/" + timestamp + ".pdf")
    pdf.load()
    pdf.tree.write(packagepath + "pdfs/pdfxml.xml", pretty_print=True, encoding="utf-8")

    xml_doc = etree.parse(packagepath + "pdfs/pdfxml.xml")
    print(xml_doc)
    root = xml_doc.getroot()

    bbox_of_header_key = {}
    for elem in root.iter():
        text = elem.text
        # print("header text - " + str(text))
        try:
            if (header in text):
                print("header text - " + str(text))
                box = elem.attrib.get("bbox")
                print("header box - " + str(box))
                # bbox_of_key.append(box)
                pageno = getpagno(box, root)
                print("header pageno " + str(pageno))
                bbox_of_header_key[pageno] = box
        except Exception as e:
            # print (e)
            continue
    print("bbox_of_header_key - " + str(bbox_of_header_key))

    bbox_of_footer_key = {}
    for elem in root.iter():
        text = elem.text
        # print("footer text - " + str(text))
        try:
            if (footer in text):
                box = elem.attrib.get("bbox")
                print("footer box - " + str(box))
                # bbox_of_key.append(box)
                pageno = getpagno(box, root)
                print("footer pageno " + str(pageno))
                bbox_of_footer_key[pageno] = box
        except Exception as e:
            # print (e)
            continue
    print("bbox_of_footer_key - " + str(bbox_of_footer_key))

    bbox_of_header_key_list = [(k, v) for k, v in bbox_of_header_key.items()]
    bbox_of_footer_key_list = [(k, v) for k, v in bbox_of_footer_key.items()]

    print("bbox_of_header_key_list " + str(bbox_of_header_key_list))
    print("bbox_of_footer_key_list " + str(bbox_of_footer_key_list))

    os.mkdir(packagepath + "tableoutputs")

    # multi page table
    if (is_multipage == "True"):
        if ((header_availability == "True") & (footer_availability == "False")):
            print("header available")

            fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

            final_df_list = []

            for i in range(len(bbox_of_header_key_list)):
                header_block = bbox_of_header_key_list[i]
                header_page = int(header_block[0])
                header_rect = header_block[1]
                header_rect = header_rect.strip('][').split(', ')
                # print("header_page " + str(header_page))

                # footer_box = bbox_of_footer_key_list[header_page]
                # print (footer_box)

                header_starting_point = header_rect[1]
                header_ending_point = header_rect[3]
                for footer_elements in bbox_of_footer_key_list:
                    footer_page = int(footer_elements[0])
                    # print (type(footer_page))
                    if (footer_page == header_page):
                        print("footer_page == header_page")
                        footer_rect = footer_elements[1]
                        footer_rect_list = footer_rect.strip('][').split(', ')
                        footer_starting_point = footer_rect_list[1]
                        footer_ending_point = footer_rect_list[3]
                        print("footer_rect " + str(type(footer_rect)))
                        print("footer_starting_point " + str(footer_starting_point))
                        print("footer_ending_point " + str(footer_ending_point))
                    else:
                        footer_starting_point = int(bottom_margin) + 10
                        footer_ending_point = int(bottom_margin)

                # print ("footer_rect " + str(footer_rect))
                # print ("footer_starting_point " + str(footer_starting_point))

                final_area = []
                final_area.append(0)
                if (header_included_in_table == "True"):
                    final_area.append(float(header_starting_point) - 200)
                else:
                    final_area.append(float(header_ending_point) + 10)
                # if (footer_included_in_table == "True"):
                #     final_area.append(float(footer_starting_point) - 10)
                # else:
                #     final_area.append(float(footer_ending_point) + 10)
                final_area.append(5000)
                if (footer_included_in_table == "True"):
                    final_area.append(float(footer_ending_point) - 200)
                else:
                    final_area.append(float(footer_starting_point) + 10)
                # if (header_included_in_table == "True"):
                #     final_area.append(float(header_starting_point) - 10)
                # else:
                #     final_area.append(float(header_ending_point) + 10)

                print("final_area " + str(final_area))

                final_area = str(final_area)
                final_area = final_area.replace("[", "")
                final_area = final_area.replace("]", "")
                final_area_list = []
                final_area_list.append(final_area)
                print("final_area_list " + str(final_area_list))

                pdf_file_path = root_path + "/static/initialreceivepdf/" + str(
                    timestamp) + "/pdfs/" + str(timestamp) + ".pdf"

                try:
                    print("------------")
                    print(header_page)
                    print(final_area_list)
                    print(pdf_file_path)
                    print("------------")

                    # table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                    #                               pages=str(header_page), split_text=cut_text,
                    #                               table_areas=final_area_list,
                    #                               flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                    #                               row_tol=int(group_into_row))

                    table_list = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n', pages=str(header_page), split_text=cuttext, table_areas=final_area_list, flag_size=detect_superscripts, process_background=process_bg_lines, line_scale=int(smalllines))

                    nooftables = len(table_list)
                    print("nooftables_firstpage")
                    print(nooftables)

                    for x in range(nooftables):
                        df = table_list[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        df = df.loc[:, df.isin([' ', 'NULL', '', math.nan]).mean() < .5]
                        # df.dropna(axis=1, how='any')
                        final_df_list.append(df)

                except Exception as e:
                    print(e)
                    print("e")
                    continue

                # footer_block = bbox_of_footer_key_list[0]
                # footer_page = int(footer_block[0])
                # footer_rect = footer_block[1]
                # footer_rect = footer_rect.strip('][').split(', ')
                # print("footer_page " + str(footer_page))
                #
                # header_starting_point = header_rect[1]
                # header_ending_point = header_rect[3]
                #
                # footer_starting_point = footer_rect[1]
                # footer_ending_point = footer_rect[3]
                #
                # final_area = []
                #
                # if (header_included_in_table == "True"):
                #     final_area.append(0)
                #     final_area.append(float(header_starting_point) - 10)
                #     if (header_page == footer_page):
                #         final_area.append(2000)
                #         final_area.append(int(footer_starting_point))
                #     else:
                #         final_area.append(2000)
                #         final_area.append(int(bottom_margin))
                # else:
                #     final_area.append(0)
                #     final_area.append(float(header_ending_point) + 10)
                #     if (header_page == footer_page):
                #         final_area.append(2000)
                #         final_area.append(int(float(footer_starting_point)))
                #     else:
                #         final_area.append(2000)
                #         final_area.append(int(bottom_margin))
                #
                # print("final_area " + str(final_area))
                #
                # final_area = str(final_area)
                # final_area = final_area.replace("[", "")
                # final_area = final_area.replace("]", "")
                # final_area_list = []
                # final_area_list.append(final_area)
                # print("final_area_first_page_list " + str(final_area_list))

            print(len(final_df_list))

            result_df = pd.concat(final_df_list)

            for column in result_df:
                # print(result_df[column])
                result_df[column] = result_df[column].replace('<s>', '', regex=True)

            for column in result_df:
                # print(result_df[column])
                result_df[column] = result_df[column].replace('</s>', '', regex=True)

            result_df = result_df.apply(lambda x: sorted(x, key=pd.isnull), 1)

            os.mkdir(packagepath + "tableoutputs")

            csv_file = packagepath + "tableoutputs/" + timestamp + '.csv'
            csv_file_cleaned = packagepath + "tableoutputs/" + timestamp + '_cleaned.csv'

            result_df.to_csv(csv_file)
            with open(csv_file, 'r', encoding="utf8") as infile, \
                    open(csv_file_cleaned, 'w', encoding="utf8") as outfile:
                data = infile.read()
                data = data.replace("[", "")
                data = data.replace("]", "")
                data = data.replace("'", "")
                data = data.replace(", nan", " ")
                data = data.replace("\\n", "")

                outfile.write(data)
            # print (result_df)
            html = pd.DataFrame(result_df).to_html()
            # fullhtml += html
            fullhtml += html

            with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                file.write(fullhtml)

            return fullhtml

        elif ((header_availability == "False") & (footer_availability == "False")):

            if (header != ""):
                print("header available")
                fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"
                final_df_list = []

                for i in range(len(bbox_of_header_key_list)):
                    header_block = bbox_of_header_key_list[i]
                    header_page = int(header_block[0])
                    header_rect = header_block[1]
                    header_rect = header_rect.strip('][').split(', ')

    # single page table
    if (is_multipage == "False"):
        if (len(bbox_of_header_key_list) == len(bbox_of_footer_key_list)):

            full_html_for_multiple_page_table = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

            for i in range(len(bbox_of_header_key_list)):
                header_block = bbox_of_header_key_list[i]
                header_page = int(header_block[0])
                header_rect = header_block[1]
                header_rect = header_rect.strip('][').split(', ')
                print("header_page " + str(header_page))

                footer_block = bbox_of_footer_key_list[i]
                footer_page = int(footer_block[0])
                footer_rect = footer_block[1]
                footer_rect = footer_rect.strip('][').split(', ')
                print("footer_page " + str(footer_page))

                if (header_page == footer_page):
                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]

                    print(header_starting_point)
                    print(header_ending_point)

                    footer_starting_point = footer_rect[1]
                    footer_ending_point = footer_rect[3]

                    print(footer_starting_point)
                    print(footer_ending_point)

                    final_area = []

                    if (header_included_in_table == "True"):
                        final_area.append(0)
                        final_area.append(float(header_starting_point) - 10)
                    else:
                        final_area.append(0)
                        final_area.append(float(header_ending_point) + 10)

                    if (footer_included_in_table == "True"):
                        final_area.append(5000)
                        final_area.append(float(footer_ending_point))
                    else:
                        final_area.append(5000)
                        final_area.append(float(footer_starting_point))

                    final_area = str(final_area)
                    final_area = final_area.replace("[", "")
                    final_area = final_area.replace("]", "")
                    final_area_list = []
                    final_area_list.append(final_area)
                    print(type(final_area))
                    print("final_area " + str(final_area_list))
                    # tablepositionlist = ['0, 1064.808, 2000, 44.658']
                    # print (type(tablepositionlist[0]))

                    pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"

                    print("here")

                    # table_list = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n', pages=str(header_page),
                    #                               split_text=cut_text, table_areas=final_area_list,
                    #                               flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                    #                               row_tol=int(group_into_row))

                    table_list = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                  pages=str(header_page), split_text=cuttext,
                                                  table_areas=final_area_list,
                                                  flag_size=detect_superscripts,
                                                  process_background=process_bg_lines, line_scale=int(smalllines))

                    nooftables = len(table_list)

                    print("nooftables")
                    print(nooftables)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

                    for x in range(nooftables):

                        df = table_list[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])

                        for column in df:
                            # print(result_df[column])
                            df[column] = df[column].replace('<s>', '', regex=True)

                        for column in df:
                            # print(result_df[column])
                            df[column] = df[column].replace('</s>', '', regex=True)

                        # df = df.apply(lambda x: sorted(x, key=pd.isnull), 1)

                        # csv_file = root_path + 'static/finaloutputtables/' + timestamp + '.csv'
                        # csv_file_cleaned = root_path + 'static/finaloutputtables/' + timestamp + '_cleaned.csv'

                        # os.mkdir(packagepath + "tableoutputs")

                        csv_file = packagepath + "tableoutputs/" + timestamp + '.csv'
                        csv_file_cleaned = packagepath + "tableoutputs/" + timestamp + '_cleaned.csv'

                        df.to_csv(csv_file)
                        with open(csv_file, 'r') as infile, \
                                open(csv_file_cleaned, 'w') as outfile:
                            data = infile.read()
                            data = data.replace("[", "")
                            data = data.replace("]", "")
                            data = data.replace("'", "")
                            outfile.write(data)

                        if (x == 0):
                            html = pd.DataFrame(df).to_html()
                            fullhtml += html
                        else:
                            tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                            fullhtml += tabletitle
                            html = pd.DataFrame(df).to_html()
                            fullhtml += html

                    with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                        file.write(fullhtml)

                    return fullhtml

                elif (footer_page == (header_page + 1)):
                    print("table spawns in 2 pages")

                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]
                    footer_starting_point = footer_rect[1]
                    footer_ending_point = footer_rect[3]

                    final_area_first_page = []
                    final_area_second_page = []

                    if (header_included_in_table == "True"):
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_starting_point) - 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))
                    else:
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_ending_point) + 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))

                    if (footer_included_in_table == "True"):
                        final_area_second_page.append(0)
                        final_area_second_page.append(int(top_margin))
                        final_area_second_page.append(5000)
                        final_area_second_page.append(float(footer_ending_point))
                    else:
                        final_area_second_page.append(0)
                        final_area_second_page.append(originalheight - (int(top_margin)))
                        final_area_second_page.append(5000)
                        final_area_second_page.append(float(footer_starting_point))

                    print(final_area_first_page)
                    print(final_area_second_page)

                    # print ("tick")

                    final_area_first_page = str(final_area_first_page)
                    final_area_first_page = final_area_first_page.replace("[", "")
                    final_area_first_page = final_area_first_page.replace("]", "")
                    final_area_first_page_list = []
                    final_area_first_page_list.append(final_area_first_page)
                    print(final_area_first_page)

                    final_area_second_page = str(final_area_second_page)
                    final_area_second_page = final_area_second_page.replace("[", "")
                    final_area_second_page = final_area_second_page.replace("]", "")
                    final_area_second_page_list = []
                    final_area_second_page_list.append(final_area_second_page)
                    print(final_area_second_page)

                    pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"

                    # table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                    #                                          pages=str(header_page), split_text=cut_text,
                    #                                          table_areas=final_area_first_page_list,
                    #                                          flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                    #                                          row_tol=int(group_into_row))
                    table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                             pages=str(header_page), split_text=cuttext,
                                                             table_region=final_area_first_page_list,
                                                             flag_size=detect_superscripts,
                                                             process_background=eval(process_bg_lines),
                                                             line_scale=int(smalllines))
                    nooftables_firstpage = len(table_list_first_page)

                    # table_list_second_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                    #                                           pages=str(footer_page), split_text=cut_text,
                    #                                           table_areas=final_area_second_page_list,
                    #                                           flag_size=detect_superscripts,
                    #                                           edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                    table_list_second_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                              pages=str(footer_page), split_text=cuttext,
                                                              table_region=final_area_second_page_list,
                                                              flag_size=detect_superscripts,
                                                              process_background=eval(process_bg_lines),
                                                              line_scale=int(smalllines))
                    nooftables_secondpage = len(table_list_second_page)

                    print("tick")

                    print("nooftables_firstpage")
                    print(nooftables_firstpage)

                    print("nooftables_secondpage")
                    print(nooftables_secondpage)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Captured Table</h1></body>"

                    df_list = []

                    for x in range(nooftables_firstpage):
                        df = table_list_first_page[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        df_list.append(df)
                        # if (x == 0):
                        #     html = df.to_html()
                        #     fullhtml += html
                        # else:
                        #     tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                        #     fullhtml += tabletitle
                        #     html = df.to_html()
                        #     fullhtml += html

                    for y in range(nooftables_secondpage):
                        df = table_list_second_page[y].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        df_list.append(df)
                        # if (y == 0):
                        #     html = df.to_html()
                        #     fullhtml += html
                        # else:
                        #     tabletitle = "<h1>Table " + str(y - 1) + "</h1>"
                        #     fullhtml += tabletitle
                        #     html = df.to_html()
                        #     fullhtml += html

                    # for x in range(nooftables_firstpage):
                    #
                    #     df = table_list_first_page[x].df
                    #     df.rename(columns=df.iloc[0]).drop(df.index[0])
                    #     df_list.append(df)
                    #     if (x == 0):
                    #         html = df.to_html()
                    #         fullhtml += html
                    #     else:
                    #         tabletitle = "<h1>Table " + str(x - 1) + "</h1>"
                    #         fullhtml += tabletitle
                    #         html = df.to_html()
                    #         fullhtml += html

                    result_df = pd.concat(df_list)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('<s>', '', regex=True)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('</s>', '', regex=True)

                    # print (result_df)
                    html = result_df.to_html()
                    fullhtml += html

                    with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                        file.write(fullhtml)

                    return fullhtml

                else:
                    print("table spawns in more than 2 pages")

                    header_starting_point = header_rect[1]
                    header_ending_point = header_rect[3]
                    footer_starting_point = footer_rect[1]
                    footer_ending_point = footer_rect[3]

                    final_area_first_page = []
                    final_area_last_page = []

                    if (header_included_in_table == "True"):
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_starting_point) - 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))
                    else:
                        final_area_first_page.append(0)
                        final_area_first_page.append(float(header_ending_point) + 10)
                        final_area_first_page.append(5000)
                        final_area_first_page.append(int(bottom_margin))

                    if (footer_included_in_table == "True"):
                        final_area_last_page.append(0)
                        final_area_last_page.append(int(top_margin))
                        final_area_last_page.append(5000)
                        final_area_last_page.append(float(footer_ending_point))
                    else:
                        final_area_last_page.append(0)
                        final_area_last_page.append(originalheight - (int(top_margin)))
                        final_area_last_page.append(5000)
                        final_area_last_page.append(float(footer_starting_point))

                    print("final_area_first_page " + str(final_area_first_page))
                    print("final_area_last_page " + str(final_area_last_page))

                    print("tick")

                    final_df_list = []

                    final_area_first_page = str(final_area_first_page)
                    final_area_first_page = final_area_first_page.replace("[", "")
                    final_area_first_page = final_area_first_page.replace("]", "")
                    final_area_first_page_list = []
                    final_area_first_page_list.append(final_area_first_page)
                    print("final_area_first_page_list " + str(final_area_first_page_list))

                    final_area_last_page = str(final_area_last_page)
                    final_area_last_page = final_area_last_page.replace("[", "")
                    final_area_last_page = final_area_last_page.replace("]", "")
                    final_area_last_page_list = []
                    final_area_last_page_list.append(final_area_last_page)
                    print("final_area_last_page_list " + str(final_area_last_page_list))

                    pdf_file_path = packagepath + "/pdfs/" + timestamp + ".pdf"

                    # table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                    #                                          pages=str(header_page), split_text=cut_text,
                    #                                          table_areas=final_area_first_page_list,
                    #                                          flag_size=detect_superscripts, edge_tol=int(text_edge_tol),
                    #                                          row_tol=int(group_into_row))
                    table_list_first_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                             pages=str(header_page), split_text=cuttext,
                                                             table_region=final_area_first_page_list,
                                                             flag_size=detect_superscripts,
                                                             process_background=eval(process_bg_lines),
                                                             line_scale=int(smalllines))
                    nooftables_firstpage = len(table_list_first_page)
                    print("nooftables_firstpage")
                    print(nooftables_firstpage)

                    # table_list_last_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                    #                                         pages=str(footer_page), split_text=cut_text,
                    #                                         table_areas=final_area_last_page_list,
                    #                                         flag_size=detect_superscripts,
                    #                                         edge_tol=int(text_edge_tol), row_tol=int(group_into_row))
                    table_list_last_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                            pages=str(footer_page), split_text=cuttext,
                                                            table_region=final_area_last_page_list,
                                                            flag_size=detect_superscripts,
                                                            process_background=eval(process_bg_lines),
                                                            line_scale=int(smalllines))
                    nooftables_lastpage = len(table_list_last_page)
                    print("nooftables_secondpage")
                    print(nooftables_lastpage)

                    fullhtml = "<head><link rel='stylesheet' href='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/styles/outputpagestyle.css'></head><body background='https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/images/website-bg.png'><h1>Table 1</h1></body>"

                    for x in range(nooftables_firstpage):
                        df = table_list_first_page[x].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        final_df_list.append(df)

                    pages_in_between_list = []
                    for pages_in_between in range(header_page, footer_page + 1):
                        if (pages_in_between != header_page):
                            if (pages_in_between != (footer_page)):
                                # print ("pages_in_between " + str(pages_in_between))
                                pages_in_between_list.append(pages_in_between)

                    print(pages_in_between_list)

                    table_area_in_between_pages = []
                    table_area_in_between_pages.append(0)
                    table_area_in_between_pages.append(originalheight - (int(top_margin)))
                    table_area_in_between_pages.append(5000)
                    table_area_in_between_pages.append(int(bottom_margin))
                    table_area_in_between_pages = str(table_area_in_between_pages)
                    table_area_in_between_pages = table_area_in_between_pages.replace("[", "")
                    table_area_in_between_pages = table_area_in_between_pages.replace("]", "")
                    table_area_in_between_pages_list = []
                    table_area_in_between_pages_list.append(table_area_in_between_pages)
                    print(table_area_in_between_pages_list)

                    for pages_in_between_for_df in pages_in_between_list:
                        # table_list_in_between_page = camelot.read_pdf(pdf_file_path, flavor='stream', strip=' \n',
                        #                                               pages=str(pages_in_between_for_df),
                        #                                               split_text=cut_text,
                        #                                               table_areas=table_area_in_between_pages_list,
                        #                                               flag_size=detect_superscripts,
                        #                                               edge_tol=int(text_edge_tol),
                        #                                               row_tol=int(group_into_row))
                        table_list_in_between_page = camelot.read_pdf(pdf_file_path, flavor='lattice', strip=' \n',
                                                                      pages=str(pages_in_between_for_df),
                                                                      split_text=cuttext,
                                                                      table_region=table_area_in_between_pages_list,
                                                                      flag_size=detect_superscripts,
                                                                      process_background=eval(process_bg_lines),
                                                                      line_scale=int(smalllines))
                        nooftables_inbetweenpage = len(table_list_in_between_page)
                        print(nooftables_inbetweenpage)

                        for x in range(nooftables_inbetweenpage):
                            df = table_list_in_between_page[x].df
                            df.rename(columns=df.iloc[0]).drop(df.index[0])
                            final_df_list.append(df)

                    for y in range(nooftables_lastpage):
                        df = table_list_last_page[y].df
                        df.rename(columns=df.iloc[0]).drop(df.index[0])
                        final_df_list.append(df)

                    result_df = pd.concat(final_df_list)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('<s>', '', regex=True)

                    for column in result_df:
                        # print(result_df[column])
                        result_df[column] = result_df[column].replace('</s>', '', regex=True)

                    # print (result_df)
                    html = result_df.to_html()
                    # fullhtml += html
                    full_html_for_multiple_page_table += html

            with open(packagepath + "tableoutputs/" + timestamp + '.html', 'w', encoding="utf8") as file:
                file.write(full_html_for_multiple_page_table)

            return full_html_for_multiple_page_table

















def add(captured_text, condi2, condi3):
    print ("hi")

def remove(captured_text, condi2, condi3):
    if (condi2 == "characters"):
        res = captured_text.replace(condi3, "")
        return res


def removeall(captured_text, condi2, condi3):
    if (condi2 == "characters"):
        res = captured_text.replace(condi3, "")
        return res
    elif (condi2 == "numbers"):
        res = ''.join([i for i in captured_text if not i.isdigit()])
        return res
    elif (condi2 == "specialcharacters"):
        res = captured_text.translate(str.maketrans('', '', string.punctuation))
        return res
    elif (condi2 == "alphabets"):
        res = re.sub(r'[a-z]', '', captured_text.lower())
        return res
    elif (condi2 == "alphanumerics"):
        # res = re.sub(r'\w*\d\w*', '', captured_text)
        # return res
        list = []
        temp = captured_text.split()
        for idx in temp:
            if any(chr.isalpha() for chr in idx) and any(chr.isdigit() for chr in idx):
                list.append(idx)
        print(list)
        for i in range(len(list)):
            captured_text = captured_text.replace(list[i], "")
        return captured_text

    elif (condi2 == "allspaces"):
        res = captured_text.replace(" ", "")
        return res
    elif (condi2 == "trailingspaces"):
        res = captured_text.strip()
        return res

def replace(captured_text, condi2, condi3):
    print("hi")


def replaceall(captured_text, condi2, condi3):
    print("hi")


def translate(captured_text, condi2, condi3):
    print("hi")


def runregex(captured_text, condi2, condi3):
    print("hi")


def extractonly(captured_text, condi2, condi3):
    if (condi2 == "dates"):
        if (condi3 == "dd/mm/yyyy"):
            match = re.findall(r'\d{2}/\d{2}/\d{4}', captured_text)
            return match
        elif (condi3 == "dd-mm-yyyy"):
            match = re.findall(r'\d{2}-\d{2}-\d{4}', captured_text)
            return match
        elif (condi3 == "mm/dd/yyyy"):
            match = re.findall(r'\d{2}/\d{2}/\d{4}', captured_text)
            return match
        elif (condi3 == "mm-dd-yyyy"):
            match = re.findall(r'\d{2}-\d{2}-\d{4}', captured_text)
            return match
        elif (condi3 == "dd/mmm/yyyy"):
            match = re.findall(r'\d{2}/[a-zA-Z]{3}/\d{4}', captured_text)
            return match
        elif (condi3 == "dd-mmm-yyyy"):
            match = re.findall(r'\d{2}-[a-zA-Z]{3}-\d{4}', captured_text)
            return match
        elif (condi3 == "dd/mmmm/yyy"):
            # match = re.findall(r'\d{2}-[a-zA-Z]{3}-\d{4}', captured_text)
            match_list = []
            for i in range(3, 10):
                match = re.findall(r'\d{2}/[a-zA-Z]{' + str(i) + '}/\d{4}', captured_text)
                match_list.append(match)
            match_list = [ele for ele in match_list if ele != []]
            return match_list
        elif (condi3 == "dd-mmmm-yyyy"):
            # match = re.findall(r'\d{2}-[a-zA-Z]{3}-\d{4}', captured_text)
            match_list = []
            for i in range(3, 10):
                match = re.findall(r'\d{2}/[a-zA-Z]{' + str(i) + '}/\d{4}', captured_text)
                match_list.append(match)
            match_list = [ele for ele in match_list if ele != []]
            return match_list
        elif (condi3 == "dd.mm.yyyy"):
            match = re.findall(r'\d{2}\.\d{2}\.\d{4}', captured_text)
            return match
        elif (condi3 == "dd/mm/yy"):
            match = re.findall(r'\d{2}/\d{2}/\d{2}', captured_text)
            return match
    elif (condi2 == "numbers"):
        operator = condi3[0]
        if (operator == "<"):
            condi3_operator_removed = condi3.replace("<", "")
            list = re.findall("\d+", captured_text)
            match_list = []
            for i in range(len(list)):
                if (len(list[i]) < int(condi3_operator_removed)):
                    match_list.append(list[i])
            return match_list
        elif (operator == ">"):
            condi3_operator_removed = condi3.replace(">","")
            list = re.findall("\d+", captured_text)
            match_list = []
            for i in range(len(list)):
                if (len(list[i]) > int(condi3_operator_removed)):
                    match_list.append(list[i])
            return match_list
        else:
            list = re.findall("\d+", captured_text)
            match_list = []
            for i in range(len(list)):
                if (len(list[i]) == int(condi3)):
                    match_list.append(list[i])
            return match_list
    elif(condi2 == "alphanumerics"):
        print("alphanumerics")
        list = []
        match_list = []

        temp = captured_text.split()
        for idx in temp:
            if any(chr.isalpha() for chr in idx) and any(chr.isdigit() for chr in idx):
                list.append(idx)
        print(list)

        operator = condi3[0]

        if (operator == "<"):
            condi3_operator_removed = condi3.replace("<", "")
            for i in range(len(list)):
                if (len(list[i]) < int(condi3_operator_removed)):
                    match_list.append(list[i])
        elif (operator == ">"):
            condi3_operator_removed = condi3.replace(">", "")
            for i in range(len(list)):
                if (len(list[i]) > int(condi3_operator_removed)):
                    match_list.append(list[i])
        else:
            for i in range(len(list)):
                if (len(list[i]) == int(condi3)):
                    match_list.append(list[i])

        return match_list
    elif(condi2 == "textbefore"):
        print("textbefore")
        extracted_text = captured_text.split(condi3)[0]
        return extracted_text
    elif (condi2 == "textafter"):
        print("textafter")
        extracted_text = str(captured_text.split(condi3)[1:])
        extracted_text = extracted_text.replace("\"", "")
        return extracted_text










def getupdatedxml(pdfxmlpath, position, x0, y0, x1, y1):
    # keywordposition = 100
    # position = "right"

    # tree = ET.parse('C:/Users/ET437GL/OneDrive - EY/Documents/TAFE/Sample_inputs/nelcast/pdfxml.xml')
    xmlfile = open(pdfxmlpath + "/pdfxml.xml", "r")
    xml = xmlfile.read()
    tree = ET.fromstring(xml)
    # tree = tree.getroot()

    if (position == "top"):
        for elem in tree.iter():
            if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                try:
                    ytop = int(float(elem.attrib.get("y1")))
                    if (ytop > y1):
                        if (elem.text != None):
                            for element in tree.iter():
                                for child in list(element):
                                    if child.text == elem.text:
                                        element.remove(child)

                except Exception as e:
                    print(e)
                    return "failed"

        tree_out = ET.tostring(tree)
        print(tree_out.decode("utf-8"))
        xmlfile.write(tree_out)
        xmlfile.close()
        return "done"

    elif (position == "bottom"):
        for elem in tree.iter():
            if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                try:
                    ybottom = int(float(elem.attrib.get("y0")))
                    if (ybottom < y0):
                        if (elem.text != None):
                            for element in tree.iter():
                                for child in list(element):
                                    if child.text == elem.text:
                                        element.remove(child)

                except Exception as e:
                    print(e)
                    return "failed"

        tree_out = ET.tostring(tree)
        print(tree_out.decode("utf-8"))
        xmlfile.write(tree_out, "w")
        xmlfile.close()
        return "done"

    elif (position == "left"):
        for elem in tree.iter():
            if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                try:
                    xleft = int(float(elem.attrib.get("x0")))
                    if (xleft < x0):
                        if (elem.text != None):
                            for element in tree.iter():
                                for child in list(element):
                                    if child.text == elem.text:
                                        element.remove(child)

                except Exception as e:
                    print(e)
                    return "failed"

        tree_out = ET.tostring(tree)
        print(tree_out.decode("utf-8"))
        xmlfile.write(tree_out, "w")
        xmlfile.close()
        return "done"

    elif (position == "right"):
        for elem in tree.iter():
            if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                try:
                    xleft = int(float(elem.attrib.get("x1")))
                    if (xleft > x1):
                        if (elem.text != None):
                            for element in tree.iter():
                                for child in list(element):
                                    if child.text == elem.text:
                                        element.remove(child)

                except Exception as e:
                    print(e)
                    return "failed"

        tree_out = ET.tostring(tree)
        print(tree_out.decode("utf-8"))
        xmlfile.write(tree_out, "w")
        xmlfile.close()
        return "done"


def get_final_text_found(pdf, root, conditions, master_keyword, master_position, width, height):
    threshold_list = []
    for i in range(len(conditions)):
        #     print (i)
        condition = conditions[i]
        text = condition[0]
        position = condition[1]
        if (text != "-"):
            threshold_line = get_threshold_cut_line(root, text, position)
            # print (threshold_line.split(',')[0])
            threshold_line = threshold_line.split(',')[0]
            threshold_list.append(threshold_line)
        else:
            threshold_list.append("-")
    print("threshold_list")
    print (threshold_list)

    boxes_found_with_required_text = get_all_boxes_with_text(root, master_keyword)
    boxes_found_with_required_text

    final_box = get_final_box(threshold_list, boxes_found_with_required_text)
    final_box = final_box[0]

    print ("final_box")
    print (final_box)
    final_box_float = []
    for i in range(len(final_box)):
        x = float(final_box[i])
        final_box_float.append(x)
    print (final_box_float)

    key_text = get_text_in_a_box(pdf, final_box_float)
    print(key_text)
    text_for_master_keyword = get_text_for_master_keyword(pdf, root, key_text, final_box, master_keyword, master_position, width,
                                                          height)
    print ("+++++++++++++")
    print (text_for_master_keyword)
    print ("+++++++++++++")
    return (text_for_master_keyword)


def get_text_for_master_keyword(pdf, root, key_text, final_box, master_keyword, position, width, height):
    #     print ("----------")
    #     print (key_text)
    #     print (final_box)
    #     print (master_keyword)
    #     print (position)
    #     print (width)
    #     print (height)
    #     print ("----------")

    x0_final_box = int(float(final_box[0]))
    y0_final_box = int(float(final_box[1]))
    x1_final_box = int(float(final_box[2]))
    y1_final_box = int(float(final_box[3]))

    right_final_box = []
    left_final_box = []
    bottom_final_box = []
    top_final_box = []

    print("here")

    similarity_ratio = int(similar(key_text, master_keyword))
    print (similarity_ratio)

    if (position == "right"):
        try:
            if (similarity_ratio > 0):
                found_text = key_text.replace(master_keyword, "")
                print ("found_text")
                return (found_text)
            else:
                for elem in root.iter():
                    try:
                        x0 = int(float(elem.attrib.get("x0")))
                        if (x0 > x1_final_box):
                            # print(type(x1_final_box))
                            # print(type(width))
                            if (x0 < (x1_final_box + int(width))):
                                y0 = int(float(elem.attrib.get("y0")))
                                # print (type(y0_final_box))
                                # print (type(height))
                                lower_height = y0_final_box - int(height)
                                upper_height = y0_final_box + int(height)
                                if (y0 in range(lower_height, upper_height)):
                                    if (elem.tag == "LTTextBoxHorizontal"):
                                        right_final_box.append(elem.attrib.get("bbox"))
                                        #                                         print (get_text_in_a_box(elem.attrib.get("bbox")))
                                        return (get_text_in_a_box(pdf, elem.attrib.get("bbox")))
                    except Exception as e:
                        print("error")
                        print(e)
        except Exception as e:
            print(e)

    elif (position == "left"):
        try:
            if (similarity_ratio > 0):
                found_text = key_text.replace(master_keyword, "")
                #                     print (found_text)
                return (found_text)
            else:
                for elem in root.iter():
                    try:
                        x1 = int(float(elem.attrib.get("x1")))
                        if (x1 < x0_final_box):
                            if (x1 > (x0_final_box - width)):
                                y0 = int(float(elem.attrib.get("y0")))
                                lower_height = y0_final_box - height
                                upper_height = y0_final_box + height
                                if (y0 in range(lower_height, upper_height)):
                                    if (elem.tag == "LTTextBoxHorizontal"):
                                        left_final_box.append(elem.attrib.get("bbox"))
                                        #                                             print (get_text_in_a_box(elem.attrib.get("bbox")))
                                        return (get_text_in_a_box(pdf, elem.attrib.get("bbox")))
                    except Exception as e:
                        print(e)
        except Exception as e:
            print(e)

    elif (position == "bottom"):
        try:
            width_vision_start = x0_final_box - width
            width_vision_end = x0_final_box + width
            height_vision_start = y0_final_box
            height_vision_end = abs(y0_final_box - height)
            #                 print (width_vision_start)
            #                 print (width_vision_end)
            #                 print (height_vision_start)
            #                 print (height_vision_end)
            for elem in root.iter():
                try:
                    y1 = int(float(elem.attrib.get("y1")))
                    x0 = int(float(elem.attrib.get("x0")))

                    if (x0 > width_vision_start):
                        if (x0 < width_vision_end):
                            if (y1 < height_vision_start):
                                if (y1 > height_vision_end):
                                    if (elem.tag == "LTTextBoxHorizontal"):
                                        bottom_final_box.append(elem.attrib.get("bbox"))
                                        #                                             print (get_text_in_a_box(elem.attrib.get("bbox")))
                                        return (get_text_in_a_box(pdf, elem.attrib.get("bbox")))

                except Exception as e:
                    print(e)
        except Exception as e:
            print(e)

    elif (position == "top"):
        try:
            width_vision_start = x0_final_box - width
            width_vision_end = x0_final_box + width
            height_vision_start = y0_final_box
            height_vision_end = abs(y0_final_box + height)
            #                 print (width_vision_start)
            #                 print (width_vision_end)
            #                 print (height_vision_start)
            #                 print (height_vision_end)
            for elem in root.iter():
                try:
                    y1 = int(float(elem.attrib.get("y1")))
                    x0 = int(float(elem.attrib.get("x0")))

                    if (x0 > width_vision_start):
                        if (x0 < width_vision_end):
                            if (y0 > height_vision_start):
                                if (y0 < height_vision_end):
                                    if (elem.tag == "LTTextBoxHorizontal"):
                                        top_final_box.append(elem.attrib.get("bbox"))
                                        #                                             print (get_text_in_a_box(elem.attrib.get("bbox")))
                                        return (get_text_in_a_box(pdf, elem.attrib.get("bbox")))

                except Exception as e:
                    print(e)
        except Exception as e:
            print(e)


def get_threshold_cut_line(root, keyword, position):
    threshold = []
    for elem in root.iter():
        text = elem.text
        try:
            if (keyword in text):
                box = elem.attrib.get("bbox")
#                 print (box)
                if (position == "top"):
                    y = int(float(elem.attrib.get("y1")))
                    threshold.append(y)
                elif (position == "bottom"):
                    y = int(float(elem.attrib.get("y0")))
                    threshold.append(y)
                elif (position == "left"):
                    x = int(float(elem.attrib.get("x0")))
                    threshold.append(x)
                elif (position == "right"):
                    x = int(float(elem.attrib.get("x1")))
                    threshold.append(x)
        except Exception as e:
            continue
#     print (threshold)
    threshold = str(threshold).replace("[", "")
    threshold = threshold.replace("]", "")
    return threshold


def get_all_boxes_with_text(root, texttofind):
    bbox_of_key = []
    for elem in root.iter():
        text = elem.text
        try:
            if (texttofind in text):
                box = elem.attrib.get("bbox")
                bbox_of_key.append(box)
        except Exception as e:
            continue

#     print (bbox_of_key)
    return bbox_of_key


def get_text_in_a_box(pdf, box):
    if (type(box) is str):
#         print ("str")
        box = box.strip('][').split(', ')
#     print (box)
    x0 = str(box[0])
    y0 = str(box[1])
    x1 = str(box[2])
    y1 = str(box[3])
#     print (x0)
#     print (y0)
#     print (x1)
#     print (y1)
#     name = pdf.pq('LTTextBoxHorizontal:overlaps_bbox("%s, %s, %s, %s")' % (x0, y0, x1, y1)).text()
    name = pdf.pq('LTTextLineHorizontal:in_bbox("%s, %s, %s, %s")' % (x0, y0, x1, y1)).text()
#     print (name)
    if (name == ""):
        name = pdf.pq('LTTextBoxHorizontal:in_bbox("%s, %s, %s, %s")' % (x0, y0, x1, y1)).text()
#     print (name)
    return name


def similar(a, b):
    a = a.replace(" ", "")
    b = b.replace(" ", "")
    len_of_a = len(a)
    len_of_b = len(b)
    diff_value = abs(len_of_a - len_of_b)
    return diff_value


def get_final_box(threshold_list, boxes_found_with_required_text):
    final_box = []
    top = threshold_list[0]
    bottom = threshold_list[1]
    left = threshold_list[2]
    right = threshold_list[3]

    for i in range(len(boxes_found_with_required_text)):
        box_found_with_required_text = boxes_found_with_required_text[i]
        box_found_with_required_text = box_found_with_required_text.strip('][').split(', ')
        x0 = int(float(box_found_with_required_text[0]))
        y0 = int(float(box_found_with_required_text[1]))
        x1 = int(float(box_found_with_required_text[2]))
        y1 = int(float(box_found_with_required_text[3]))

    #     only top, bottom, left, right
        if ((top != "-") & (bottom != "-") & (left != "-") & (right != "-")):
            top = int(top)
            bottom = int(bottom)
            left = int(left)
            right = int(right)
            if ((y0 > top) & (y0 < bottom) & (x0 < left) & (x0 > right)):
                final_box.append(box_found_with_required_text)
    #     only top, bottom, left
        elif ((top != "-") & (bottom != "-") & (left != "-") & (right == "-")):
            top = int(top)
            bottom = int(bottom)
            left = int(left)
            if ((y0 > top) & (y0 < bottom) & (x0 < left)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only top, bottom
        elif ((top != "-") & (bottom != "-") & (left == "-") & (right == "-")):
            top = int(top)
            bottom = int(bottom)
            if ((y0 > top) & (y0 < bottom)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only top
        elif ((top != "-") & (bottom == "-") & (left == "-") & (right == "-")):
            top = int(top)
            if ((y0 > top)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only bottom
        elif ((top == "-") & (bottom != "-") & (left == "-") & (right == "-")):
            bottom = int(bottom)
            if ((y0 < bottom)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only left
        elif ((top == "-") & (bottom == "-") & (left != "-") & (right == "-")):
            left = int(left)
            if ((x0 < left)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only right
        elif ((top == "-") & (bottom == "-") & (left == "-") & (right != "-")):
            right = int(right)
            if ((x0 > right)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only bottom, left
        elif ((top == "-") & (bottom != "-") & (left != "-") & (right == "-")):
            bottom = int(bottom)
            left = int(left)
            if ((y0 < bottom) & (x0 < left)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only left, right
        elif ((top == "-") & (bottom == "-") & (left != "-") & (right != "-")):
            left = int(left)
            right = int(right)
            if ((x0 < left) & (x0 > right)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only top, right
        elif ((top != "-") & (bottom == "-") & (left == "-") & (right != "-")):
            top = int(top)
            right = int(right)
            if ((y0 > top) & (x0 > right)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only top, left
        elif ((top != "-") & (bottom == "-") & (left != "-") & (right == "-")):
            top = int(top)
            left = int(left)
            if ((y0 > top) & (x0 < left)):
                final_box.append(box_found_with_required_text)
    #         continue
    #     only bottom, right
        elif ((top == "-") & (bottom != "-") & (left == "-") & (right != "-")):
            bottom = int(bottom)
            right = int(right)
            if ((y0 < bottom) & (x0 > right)):
                final_box.append(box_found_with_required_text)
    #         continue
        else:
            print ("no conditions")

#     print (final_box)
    return final_box


def resize_pages(pdf_path, pdf_file, pdfstoresizepath, page_dims, consolidatedpdffilename):

    try:
        print ("pdf_path - " + str(pdf_path))
        print ("pdf_file - " + str(pdf_file))
        print ("pdfstoresizepath - " + str(pdfstoresizepath))
        print ("page_dims - " + str(page_dims))
        print ("consolidatedpdffilename - " + str(consolidatedpdffilename))

        pdf = PyPDF2.PdfFileReader(pdf_path + pdf_file)
        os.mkdir(pdfstoresizepath)

        object = PyPDF2.PdfFileReader(pdf_path + pdf_file)
        NumPages = object.getNumPages()
        print ("NumPages - " + str(NumPages))

        for i in range(0, NumPages):
            print (page_dims[i])
            model_width = page_dims[i][0]
            model_height = page_dims[i][1]

            page = pdf.getPage(i).mediaBox
            orientation = pdf.getPage(i).get('/Rotate')

            page0 = pdf.getPage(i)

            if (orientation == None):
                width = int(page[2])
                height = int(page[3])
            else:
                width = int(page[3])
                height = int(page[2])

            width_ratio = (model_width/width)
            print ("width_ratio - " + str(width_ratio))

            height_ratio = (model_height / width)
            print("height_ratio - " + str(height_ratio))

            page0.scaleBy(width_ratio)

            writer = PyPDF2.PdfFileWriter()
            writer.addPage(page0)

            filename = "resized" + str(i) + ".pdf"
            print (pdfstoresizepath + filename)
            with open(pdfstoresizepath + filename, "wb+") as f:
                writer.write(f)
            # f.close()

    except Exception as e:
        print("error - " + str(e))
        return (e)


    try:
        merge_pdfs(pdfstoresizepath, consolidatedpdffilename)
        shutil.copy(pdfstoresizepath + consolidatedpdffilename, pdf_path + consolidatedpdffilename)
        return ("done")
    #     shutil.rmtree(pdfstoresizepath)
    except Exception as e:
        print(e)
        return (e)









def convert_img_to_pdf(img_path, pdf_path, timestamp):

    try:
        image = Image.open(img_path)
        # image = remove_transparency(image, bg_colour=(255, 255, 255))

        if image.mode == "RGBA":
            # If in 'RGBA' Mode Convert to 'RGB' Mode
            image = image.convert("RGB")
            image.save(img_path)

        image = Image.open(img_path)

        pdf_bytes = img2pdf.convert(image.filename)

        file = open(pdf_path + timestamp + ".pdf", "wb")
        file.write(pdf_bytes)

        image.close()
        file.close()

        print("Successfully made pdf file")
        return "done"

    except Exception as e:
        print (e)
        # return sendresponse("failed", 200)


def is_pdf_txt_or_img(fname):
    try:
        print ("reached here")
        searchable_pages = []
        non_searchable_pages = []
        page_num = 0
        with open(fname, 'rb') as infile:
            print("reached here1")
            for page in PDFPage.get_pages(infile):
                print("reached here2")
                page_num += 1
                print ("page.resources.keys() - " + str(page.resources.keys()))
                if 'Font' in page.resources.keys():
                    searchable_pages.append(page_num)
                    return "txt"
                else:
                    non_searchable_pages.append(page_num)
                    return "img"
    except Exception as e:
        print (e)
        return e


def convert_pdf_to_img(pdfpath, savepath):
    try:
        print ("pdfpath - " + pdfpath)
        # pages = pdf2image.convert_from_path(pdf_path = pdfpath, dpi=300, grayscale=True, size=(2480, 3505))
        pages = pdf2image.convert_from_path(pdf_path = pdfpath, grayscale=True)
        global noofpagesinpdf
        noofpagesinpdf = len(pages)
        print ("no of pages " + str(noofpagesinpdf))
        for i in range(len(pages)):
            pages[i].save(savepath + "output" + str(i) + '.jpg')
        return "done"
    except Exception as e:
        print (e)
        return e


def binarize_imgs(imgspath):
    try:
        threads = []
        for filename in os.listdir(imgspath):
            if filename.endswith(".jpg") or filename.endswith(".png"):
                process = Thread(target=threded_binarize_img, args=(imgspath, filename,))
                time.sleep(1)
                process.start()
                threads.append(process)
            else:
                continue
        for process in threads:
            process.join()
        return "done"
    except Exception as e:
        print("f here")
        print (e)
        return e


def threded_binarize_img(imgspath, filename):
    # print("reached here 0")
    fullpath = os.path.join(imgspath, filename)
    image = cv2.imread(fullpath)
    # image = remove_transparency(image, bg_colour=(255, 255, 255))
    # ret, image = cv2.threshold(image, 150, 255, cv2.THRESH_BINARY)
    # ret, image = cv2.threshold(image, 120, 255, cv2.THRESH_BINARY)
    ret, image = cv2.threshold(image, 150, 255, cv2.THRESH_BINARY)
    image = Image.fromarray(image)
    # image = binarization.nlbin(image, threshold = 0.5)
    image.save(fullpath)
#     display(image)


def skew_correct(imgspath):
    try:
        threads = []
        for filename in os.listdir(imgspath):
            if filename.endswith(".jpg") or filename.endswith(".png"):
                process = Thread(target=threaded_skew_correct, args=(imgspath, filename,))
                time.sleep(1)
                process.start()
                threads.append(process)
            else:
                continue
        for process in threads:
            process.join()
        return "done"
    except Exception as e:
        print (e)
        return e


def threaded_skew_correct(imgspath, filename):
    fullpath = os.path.join(imgspath, filename)
    image = cv2.imread(fullpath)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray = cv2.bitwise_not(gray)
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

    coords = np.column_stack(np.where(thresh > 0))
    angle = cv2.minAreaRect(coords)[-1]

    if angle < -45:
        angle = -(90 + angle)

    else:
        angle = -angle

    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, angle, 1.0)
    rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    print("[INFO] angle: {:.3f}".format(angle))
    rotated = Image.fromarray(rotated)
    # rotated.save(fullpath)


def remove_lines(imgspath):
    try:
        lines_removes_path = imgspath + "linesremoved/"
        threads = []
        for filename in os.listdir(imgspath):
            if filename.endswith(".jpg") or filename.endswith(".png"):
                if filename.startswith("o"):
                    process = Thread(target=threaded_remove_lines, args=(imgspath, filename, lines_removes_path,))
                    time.sleep(1)
                    process.start()
                    threads.append(process)
            else:
                continue
        for process in threads:
            process.join()
        return "done"
    except Exception as e:
        print (e)
        return e


def threaded_remove_lines(imgspath, filename, lines_removes_path):
    fullpath = os.path.join(imgspath, filename)
    image = cv2.imread(fullpath)
    result = image.copy()
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    kernel = np.ones((5, 5), np.uint8)
    img_erosion = cv2.erode(gray, kernel, iterations=1)

    thresh = cv2.threshold(img_erosion, 125, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

    # edges = cv2.Canny(thresh, 100, 150, apertureSize=3)
    #
    # minLineLength = 1
    # maxLineGap = 1
    # print(np.pi)
    # lines = cv2.HoughLinesP(edges, 1, np.pi / 180, 350, minLineLength, maxLineGap, 300)
    # for line in lines:
    #     for x1, y1, x2, y2 in line:
    #         cv2.line(image, (x1, y1), (x2, y2), (255, 255, 255), 3)
    #
    # cv2.imwrite(lines_removes_path + filename, image)

    # Remove horizontal lines
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (80, 1))
    remove_horizontal = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
    cnts = cv2.findContours(remove_horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]
    for c in cnts:
        cv2.drawContours(result, [c], -1, (255, 255, 255), 5)

    # Remove vertical lines
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 80))
    remove_vertical = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
    cnts = cv2.findContours(remove_vertical, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]
    for c in cnts:
        cv2.drawContours(result, [c], -1, (255, 255, 255), 5)

    result = Image.fromarray(result)
    result.save(lines_removes_path + filename)
    print ("done")


def get_txt_boxes(lines_removed_path, main_images_path, language):
    try:
        pytesseract.pytesseract.tesseract_cmd = 'C:\\Users\\ET437GL\\AppData\\Local\\Tesseract-OCR\\tesseract.exe'
        final_json = {}
        threads = []
        for filename in os.listdir(main_images_path):
            if filename.endswith(".jpg") or filename.endswith(".png"):
                process = Thread(target=threaded_get_txt_boxes, args=(lines_removed_path, main_images_path, filename, final_json, language,))
                time.sleep(1)
                process.start()
                threads.append(process)
        # print (final_json)
        for process in threads:
            process.join()
        final_json_file_path = main_images_path + "final_json.json"
        with open(final_json_file_path, 'w') as json_file:
            json.dump(final_json, json_file)
        return "done"
    except Exception as e:
        print (e)
        return e


def threaded_get_txt_boxes(lines_removed_path, main_images_path, filename, final_json, language):
    img1 = cv2.imread(lines_removed_path + filename)
    img2 = cv2.imread(main_images_path + filename)
    boxes = pytesseract.image_to_data(img1, config='--psm 4', lang = language)
    json_for_file = {}
    for a,b in enumerate(boxes.splitlines()):
        if a!=0:
            b = b.split()
            if len(b)==12:
                x,y,w,h = int(b[6]),int(b[7]),int(b[8]),int(b[9])
                # cv2.putText(img2,b[11],(x,y-5),cv2.FONT_HERSHEY_SIMPLEX,0.5,(50,50,255),2)
                cv2.rectangle(img2, (x,y), (x+w, y+h), (50, 50, 255), 2)
                bounding_box = "(" + str(x) + "," + str(y) + "," + str(x+w) + "," + str(y+h) + ")"
                found_text = b[11]
                json_for_file[bounding_box] = found_text
    final_json[filename] = json_for_file
    orig = Image.fromarray(img2)
    orig.save(main_images_path + filename)


def convert_imgs_into_pdf(pdf_path, main_images_path, language):
    try:
        # folderpath = os.path.join(main_images_path, "pdfs")
        # os.mkdir(folderpath)
        threads = []
        for filename in os.listdir(main_images_path):
            if filename.endswith(".jpg") or filename.endswith(".png"):
                if filename.startswith("o"):
                    print ('here')
                    process = Thread(target=threaded_imgs_into_pdf, args=(main_images_path, filename, pdf_path, language))
                    time.sleep(1)
                    process.start()
                    threads.append(process)
            else:
                continue
        for process in threads:
            process.join()
        return "done"
    except Exception as e:
        print (e)
        return e


def threaded_imgs_into_pdf(main_images_path, filename, pdf_path, language):

    im = Image.open(main_images_path + filename)
    width, height = im.size

    pdf = pytesseract.image_to_pdf_or_hocr(main_images_path + filename, extension='pdf', lang = language)
    with open(pdf_path + filename.split(".")[0] + ".pdf", 'w+b') as f:
        f.write(pdf)

    pdffile = PyPDF2.PdfFileReader(pdf_path + filename.split(".")[0] + ".pdf")
    page = pdffile.getPage(0)
    page.scaleTo(width, height)
    writer = PyPDF2.PdfFileWriter()  # create a writer to save the updated results
    writer.addPage(page)
    with open(pdf_path + filename.split(".")[0] + ".pdf", "wb+") as f:
        writer.write(f)
        print("done")


def merge_pdfs(pdf_path, consolidatedpdffilename):
    try:
        os.chdir(pdf_path)
        pdf2merge = []
        for filename in os.listdir('.'):
            if filename.endswith('.pdf'):
                pdf2merge.append(filename)

        pdfWriter = PyPDF2.PdfFileWriter()

        print ('files to merge - ')
        print (pdf2merge)

        for filename in pdf2merge:
            pdfFileObj = open(filename,'rb')
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            for pageNum in range(pdfReader.numPages):
                pageObj = pdfReader.getPage(pageNum)
                pdfWriter.addPage(pageObj)
        pdfOutput = open(consolidatedpdffilename, 'wb')
        pdfWriter.write(pdfOutput)
        pdfOutput.close()
        return "done"
    except Exception as e:
        print (e)
        return e


def getpagno(bbox, root):

    try:

        parent_map = {c: p for p in root.iter() for c in p}

        # print ('hi')

        parents = root.findall('.//LTTextLineHorizontal[@bbox="' + bbox + '"]...')
        immediate_parent = parents[0]
        # print ("immediate_parent " + str(immediate_parent))


        immediate_parent_type = immediate_parent.tag
        if (immediate_parent_type != "LTPage"):
            immediate_parent_2 = parent_map[immediate_parent]
            immediate_parent_2_type = immediate_parent_2.tag
            # print("immediate_parent_2_type " + str(immediate_parent_2_type))
            if (immediate_parent_2_type != "LTPage"):
                immediate_parent_3 = parent_map[immediate_parent_2]
                immediate_parent_3_type = immediate_parent_3.tag
                # print("immediate_parent_3_type " + str(immediate_parent_3_type))
                if (immediate_parent_3_type != "LTPage"):
                    immediate_parent_4 = parent_map[immediate_parent_3]
                    immediate_parent_4_type = immediate_parent_4.tag
                    # print("immediate_parent_4_type " + str(immediate_parent_4_type))
                    if (immediate_parent_4_type != "LTPage"):
                        immediate_parent_5 = parent_map[immediate_parent_4]
                        immediate_parent_5_type = immediate_parent_5.tag
                        # print("immediate_parent_5_type " + str(immediate_parent_5_type))
                        if (immediate_parent_5_type != "LTPage"):
                            immediate_parent_6 = parent_map[immediate_parent_5]
                            immediate_parent_6_type = immediate_parent_6.tag
                            # print("immediate_parent_6_type " + str(immediate_parent_6_type))
                            if (immediate_parent_6_type != "LTPage"):
                                immediate_parent_7 = parent_map[immediate_parent_6]
                                immediate_parent_7_type = immediate_parent_7.tag
                                # print("immediate_parent_7_type " + str(immediate_parent_7_type))
                                if (immediate_parent_7_type != "LTPage"):
                                    immediate_parent_8 = parent_map[immediate_parent_7]
                                    immediate_parent_8_type = immediate_parent_8.tag
                                    # print("immediate_parent_8_type " + str(immediate_parent_8_type))
                                else:
                                    pageno = immediate_parent_7.attrib.get('page_label')
                                    if (pageno == ""):
                                        pageno = immediate_parent_7.attrib.get('pageid')
                                    # print("pageno " + str(pageno))
                                    return pageno
                            else:
                                pageno = immediate_parent_6.attrib.get('page_label')
                                if (pageno == ""):
                                    pageno = immediate_parent_6.attrib.get('pageid')
                                # print("pageno " + str(pageno))
                                return pageno
                        else:
                            immediate_parent_6 = parent_map[immediate_parent_5]
                            pageno = immediate_parent_6.attrib.get('page_label')
                            if (pageno == ""):
                                pageno = immediate_parent_6.attrib.get('pageid')
                            # print("pageno " + str(pageno))
                            return pageno
                    else:
                        pageno = immediate_parent_4.attrib.get('page_label')
                        if (pageno == ""):
                            pageno = immediate_parent_4.attrib.get('pageid')
                        # print("pageno " + str(pageno))
                        return pageno
                else:
                    pageno = immediate_parent_3.attrib.get('page_label')
                    if (pageno == ""):
                        pageno = immediate_parent_3.attrib.get('pageid')
                    # print("pageno " + str(pageno))
                    return pageno
            else:
                pageno = immediate_parent_2.attrib.get('page_label')
                if (pageno == ""):
                    pageno = immediate_parent_2.attrib.get('pageid')
                # print("pageno " + str(pageno))
                return pageno
        else:
            pageno = immediate_parent.attrib.get('page_label')
            if (pageno == ""):
                pageno = immediate_parent.attrib.get('pageid')
            # print("pageno " + str(pageno))
            return pageno





        # if (immediate_parent_type == "LTRect"):
        #     immediate_parent_2 = parent_map[immediate_parent]
        #     immediate_parent_2_type = immediate_parent_2.tag
        #     print ("immediate_parent_2_type " + str(immediate_parent_2_type))
        #     if (immediate_parent_2_type == "LTPage"):
        #         pageno = immediate_parent_2.attrib.get('page_label')
        #         if (pageno == ""):
        #             pageno = immediate_parent_2.attrib.get('pageid')
        #         print ("pageno " + str(pageno))
        #         return pageno
        # elif (immediate_parent_type == "LTTextBoxHorizontal"):
        #
        #     immediate_parent_2 = parent_map[immediate_parent]
        #     immediate_parent_2_type = immediate_parent_2.tag
        #     print(immediate_parent_2_type)
        #     if (immediate_parent_2_type == "LTRect"):
        #         immediate_parent_3 = parent_map[immediate_parent_2]
        #         immediate_parent_3_type = immediate_parent_3.tag
        #         if (immediate_parent_3_type == "LTPage"):
        #             pageno = immediate_parent_3.attrib.get('page_label')
        #             if (pageno == ""):
        #                 pageno = immediate_parent_3.attrib.get('pageid')
        #             print("pageno " + str(pageno))
        #             return pageno
        #     elif (immediate_parent_2_type == "LTTextBoxHorizontal"):
        #         immediate_parent_3 = parent_map[immediate_parent_2]
        #         immediate_parent_3_type = immediate_parent_3.tag
        #         if (immediate_parent_3_type == "LTRect"):
        #             immediate_parent_4 = parent_map[immediate_parent_3]
        #             immediate_parent_4_type = immediate_parent_4.tag
        #             if (immediate_parent_4_type == "LTPage"):
        #                 pageno = immeediate_parent_4.attrib.get('page_label')
        #                 if (pageno == ""):
        #                     pageno = immeediate_parent_4.attrib.get('pageid')
        #                 print("pageno " + str(pageno))
        #                 return pageno
        #     elif (immediate_parent_2_type == "LTFigure"):
        #         immediate_parent_3 = parent_map[immediate_parent_2]
        #         immediate_parent_3_type = immediate_parent_3.tag
        #
        #         if (immediate_parent_3_type == "LTImage"):
        #             immediate_parent_4 = parent_map[immediate_parent_3]
        #             immediate_parent_4_type = immediate_parent_4.tag
        #
        #             if (immediate_parent_4_type == "LTPage"):
        #                 pageno = immediate_parent_4.attrib.get('pageid')
        #                 if (pageno == ""):
        #                     pageno = immeediate_parent_4.attrib.get('page_label')
        #                 print(pageno)
        #                 return pageno
        #     elif (immediate_parent_2_type == "LTPage"):
        #         pageno = immediate_parent_2.attrib.get('pageid')
        #         if (pageno == ""):
        #             pageno = immediate_parent_2.attrib.get('pageid')
        #         return pageno
        #
        # elif (immediate_parent_type == "LTTextLineHorizontal"):
        #     print(immediate_parent_type)
        # elif (immediate_parent_type == "LTPage"):
        #     # print(immediate_parent_type)
        #     # immediate_parent_4_type = immediate_parent_4.tag
        #     pageno = immediate_parent.attrib.get('pageid')
        #     if (pageno == ""):
        #         pageno = immediate_parent.attrib.get('page_label')
        #     return pageno
        # elif (immediate_parent_type == "pdfxml"):
        #     print(immediate_parent_type)
        # elif (immediate_parent_type == "LTFigure"):
        #
        #     immediate_parent_2 = parent_map[immediate_parent]
        #     immediate_parent_2_type = immediate_parent_2.tag
        #
        #     if (immediate_parent_2_type == "LTImage"):
        #         immediate_parent_3 = parent_map[immediate_parent_2]
        #         immediate_parent_3_type = immediate_parent_3.tag
        #
        #         if (immediate_parent_3_type == "LTPage"):
        #             pageno = immediate_parent_3.attrib.get('pageid')
        #             if (pageno == ""):
        #                 pageno = immediate_parent_3.attrib.get('page_label')
        #             print(pageno)
        #             return pageno
        # elif (immediate_parent_type == "LTCurve"):
        #     immediate_parent_2 = parent_map[immediate_parent]
        #     immediate_parent_2_type = immediate_parent_2.tag
        #     print("immediate_parent_2_type " + str(immediate_parent_2_type))
        #     if (immediate_parent_2_type == "LTCurve"):
        #         immediate_parent_3 = parent_map[immediate_parent_2]
        #         immediate_parent_3_type = immediate_parent_3.tag
        #         print("immediate_parent_3_type " + str(immediate_parent_3_type))




    except Exception as e:
        print (e)


def remove_transparency(im, bg_colour=(255, 255, 255)):

    print ("reached here")

    # Only process if image has transparency (http://stackoverflow.com/a/1963146)
    if im.mode in ('RGBA', 'LA') or (im.mode == 'P' and 'transparency' in im.info):
        print("reached here 1")

        # Need to convert to RGBA if LA format due to a bug in PIL (http://stackoverflow.com/a/1963146)
        alpha = im.convert('RGBA').split()[-1]

        # Create a new background image of our matt color.
        # Must be RGBA because paste requires both images have the same format
        # (http://stackoverflow.com/a/8720632  and  http://stackoverflow.com/a/9459208)
        bg = Image.new("RGBA", im.size, bg_colour + (255,))
        bg.paste(im, mask=alpha)
        return bg

    else:
        print("reached here 2")
        return im











# Al-Kabli custom solutions

@app.route('/getdatafromdsd', methods = ['POST'])
def getdatafromdsd():
    try:
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            folderpath = root_path + 'static/initialreceivepdf/alkabli'
            foldername = timestamp
            output_foldername = "outputs"
            filename = timestamp + '.pdf'
            consolidatedpdffilename = timestamp + '.pdf'

            folderpath = os.path.join(folderpath, foldername)
            folderpath = folderpath + "/"
            # folderpath = savepath
            output_folderpath = os.path.join(folderpath, output_foldername)
            inputpdfpath = folderpath + "/" + timestamp + '.pdf'

            linesremovedpath = folderpath + '/linesremoved/'
            pdfpath = folderpath + '/pdfs/'
            pdfstoresizepath = pdfpath + '/pdfstoresize/'

            xml_path = output_folderpath + "pdfxml.xml"

            print(folderpath)
            print(output_folderpath)

            try:
                os.mkdir(folderpath)
                os.mkdir(output_folderpath)
                os.mkdir(linesremovedpath)
                os.mkdir(pdfpath)
                # os.mkdir(pdfstoresizepath)
            except Exception as e:
                print("create folder")
                print(e)

            file.save(folderpath + "/" + filename)
            print("file received")

            file_path = folderpath + "/pdfs/" + filename
            # output_path = 'C:/Users/ET437GL/OneDrive - EY/Documents/EYESIGHT/Sample inputs/tamimi/outputs/' + str(ts) + "/"
            footer_text = "Total number"
            footer_text_2 = "Products will"
            header_text = "Article"
            footer_page = 0
            footer_y1 = 0

            pdf_format = is_pdf_txt_or_img(inputpdfpath)

            print(pdf_format)
            if (pdf_format == "img"):
                convert_pdf_to_img_status = convert_pdf_to_img(inputpdfpath, folderpath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs(folderpath)
                    print("binarize_imgs_status - " + "done")
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(folderpath)
                        print("skew_correct_status - " + "done")
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(folderpath)
                            print("remove_lines_status - " + "done")
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, folderpath, "eng")
                                print("get_txt_boxes_status - " + "done")
                                if (get_txt_boxes_status == "done"):
                                    convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath,
                                                                                         language)
                                    if (convert_imgs_into_pdf_status == "done"):
                                        print("convert_imgs_into_pdf_status - " + "done")
                                        merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                        if (merge_pdfs_status == "done"):

                                            page_dims = []

                                            for filename_in_path in os.listdir(folderpath):
                                                if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
                                                        ".png"):
                                                    if filename_in_path.startswith("o"):
                                                        current_page_dim = []
                                                        im = Image.open(folderpath + filename_in_path)
                                                        width, height = im.size
                                                        current_page_dim.append(width)
                                                        current_page_dim.append(height)
                                                        page_dims.append(current_page_dim)

                                            print("page_dims - " + str(page_dims))

                                            resize_pages_status = resize_pages(pdfpath, filename, pdfstoresizepath,
                                                                               page_dims, filename)

                                            if (resize_pages_status != "done"):
                                                return "failed in rescaling pdf pages"

                                            # get total pages
                                            doc = fitz.open(file_path)
                                            total_pages = doc.pageCount
                                            print("total_pages - " + str(total_pages))

                                            # making xml tree
                                            pdf = pdfquery.PDFQuery(file_path)
                                            pdf.load()
                                            pdf.tree.write(xml_path, pretty_print=True, encoding="utf-8")
                                            xml_doc = etree.parse(xml_path)
                                            root = xml_doc.getroot()

                                            # assigning footer location and footer page
                                            for elem in root.iter():
                                                if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                                                    text_in_box = elem.text
                                                    if (text_in_box is not None):
                                                        if (footer_text in text_in_box):
                                                            footer_y1 = elem.attrib.get("y1")
                                                            bbox = elem.attrib.get("bbox")
                                                            footer_page = getpagno(bbox, root)
                                                            footer_page = int(footer_page) - 1
                                                            footer_y1 = str(int(float(footer_y1)))
                                                            print("footer_y1 - " + str(footer_y1))
                                                            print("footer_page - " + str(footer_page))
                                            if (footer_y1 == 0):
                                                for elem in root.iter():
                                                    if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                                                        text_in_box = elem.text
                                                        if (text_in_box is not None):
                                                            if (footer_text_2 in text_in_box):
                                                                footer_y1 = elem.attrib.get("y1")
                                                                bbox = elem.attrib.get("bbox")
                                                                footer_page = getpagno(bbox, root)
                                                                footer_page = int(footer_page) - 1
                                                                footer_y1 = str(int(float(footer_y1)))
                                                                print("footer_y1 - " + str(footer_y1))
                                                                print("footer_page - " + str(footer_page))

                                            if (footer_y1 == 0):
                                                footer_page = total_pages
                                                footer_y1 = 100

                                            # main execution
                                            try:
                                                values_extract_status = get_values(pdf, timestamp,
                                                                                   output_folderpath)
                                                tables_extract_status = get_tables(pdf, total_pages, timestamp,
                                                                                   footer_page, footer_y1, root,
                                                                                   header_text, file_path,
                                                                                   output_folderpath)

                                                if (values_extract_status != "failed"):
                                                    if (tables_extract_status != "failed"):

                                                        output_zip = zipfile.ZipFile(
                                                            output_folderpath + '_eyesight.zip', 'w')
                                                        for folder, subfolders, files in os.walk(output_folderpath):
                                                            for file in files:
                                                                output_zip.write(os.path.join(folder, file),
                                                                                 os.path.relpath(
                                                                                     os.path.join(folder, file),
                                                                                     output_folderpath),
                                                                                 compress_type=zipfile.ZIP_DEFLATED)

                                                        output_zip.close()

                                                        print("output saved successfully")

                                                        download_link = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/initialreceivepdf/alkabli/" + timestamp + "/outputs_eyesight.zip"

                                                        return sendresponse(download_link, 200)

                                                    else:
                                                        print("failed to extract table")
                                                        return sendresponse("failed to extract table", 201)
                                                else:
                                                    print("failed to extract values")
                                                    return sendresponse("failed to extract values", 201)

                                            except Exception as e:
                                                print(e)
                                                return sendresponse(e, 201)

                                            # return "all done well"



                                        else:
                                            print("failed in merge_pdfs")
                                            return "failed in merge_pdfs"
                                    else:
                                        print("failed in convert_imgs_into_pdf")
                                        return "failed in convert_imgs_into_pdf"
                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"
            else:
                # get total pages
                file_path = folderpath + "/" + filename
                doc = fitz.open(file_path)
                total_pages = doc.pageCount
                print("total_pages - " + str(total_pages))

                # making xml tree
                pdf = pdfquery.PDFQuery(file_path)
                pdf.load()
                pdf.tree.write(xml_path, pretty_print=True, encoding="utf-8")
                xml_doc = etree.parse(xml_path)
                root = xml_doc.getroot()

                # assigning footer location and footer page
                for elem in root.iter():
                    if (
                            elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                        text_in_box = elem.text
                        if (text_in_box is not None):
                            if (footer_text in text_in_box):
                                footer_y1 = elem.attrib.get("y1")
                                bbox = elem.attrib.get("bbox")
                                footer_page = getpagno(bbox, root)
                                footer_page = int(footer_page) - 1
                                footer_y1 = str(int(float(footer_y1)))
                                print("footer_y1 - " + str(footer_y1))
                                print("footer_page - " + str(footer_page))
                if (footer_y1 == 0):
                    footer_page = total_pages
                    footer_y1 = 100

                # main execution
                try:
                    values_extract_status = get_values(pdf, timestamp,
                                                       output_folderpath)
                    tables_extract_status = get_tables(pdf, total_pages, timestamp,
                                                       footer_page, footer_y1, root,
                                                       header_text, file_path,
                                                       output_folderpath)

                    if (values_extract_status != "failed"):
                        if (tables_extract_status != "failed"):

                            output_zip = zipfile.ZipFile(
                                output_folderpath + '_eyesight.zip', 'w')
                            for folder, subfolders, files in os.walk(output_folderpath):
                                for file in files:
                                    output_zip.write(os.path.join(folder, file),
                                                     os.path.relpath(
                                                         os.path.join(folder, file),
                                                         output_folderpath),
                                                     compress_type=zipfile.ZIP_DEFLATED)

                            output_zip.close()

                            print("output saved successfully")

                            download_link = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/initialreceivepdf/alkabli/" + timestamp + "/outputs_eyesight.zip"

                            return sendresponse(download_link, 200)

                        else:
                            print("failed to extract table")
                            return sendresponse("failed to extract table", 201)
                    else:
                        print("failed to extract values")
                        return sendresponse("failed to extract values", 201)

                except Exception as e:
                    print(e)
                    return sendresponse(e, 201)

    except Exception as e:
        print (e)
        return sendresponse("failed", 200)


@app.route('/getdatafromfax', methods = ['POST'])
def getdatafromfax():
    try:
        file = request.files['file']
        if 'file' not in request.files:
            print('No file part')
            return sendresponse("file not received", 201)
        else:
            timestamp = time.time()
            timestamp = str(timestamp).split('.')
            timestamp = str(timestamp[0])
            print(timestamp)

            folderpath = root_path + 'static/initialreceivepdf/alkabli'
            foldername = timestamp
            output_foldername = "outputs"
            filename = timestamp + '.pdf'
            consolidatedpdffilename = timestamp + '.pdf'

            folderpath = os.path.join(folderpath, foldername)
            folderpath = folderpath + "/"
            # folderpath = savepath
            output_folderpath = os.path.join(folderpath, output_foldername)
            inputpdfpath = folderpath + "/" + timestamp + '.pdf'
            txtfilepath = folderpath + "/" + timestamp + '.txt'

            linesremovedpath = folderpath + '/linesremoved/'
            pdfpath = folderpath + '/pdfs/'
            pdfstoresizepath = pdfpath + '/pdfstoresize/'

            xml_path = output_folderpath + "pdfxml.xml"

            print(folderpath)
            print(output_folderpath)
            
            
            
            

            try:
                os.mkdir(folderpath)
                os.mkdir(output_folderpath)
                os.mkdir(linesremovedpath)
                os.mkdir(pdfpath)
                # os.mkdir(pdfstoresizepath)
            except Exception as e:
                print(e)

            file.save(folderpath + "/" + filename)
            print("file received")

            file_path = folderpath + "/pdfs/" + filename
            # output_path = 'C:/Users/ET437GL/OneDrive - EY/Documents/EYESIGHT/Sample inputs/tamimi/outputs/' + str(ts) + "/"
            footer_text = "Purchase order"
            header_text = "FAM."
            header_text_2 = "SUPPLIER "
            footer_page = 0
            footer_y1 = 0

            pdf_format = is_pdf_txt_or_img(inputpdfpath)

            print(pdf_format)
            if (pdf_format == "img"):
                convert_pdf_to_img_status = convert_pdf_to_img_alkabli(inputpdfpath, folderpath)
                if (convert_pdf_to_img_status == "done"):
                    binarize_imgs_status = binarize_imgs_alkabli(folderpath)
                    print("binarize_imgs_status - " + "done")
                    if (binarize_imgs_status == "done"):
                        skew_correct_status = skew_correct(folderpath)
                        print("skew_correct_status - " + "done")
                        if (skew_correct_status == "done"):
                            remove_lines_status = remove_lines(folderpath)
                            print("remove_lines_status - " + "done")
                            if (remove_lines_status == "done"):
                                get_txt_boxes_status = get_txt_boxes(linesremovedpath, folderpath, "eng")
                                print("get_txt_boxes_status - " + "done")
                                if (get_txt_boxes_status == "done"):
                                    convert_imgs_into_pdf_status = convert_imgs_into_pdf(pdfpath, linesremovedpath,
                                                                                         language)
                                    if (convert_imgs_into_pdf_status == "done"):
                                        print("convert_imgs_into_pdf_status - " + "done")
                                        merge_pdfs_status = merge_pdfs(pdfpath, consolidatedpdffilename)
                                        if (merge_pdfs_status == "done"):

                                            page_dims = []

                                            for filename_in_path in os.listdir(folderpath):
                                                if filename_in_path.endswith(".jpg") or filename_in_path.endswith(
                                                        ".png"):
                                                    if filename_in_path.startswith("o"):
                                                        current_page_dim = []
                                                        im = Image.open(folderpath + filename_in_path)
                                                        width, height = im.size
                                                        current_page_dim.append(width)
                                                        current_page_dim.append(height)
                                                        page_dims.append(current_page_dim)

                                            print("page_dims - " + str(page_dims))

                                            resize_pages_status = resize_pages(pdfpath, filename, pdfstoresizepath,
                                                                               page_dims, filename)

                                            if (resize_pages_status != "done"):
                                                return "failed in rescaling pdf pages"

                                            # get total pages
                                            doc = fitz.open(file_path)
                                            total_pages = doc.pageCount
                                            print("total_pages - " + str(total_pages))

                                            # make txt file
                                            try:
                                                text_found_in_pdf = ""
                                                for page in doc:
                                                    text_found_in_pdf += page.getText()
                                                # with fitz.open(inputpdfpath) as doc:
                                                #     text_found_in_pdf = ""
                                                #     for page in doc:
                                                #         text_found_in_pdf += page.getText()

                                                print(text_found_in_pdf, file=open(txtfilepath, 'w'))
                                            except Exception as e:
                                                print("cannot create txt file")
                                                print(e)
                                            
                                            
                                            

                                            # making xml tree
                                            pdf = pdfquery.PDFQuery(file_path)
                                            pdf.load()
                                            pdf.tree.write(xml_path, pretty_print=True, encoding="utf-8")
                                            xml_doc = etree.parse(xml_path)
                                            root = xml_doc.getroot()

                                            # assigning footer location and footer page
                                            for elem in root.iter():
                                                if (
                                                        elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                                                    text_in_box = elem.text
                                                    if (text_in_box is not None):
                                                        if (footer_text in text_in_box):
                                                            footer_y1 = elem.attrib.get("y1")
                                                            bbox = elem.attrib.get("bbox")
                                                            footer_page = getpagno(bbox, root)
                                                            footer_page = int(footer_page) - 1
                                                            footer_y1 = str(int(float(footer_y1)))
                                                            print("footer_y1 - " + str(footer_y1))
                                                            print("footer_page - " + str(footer_page))
                                            if (footer_y1 == 0):
                                                footer_page = total_pages
                                                footer_y1 = 100

                                            # main execution
                                            try:
                                                values_extract_status = get_values_fax(pdf, timestamp,
                                                                                   output_folderpath, total_pages, txtfilepath)
                                                tables_extract_status = get_tables_fax(pdf, total_pages, timestamp,
                                                                                   footer_page, footer_y1, root,
                                                                                   header_text, file_path,
                                                                                   output_folderpath, header_text_2)

                                                if (values_extract_status != "failed"):
                                                    if (tables_extract_status != "failed"):

                                                        output_zip = zipfile.ZipFile(
                                                            output_folderpath + '_eyesight.zip', 'w')
                                                        for folder, subfolders, files in os.walk(output_folderpath):
                                                            for file in files:
                                                                output_zip.write(os.path.join(folder, file),
                                                                                 os.path.relpath(
                                                                                     os.path.join(folder, file),
                                                                                     output_folderpath),
                                                                                 compress_type=zipfile.ZIP_DEFLATED)

                                                        output_zip.close()

                                                        print("output saved successfully")

                                                        download_link = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/initialreceivepdf/alkabli/" + timestamp + "/outputs_eyesight.zip"

                                                        return sendresponse(download_link, 200)

                                                    else:
                                                        print("failed to extract table")
                                                        return sendresponse("failed to extract table", 201)
                                                else:
                                                    print("failed to extract values")
                                                    return sendresponse("failed to extract values", 201)

                                            except Exception as e:
                                                print(e)
                                                return sendresponse(e, 201)

                                            # return "all done well"



                                        else:
                                            print("failed in merge_pdfs")
                                            return "failed in merge_pdfs"
                                    else:
                                        print("failed in convert_imgs_into_pdf")
                                        return "failed in convert_imgs_into_pdf"
                                else:
                                    print("failed in get_txt_boxes")
                                    return "failed in get_txt_boxes"
                            else:
                                print("failed in remove_lines")
                                return "failed in remove_lines"
                        else:
                            print("failed in skew_correct")
                            return "failed in skew_correct"
                    else:
                        print("failed in binarize_imgs")
                        return "failed in binarize_imgs"
                else:
                    print("failed in convert_pdf_to_img")
                    return "failed in convert_pdf_to_img"
            else:
                # get total pages
                file_path = folderpath + "/" + filename
                doc = fitz.open(file_path)
                total_pages = doc.pageCount
                print("total_pages - " + str(total_pages))

                # making xml tree
                pdf = pdfquery.PDFQuery(file_path)
                pdf.load()
                pdf.tree.write(xml_path, pretty_print=True, encoding="utf-8")
                xml_doc = etree.parse(xml_path)
                root = xml_doc.getroot()

                # assigning footer location and footer page
                for elem in root.iter():
                    if (
                            elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
                        text_in_box = elem.text
                        if (text_in_box is not None):
                            if (footer_text in text_in_box):
                                footer_y1 = elem.attrib.get("y1")
                                bbox = elem.attrib.get("bbox")
                                footer_page = getpagno(bbox, root)
                                footer_page = int(footer_page) - 1
                                footer_y1 = str(int(float(footer_y1)))
                                print("footer_y1 - " + str(footer_y1))
                                print("footer_page - " + str(footer_page))
                if (footer_y1 == 0):
                    footer_page = total_pages
                    footer_y1 = 100

                # main execution
                try:
                    values_extract_status = get_values(pdf, timestamp,
                                                       output_folderpath)
                    tables_extract_status = get_tables(pdf, total_pages, timestamp,
                                                       footer_page, footer_y1, root,
                                                       header_text, file_path,
                                                       output_folderpath)

                    if (values_extract_status != "failed"):
                        if (tables_extract_status != "failed"):

                            output_zip = zipfile.ZipFile(
                                output_folderpath + '_eyesight.zip', 'w')
                            for folder, subfolders, files in os.walk(output_folderpath):
                                for file in files:
                                    output_zip.write(os.path.join(folder, file),
                                                     os.path.relpath(
                                                         os.path.join(folder, file),
                                                         output_folderpath),
                                                     compress_type=zipfile.ZIP_DEFLATED)

                            output_zip.close()

                            print("output saved successfully")

                            download_link = "https://aspace-web.eastus.cloudapp.azure.com/eyesight/static/initialreceivepdf/alkabli/" + timestamp + "/outputs_eyesight.zip"

                            return sendresponse(download_link, 200)

                        else:
                            print("failed to extract table")
                            return sendresponse("failed to extract table", 201)
                    else:
                        print("failed to extract values")
                        return sendresponse("failed to extract values", 201)

                except Exception as e:
                    print(e)
                    return sendresponse(e, 201)

    except Exception as e:
        print (e)
        return sendresponse("failed", 200)


def get_header_location(page, root, header_text):
    # print ("page - " + str(page))
    for elem in root.iter():
        if (elem.tag == "LTTextBoxHorizontal" or elem.tag == "LTTextLineHorizontal"):
            text_in_box = elem.text
            if(text_in_box is not None):
                if (header_text in text_in_box):
                    try:
                        header_y0 = elem.attrib.get("y0")
                        bbox = elem.attrib.get("bbox")
                        header_page = getpagno(bbox, root)
                        header_page = int(header_page)-1
                        header_y0 = str(int(float(header_y0)))
                        # print ("header_y0 - " + str(header_y0))
                        if (header_page == page):
                            return header_y0
                    except Exception as e:
                        print ("error from method")
                        print (e)


def get_values(pdf, ts, output_path):
    try:
        pdf.load(0)
        key_values = {}
        vendor_No = get_key_value("Vendor No", pdf)
        telephone = get_key_value("Telephone", pdf)
        fax = get_key_value("Fax Number", pdf)
        discount  = get_key_value("Vendor Discount", pdf)
        vat = get_key_value("VAT Number", pdf)
        email = get_key_value("E-mail Address: ", pdf)

        key_values["vendor_No"] = vendor_No
        key_values["telephone"] = telephone
        key_values["fax"] = fax
        key_values["discount"] = discount
        key_values["vat"] = vat
        key_values["email"] = email

        csv_file = "/values.csv"
        csv_path = output_path + csv_file
        with open(csv_path, 'w') as csvfile:
            for key in key_values.keys():
                csvfile.write("%s,%s\n" % (key, key_values[key]))
        return "success"
    except Exception as e:
        return "failed"

def get_values_fax(pdf, ts, output_path, total_pages, txtfilepath):
    try:
        # pdf.load(int(total_pages) - 1)
        pdf.load(0)
        key_values = {}
        order = get_key_value_fax("ORDER ", pdf)
        order_date = get_key_value_fax("ORDER DATE ", pdf)
        # department = get_key_value_fax("DEPARTMENT ", pdf)
        # section = get_key_value_fax("SECTION ", pdf)
        delivered_to = get_key_value_fax("DELIVERED TO ", pdf)
        # delivery_date = get_key_value_fax("DELIVERY DATE ", pdf)
        # deadline = get_key_value_fax("DEADLINE ", pdf)

        # pre-processing
        list = re.findall("\d+", order)
        order_list = []
        for i in range(len(list)):
            print ("len(list[i] - " + str(len(list[i])))
            if (len(list[i]) > 5):
                order_list.append(list[i])
        order_final = order_list[0]

        order_date_final = re.search("(\s)*\d{1,}/\d{1,}/\d{1,}(\s)*", order_date).group(0)
        order_date_final = order_date_final.strip()

        try:
            delivered_to_final = delivered_to.split(":", 1)[1]
            delivered_to_final = delivered_to_final.replace(">", "")
            delivered_to_final = delivered_to_final.replace("<", "")
            delivered_to_final = delivered_to_final.strip()
        except Exception as e:
            print (e)
            delivered_to_final = delivered_to.strip()

        order_type = ""
        txt_file = open(txtfilepath)
        strings = txt_file.read()
        if ("promotion" in strings):
            order_type = "promo"
        else:
            order_type = "regular"




        key_values["order"] = order_final
        key_values["order_date"] = order_date_final
        # key_values["department"] = department
        # key_values["section"] = section
        key_values["delivered_to"] = delivered_to_final
        key_values["order_type"] = order_type
        # key_values["delivery_date"] = delivery_date
        # key_values["deadline"] = deadline

        csv_file = "/values.csv"
        csv_path = output_path + csv_file
        with open(csv_path, 'w') as csvfile:
            for key in key_values.keys():
                csvfile.write("%s,%s\n" % (key, key_values[key]))
        return "success"
    except Exception as e:
        print (e)
        return "failed"


def get_key_value(keyword, pdf):
    try:
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        txt = label.text()
        print(txt)
        x0 = float(label.attr('x0'))
        x1 = float(label.attr('x1'))
        y0 = float(label.attr('y0'))
        y1 = float(label.attr('y1'))
        foundtext = pdf.pq('LTTextLineHorizontal:in_bbox("%s, %s, %s, %s")' % (
            x1, y0 - 30, x1 + 5000, y1 + 30)).text()
        foundtext = foundtext.replace(keyword, "")
        foundtext = foundtext.replace(":", "")
        foundtext = foundtext.strip()
        if (foundtext == ""):
            foundtext = txt.replace(keyword, "")
            foundtext = foundtext.replace("—", "")
            foundtext = foundtext.strip()
    except Exception as e:
        foundtext = "-"
    #         print (e)

    return foundtext

def get_key_value_fax(keyword, pdf):
    try:
        parameter = 'LTTextLineHorizontal:contains("' + keyword + '")'
        label = pdf.pq(parameter)
        txt = label.text()
        #         print (txt)
        x0 = float(label.attr('x0'))
        x1 = float(label.attr('x1'))
        y0 = float(label.attr('y0'))
        y1 = float(label.attr('y1'))
        foundtext = pdf.pq('LTTextLineHorizontal:in_bbox("%s, %s, %s, %s")' % (
            x1, y0 - 10, x1 + 1000, y1 + 10)).text()
        foundtext = foundtext.replace(keyword, "")
        foundtext = foundtext.replace(":", "")
        foundtext = foundtext.strip()
        if (foundtext == ""):
            foundtext = txt.replace(keyword, "")
            foundtext = foundtext.replace("—", "")
            foundtext = foundtext.strip()
        print(foundtext)
    except Exception as e:
        foundtext = "-"
    #         print (e)

    return foundtext



def get_tables(pdf, total_pages, ts, footer_page, footer_y1, root, header_text, file_path, output_path):
    try:
        df_list = []
        footer_keyword = "Total number"

        print ("footer_page - " + str(footer_page))
        print ("footer_y1 - " + str(footer_y1))
        footer_page = int(footer_page)
        # footer_y1 = int(footer_y1)

        for i in range(total_pages):

            print ("loop - " + str(i))

            if (i == 0):
                header_y1 = get_header_location(i, root, header_text)
                if(header_y1 == None):
                    header_y1 = 1484
                print ("header_y1 - " + str(header_y1))
                # table_location = '109,' + str(int(header_y1)-80) +',1647,180'
                table_location = '0,' + str(int(header_y1) - 80) + ',1647,180'
                print ("table_location - " + str(table_location))
                tableregionlist = [(table_location)]
                group_into_row = 16
                columns_list = ['200, 355, 525, 1150, 1250, 1450']
                tables = camelot.read_pdf(file_path, strip='\n', flavor='stream', pages = "1", row_tol=group_into_row, table_areas=tableregionlist, columns=columns_list)
                tables_len = len(tables)
                print ("no of tables detected in page " + str(i+1) + " - " + str(tables_len))
                if (tables_len != 0):
                    print("here")
                    df = tables[0].df
                    df = df.replace('\n',' ', regex=True)
                    df_list.append(df)
                    print("here1")


            if ((i != 0) & (i < footer_page)):
                print("here2")
                print ("page - " + str(i+1))
                header_y1 = get_header_location(i, root, header_text)
                if(header_y1 == None):
                    header_y1 = 1817
                print("header_y1 - " + str(header_y1))
                # table_location = '109,' + str(header_y1) + ',1647,180'
                table_location = '0,' + str(header_y1) + ',1647,180'
                print ("table_location - " + str(table_location))
                tableregionlist = [(table_location)]
                group_into_row = 16
                page = str(i+1)
                tables = camelot.read_pdf(file_path, strip='\n', flavor='stream', pages = page, row_tol=group_into_row, table_areas=tableregionlist, columns=columns_list)
                tables_len = len(tables)
                print ("no of tables detected in page " + str(i+1) + " " + str(tables_len))
                if (tables_len != 0):
                    print ("here3")
                    df = tables[0].df
                    df = df.replace('\n', ' ', regex=True)
                    df_list.append(df)
                    print("here4")


            if (i == footer_page):
                header_y1 = get_header_location(i, root, header_text)
                if(header_y1 == None):
                    header_y1 = 1817
                print("header_y1 - " + str(header_y1))
                # table_location = '109,' + str(header_y1) + ',1647,' + footer_y1
                table_location = '0,' + str(header_y1) + ',1647,' + footer_y1
                print ("table_location - " + str(table_location))
                tableregionlist = [(table_location)]
                group_into_row = 16
                page = str(i+1)
                tables = camelot.read_pdf(file_path, strip='\n', flavor='stream', pages = page, row_tol=group_into_row, table_areas=tableregionlist, columns=columns_list)
                tables_len = len(tables)
                print ("no of tables detected in page " + str(i+1) + " " + str(tables_len))
                if (tables_len != 0):
                    df = tables[0].df
                    df = df.replace('\n',' ', regex=True)
                    df_list.append(df)

        result_df = pd.concat(df_list)
        result_df
        result_df.to_csv(output_path + '/table.csv')
        return "success"
    except Exception as e:
        print (e)
        return "failed"

def get_tables_fax(pdf, total_pages, ts, footer_page, footer_y1, root, header_text, file_path, output_path, header_text_2):
    try:
        df_list = []
        footer_keyword = "Total number"


        print ("footer_y1 - " + str(footer_y1))
        footer_page = int(footer_page)
        print("footer_page - " + str(footer_page))
        # footer_y1 = int(footer_y1)

        for i in range(total_pages):

            if (i == 0):
                print("i - " + str(i))
                header_y1 = get_header_location(0, root, header_text)
                # print ("header_y1 - " + str(header_y1))
                if (header_y1 == None):
                    header_y1 = get_header_location(0, root, header_text_2)
                    # print("header_y1 - " + str(header_y1))
                if (header_y1 == None):
                    header_y1 = 1557
                print("header_y1 - " + str(header_y1))
                table_location = '109,' + str(int(header_y1)) +',5000,180'
                print ("table_location - " + str(table_location))
                tableregionlist = [(table_location)]
                # columns_list = ['320, 570, 670, 1270, 1420, 1630, 1855, 2055']
                columns_list = ['570, 1000, 1180, 2250, 2500, 2880, 3285, 3625']
                group_into_row = 16
                tables = camelot.read_pdf(file_path, strip='\n', flavor='stream', pages = "1", row_tol=group_into_row, table_areas=tableregionlist, columns=columns_list)
                tables_len = len(tables)
                print ("no of tables detected in page " + str(i+1) + " " + str(tables_len))
                if (tables_len != 0):
                    df = tables[0].df
                    print (df.size)
                    rows_to_drop = []
                    col_len = len(df[0])
                    for i in range(col_len):
                        cell_text = df[0][i]
                        if ((cell_text == "") or (cell_text.upper().isupper())):
                            rows_to_drop.append(i)
                    df_cleaned = df.drop(labels=rows_to_drop,axis=0,)
                    df_cleaned = df_cleaned.replace('\n',' ', regex=True)
                    for i in range(len(df_cleaned.columns)):
                        df_cleaned[i] = df_cleaned[i].str.replace('|', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(']', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(':', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(')', '')
                        df_cleaned[i] = df_cleaned[i].str.replace('(', '')
                    df_cleaned = df_cleaned.replace('}', '', regex=True)
                    df_list.append(df_cleaned)
                    # df_list.append(df)


            if ((i != 0) & (i < footer_page)):
                print("i - " + str(i))
                header_y1 = get_header_location(i, root, header_text)
                # print("header_y1 - " + str(header_y1))
                if (str(header_y1) == "None"):
                    header_y1 = get_header_location(i, root, header_text_2)
                if (str(header_y1) == "None"):
                    header_y1 = 1697
                print("header_y1_mid_page - " + str(header_y1))
                table_location = '109,' + str(int(header_y1)) +',5000,180'
                print ("table_location - " + str(table_location))
                tableregionlist = [(table_location)]
                # columns_list = ['320, 570, 670, 1270, 1420, 1630, 1855, 2055']
                columns_list = ['570, 1000, 1180, 2250, 2500, 2880, 3285, 3625']
                group_into_row = 16
                page = str(i+1)
                tables = camelot.read_pdf(file_path, strip='\n', flavor='stream', pages = page, row_tol=group_into_row, table_areas=tableregionlist, columns=columns_list)
                tables_len = len(tables)
                print ("no of tables detected in page " + str(i+1) + " " + str(tables_len))
                if (tables_len != 0):
                    df = tables[0].df
                    print(df.size)
                    rows_to_drop = []
                    col_len = len(df[0])
                    for i in range(col_len):
                        cell_text = df[0][i]
                        if ((cell_text == "") or (cell_text.upper().isupper())):
                            rows_to_drop.append(i)
                    df_cleaned = df.drop(labels=rows_to_drop,axis=0,)
                    df_cleaned = df_cleaned.replace('\n',' ', regex=True)
                    for i in range(len(df_cleaned.columns)):
                        df_cleaned[i] = df_cleaned[i].str.replace('|', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(']', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(':', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(')', '')
                        df_cleaned[i] = df_cleaned[i].str.replace('(', '')
                    df_cleaned = df_cleaned.replace('}', '', regex=True)
                    df_list.append(df_cleaned)


            if (i == footer_page):
                header_y1 = get_header_location(i, root, header_text)
                if (header_y1 == None):
                    header_y1 = get_header_location(i, root, header_text_2)
                if (header_y1 == None):
                    header_y1 = 1697
                # print ("i - " + str(i))
                print ("header_y1_last_page - " + str(header_y1))
                print ("footer_y1 - " + str(footer_y1))
                table_location = '109,' + str(int(header_y1)) + ',5000,' + str(int(footer_y1)+100)
                print ("table_location - " + str(table_location))
                tableregionlist = [(table_location)]
                # columns_list = ['320, 570, 670, 1270, 1420, 1630, 1855, 2055']
                columns_list = ['570, 1000, 1180, 2250, 2500, 2880, 3285, 3625']
                group_into_row = 16
                page = str(i+1)
                tables = camelot.read_pdf(file_path, strip='\n', flavor='stream', pages = page, row_tol=group_into_row, table_areas=tableregionlist, columns=columns_list)
                tables_len = len(tables)
                print ("no of tables detected in page " + str(i+1) + " " + str(tables_len))
                if (tables_len != 0):
                    df = tables[0].df
                    print(df.size)
                    rows_to_drop = []
                    col_len = len(df[0])
                    for i in range(col_len):
                        cell_text = df[0][i]
                        if ((cell_text == "") or (cell_text.upper().isupper())):
                            rows_to_drop.append(i)
                    df_cleaned = df.drop(labels=rows_to_drop,axis=0,)
                    df_cleaned = df_cleaned.replace('\n',' ', regex=True)
                    for i in range(len(df_cleaned.columns)):
                        df_cleaned[i] = df_cleaned[i].str.replace('|', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(']', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(':', '')
                        df_cleaned[i] = df_cleaned[i].str.replace(')', '')
                        df_cleaned[i] = df_cleaned[i].str.replace('(', '')
                    df_cleaned = df_cleaned.replace('}', '', regex=True)
                    df_list.append(df_cleaned)

        result_df = pd.concat(df_list)
        result_df[1] = result_df[1].str.replace('o', '0')
        result_df[1] = result_df[1].str.replace('O', '0')
        print (result_df[1])

        # rows_to_drop_from_final = []
        # col_len_final = len(result_df[3])
        # for i in range(col_len_final):
        #     cell_text_final = str(result_df[2][i])
        #     if (cell_text_final.upper().isupper()):
        #         rows_to_drop_from_final.append(i)
        # result_df_final = result_df.drop(labels=rows_to_drop_from_final, axis=0, )

        result_df.to_csv(output_path + '/table.csv')
        return "success"
    except Exception as e:
        print (e)
        return "failed"


def zipdir(path, ziph):
    # ziph is zipfile handle
    for root, dirs, files in os.walk(path):
        for file in files:
            ziph.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), os.path.join(path, '..')))


def binarize_imgs_alkabli(imgspath):
    try:
        threads = []
        for filename in os.listdir(imgspath):
            if filename.endswith(".jpg") or filename.endswith(".png"):
                process = Thread(target=threded_binarize_img_alkabli, args=(imgspath, filename,))
                time.sleep(1)
                process.start()
                threads.append(process)
            else:
                continue
        for process in threads:
            process.join()
        return "done"
    except Exception as e:
        print (e)
        return e


def threded_binarize_img_alkabli(imgspath, filename):
    # print("reached here 0")
    fullpath = os.path.join(imgspath, filename)
    initial_image = cv2.imread(fullpath)

    # kernel = np.ones((5, 5), np.uint8)
    # erosion = cv2.erode(initial_image, kernel, iterations=1)
    # dilation = cv2.dilate(erosion, kernel, iterations=1)
    # image = Image.fromarray(dilation)

    # image = remove_transparency(image, bg_colour=(255, 255, 255))
    ret, image = cv2.threshold(initial_image, 125, 255, cv2.THRESH_BINARY)
    image = Image.fromarray(image)
    # image = binarization.nlbin(image, threshold = 0.5)
    image.save(fullpath)
#     display(image)


def convert_pdf_to_img_alkabli(pdfpath, savepath):
    try:
        print ("pdfpath - " + pdfpath)
        pages = pdf2image.convert_from_path(pdf_path = pdfpath, grayscale=True, size=(3000, None))
        global noofpagesinpdf
        noofpagesinpdf = len(pages)
        print ("no of pages " + str(noofpagesinpdf))
        for i in range(len(pages)):
            pages[i].save(savepath + "output" + str(i) + '.jpg')
        return "done"
    except Exception as e:
        print (e)
        return e






# QR code solutions

@app.route('/extractqr', methods = ['POST'])
def extractqr():
    file = request.files['file']
    if 'file' not in request.files:
        print('No file part')
        return sendresponse("file not received", 201)
    else:
        timestamp = time.time()
        timestamp = str(timestamp).split('.')
        timestamp = str(timestamp[0])
        print(timestamp)

        folderpath = root_path + 'static/qr/'
        foldername = timestamp
        filename = file.filename

        folderpath = os.path.join(folderpath, foldername)
        print(folderpath)
        try:
            os.mkdir(folderpath)
            os.mkdir(folderpath + "/snapshots/")
        except Exception as e:
            print("create folder")
            print(e)

        file.save(folderpath + "/" + filename)
        print("file received")

        if filename.endswith(".pdf"):
            pages = pdf2image.convert_from_path(pdf_path=folderpath + "/" + filename, grayscale=True)
            for i in range(len(pages)):
                pages[i].save(folderpath + "/snapshots/" + "output" + str(i) + '.jpg')

            final_dict = {}

            count = 1
            for filename in os.listdir(folderpath + "/snapshots/"):
                image = cv2.imread(folderpath + "/snapshots/" + filename)
                barcodes = pyzbar.decode(image)
                # print ("No of QR codes detected in page " + str(count) + " is " + str(len(barcodes)))
                code_list = []
                for barcode in barcodes:
                    (x, y, w, h) = barcode.rect
                    cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    barcodeData = barcode.data.decode("utf-8")
                    barcodeType = barcode.type
                    text = "{} ({})".format(barcodeData, barcodeType)
                    cv2.putText(image, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                    # print("[INFO] Found {} barcode: {}".format(barcodeType, barcodeData))
                    code_list.append(barcodeData)
                final_dict["page " + str(count)] = code_list
                count = count + 1
            print(final_dict)
            return (final_dict)
        else:
            count = 1
            final_dict = {}
            # print (folderpath + filename)
            image = cv2.imread(folderpath + "/" + filename)
            barcodes = pyzbar.decode(image)
            # print ("No of QR codes detected in page " + str(count) + " is " + str(len(barcodes)))
            code_list = []
            for barcode in barcodes:
                (x, y, w, h) = barcode.rect
                cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 255), 2)
                barcodeData = barcode.data.decode("utf-8")
                barcodeType = barcode.type
                text = "{} ({})".format(barcodeData, barcodeType)
                cv2.putText(image, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                # print("[INFO] Found {} barcode: {}".format(barcodeType, barcodeData))
                code_list.append(barcodeData)
            final_dict["page " + str(count)] = code_list
            count = count + 1
            print(final_dict)
            return (final_dict)















@app.errorhandler(400)
def bad_request(error=None):
    message = {
        'status': 400,
        'message': 'Bad Request: ' + request.url + '--> Please check your data payload...'
    }
    resp = jsonify(message)
    resp.status_code = 400

    return resp


if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True, port=5004, threaded=True)




